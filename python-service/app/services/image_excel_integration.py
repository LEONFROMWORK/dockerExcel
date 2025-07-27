"""
Image to Excel Integration Service
이미지를 Excel로 변환하고 오류 감지 시스템과 통합
"""

from typing import Dict, Any, List, Optional, Tuple
import base64
import io
import os
import re
import unicodedata
import numpy as np
from PIL import Image
import pandas as pd
import openpyxl
from datetime import datetime
import logging
import asyncio
from dataclasses import dataclass

# 조건부 imports
try:
    from scipy import ndimage
except ImportError:
    ndimage = None
    
try:
    from dateutil import parser as date_parser
except ImportError:
    date_parser = None

from app.services.multilingual_two_tier_service import MultilingualTwoTierService
from app.services.detection.integrated_error_detector import IntegratedErrorDetector
from app.services.openai_service import OpenAIService
from app.core.exceptions import WorkbookLoadError

logger = logging.getLogger(__name__)

@dataclass
class ImageAnalysisResult:
    """이미지 분석 결과"""
    image_type: str  # 'table', 'chart', 'form', 'mixed'
    confidence: float
    extracted_data: List[List[Any]]
    headers: Optional[List[str]]
    metadata: Dict[str, Any]
    ocr_method: str  # 'tier2', 'tier3', 'vision'
    
@dataclass
class ExcelConversionResult:
    """Excel 변환 결과"""
    file_path: str
    sheet_count: int
    total_cells: int
    error_count: int
    warnings: List[str]

class ImageExcelIntegrationService:
    """이미지를 Excel로 변환하고 오류를 감지하는 통합 서비스"""
    
    def __init__(self):
        self.ocr_service = MultilingualTwoTierService()
        self.openai_service = OpenAIService()
        self.error_detector = IntegratedErrorDetector()
        
    async def process_image_with_error_detection(
        self, 
        image_path: str,
        output_format: str = "xlsx",
        detect_errors: bool = True
    ) -> Dict[str, Any]:
        """이미지를 처리하고 Excel로 변환 후 오류 감지"""
        
        try:
            # 1단계: 이미지 분석
            logger.info(f"이미지 분석 시작: {image_path}")
            analysis_result = await self._analyze_image_content(image_path)
            
            # 2단계: Excel 파일 생성
            logger.info("Excel 파일 생성 중...")
            excel_result = await self._create_excel_from_analysis(
                analysis_result,
                output_format
            )
            
            # 3단계: 오류 감지 (선택적)
            error_detection_result = None
            if detect_errors and excel_result.file_path:
                logger.info("Excel 오류 감지 중...")
                error_detection_result = await self.error_detector.detect_all_errors(
                    excel_result.file_path
                )
            
            # 4단계: 결과 통합
            return {
                "status": "success",
                "image_analysis": {
                    "type": analysis_result.image_type,
                    "confidence": analysis_result.confidence,
                    "ocr_method": analysis_result.ocr_method,
                    "metadata": analysis_result.metadata
                },
                "excel_conversion": {
                    "file_path": excel_result.file_path,
                    "sheet_count": excel_result.sheet_count,
                    "total_cells": excel_result.total_cells,
                    "warnings": excel_result.warnings
                },
                "error_detection": error_detection_result,
                "processing_time": analysis_result.metadata.get("processing_time", 0)
            }
            
        except FileNotFoundError as e:
            logger.error(f"파일을 찾을 수 없음: {image_path}")
            return {
                "status": "error",
                "error": f"파일을 찾을 수 없습니다: {image_path}",
                "details": {
                    "image_path": image_path,
                    "detect_errors": detect_errors
                }
            }
        except PermissionError as e:
            logger.error(f"파일 접근 권한 없음: {image_path}")
            return {
                "status": "error",
                "error": f"파일 접근 권한이 없습니다: {image_path}",
                "details": {
                    "image_path": image_path,
                    "detect_errors": detect_errors
                }
            }
        except Exception as e:
            logger.error(f"이미지 처리 오류: {str(e)}", exc_info=True)
            return {
                "status": "error",
                "error": str(e),
                "details": {
                    "image_path": image_path,
                    "timestamp": datetime.now().isoformat()
                }
            }
    
    async def _analyze_image_content(self, image_path: str) -> ImageAnalysisResult:
        """이미지 내용 분석 - Vision API 직접 사용"""
        start_time = datetime.now()
        
        # 이미지 로드 및 전처리
        image = Image.open(image_path)
        image_array = np.array(image)
        
        # 이미지 타입 판별
        image_type = await self._detect_image_type(image_array)
        
        # Vision API로 데이터 추출 (모든 이미지 타입에 대해)
        logger.info("Vision API를 사용한 이미지 분석")
        vision_result = await self._analyze_with_vision_api(image_path)
        
        extracted_data = vision_result.get("data", [])
        headers = vision_result.get("headers")
        ocr_method = vision_result.get("method", "vision_api")
        confidence = vision_result.get("confidence", 0.95)
        additional_info = vision_result.get("additional_info", {})
        
        # 결과 로깅
        if extracted_data:
            logger.info(f"Vision API 분석 성공: {len(extracted_data)} 행 추출")
        else:
            logger.warning("Vision API 분석 결과가 비어있음")
        
        processing_time = (datetime.now() - start_time).total_seconds()
        
        return ImageAnalysisResult(
            image_type=image_type,
            confidence=confidence,
            extracted_data=extracted_data,
            headers=headers,
            metadata={
                "processing_time": processing_time,
                "image_size": image.size,
                "image_mode": image.mode,
                "additional_info": additional_info
            },
            ocr_method=ocr_method
        )
    
    async def _detect_image_type(self, image_array: np.ndarray) -> str:
        """이미지 타입 감지 (테이블, 차트, 폼 등)"""
        # 간단한 휴리스틱 기반 감지
        # 실제로는 더 정교한 컴퓨터 비전 기법 사용 가능
        
        # 엣지 검출로 선 구조 파악
        if ndimage is None:
            # scipy가 없으면 간단한 방법 사용
            return "table"  # 기본값
        
        edges = ndimage.sobel(image_array.mean(axis=2) if len(image_array.shape) == 3 else image_array)
        
        # 수평/수직 선의 비율로 테이블 감지
        horizontal_lines = np.sum(np.abs(edges[1:] - edges[:-1]) > 50, axis=1)
        vertical_lines = np.sum(np.abs(edges[:, 1:] - edges[:, :-1]) > 50, axis=0)
        
        if len(horizontal_lines) > 3 and len(vertical_lines) > 2:
            return "table"
        elif np.max(edges) > 100:  # 강한 엣지가 있으면 차트일 가능성
            return "chart"
        else:
            return "form"
    
    
    async def _analyze_with_vision_api(self, image_path: str) -> Dict[str, Any]:
        """Vision API를 사용한 고급 분석"""
        with open(image_path, 'rb') as f:
            image_content = base64.b64encode(f.read()).decode('utf-8')
        
        prompt = """
        이 이미지를 분석하여 Excel 스프레드시트로 변환할 수 있는 데이터를 추출해주세요.
        
        중요: 병합된 셀 감지
        - 여러 열에 걸쳐 있는 헤더가 있다면 병합된 셀입니다
        - 같은 카테고리가 여러 행에 걸쳐 있고 일부 셀이 비어있다면 병합된 셀입니다
        - 예: "매출원가"가 첫 번째 행에만 있고 다음 행들이 비어있다면 A2:A4 같은 형태로 병합된 것입니다
        
        다음 형식으로 응답해주세요:
        {
            "type": "table|chart|form",
            "headers": ["열1", "열2", ...],
            "data": [
                ["값1", "값2", ...],
                ...
            ],
            "merged_cells": [
                {"range": "A1:C1", "value": "병합된 셀 내용"},
                {"range": "A2:A5", "value": "카테고리명"}
            ],
            "formulas": ["수식이 있다면 여기에"],
            "formatting": {
                "bold_cells": ["A1", "B1"],
                "colored_cells": {"A2": "#FF0000"}
            },
            "chart_info": {차트 정보가 있다면}
        }
        
        데이터가 숫자인 경우 문자열이 아닌 숫자로 표현해주세요.
        병합된 셀이 있다면 반드시 merged_cells에 포함해주세요.
        비어있는 셀도 data 배열에 빈 문자열("")로 포함시켜주세요.
        """
        
        try:
            logger.info("Vision API 호출 시작")
            response = await self.openai_service.analyze_image_structured(
                image_content,
                prompt
            )
            
            logger.info(f"Vision API 응답: {response}")
            
            if response and isinstance(response, dict):
                data = response.get("data", [])
                headers = response.get("headers")
                
                logger.info(f"추출된 데이터: {len(data)} 행, 헤더: {headers}")
                
                return {
                    "data": data,
                    "headers": headers,
                    "method": "vision_api",
                    "confidence": 0.95,
                    "additional_info": {
                        "type": response.get("type"),
                        "formulas": response.get("formulas", []),
                        "chart_info": response.get("chart_info"),
                        "merged_cells": response.get("merged_cells", []),
                        "formatting": response.get("formatting", {})
                    }
                }
        except Exception as e:
            logger.error(f"Vision API 오류: {str(e)}", exc_info=True)
        
        return {
            "data": [],
            "headers": None,
            "method": "vision_api_failed",
            "confidence": 0.0
        }
    
    async def _create_excel_from_analysis(
        self,
        analysis: ImageAnalysisResult,
        output_format: str
    ) -> ExcelConversionResult:
        """분석 결과를 Excel 파일로 변환"""
        
        # 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Imported_{analysis.image_type.title()}"
        
        warnings = []
        total_cells = 0
        
        # 헤더 추가
        if analysis.headers:
            for col, header in enumerate(analysis.headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(
                    start_color="E0E0E0",
                    end_color="E0E0E0",
                    fill_type="solid"
                )
            total_cells += len(analysis.headers)
        
        # 데이터 추가
        start_row = 2 if analysis.headers else 1
        for row_idx, row_data in enumerate(analysis.extracted_data, start_row):
            for col_idx, value in enumerate(row_data, 1):
                # 데이터 타입 자동 변환
                converted_value = self._convert_cell_value(value)
                ws.cell(row=row_idx, column=col_idx, value=converted_value)
                total_cells += 1
        
        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 차트 정보가 있으면 차트 시트 추가
        if hasattr(analysis, 'metadata') and 'chart_info' in analysis.metadata:
            self._add_chart_sheet(wb, analysis.metadata['chart_info'])
        
        # 파일 저장 (보안 강화)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # 파일명 sanitize (path traversal 방지)
        safe_type = re.sub(r'[^a-zA-Z0-9_-]', '', analysis.image_type)
        safe_format = re.sub(r'[^a-zA-Z0-9]', '', output_format)
        filename = f"converted_{safe_type}_{timestamp}.{safe_format}"
        
        # 안전한 경로 생성
        upload_dir = os.path.abspath("uploads")
        file_path = os.path.join(upload_dir, filename)
        
        # 경로가 upload_dir 내부인지 확인
        if not os.path.abspath(file_path).startswith(upload_dir):
            raise ValueError("Invalid file path")
        
        os.makedirs(upload_dir, exist_ok=True, mode=0o755)
        wb.save(file_path)
        
        # 경고 확인
        if total_cells == 0:
            warnings.append("데이터를 추출할 수 없었습니다")
        if analysis.confidence < 0.7:
            warnings.append(f"낮은 신뢰도: {analysis.confidence:.2%}")
        
        return ExcelConversionResult(
            file_path=file_path,
            sheet_count=len(wb.sheetnames),
            total_cells=total_cells,
            error_count=0,  # 생성 단계에서는 오류 없음
            warnings=warnings
        )
    
    def _looks_like_header(self, cells: List[str]) -> bool:
        """헤더 행인지 판단"""
        # 모든 셀이 텍스트이고 숫자가 아닌 경우
        for cell in cells:
            try:
                float(cell)
                return False
            except ValueError:
                continue
        return True
    
    def _convert_cell_value(self, value: Any) -> Any:
        """셀 값을 적절한 타입으로 변환"""
        if isinstance(value, (int, float)):
            return value
        
        if isinstance(value, str):
            # 빈 문자열
            if not value.strip():
                return None
            
            # 숫자 변환 시도
            try:
                if '.' in value:
                    return float(value)
                else:
                    return int(value)
            except ValueError:
                pass
            
            # 날짜 변환 시도
            try:
                if date_parser:
                    return date_parser.parse(value)
                else:
                    # 기본 날짜 파싱 시도
                    for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"]:
                        try:
                            return datetime.strptime(value, fmt)
                        except ValueError:
                            continue
            except (ValueError, AttributeError):
                pass
            
            # 수식인지 확인
            if value.startswith('='):
                return value
            
            # 백분율 확인
            if value.endswith('%'):
                try:
                    return float(value[:-1]) / 100
                except ValueError:
                    pass
        
        return value
    
    def _add_chart_sheet(self, workbook: openpyxl.Workbook, chart_info: Dict[str, Any]):
        """차트 정보를 별도 시트에 추가"""
        ws = workbook.create_sheet("Chart_Data")
        
        # 차트 데이터 추가
        if "series" in chart_info:
            for i, series in enumerate(chart_info["series"]):
                ws.cell(row=1, column=i+1, value=series.get("name", f"Series{i+1}"))
                for j, value in enumerate(series.get("data", [])):
                    ws.cell(row=j+2, column=i+1, value=value)
    
    async def batch_process_images(
        self,
        image_paths: List[str],
        merge_strategy: str = "separate_sheets",
        detect_errors: bool = True
    ) -> Dict[str, Any]:
        """여러 이미지를 일괄 처리"""
        results = []
        
        # 병렬 처리
        tasks = [
            self.process_image_with_error_detection(path, detect_errors=False)
            for path in image_paths
        ]
        
        processing_results = await asyncio.gather(*tasks, return_exceptions=True)
        
        # 결과 수집
        successful_conversions = []
        failed_conversions = []
        
        for i, result in enumerate(processing_results):
            if isinstance(result, Exception):
                failed_conversions.append({
                    "image": image_paths[i],
                    "error": str(result)
                })
            elif result["status"] == "success":
                successful_conversions.append(result)
            else:
                failed_conversions.append({
                    "image": image_paths[i],
                    "error": result.get("error", "Unknown error")
                })
        
        # 병합 전략에 따라 Excel 파일 생성
        if successful_conversions and merge_strategy != "separate_files":
            merged_file = await self._merge_excel_files(
                successful_conversions,
                merge_strategy
            )
            
            # 병합된 파일에 대해 오류 감지
            if detect_errors and merged_file:
                error_detection = await self.error_detector.detect_all_errors(
                    merged_file
                )
            else:
                error_detection = None
        else:
            merged_file = None
            error_detection = None
        
        return {
            "status": "completed",
            "total_images": len(image_paths),
            "successful": len(successful_conversions),
            "failed": len(failed_conversions),
            "merge_strategy": merge_strategy,
            "merged_file": merged_file,
            "error_detection": error_detection,
            "individual_results": processing_results,
            "failures": failed_conversions
        }
    
    async def _merge_excel_files(
        self,
        conversions: List[Dict[str, Any]],
        strategy: str
    ) -> Optional[str]:
        """여러 Excel 파일을 병합"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 기본 시트 제거
        
        if strategy == "separate_sheets":
            # 각 변환을 별도 시트로
            for i, conv in enumerate(conversions):
                file_path = conv["excel_conversion"]["file_path"]
                source_wb = openpyxl.load_workbook(file_path)
                
                for sheet_name in source_wb.sheetnames:
                    source_sheet = source_wb[sheet_name]
                    new_sheet_name = f"Image{i+1}_{sheet_name}"
                    new_sheet = wb.create_sheet(new_sheet_name)
                    
                    # 데이터 복사
                    for row in source_sheet.iter_rows():
                        for cell in row:
                            new_sheet[cell.coordinate].value = cell.value
                            if cell.has_style:
                                new_sheet[cell.coordinate].font = cell.font
                                new_sheet[cell.coordinate].fill = cell.fill
        
        elif strategy == "single_sheet":
            # 모든 데이터를 하나의 시트에
            ws = wb.create_sheet("Combined_Data")
            current_row = 1
            
            for i, conv in enumerate(conversions):
                file_path = conv["excel_conversion"]["file_path"]
                source_wb = openpyxl.load_workbook(file_path)
                
                # 첫 번째 시트의 데이터만 가져오기
                source_sheet = source_wb.active
                
                # 소스 정보 추가
                ws.cell(row=current_row, column=1, value=f"Source: Image {i+1}")
                current_row += 1
                
                # 데이터 복사
                for row in source_sheet.iter_rows():
                    for col_idx, cell in enumerate(row, 1):
                        ws.cell(row=current_row, column=col_idx, value=cell.value)
                    current_row += 1
                
                # 구분을 위한 빈 행
                current_row += 1
        
        # 병합된 파일 저장
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        merged_filename = f"merged_{strategy}_{timestamp}.xlsx"
        merged_path = os.path.join("uploads", merged_filename)
        
        wb.save(merged_path)
        return merged_path