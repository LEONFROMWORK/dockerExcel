#!/usr/bin/env python3
"""Test script to analyze Excel file structure for rendering issues"""

import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
import json
from pprint import pprint

def analyze_excel_structure(file_path):
    """Analyze Excel file structure including merged cells, styles, and dimensions"""
    
    wb = openpyxl.load_workbook(file_path, data_only=False)
    analysis = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info = {
            "merged_cells": [],
            "column_widths": {},
            "row_heights": {},
            "cell_styles": {},
            "borders": {},
            "max_row": ws.max_row,
            "max_column": ws.max_column
        }
        
        # Analyze merged cells
        for merged_range in ws.merged_cells.ranges:
            sheet_info["merged_cells"].append({
                "range": str(merged_range),
                "start": merged_range.min_col,
                "end": merged_range.max_col,
                "start_row": merged_range.min_row,
                "end_row": merged_range.max_row
            })
        
        # Analyze column widths
        for col in ws.column_dimensions:
            if ws.column_dimensions[col].width:
                sheet_info["column_widths"][col] = ws.column_dimensions[col].width
        
        # Analyze row heights
        for row in ws.row_dimensions:
            if ws.row_dimensions[row].height:
                sheet_info["row_heights"][row] = ws.row_dimensions[row].height
        
        # Analyze cell styles (sample first 30 rows)
        for row in range(1, min(31, ws.max_row + 1)):
            for col in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                cell_coord = cell.coordinate
                
                if cell.value is not None or cell.fill.start_color.index not in [None, '00000000'] or cell.border.left.style:
                    style_info = {
                        "value": str(cell.value) if cell.value else None,
                        "fill": None,
                        "font": None,
                        "border": None,
                        "alignment": None
                    }
                    
                    # Fill color
                    if cell.fill.start_color.index not in [None, '00000000']:
                        style_info["fill"] = {
                            "color": cell.fill.start_color.index,
                            "type": cell.fill.fill_type
                        }
                    
                    # Font
                    if cell.font:
                        style_info["font"] = {
                            "bold": cell.font.bold,
                            "color": cell.font.color.index if cell.font.color else None,
                            "size": cell.font.size
                        }
                    
                    # Border
                    if any([cell.border.left.style, cell.border.right.style, 
                           cell.border.top.style, cell.border.bottom.style]):
                        style_info["border"] = {
                            "left": {"style": cell.border.left.style, 
                                   "color": cell.border.left.color.index if cell.border.left.color else None},
                            "right": {"style": cell.border.right.style,
                                    "color": cell.border.right.color.index if cell.border.right.color else None},
                            "top": {"style": cell.border.top.style,
                                  "color": cell.border.top.color.index if cell.border.top.color else None},
                            "bottom": {"style": cell.border.bottom.style,
                                     "color": cell.border.bottom.color.index if cell.border.bottom.color else None}
                        }
                    
                    # Alignment
                    if cell.alignment:
                        style_info["alignment"] = {
                            "horizontal": cell.alignment.horizontal,
                            "vertical": cell.alignment.vertical,
                            "wrap_text": cell.alignment.wrap_text
                        }
                    
                    sheet_info["cell_styles"][cell_coord] = style_info
        
        analysis[sheet_name] = sheet_info
    
    return analysis

if __name__ == "__main__":
    file_path = "/Users/kevin/Downloads/(공표기준)결산보고_25.2Q.xlsx"
    analysis = analyze_excel_structure(file_path)
    
    # Print summary
    for sheet_name, info in analysis.items():
        print(f"\n=== Sheet: {sheet_name} ===")
        print(f"Dimensions: {info['max_row']} rows x {info['max_column']} columns")
        print(f"Merged cells: {len(info['merged_cells'])}")
        print(f"Column widths defined: {len(info['column_widths'])}")
        print(f"Row heights defined: {len(info['row_heights'])}")
        print(f"Styled cells: {len(info['cell_styles'])}")
        
        # Print first few merged cells
        if info['merged_cells']:
            print("\nFirst 5 merged cells:")
            for mc in info['merged_cells'][:5]:
                print(f"  - {mc['range']}")
        
        # Print column widths
        if info['column_widths']:
            print("\nColumn widths:")
            for col, width in sorted(info['column_widths'].items())[:10]:
                print(f"  - Column {col}: {width}")
    
    # Save detailed analysis
    with open('/Users/kevin/Desktop/excel_structure_analysis.json', 'w', encoding='utf-8') as f:
        json.dump(analysis, f, ensure_ascii=False, indent=2)
    
    print("\nDetailed analysis saved to: /Users/kevin/Desktop/excel_structure_analysis.json")