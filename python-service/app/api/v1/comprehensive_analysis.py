"""
Comprehensive Analysis API Endpoint
Provides unified interface for comprehensive Excel error detection
"""
from fastapi import APIRouter, HTTPException, UploadFile, File, BackgroundTasks, Query
from typing import Dict, Any, Optional, List
import tempfile
import os
import logging
from datetime import datetime

from ...services.comprehensive_error_detector import ComprehensiveErrorDetector
from ...services.ai_analysis_prompts import AIAnalysisPrompts
from ...core.database import get_db
from ...core.config import settings

logger = logging.getLogger(__name__)
router = APIRouter()


@router.post("/comprehensive-analysis")
async def comprehensive_excel_analysis(
    file: UploadFile = File(...),
    include_ai: bool = Query(True, description="Include AI-based analysis"),
    include_patterns: bool = Query(True, description="Include pattern-based analysis"),
    include_vba: bool = Query(True, description="Include VBA analysis if applicable"),
    confidence_level: str = Query("balanced", description="Analysis confidence level: conservative, balanced, aggressive"),
    focus_areas: Optional[List[str]] = Query(None, description="Specific areas to focus on"),
    user_concern: Optional[str] = Query(None, description="User's specific concern about the file")
):
    """
    Perform comprehensive Excel analysis using multiple detection methods
    
    This endpoint:
    1. Runs basic tool analysis (ExcelJS/openpyxl)
    2. Runs AI analysis in parallel (always runs if enabled)
    3. Runs pattern matching in parallel
    4. Combines and prioritizes all results
    
    Returns:
        Comprehensive analysis results with errors from all detection methods
    """
    
    # Validate file
    if not file.filename:
        raise HTTPException(status_code=400, detail="No filename provided")
    
    if not file.filename.lower().endswith(('.xlsx', '.xlsm', '.xls', '.xlsb')):
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Supported formats: .xlsx, .xlsm, .xls, .xlsb"
        )
    
    # Check file size
    file_content = await file.read()
    file_size = len(file_content)
    
    if file_size > settings.MAX_FILE_SIZE:
        raise HTTPException(
            status_code=413,
            detail=f"File too large. Maximum size: {settings.MAX_FILE_SIZE / 1024 / 1024}MB"
        )
    
    # Save to temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file.filename)[1]) as tmp_file:
        tmp_file.write(file_content)
        tmp_path = tmp_file.name
    
    try:
        # Initialize detector
        detector = ComprehensiveErrorDetector()
        
        # Prepare analysis options
        analysis_options = {
            "include_basic": True,  # Always include basic analysis
            "include_ai": include_ai,
            "include_patterns": include_patterns,
            "include_vba": include_vba and file.filename.lower().endswith(('.xlsm', '.xlsb'))
        }
        
        # If user has specific concerns or focus areas, enhance AI analysis
        if include_ai and (focus_areas or user_concern):
            # Get file context for AI prompts
            file_context = {
                "filename": file.filename,
                "focus_areas": focus_areas or [],
                "user_concern": user_concern
            }
            
            # AI prompts will be automatically enhanced based on context
            analysis_options["ai_context"] = file_context
            analysis_options["confidence_level"] = confidence_level
        
        # Run comprehensive analysis
        result = await detector.analyze_excel_file(tmp_path, analysis_options)
        
        # Add request metadata
        result["metadata"] = {
            "filename": file.filename,
            "file_size": file_size,
            "analysis_options": analysis_options,
            "confidence_level": confidence_level,
            "timestamp": datetime.now().isoformat()
        }
        
        # Log analysis completion
        logger.info(f"Comprehensive analysis completed for {file.filename}: {result['total_errors']} errors found")
        
        return {
            "success": True,
            "data": result
        }
    
    except Exception as e:
        logger.error(f"Comprehensive analysis failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
    
    finally:
        # Clean up temporary file
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


@router.post("/analyze-with-context")
async def analyze_with_business_context(
    file: UploadFile = File(...),
    business_context: str = Query(..., description="Business context: financial, inventory, sales, hr, general"),
    custom_rules: Optional[str] = Query(None, description="Custom business rules in JSON format"),
    language: str = Query("ko", description="Language for analysis: ko, en"),
    export_format: Optional[str] = Query(None, description="Export format: json, excel, pdf")
):
    """
    Analyze Excel file with specific business context
    
    This endpoint allows users to specify the business domain for more accurate analysis
    """
    
    # Validate business context
    valid_contexts = ["financial", "inventory", "sales", "hr", "general", "scientific", "project"]
    if business_context not in valid_contexts:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid business context. Valid options: {', '.join(valid_contexts)}"
        )
    
    # Parse custom rules if provided
    custom_rules_dict = None
    if custom_rules:
        try:
            import json
            custom_rules_dict = json.loads(custom_rules)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="Invalid JSON format for custom rules")
    
    # Save uploaded file
    file_content = await file.read()
    with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file.filename)[1]) as tmp_file:
        tmp_file.write(file_content)
        tmp_path = tmp_file.name
    
    try:
        # Initialize components
        detector = ComprehensiveErrorDetector()
        prompt_manager = AIAnalysisPrompts()
        
        # Add custom rules if provided
        if custom_rules_dict:
            for rule_name, rule_content in custom_rules_dict.items():
                prompt_manager.add_custom_rule(rule_name, rule_content, business_context)
        
        # Prepare context-aware options
        analysis_options = {
            "include_basic": True,
            "include_ai": True,
            "include_patterns": True,
            "include_vba": file.filename.lower().endswith(('.xlsm', '.xlsb')),
            "business_context": business_context,
            "language": language
        }
        
        # Run analysis
        result = await detector.analyze_excel_file(tmp_path, analysis_options)
        
        # Format based on requested export format
        if export_format == "excel":
            # Generate Excel report
            report_path = await _generate_excel_report(result, file.filename)
            return {
                "success": True,
                "data": result,
                "report_url": f"/download/report/{os.path.basename(report_path)}"
            }
        elif export_format == "pdf":
            # Generate PDF report (would need additional implementation)
            raise HTTPException(status_code=501, detail="PDF export not yet implemented")
        else:
            # Return JSON
            return {
                "success": True,
                "data": result
            }
    
    except Exception as e:
        logger.error(f"Context-aware analysis failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
    
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


@router.get("/analysis-status/{analysis_id}")
async def get_analysis_status(analysis_id: str):
    """
    Get the status of a running analysis
    
    For long-running analyses, this endpoint allows checking progress
    """
    # This would connect to a job queue system like Celery or RQ
    # For now, return a mock response
    return {
        "analysis_id": analysis_id,
        "status": "completed",
        "progress": 100,
        "errors_found": 15,
        "estimated_time_remaining": 0
    }


@router.post("/quick-validation")
async def quick_excel_validation(
    file: UploadFile = File(...),
    check_formulas: bool = Query(True, description="Check formula errors"),
    check_structure: bool = Query(True, description="Check structural issues"),
    max_errors: int = Query(50, description="Maximum errors to return")
):
    """
    Quick validation endpoint for rapid feedback
    
    This runs only basic checks without AI analysis for faster response
    """
    
    file_content = await file.read()
    
    if len(file_content) > 10 * 1024 * 1024:  # 10MB limit for quick validation
        raise HTTPException(
            status_code=413,
            detail="File too large for quick validation. Please use comprehensive analysis."
        )
    
    with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file.filename)[1]) as tmp_file:
        tmp_file.write(file_content)
        tmp_path = tmp_file.name
    
    try:
        detector = ComprehensiveErrorDetector()
        
        # Run only basic analysis for speed
        analysis_options = {
            "include_basic": True,
            "include_ai": False,  # Skip AI for quick validation
            "include_patterns": check_structure,
            "include_vba": False  # Skip VBA for speed
        }
        
        result = await detector.analyze_excel_file(tmp_path, analysis_options)
        
        # Limit errors returned
        if result["total_errors"] > max_errors:
            result["errors"] = result["errors"][:max_errors]
            result["truncated"] = True
        
        return {
            "success": True,
            "data": result,
            "quick_validation": True
        }
    
    except Exception as e:
        logger.error(f"Quick validation failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
    
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


@router.post("/compare-analyses")
async def compare_analysis_results(
    file1: UploadFile = File(...),
    file2: UploadFile = File(...),
    comparison_type: str = Query("errors", description="Type of comparison: errors, structure, all")
):
    """
    Compare analysis results between two Excel files
    
    Useful for before/after comparisons or version differences
    """
    
    # Save both files
    files = []
    for upload_file in [file1, file2]:
        content = await upload_file.read()
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(upload_file.filename)[1]) as tmp:
            tmp.write(content)
            files.append((tmp.name, upload_file.filename))
    
    try:
        detector = ComprehensiveErrorDetector()
        
        # Analyze both files
        results = []
        for file_path, filename in files:
            result = await detector.analyze_excel_file(file_path, {
                "include_basic": True,
                "include_ai": comparison_type == "all",
                "include_patterns": comparison_type in ["errors", "all"],
                "include_vba": False
            })
            result["filename"] = filename
            results.append(result)
        
        # Compare results
        comparison = _compare_results(results[0], results[1], comparison_type)
        
        return {
            "success": True,
            "file1": file1.filename,
            "file2": file2.filename,
            "comparison": comparison
        }
    
    except Exception as e:
        logger.error(f"Comparison failed: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))
    
    finally:
        for file_path, _ in files:
            if os.path.exists(file_path):
                os.unlink(file_path)


async def _generate_excel_report(analysis_result: Dict[str, Any], original_filename: str) -> str:
    """Generate Excel report from analysis results"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = openpyxl.Workbook()
    
    # Summary sheet
    summary_sheet = wb.active
    summary_sheet.title = "Summary"
    
    # Add headers
    headers = ["Metric", "Value"]
    for col, header in enumerate(headers, 1):
        cell = summary_sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
    
    # Add summary data
    summary_data = [
        ["Total Errors", analysis_result["total_errors"]],
        ["Critical Errors", analysis_result["summary"]["critical_count"]],
        ["Analysis Date", analysis_result["analysis_date"]],
        ["Original File", original_filename]
    ]
    
    for row_idx, row_data in enumerate(summary_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            summary_sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # Errors detail sheet
    if analysis_result["errors"]:
        errors_sheet = wb.create_sheet("Error Details")
        
        # Headers
        error_headers = ["ID", "Category", "Severity", "Location", "Description", "Fix Suggestion"]
        for col, header in enumerate(error_headers, 1):
            cell = errors_sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        # Error data
        for row_idx, error in enumerate(analysis_result["errors"], 2):
            errors_sheet.cell(row=row_idx, column=1, value=error.get("id", ""))
            errors_sheet.cell(row=row_idx, column=2, value=error.get("category", ""))
            errors_sheet.cell(row=row_idx, column=3, value=error.get("severity", ""))
            errors_sheet.cell(row=row_idx, column=4, value=str(error.get("location", {})))
            errors_sheet.cell(row=row_idx, column=5, value=error.get("description", ""))
            errors_sheet.cell(row=row_idx, column=6, value=error.get("suggested_fix", ""))
    
    # Save report
    report_filename = f"analysis_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    report_path = os.path.join(tempfile.gettempdir(), report_filename)
    wb.save(report_path)
    
    return report_path


def _compare_results(result1: Dict[str, Any], result2: Dict[str, Any], comparison_type: str) -> Dict[str, Any]:
    """Compare two analysis results"""
    comparison = {
        "total_errors_diff": result2["total_errors"] - result1["total_errors"],
        "new_errors": [],
        "fixed_errors": [],
        "common_errors": []
    }
    
    # Create error maps for comparison
    errors1 = {f"{e['category']}_{e['location']}": e for e in result1.get("errors", [])}
    errors2 = {f"{e['category']}_{e['location']}": e for e in result2.get("errors", [])}
    
    # Find differences
    for key, error in errors2.items():
        if key not in errors1:
            comparison["new_errors"].append(error)
        else:
            comparison["common_errors"].append(error)
    
    for key, error in errors1.items():
        if key not in errors2:
            comparison["fixed_errors"].append(error)
    
    # Add summary
    comparison["summary"] = {
        "improvement": len(comparison["fixed_errors"]) > len(comparison["new_errors"]),
        "error_reduction_rate": (len(comparison["fixed_errors"]) / max(len(errors1), 1)) * 100
    }
    
    return comparison