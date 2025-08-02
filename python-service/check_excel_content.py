#!/usr/bin/env python3
"""
Check actual content of pic.xlsx file using openpyxl
"""
import openpyxl

def check_excel_content():
    print("=== CHECKING PIC.XLSX CONTENT ===")
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook('pic.xlsx', data_only=True)
        print(f"Workbook loaded successfully")
        print(f"Sheet names: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n=== Sheet: {sheet_name} ===")
            print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")
            
            # Check first 10x10 cells
            print("\nFirst 10x10 cells:")
            has_any_content = False
            
            for row in range(1, min(11, ws.max_row + 1)):
                row_values = []
                for col in range(1, min(11, ws.max_column + 1)):
                    cell = ws.cell(row=row, column=col)
                    value = cell.value
                    
                    if value is not None and value != "":
                        has_any_content = True
                    
                    # Show abbreviated value
                    if value is None:
                        display_value = "None"
                    elif isinstance(value, str) and len(value) > 10:
                        display_value = f'"{value[:10]}..."'
                    else:
                        display_value = f'"{value}"' if isinstance(value, str) else str(value)
                    
                    row_values.append(display_value)
                
                print(f"Row {row}: {row_values}")
            
            if not has_any_content:
                print("\n⚠️ No content found in first 10x10 cells!")
                
                # Check for any cell with content
                print("\nSearching for any cell with content...")
                found_content = False
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        if cell.value is not None and cell.value != "":
                            print(f"Found content at {cell.coordinate}: {cell.value}")
                            found_content = True
                            break
                    if found_content:
                        break
                
                if not found_content:
                    print("❌ No content found in entire sheet!")
            
            # Check for merged cells
            if ws.merged_cells.ranges:
                print(f"\nMerged cells: {len(ws.merged_cells.ranges)} ranges")
                for merged_range in list(ws.merged_cells.ranges)[:5]:
                    print(f"  {merged_range}")
            
            # Check for images
            if hasattr(ws, '_images') and ws._images:
                print(f"\nImages: {len(ws._images)} found")
        
        wb.close()
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    check_excel_content()