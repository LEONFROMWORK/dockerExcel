"""
Complete workflow test with job status checking
"""
import requests
import asyncio
import json
import os
from openpyxl import Workbook
import tempfile
import time

# Test configuration
RAILS_BASE_URL = "http://localhost:3000"
PYTHON_BASE_URL = "http://localhost:8000"

def create_test_excel_file():
    """Create test Excel file with errors"""
    wb = Workbook()
    ws = wb.active
    ws.title = "테스트시트"

    # Add various error scenarios
    ws['A1'] = "테스트 데이터"
    ws['B1'] = 10
    ws['C1'] = 0

    # DIV/0 error
    ws['A2'] = "=B1/C1"

    # #N/A error
    ws['A3'] = "=VLOOKUP(\"없는값\",E1:F10,2,FALSE)"

    # #NAME? error
    ws['A4'] = "=존재하지않는함수()"

    # #REF! error
    ws['A5'] = "=Sheet2!A1"

    # #VALUE! error
    ws['A6'] = "=A1+B1"  # Text + number

    # Circular reference
    ws['A7'] = "=A8"
    ws['A8'] = "=A7"

    # Additional data
    ws['E1'] = "찾을값"
    ws['F1'] = "결과값"

    # Save to temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        return tmp.name

def test_file_upload(file_path):
    """1. File upload test"""
    print("\n=== 1. 파일 업로드 ===")

    url = f"{RAILS_BASE_URL}/api/v1/excel_analysis/files"

    with open(file_path, 'rb') as f:
        files = {'file': ('test_errors.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        response = requests.post(url, files=files)

    if response.status_code == 200:
        result = response.json()
        print(f"✅ 파일 업로드 성공: ID={result.get('id')}")
        return result.get('id')
    else:
        print(f"❌ 파일 업로드 실패: {response.status_code}")
        print(response.text)
        return None

def test_error_detection_with_job(file_id):
    """2. Error detection test with job status checking"""
    print("\n=== 2. 오류 감지 (Background Job) ===")

    # Start analysis
    url = f"{RAILS_BASE_URL}/api/v1/excel_analysis/analyze"

    data = {
        'file_id': file_id,
        'analysis_type': 'parallel',
        'include_all_tiers': True
    }

    response = requests.post(url, json=data)

    if response.status_code != 200:
        print(f"❌ 분석 시작 실패: {response.status_code}")
        return []

    result = response.json()
    session_id = result.get('session_id')
    print(f"✅ 분석 시작: session_id={session_id}")
    print(f"상태: {result.get('status')}")

    # Poll for job completion
    status_url = f"{RAILS_BASE_URL}/api/v1/excel_analysis/status/{session_id}"

    for i in range(15):  # Max 15 seconds
        time.sleep(1)
        response = requests.get(status_url)

        if response.status_code == 200:
            status_result = response.json()
            print(f"진행 {i+1}초: 상태={status_result.get('status')}, 오류={status_result.get('total_errors', 0)}개")

            if status_result.get('analysis_complete'):
                errors = status_result.get('errors', [])
                print(f"\n✅ 분석 완료: {len(errors)}개 오류 발견")

                for i, error in enumerate(errors[:10]):  # Show first 10
                    print(f"\n오류 {i+1}:")
                    print(f"  - 타입: {error['type']}")
                    print(f"  - 위치: {error['sheet']}!{error['cell']}")
                    print(f"  - 메시지: {error['message']}")
                    print(f"  - 수식: {error.get('formula', 'N/A')}")
                    print(f"  - 자동수정가능: {error.get('auto_fix', False)}")

                return errors

    print("⏱️ 분석 시간 초과")
    return []

def test_single_error_fix(file_id, error):
    """3. Single error fix test"""
    print(f"\n=== 3. 오류 수정: {error['type']} at {error['cell']} ===")

    url = f"{RAILS_BASE_URL}/api/v1/excel/fixes/apply-single-fix"

    data = {
        'file_id': file_id,
        'error_id': error['id'],
        'auto_fix': True
    }

    response = requests.post(url, json=data)

    if response.status_code == 200:
        result = response.json()
        if result.get('success'):
            print(f"✅ 수정 성공:")
            print(f"  - 원본: {result.get('original_formula')}")
            print(f"  - 수정: {result.get('fixed_formula')}")
            print(f"  - 신뢰도: {result.get('confidence', 0) * 100}%")
        else:
            print(f"⚠️ 수정 실패: {result.get('message')}")
        return result
    else:
        print(f"❌ API 호출 실패: {response.status_code}")
        return None

def test_batch_fix(file_id, errors):
    """4. Batch fix test"""
    print("\n=== 4. 일괄 오류 수정 ===")

    url = f"{RAILS_BASE_URL}/api/v1/excel/fixes/apply-batch-fixes"

    # Select auto-fixable errors
    fixable_errors = [e for e in errors if e.get('auto_fix', False)][:5]  # Max 5

    if not fixable_errors:
        print("자동 수정 가능한 오류가 없습니다")
        return None

    data = {
        'file_id': file_id,
        'error_ids': [e['id'] for e in fixable_errors],
        'strategy': 'safe'
    }

    response = requests.post(url, json=data)

    if response.status_code == 200:
        result = response.json()
        print(f"✅ 일괄 수정 완료:")
        print(f"  - 전체: {result.get('total', 0)}개")
        print(f"  - 성공: {result.get('success', 0)}개")
        print(f"  - 실패: {result.get('failed', 0)}개")
        print(f"  - 건너뜀: {result.get('skipped', 0)}개")
        return result
    else:
        print(f"❌ 일괄 수정 실패: {response.status_code}")
        return None

def main():
    """Main test execution"""
    print("=== Excel 오류 검출 및 수정 완전 테스트 V2 ===")

    # Create test file
    test_file = create_test_excel_file()
    print(f"테스트 파일 생성: {test_file}")

    try:
        # 1. File upload
        file_id = test_file_upload(test_file)
        if not file_id:
            print("파일 업로드 실패로 테스트 중단")
            return

        # 2. Error detection with job status
        errors = test_error_detection_with_job(file_id)
        if not errors:
            print("오류가 감지되지 않아 테스트 중단")
            return

        # 3. Single error fix (first auto-fixable error)
        auto_fixable = [e for e in errors if e.get('auto_fix', False)]
        if auto_fixable:
            test_single_error_fix(file_id, auto_fixable[0])
        else:
            print("\n자동 수정 가능한 오류가 없어 단일 수정 테스트 건너뜀")

        # 4. Batch fix
        if len(auto_fixable) > 1:
            test_batch_fix(file_id, auto_fixable[1:])
        else:
            print("\n추가 자동 수정 가능한 오류가 없어 일괄 수정 테스트 건너뜀")

        # 5. Re-analyze after fixes
        print("\n=== 5. 수정 후 재검증 ===")
        time.sleep(2)  # Wait for fixes to be applied
        remaining_errors = test_error_detection_with_job(file_id)

        if len(remaining_errors) < len(errors):
            print(f"\n✅ 오류가 {len(errors)}개에서 {len(remaining_errors)}개로 감소했습니다!")
        else:
            print(f"\n현재 남은 오류: {len(remaining_errors)}개")

    finally:
        # Clean up
        if os.path.exists(test_file):
            os.unlink(test_file)
            print(f"\n테스트 파일 삭제: {test_file}")

if __name__ == "__main__":
    main()
    print("\n테스트 완료!")
