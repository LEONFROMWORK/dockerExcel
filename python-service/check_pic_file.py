#!/usr/bin/env python3
"""
Check pic.xlsx file contents directly
"""
from openpyxl import load_workbook

# Load the file
wb = load_workbook('pic.xlsx', data_only=True)
ws = wb.active

print(f"Sheet name: {ws.title}")
print(f"Max row: {ws.max_row}")
print(f"Max column: {ws.max_column}")

# Check first 10x10 cells
print("\nFirst 10x10 cells:")
for row in range(1, min(11, ws.max_row + 1)):
    for col in range(1, min(11, ws.max_column + 1)):
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            print(f"Cell({row},{col}): {repr(cell.value)}")

# Check specific cells mentioned in the logs
print("\n\nSpecific cells from logs:")
# Row 4 (index 3), Col 1 (index 0) - "(1) 손익 실적"
cell = ws.cell(row=4, column=1)
print(f"Cell(4,1): value={repr(cell.value)}, has_style={cell.has_style}")

# Row 5 (index 4), Col 2 (index 1) - "과목"
cell = ws.cell(row=5, column=2)
print(f"Cell(5,2): value={repr(cell.value)}, has_style={cell.has_style}")

# Row 8 (index 7), Col 2 (index 1) - "매출액"
cell = ws.cell(row=8, column=2)
print(f"Cell(8,2): value={repr(cell.value)}, has_style={cell.has_style}")

# Row 8 (index 7), Col 4 (index 3) - 6856.426353
cell = ws.cell(row=8, column=4)
print(f"Cell(8,4): value={repr(cell.value)}, has_style={cell.has_style}")

print("\n\nChecking ALL non-empty cells:")
non_empty_count = 0
for row in ws.iter_rows():
    for cell in row:
        if cell.value is not None:
            non_empty_count += 1
            if non_empty_count <= 20:  # Print first 20
                print(f"Cell({cell.row},{cell.column}): {repr(cell.value)}")

print(f"\nTotal non-empty cells: {non_empty_count}")