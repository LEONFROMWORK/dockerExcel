"""
완전한 워크플로우 테스트
파일 업로드부터 오류 검출 및 수정까지
"""

import requests
import asyncio
import json
import os
from openpyxl import Workbook
import tempfile
import time

# 테스트 설정
RAILS_BASE_URL = "http://localhost:3000"
PYTHON_BASE_URL = "http://localhost:8000"

def create_test_excel_file():
    """테스트용 Excel 파일 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = "테스트시트"

    # 다양한 오류 시나리오
    ws['A1'] = "테스트 데이터"
    ws['B1'] = 10
    ws['C1'] = 0

    # DIV/0 오류
    ws['A2'] = "=B1/C1"

    # #N/A 오류
    ws['A3'] = "=VLOOKUP(\"없는값\",E1:F10,2,FALSE)"

    # #NAME? 오류
    ws['A4'] = "=존재하지않는함수()"

    # #REF! 오류
    ws['A5'] = "=Sheet2!A1"

    # #VALUE! 오류
    ws['A6'] = "=A1+B1"  # 텍스트와 숫자 더하기

    # 순환 참조
    ws['A7'] = "=A8"
    ws['A8'] = "=A7"

    # 데이터 추가
    ws['E1'] = "찾을값"
    ws['F1'] = "결과값"

    # 임시 파일로 저장
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        return tmp.name

def test_file_upload(file_path):
    """1. 파일 업로드 테스트"""
    print("\n=== 1. 파일 업로드 ===")

    url = f"{RAILS_BASE_URL}/api/v1/excel_analysis/files"

    with open(file_path, 'rb') as f:
        files = {'file': ('test_errors.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}

        # CSRF 토큰 가져오기 (실제 환경에서는 로그인 후 세션에서)
        # 여기서는 테스트를 위해 스킵
        headers = {
            'Accept': 'application/json'
        }

        response = requests.post(url, files=files, headers=headers)

        if response.status_code == 200:
            result = response.json()
            print(f"✅ 파일 업로드 성공: ID={result.get('id')}")
            return result.get('id')
        else:
            print(f"❌ 파일 업로드 실패: {response.status_code}")
            print(response.text)
            return None

def test_error_detection(file_id):
    """2. 오류 감지 테스트"""
    print("\n=== 2. 오류 감지 ===")

    url = f"{RAILS_BASE_URL}/api/v1/excel_analysis/analyze"

    data = {
        'file_id': file_id,
        'analysis_type': 'parallel',
        'include_all_tiers': True
    }

    response = requests.post(url, json=data)

    if response.status_code == 200:
        result = response.json()
        errors = result.get('errors', [])
        print(f"✅ 오류 감지 완료: {len(errors)}개 오류 발견")

        for i, error in enumerate(errors):
            print(f"\n오류 {i+1}:")
            print(f"  - 타입: {error['type']}")
            print(f"  - 위치: {error['sheet']}!{error['cell']}")
            print(f"  - 메시지: {error['message']}")
            print(f"  - 수식: {error.get('formula', 'N/A')}")
            print(f"  - 자동수정가능: {error.get('is_auto_fixable', False)}")

        return errors
    else:
        print(f"❌ 오류 감지 실패: {response.status_code}")
        print(response.text)
        return []

def test_single_error_fix(file_id, error):
    """3. 단일 오류 수정 테스트"""
    print(f"\n=== 3. 오류 수정: {error['type']} at {error['cell']} ===")

    url = f"{RAILS_BASE_URL}/api/v1/excel/fixes/apply-single-fix"

    data = {
        'file_id': file_id,
        'error_id': error['id'],
        'auto_fix': True
    }

    response = requests.post(url, json=data)

    if response.status_code == 200:
        result = response.json()
        if result.get('success'):
            print(f"✅ 수정 성공:")
            print(f"  - 원본: {result.get('original_formula')}")
            print(f"  - 수정: {result.get('fixed_formula')}")
            print(f"  - 신뢰도: {result.get('confidence', 0) * 100}%")
        else:
            print(f"⚠️ 수정 실패: {result.get('message')}")
        return result
    else:
        print(f"❌ API 호출 실패: {response.status_code}")
        return None

def test_batch_fix(file_id, errors):
    """4. 일괄 수정 테스트"""
    print("\n=== 4. 일괄 오류 수정 ===")

    url = f"{RAILS_BASE_URL}/api/v1/excel/fixes/apply-batch-fixes"

    # 자동 수정 가능한 오류만 선택
    fixable_errors = [e for e in errors if e.get('is_auto_fixable', False)]

    data = {
        'file_id': file_id,
        'error_ids': [e['id'] for e in fixable_errors],
        'strategy': 'safe'
    }

    response = requests.post(url, json=data)

    if response.status_code == 200:
        result = response.json()
        print(f"✅ 일괄 수정 완료:")
        print(f"  - 전체: {result.get('total', 0)}개")
        print(f"  - 성공: {result.get('success', 0)}개")
        print(f"  - 실패: {result.get('failed', 0)}개")
        print(f"  - 건너뜀: {result.get('skipped', 0)}개")
        return result
    else:
        print(f"❌ 일괄 수정 실패: {response.status_code}")
        return None

def test_websocket_progress():
    """5. WebSocket 진행 상황 추적 테스트"""
    print("\n=== 5. WebSocket 실시간 진행 상황 ===")

    import websocket
    import threading

    ws_url = f"ws://localhost:8000/ws/excel/test_session_123"

    def on_message(ws, message):
        data = json.loads(message)
        msg_type = data.get('type')

        if msg_type == 'progress':
            progress = data['data']
            print(f"📊 진행률: {progress['percentage']}% - {progress['message']}")
        elif msg_type == 'error_detected':
            print(f"🔍 오류 감지: {data['data']}")
        elif msg_type == 'error_fixed':
            print(f"✅ 오류 수정: {data['data']}")

    def on_error(ws, error):
        print(f"❌ WebSocket 오류: {error}")

    def on_close(ws):
        print("WebSocket 연결 종료")

    def on_open(ws):
        print("WebSocket 연결 성공")

    # WebSocket 연결
    ws = websocket.WebSocketApp(ws_url,
                              on_open=on_open,
                              on_message=on_message,
                              on_error=on_error,
                              on_close=on_close)

    # 별도 스레드에서 실행
    wst = threading.Thread(target=ws.run_forever)
    wst.daemon = True
    wst.start()

    return ws

def main():
    """메인 테스트 실행"""
    print("=== Excel 오류 검출 및 수정 완전 테스트 ===")

    # 테스트 파일 생성
    test_file = create_test_excel_file()
    print(f"테스트 파일 생성: {test_file}")

    try:
        # 1. 파일 업로드
        file_id = test_file_upload(test_file)
        if not file_id:
            print("파일 업로드 실패로 테스트 중단")
            return

        # 잠시 대기 (서버 처리 시간)
        time.sleep(2)

        # 2. 오류 감지
        errors = test_error_detection(file_id)
        if not errors:
            print("오류가 감지되지 않아 테스트 중단")
            return

        # 3. 단일 오류 수정 (첫 번째 오류)
        if errors:
            test_single_error_fix(file_id, errors[0])

        # 4. 일괄 수정
        if len(errors) > 1:
            test_batch_fix(file_id, errors[1:])

        # 5. 수정 후 재검증
        print("\n=== 6. 수정 후 재검증 ===")
        remaining_errors = test_error_detection(file_id)

        if len(remaining_errors) < len(errors):
            print(f"✅ 오류가 {len(errors)}개에서 {len(remaining_errors)}개로 감소했습니다!")

    finally:
        # 테스트 파일 삭제
        if os.path.exists(test_file):
            os.unlink(test_file)
            print(f"\n테스트 파일 삭제: {test_file}")

if __name__ == "__main__":
    # 메인 테스트 실행
    main()

    print("\n테스트 완료!")
