"""
Debug formula detection issue
"""
import asyncio
import sys
import os
from openpyxl import load_workbook, Workbook

# Add project to path
sys.path.insert(0, '/Users/kevin/excel-unified/python-service')

async def debug_formula_detection():
    # Create a test Excel file with known errors
    import tempfile

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"

    # Add test data
    ws['A1'] = "Test Data"
    ws['B1'] = 10
    ws['C1'] = 0

    # Add various error formulas
    ws['A2'] = "=B1/C1"          # DIV/0 error
    ws['A3'] = "=VLOOKUP(\"X\",E1:F10,2,FALSE)"  # #N/A error
    ws['A4'] = "=UNKNOWNFUNC()"  # #NAME? error
    ws['A5'] = "=Sheet2!A1"      # #REF! error (no Sheet2)
    ws['A6'] = "=A1+B1"          # #VALUE! error (text + number)
    ws['A7'] = "=A8"             # Circular reference
    ws['A8'] = "=A7"             # Circular reference

    # Save to temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        temp_file = tmp.name

    print(f"Created test file: {temp_file}")

    # Load with openpyxl and check what we get
    print("\n=== Checking cell data types and values ===")
    wb_loaded = load_workbook(temp_file, data_only=False)
    ws_loaded = wb_loaded.active

    for row in range(1, 9):
        cell = ws_loaded[f'A{row}']
        print(f"A{row}:")
        print(f"  value: {cell.value}")
        print(f"  data_type: {cell.data_type}")
        print(f"  is_formula: {'f' if cell.data_type == 'f' else 'no'}")
        print(f"  internal_value: {cell._value if hasattr(cell, '_value') else 'N/A'}")

    # Now check with data_only=True to see error values
    print("\n=== Checking with data_only=True ===")
    wb_data = load_workbook(temp_file, data_only=True)
    ws_data = wb_data.active

    for row in range(1, 9):
        cell = ws_data[f'A{row}']
        print(f"A{row}: value={cell.value}, data_type={cell.data_type}")

    # Test formula detection logic
    print("\n=== Testing formula detection logic ===")
    from app.services.detection.strategies.formula_error_detector import FormulaErrorDetector

    detector = FormulaErrorDetector()

    # Manually check each cell
    for row in range(1, 9):
        cell = ws_loaded[f'A{row}']
        print(f"\nChecking A{row}:")
        print(f"  Is formula cell? {cell.data_type == 'f'}")
        print(f"  Is error value? {isinstance(cell.value, str) and cell.value.startswith('#') if cell.value else False}")

        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            print(f"  Formula detected: {cell.value}")
            # Try to check syntax
            syntax_error = detector._check_formula_syntax(cell)
            if syntax_error:
                print(f"  Syntax error: {syntax_error}")

    # Clean up
    os.unlink(temp_file)
    print(f"\nDeleted test file: {temp_file}")

if __name__ == "__main__":
    asyncio.run(debug_formula_detection())
