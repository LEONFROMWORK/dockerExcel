"""
Style converters - handles conversion from JSON to openpyxl objects
"""

from typing import Dict, Any
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


class FontConverter:
    """Converts font configuration to Font objects"""
    
    def convert(self, font_config: Dict[str, Any], color_scheme: Any) -> Font:
        """Convert font configuration to Font object"""
        config = font_config.copy()
        
        # Handle color references to ColorScheme
        if "color" in config and config["color"] == "error":
            config["color"] = color_scheme.error
        
        return Font(**config)


class FillConverter:
    """Converts fill configuration to PatternFill objects"""
    
    def convert(self, fill_config: Dict[str, Any], color_scheme: Any) -> PatternFill:
        """Convert fill configuration to PatternFill object"""
        config = fill_config.copy()
        
        # Handle color references to ColorScheme
        for color_key in ["start_color", "end_color"]:
            if color_key in config:
                color_value = config[color_key]
                if hasattr(color_scheme, color_value):
                    config[color_key] = getattr(color_scheme, color_value)
        
        return PatternFill(**config)


class BorderConverter:
    """Converts border configuration to Border objects"""
    
    def convert(self, border_config: Any, get_thin_border) -> Border:
        """Convert border configuration to Border object"""
        if border_config == "thin":
            return get_thin_border()
        elif isinstance(border_config, dict):
            border_dict = {}
            for side, style in border_config.items():
                border_dict[side] = Side(style=style)
            return Border(**border_dict)
        return None


class AlignmentConverter:
    """Converts alignment configuration to Alignment objects"""
    
    def convert(self, alignment_config: Dict[str, Any]) -> Alignment:
        """Convert alignment configuration to Alignment object"""
        return Alignment(**alignment_config)


class StyleConverter:
    """Main converter that coordinates all style conversions"""
    
    def __init__(self, color_scheme: Any):
        self.color_scheme = color_scheme
        self.font_converter = FontConverter()
        self.fill_converter = FillConverter()
        self.border_converter = BorderConverter()
        self.alignment_converter = AlignmentConverter()
    
    def convert_style(self, style_config: Dict[str, Any], get_thin_border) -> Dict[str, Any]:
        """Convert a single style configuration"""
        style_dict = {}
        
        # Convert each component
        if "font" in style_config:
            style_dict["font"] = self.font_converter.convert(
                style_config["font"], self.color_scheme
            )
        
        if "fill" in style_config:
            style_dict["fill"] = self.fill_converter.convert(
                style_config["fill"], self.color_scheme
            )
        
        if "alignment" in style_config:
            style_dict["alignment"] = self.alignment_converter.convert(
                style_config["alignment"]
            )
        
        if "border" in style_config:
            border = self.border_converter.convert(
                style_config["border"], get_thin_border
            )
            if border:
                style_dict["border"] = border
        
        # Copy other properties directly
        for key in ["number_format"]:
            if key in style_config:
                style_dict[key] = style_config[key]
        
        return style_dict
    
    def convert_all_styles(self, presets_data: Dict[str, Any], get_thin_border) -> Dict[str, Dict[str, Any]]:
        """Convert all style presets from JSON to openpyxl objects"""
        converted = {}
        
        for domain, styles in presets_data.items():
            converted[domain] = {}
            
            for style_name, style_config in styles.items():
                converted[domain][style_name] = self.convert_style(
                    style_config, get_thin_border
                )
        
        return converted