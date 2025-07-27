"""
AI 기반 지능형 Excel 수정 시스템
AI-Powered Smart Excel Fixer - 컨텍스트 인식 수정 및 개선 제안
"""

import logging
import json
import re
from typing import Dict, List, Any, Optional, Tuple
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime

from .openai_service import OpenAIService

logger = logging.getLogger(__name__)

class AISmartFixer:
    """AI 기반 지능형 Excel 수정 시스템"""

    def __init__(self):
        self.openai_service = OpenAIService()
        self.context_analysis = {}
        self.fix_history = []

    async def smart_fix_file(self, workbook: openpyxl.Workbook, file_analysis: Dict[str, Any],
                           user_intent: str = None) -> Dict[str, Any]:
        """AI 기반 종합적 스마트 수정"""

        results = {
            'ai_fixes_applied': 0,
            'context_improvements': 0,
            'data_insights': [],
            'optimization_suggestions': [],
            'fixes_applied': [],
            'ai_analysis': {},
            'confidence_scores': {}
        }

        try:
            # 1. 파일 컨텍스트 분석
            logger.info("AI 컨텍스트 분석 시작...")
            context_analysis = await self._analyze_file_context(workbook, file_analysis, user_intent)
            results['ai_analysis']['context'] = context_analysis

            # 2. 지능형 데이터 타입 추론 및 수정
            logger.info("지능형 데이터 타입 추론...")
            type_fixes = await self._smart_data_type_inference(workbook, context_analysis)
            results['ai_fixes_applied'] += type_fixes['count']
            results['fixes_applied'].extend(type_fixes['details'])
            results['confidence_scores']['data_types'] = type_fixes['confidence']

            # 3. 스마트 수식 최적화
            logger.info("스마트 수식 최적화...")
            formula_optimizations = await self._smart_formula_optimization(workbook, context_analysis)
            results['ai_fixes_applied'] += formula_optimizations['count']
            results['fixes_applied'].extend(formula_optimizations['details'])
            results['confidence_scores']['formulas'] = formula_optimizations['confidence']

            # 4. 컨텍스트 기반 데이터 정리
            logger.info("컨텍스트 기반 데이터 정리...")
            data_improvements = await self._contextual_data_cleanup(workbook, context_analysis)
            results['context_improvements'] += data_improvements['count']
            results['fixes_applied'].extend(data_improvements['details'])

            # 5. 데이터 인사이트 생성
            logger.info("데이터 인사이트 생성...")
            insights = await self._generate_data_insights(workbook, context_analysis)
            results['data_insights'] = insights

            # 6. 최적화 제안 생성
            logger.info("최적화 제안 생성...")
            suggestions = await self._generate_optimization_suggestions(workbook, context_analysis, file_analysis)
            results['optimization_suggestions'] = suggestions

            logger.info(f"AI 스마트 수정 완료: {results['ai_fixes_applied']}개 수정, {results['context_improvements']}개 개선")

        except Exception as e:
            logger.error(f"AI 스마트 수정 중 오류: {str(e)}")
            results['error'] = str(e)

        return results

    async def _analyze_file_context(self, workbook: openpyxl.Workbook, file_analysis: Dict[str, Any],
                                  user_intent: str = None) -> Dict[str, Any]:
        """파일 컨텍스트 분석"""

        # 워크시트 구조 분석
        sheets_info = []
        for sheet in workbook.worksheets:
            sheet_info = {
                'name': sheet.title,
                'dimensions': f"{sheet.max_row}x{sheet.max_column}",
                'data_sample': self._extract_sheet_sample(sheet),
                'headers': self._identify_headers(sheet),
                'data_types': self._analyze_column_types(sheet)
            }
            sheets_info.append(sheet_info)

        # AI에게 컨텍스트 분석 요청
        context_prompt = f"""
        Excel 파일의 컨텍스트를 분석해주세요:

        파일 정보:
        - 시트 수: {len(sheets_info)}
        - 시트 구조: {json.dumps(sheets_info, ensure_ascii=False, indent=2)}

        사용자 의도: {user_intent or "명시되지 않음"}

        다음 JSON 형태로 분석 결과를 제공해주세요:
        {{
            "file_purpose": "파일의 목적/용도",
            "data_domain": "데이터 영역 (재무, 판매, 인사 등)",
            "business_context": "비즈니스 컨텍스트",
            "quality_issues": ["발견된 품질 문제들"],
            "improvement_priorities": ["개선 우선순위"],
            "suggested_structure": "제안하는 구조",
            "data_relationships": ["데이터간 관계"],
            "automation_opportunities": ["자동화 가능한 영역"]
        }}
        """

        try:
            response = await self.openai_service.chat_completion(
                messages=[{"role": "user", "content": context_prompt}],
                temperature=0.3
            )

            context_analysis = json.loads(response)
            context_analysis['sheets_analyzed'] = len(sheets_info)
            context_analysis['analysis_timestamp'] = datetime.now().isoformat()

            return context_analysis

        except Exception as e:
            logger.warning(f"AI 컨텍스트 분석 실패: {str(e)}")
            return {
                "file_purpose": "분석 불가",
                "data_domain": "미확인",
                "business_context": "컨텍스트 분석 실패",
                "quality_issues": [],
                "improvement_priorities": [],
                "error": str(e)
            }

    def _extract_sheet_sample(self, sheet: openpyxl.worksheet.worksheet.Worksheet, sample_size: int = 5) -> List[List]:
        """시트에서 샘플 데이터 추출"""

        sample_data = []
        max_rows = min(sample_size, sheet.max_row)
        max_cols = min(10, sheet.max_column)  # 최대 10개 컬럼

        for row in range(1, max_rows + 1):
            row_data = []
            for col in range(1, max_cols + 1):
                cell_value = sheet.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            sample_data.append(row_data)

        return sample_data

    def _identify_headers(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> List[str]:
        """헤더 행 식별"""

        headers = []
        if sheet.max_row > 0:
            for col in range(1, min(sheet.max_column + 1, 20)):  # 최대 20개 컬럼
                cell_value = sheet.cell(row=1, column=col).value
                headers.append(str(cell_value) if cell_value is not None else f"Column_{col}")

        return headers

    def _analyze_column_types(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> Dict[str, str]:
        """컬럼별 데이터 타입 분석"""

        column_types = {}

        for col in range(1, min(sheet.max_column + 1, 20)):
            col_letter = get_column_letter(col)

            # 샘플 데이터 수집 (헤더 제외)
            sample_values = []
            for row in range(2, min(sheet.max_row + 1, 12)):  # 최대 10개 행 샘플
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value is not None:
                    sample_values.append(cell_value)

            # 타입 추론
            if not sample_values:
                column_types[col_letter] = "empty"
            elif all(isinstance(v, (int, float)) for v in sample_values):
                column_types[col_letter] = "numeric"
            elif all(isinstance(v, str) and self._is_date_like(v) for v in sample_values):
                column_types[col_letter] = "date"
            elif all(isinstance(v, str) for v in sample_values):
                column_types[col_letter] = "text"
            else:
                column_types[col_letter] = "mixed"

        return column_types

    def _is_date_like(self, text: str) -> bool:
        """텍스트가 날짜 형식인지 확인"""
        date_patterns = [
            r'\d{4}-\d{2}-\d{2}',
            r'\d{2}/\d{2}/\d{4}',
            r'\d{4}/\d{2}/\d{2}',
            r'\d{2}-\d{2}-\d{4}',
        ]

        for pattern in date_patterns:
            if re.match(pattern, text.strip()):
                return True
        return False

    async def _smart_data_type_inference(self, workbook: openpyxl.Workbook,
                                       context: Dict[str, Any]) -> Dict[str, Any]:
        """지능형 데이터 타입 추론 및 변환"""

        fixes = {
            'count': 0,
            'details': [],
            'confidence': 0.0
        }

        # 비즈니스 컨텍스트 기반 타입 추론
        business_context = context.get('business_context', '')
        data_domain = context.get('data_domain', '')

        for sheet in workbook.worksheets:
            try:
                # AI에게 컬럼별 적절한 데이터 타입 추천 요청
                headers = self._identify_headers(sheet)
                sample_data = self._extract_sheet_sample(sheet)

                type_inference_prompt = f"""
                비즈니스 컨텍스트: {business_context}
                데이터 영역: {data_domain}

                시트 '{sheet.title}'의 컬럼별 데이터 타입을 추론해주세요:
                헤더: {headers[:10]}  # 최대 10개
                샘플 데이터: {sample_data[:3]}  # 최대 3행

                각 컬럼에 대해 다음 JSON 형태로 응답해주세요:
                {{
                    "column_recommendations": {{
                        "A": {{"type": "number/text/date/currency/percentage", "confidence": 0.9, "reason": "추론 근거"}},
                        "B": {{"type": "text", "confidence": 0.8, "reason": "추론 근거"}}
                    }}
                }}
                """

                response = await self.openai_service.chat_completion(
                    messages=[{"role": "user", "content": type_inference_prompt}],
                    temperature=0.2
                )

                recommendations = json.loads(response)

                # 추천 사항에 따라 데이터 타입 변환
                for col_letter, rec in recommendations.get('column_recommendations', {}).items():
                    if rec['confidence'] > 0.7:  # 높은 신뢰도만 적용
                        col_index = column_index_from_string(col_letter)

                        if col_index <= sheet.max_column:
                            converted = self._convert_column_type(sheet, col_index, rec['type'])
                            if converted:
                                fixes['count'] += 1
                                fixes['details'].append({
                                    'type': 'smart_type_conversion',
                                    'location': f"{sheet.title}!{col_letter}",
                                    'action': f"컬럼 타입을 {rec['type']}로 변환",
                                    'confidence': rec['confidence'],
                                    'reason': rec['reason']
                                })

                # 평균 신뢰도 계산
                confidences = [rec['confidence'] for rec in recommendations.get('column_recommendations', {}).values()]
                fixes['confidence'] = sum(confidences) / len(confidences) if confidences else 0.0

            except Exception as e:
                logger.warning(f"시트 {sheet.title} 타입 추론 실패: {str(e)}")

        return fixes

    def _convert_column_type(self, sheet: openpyxl.worksheet.worksheet.Worksheet,
                           col_index: int, target_type: str) -> bool:
        """컬럼 데이터 타입 변환"""

        converted = False

        for row in range(2, sheet.max_row + 1):  # 헤더 제외
            cell = sheet.cell(row=row, column=col_index)

            if cell.value is None:
                continue

            try:
                if target_type == 'number':
                    if isinstance(cell.value, str):
                        # 문자열에서 숫자 추출
                        clean_text = re.sub(r'[^\d.-]', '', str(cell.value))
                        if clean_text:
                            numeric_value = float(clean_text)
                            if numeric_value.is_integer():
                                numeric_value = int(numeric_value)
                            cell.value = numeric_value
                            converted = True

                elif target_type == 'currency':
                    if isinstance(cell.value, str):
                        # 통화 형식에서 숫자 추출
                        clean_text = re.sub(r'[^\d.-]', '', str(cell.value))
                        if clean_text:
                            cell.value = float(clean_text)
                            # 통화 서식 적용
                            cell.number_format = '#,##0'
                            converted = True

                elif target_type == 'percentage':
                    if isinstance(cell.value, (int, float)):
                        if cell.value > 1:  # 100% 형식을 소수점 형식으로
                            cell.value = cell.value / 100
                        cell.number_format = '0.00%'
                        converted = True

                elif target_type == 'date':
                    if isinstance(cell.value, str) and self._is_date_like(cell.value):
                        from datetime import datetime
                        date_formats = ['%Y-%m-%d', '%m/%d/%Y', '%Y/%m/%d', '%m-%d-%Y']

                        for fmt in date_formats:
                            try:
                                date_obj = datetime.strptime(cell.value.strip(), fmt)
                                cell.value = date_obj
                                converted = True
                                break
                            except Exception:
                                continue

            except Exception as e:
                logger.warning(f"셀 {get_column_letter(col_index)}{row} 타입 변환 실패: {str(e)}")

        return converted

    async def _smart_formula_optimization(self, workbook: openpyxl.Workbook,
                                        context: Dict[str, Any]) -> Dict[str, Any]:
        """스마트 수식 최적화"""

        fixes = {
            'count': 0,
            'details': [],
            'confidence': 0.0
        }

        formula_cells = []

        # 모든 수식 셀 수집
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == 'f' and cell.value:  # 수식 셀
                        formula_cells.append({
                            'sheet': sheet.title,
                            'address': cell.coordinate,
                            'formula': cell.value,
                            'cell': cell
                        })

        if not formula_cells:
            return fixes

        # AI에게 수식 최적화 제안 요청
        formulas_sample = formula_cells[:10]  # 최대 10개 수식만 분석

        optimization_prompt = f"""
        비즈니스 컨텍스트: {context.get('business_context', '')}

        다음 Excel 수식들을 최적화해주세요:
        {json.dumps([{'위치': f['sheet'] + '!' + f['address'], '수식': f['formula']} for f in formulas_sample], ensure_ascii=False, indent=2)}

        각 수식에 대해 다음 JSON 형태로 최적화 제안을 해주세요:
        {{
            "optimizations": [
                {{
                    "original_location": "Sheet1!A1",
                    "original_formula": "=SUM(B1:B10)",
                    "optimized_formula": "=SUM(B:B)",
                    "optimization_type": "range_optimization/function_replacement/performance",
                    "improvement": "개선 사항 설명",
                    "confidence": 0.9
                }}
            ]
        }}
        """

        try:
            response = await self.openai_service.chat_completion(
                messages=[{"role": "user", "content": optimization_prompt}],
                temperature=0.2
            )

            optimizations = json.loads(response)

            # 최적화 적용
            for opt in optimizations.get('optimizations', []):
                if opt['confidence'] > 0.8:  # 높은 신뢰도만 적용
                    # 해당 셀 찾기
                    location = opt['original_location']
                    for formula_info in formula_cells:
                        if f"{formula_info['sheet']}!{formula_info['address']}" == location:
                            try:
                                formula_info['cell'].value = opt['optimized_formula']
                                fixes['count'] += 1
                                fixes['details'].append({
                                    'type': 'formula_optimization',
                                    'location': location,
                                    'action': f"수식 최적화: {opt['optimization_type']}",
                                    'improvement': opt['improvement'],
                                    'confidence': opt['confidence']
                                })
                            except Exception as e:
                                logger.warning(f"수식 최적화 적용 실패 {location}: {str(e)}")
                            break

            # 평균 신뢰도 계산
            confidences = [opt['confidence'] for opt in optimizations.get('optimizations', [])]
            fixes['confidence'] = sum(confidences) / len(confidences) if confidences else 0.0

        except Exception as e:
            logger.warning(f"수식 최적화 실패: {str(e)}")

        return fixes

    async def _contextual_data_cleanup(self, workbook: openpyxl.Workbook,
                                     context: Dict[str, Any]) -> Dict[str, Any]:
        """컨텍스트 기반 데이터 정리"""

        improvements = {
            'count': 0,
            'details': []
        }

        business_context = context.get('business_context', '')
        data_domain = context.get('data_domain', '')

        for sheet in workbook.worksheets:
            try:
                # 컨텍스트에 맞는 데이터 정리 규칙 적용
                if 'financial' in data_domain.lower() or '재무' in business_context:
                    # 재무 데이터 특화 정리
                    financial_improvements = self._apply_financial_cleanup(sheet)
                    improvements['count'] += financial_improvements

                elif 'sales' in data_domain.lower() or '판매' in business_context:
                    # 판매 데이터 특화 정리
                    sales_improvements = self._apply_sales_cleanup(sheet)
                    improvements['count'] += sales_improvements

                elif 'inventory' in data_domain.lower() or '재고' in business_context:
                    # 재고 데이터 특화 정리
                    inventory_improvements = self._apply_inventory_cleanup(sheet)
                    improvements['count'] += inventory_improvements

                # 일반적인 데이터 품질 개선
                general_improvements = self._apply_general_cleanup(sheet)
                improvements['count'] += general_improvements

            except Exception as e:
                logger.warning(f"시트 {sheet.title} 컨텍스트 정리 실패: {str(e)}")

        return improvements

    def _apply_financial_cleanup(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> int:
        """재무 데이터 특화 정리"""
        improvements = 0

        # 통화 형식 표준화, 음수 처리 등
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # 통화 기호 정리
                    if re.search(r'[$₩¥€£]', str(cell.value)):
                        clean_value = re.sub(r'[^\d.-]', '', str(cell.value))
                        try:
                            cell.value = float(clean_value)
                            cell.number_format = '#,##0'
                            improvements += 1
                        except Exception:
                            pass

        return improvements

    def _apply_sales_cleanup(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> int:
        """판매 데이터 특화 정리"""
        improvements = 0

        # 수량, 단가 등 표준화
        headers = self._identify_headers(sheet)

        for col_idx, header in enumerate(headers, 1):
            if any(keyword in header.lower() for keyword in ['quantity', 'qty', '수량', 'amount', '금액']):
                for row in range(2, sheet.max_row + 1):
                    cell = sheet.cell(row=row, column=col_idx)
                    if cell.value and isinstance(cell.value, str):
                        # 숫자 추출 및 표준화
                        clean_value = re.sub(r'[^\d.-]', '', str(cell.value))
                        try:
                            if clean_value:
                                cell.value = float(clean_value)
                                improvements += 1
                        except Exception:
                            pass

        return improvements

    def _apply_inventory_cleanup(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> int:
        """재고 데이터 특화 정리"""
        improvements = 0

        # SKU, 상품코드 표준화 등
        headers = self._identify_headers(sheet)

        for col_idx, header in enumerate(headers, 1):
            if any(keyword in header.lower() for keyword in ['sku', 'code', '코드', 'id']):
                for row in range(2, sheet.max_row + 1):
                    cell = sheet.cell(row=row, column=col_idx)
                    if cell.value and isinstance(cell.value, str):
                        # 코드 형식 표준화 (대문자, 특수문자 제거 등)
                        standardized = re.sub(r'[^\w-]', '', str(cell.value).upper())
                        if standardized != cell.value:
                            cell.value = standardized
                            improvements += 1

        return improvements

    def _apply_general_cleanup(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> int:
        """일반적인 데이터 품질 개선"""
        improvements = 0

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    original_value = cell.value

                    # 공백 정리
                    cleaned_value = str(cell.value).strip()
                    cleaned_value = re.sub(r'\s+', ' ', cleaned_value)  # 여러 공백을 하나로

                    if cleaned_value != original_value:
                        cell.value = cleaned_value
                        improvements += 1

        return improvements

    async def _generate_data_insights(self, workbook: openpyxl.Workbook,
                                    context: Dict[str, Any]) -> List[Dict[str, Any]]:
        """데이터 인사이트 생성"""

        insights = []

        try:
            # 각 시트별 기본 통계 수집
            sheet_stats = []
            for sheet in workbook.worksheets:
                stats = {
                    'sheet_name': sheet.title,
                    'row_count': sheet.max_row,
                    'column_count': sheet.max_column,
                    'headers': self._identify_headers(sheet)[:5],  # 최대 5개
                    'sample_data': self._extract_sheet_sample(sheet, 3)  # 3행만
                }
                sheet_stats.append(stats)

            # AI에게 인사이트 생성 요청
            insights_prompt = f"""
            다음 Excel 데이터를 분석하여 비즈니스 인사이트를 제공해주세요:

            비즈니스 컨텍스트: {context.get('business_context', '')}
            데이터 영역: {context.get('data_domain', '')}

            시트 정보:
            {json.dumps(sheet_stats, ensure_ascii=False, indent=2)}

            다음 JSON 형태로 인사이트를 제공해주세요:
            {{
                "insights": [
                    {{
                        "type": "data_quality/business_trend/optimization",
                        "title": "인사이트 제목",
                        "description": "상세 설명",
                        "impact": "high/medium/low",
                        "action_recommendation": "권장 조치사항"
                    }}
                ]
            }}
            """

            response = await self.openai_service.chat_completion(
                messages=[{"role": "user", "content": insights_prompt}],
                temperature=0.4
            )

            result = json.loads(response)
            insights = result.get('insights', [])

        except Exception as e:
            logger.warning(f"데이터 인사이트 생성 실패: {str(e)}")
            insights = [{
                'type': 'error',
                'title': '인사이트 생성 오류',
                'description': f'인사이트 생성 중 오류가 발생했습니다: {str(e)}',
                'impact': 'low'
            }]

        return insights

    async def _generate_optimization_suggestions(self, workbook: openpyxl.Workbook,
                                               context: Dict[str, Any],
                                               file_analysis: Dict[str, Any]) -> List[Dict[str, Any]]:
        """최적화 제안 생성"""

        suggestions = []

        try:
            # 현재 파일 상태 요약
            file_summary = {
                'total_sheets': len(workbook.worksheets),
                'total_errors': file_analysis.get('summary', {}).get('total_errors', 0),
                'file_size_estimate': f"{workbook.worksheets[0].max_row * workbook.worksheets[0].max_column}셀",
                'quality_issues': context.get('quality_issues', [])
            }

            # AI에게 최적화 제안 요청
            optimization_prompt = f"""
            Excel 파일 최적화 제안을 해주세요:

            현재 상태:
            {json.dumps(file_summary, ensure_ascii=False, indent=2)}

            비즈니스 컨텍스트: {context.get('business_context', '')}
            개선 우선순위: {context.get('improvement_priorities', [])}

            다음 JSON 형태로 최적화 제안을 해주세요:
            {{
                "suggestions": [
                    {{
                        "category": "performance/structure/data_quality/automation",
                        "title": "제안 제목",
                        "description": "상세 설명",
                        "priority": "high/medium/low",
                        "estimated_impact": "예상 효과",
                        "implementation_steps": ["단계1", "단계2"],
                        "automation_potential": true/false
                    }}
                ]
            }}
            """

            response = await self.openai_service.chat_completion(
                messages=[{"role": "user", "content": optimization_prompt}],
                temperature=0.3
            )

            result = json.loads(response)
            suggestions = result.get('suggestions', [])

        except Exception as e:
            logger.warning(f"최적화 제안 생성 실패: {str(e)}")
            suggestions = [{
                'category': 'error',
                'title': '제안 생성 오류',
                'description': f'최적화 제안 생성 중 오류가 발생했습니다: {str(e)}',
                'priority': 'low'
            }]

        return suggestions
