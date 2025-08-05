from fastapi import APIRouter, HTTPException, BackgroundTasks
from typing import List, Dict, Any, Optional
import openpyxl
import pandas as pd
import requests
import tempfile
import os
from datetime import datetime

from ...services.openai_service import OpenAIService

router = APIRouter()


class ExcelModifier:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = openpyxl.load_workbook(
            file_path, keep_vba=True, data_only=False
        )
        self.modifications_log = []

    def apply_modifications(self, modifications: List[Dict[str, Any]]) -> str:
        """Apply a list of modifications to the Excel file"""
        for mod in modifications:
            mod_type = mod.get("type")

            if mod_type == "formula_fix":
                self._fix_formula(mod)
            elif mod_type == "data_cleanup":
                self._cleanup_data(mod)
            elif mod_type == "add_validation":
                self._add_validation(mod)
            elif mod_type == "add_formula":
                self._add_formula(mod)
            elif mod_type == "format_cells":
                self._format_cells(mod)
            elif mod_type == "add_chart":
                self._add_chart(mod)
            elif mod_type == "vba_modification":
                self._modify_vba(mod)
            else:
                self.modifications_log.append(
                    {
                        "type": mod_type,
                        "status": "skipped",
                        "reason": "Unknown modification type",
                    }
                )

        # Save the modified file
        output_path = self._save_modified_file()
        return output_path

    def _fix_formula(self, mod: Dict[str, Any]):
        """Fix formula errors"""
        sheet_name = mod.get("sheet", self.workbook.active.title)
        cell_ref = mod.get("cell")
        new_formula = mod.get("new_formula")

        try:
            sheet = self.workbook[sheet_name]

            if cell_ref and new_formula:
                sheet[cell_ref].value = new_formula
                self.modifications_log.append(
                    {
                        "type": "formula_fix",
                        "location": f"{sheet_name}!{cell_ref}",
                        "status": "success",
                        "details": f"Formula updated to: {new_formula}",
                    }
                )
            else:
                # Auto-fix common formula errors
                for row in sheet.iter_rows():
                    for cell in row:
                        if (
                            cell.value
                            and isinstance(cell.value, str)
                            and cell.value.startswith("=")
                        ):
                            # Fix #REF! errors
                            if "#REF!" in cell.value:
                                fixed_formula = cell.value.replace(
                                    "#REF!", "A1"
                                )  # Default replacement
                                cell.value = fixed_formula
                                self.modifications_log.append(
                                    {
                                        "type": "formula_fix",
                                        "location": f"{sheet_name}!{cell.coordinate}",
                                        "status": "success",
                                        "details": "Fixed #REF! error",
                                    }
                                )

                            # Fix #DIV/0! errors by wrapping in IFERROR
                            elif "#DIV/0!" in str(cell.value):
                                cell.value = f"=IFERROR({cell.value[1:]}, 0)"
                                self.modifications_log.append(
                                    {
                                        "type": "formula_fix",
                                        "location": f"{sheet_name}!{cell.coordinate}",
                                        "status": "success",
                                        "details": "Wrapped in IFERROR to handle division by zero",
                                    }
                                )

        except Exception as e:
            self.modifications_log.append(
                {"type": "formula_fix", "status": "error", "error": str(e)}
            )

    def _cleanup_data(self, mod: Dict[str, Any]):
        """Clean up data issues"""
        sheet_name = mod.get("sheet", self.workbook.active.title)
        cleanup_type = mod.get("cleanup_type", "all")

        try:
            sheet = self.workbook[sheet_name]
            df = pd.DataFrame(sheet.values)

            if df.empty:
                return

            # Set first row as header if it looks like headers
            if cleanup_type in ["all", "headers"]:
                first_row = df.iloc[0]
                if all(isinstance(val, str) for val in first_row if val):
                    df.columns = first_row
                    df = df[1:]

            # Remove duplicates
            if cleanup_type in ["all", "duplicates"]:
                original_len = len(df)
                df = df.drop_duplicates()
                if len(df) < original_len:
                    self.modifications_log.append(
                        {
                            "type": "data_cleanup",
                            "status": "success",
                            "details": f"Removed {original_len - len(df)} duplicate rows",
                        }
                    )

            # Fill missing values
            if cleanup_type in ["all", "missing"]:
                for col in df.columns:
                    if df[col].dtype in ["float64", "int64"]:
                        df[col].fillna(0, inplace=True)
                    else:
                        df[col].fillna("", inplace=True)

                self.modifications_log.append(
                    {
                        "type": "data_cleanup",
                        "status": "success",
                        "details": "Filled missing values",
                    }
                )

            # Write back to sheet
            sheet.delete_rows(1, sheet.max_row)
            for r_idx, row in enumerate(df.values, 1):
                for c_idx, value in enumerate(row, 1):
                    sheet.cell(row=r_idx, column=c_idx, value=value)

        except Exception as e:
            self.modifications_log.append(
                {"type": "data_cleanup", "status": "error", "error": str(e)}
            )

    def _add_validation(self, mod: Dict[str, Any]):
        """Add data validation rules"""
        from openpyxl.worksheet.datavalidation import DataValidation

        sheet_name = mod.get("sheet", self.workbook.active.title)
        range_ref = mod.get("range")
        validation_type = mod.get("validation_type", "list")

        try:
            sheet = self.workbook[sheet_name]

            if validation_type == "list":
                values = mod.get("values", [])
                dv = DataValidation(
                    type="list", formula1=f'"{",".join(values)}"', allow_blank=True
                )
                dv.error = "Invalid entry"
                dv.errorTitle = "Invalid Entry"
                sheet.add_data_validation(dv)
                dv.add(range_ref)

            elif validation_type == "date":
                dv = DataValidation(
                    type="date",
                    operator="between",
                    formula1=mod.get("min_date", "2000-01-01"),
                    formula2=mod.get("max_date", "2100-12-31"),
                )
                sheet.add_data_validation(dv)
                dv.add(range_ref)

            elif validation_type == "number":
                dv = DataValidation(
                    type="whole",
                    operator="between",
                    formula1=mod.get("min_value", 0),
                    formula2=mod.get("max_value", 999999),
                )
                sheet.add_data_validation(dv)
                dv.add(range_ref)

            self.modifications_log.append(
                {
                    "type": "add_validation",
                    "status": "success",
                    "details": f"Added {validation_type} validation to {range_ref}",
                }
            )

        except Exception as e:
            self.modifications_log.append(
                {"type": "add_validation", "status": "error", "error": str(e)}
            )

    def _add_formula(self, mod: Dict[str, Any]):
        """Add new formulas"""
        sheet_name = mod.get("sheet", self.workbook.active.title)
        cell_ref = mod.get("cell")
        formula = mod.get("formula")

        try:
            sheet = self.workbook[sheet_name]
            sheet[cell_ref] = formula

            self.modifications_log.append(
                {
                    "type": "add_formula",
                    "location": f"{sheet_name}!{cell_ref}",
                    "status": "success",
                    "details": f"Added formula: {formula}",
                }
            )

        except Exception as e:
            self.modifications_log.append(
                {"type": "add_formula", "status": "error", "error": str(e)}
            )

    def _format_cells(self, mod: Dict[str, Any]):
        """Format cells"""
        from openpyxl.styles import Font, PatternFill, Alignment

        sheet_name = mod.get("sheet", self.workbook.active.title)
        range_ref = mod.get("range")

        try:
            sheet = self.workbook[sheet_name]

            # Parse formatting options
            font_dict = mod.get("font", {})
            fill_dict = mod.get("fill", {})
            alignment_dict = mod.get("alignment", {})

            # Create style objects
            font = Font(
                name=font_dict.get("name", "Calibri"),
                size=font_dict.get("size", 11),
                bold=font_dict.get("bold", False),
                italic=font_dict.get("italic", False),
                color=font_dict.get("color", "000000"),
            )

            if fill_dict:
                fill = PatternFill(
                    start_color=fill_dict.get("color", "FFFFFF"),
                    end_color=fill_dict.get("color", "FFFFFF"),
                    fill_type="solid",
                )
            else:
                fill = None

            alignment = Alignment(
                horizontal=alignment_dict.get("horizontal", "left"),
                vertical=alignment_dict.get("vertical", "center"),
            )

            # Apply to range
            for row in sheet[range_ref]:
                for cell in row:
                    cell.font = font
                    if fill:
                        cell.fill = fill
                    cell.alignment = alignment

            self.modifications_log.append(
                {
                    "type": "format_cells",
                    "status": "success",
                    "details": f"Formatted range {range_ref}",
                }
            )

        except Exception as e:
            self.modifications_log.append(
                {"type": "format_cells", "status": "error", "error": str(e)}
            )

    def _add_chart(self, mod: Dict[str, Any]):
        """Add charts to the workbook"""
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference

        sheet_name = mod.get("sheet", self.workbook.active.title)
        chart_type = mod.get("chart_type", "bar")
        mod.get("data_range")
        position = mod.get("position", "E5")

        try:
            sheet = self.workbook[sheet_name]

            # Create chart based on type
            if chart_type == "bar":
                chart = BarChart()
            elif chart_type == "line":
                chart = LineChart()
            elif chart_type == "pie":
                chart = PieChart()
            else:
                chart = BarChart()

            # Set chart properties
            chart.title = mod.get("title", "Chart")
            chart.style = 10

            # Add data
            # This is simplified - in practice, you'd parse the data_range properly
            data = Reference(sheet, min_col=2, min_row=1, max_row=10, max_col=3)
            categories = Reference(sheet, min_col=1, min_row=2, max_row=10)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)

            # Add to sheet
            sheet.add_chart(chart, position)

            self.modifications_log.append(
                {
                    "type": "add_chart",
                    "status": "success",
                    "details": f"Added {chart_type} chart at {position}",
                }
            )

        except Exception as e:
            self.modifications_log.append(
                {"type": "add_chart", "status": "error", "error": str(e)}
            )

    def _modify_vba(self, mod: Dict[str, Any]):
        """Modify VBA code (placeholder - requires more complex implementation)"""
        self.modifications_log.append(
            {
                "type": "vba_modification",
                "status": "pending",
                "details": "VBA modification requires additional implementation",
            }
        )

    def _save_modified_file(self) -> str:
        """Save the modified workbook"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"modified_{timestamp}_{os.path.basename(self.file_path)}"
        output_path = os.path.join(tempfile.gettempdir(), output_filename)

        self.workbook.save(output_path)
        self.workbook.close()

        return output_path


@router.post("/modify")
async def modify_excel(
    file_id: int,
    file_url: str,
    modifications: List[Dict[str, Any]],
    background_tasks: BackgroundTasks,
):
    """Modify an Excel file with specified changes"""
    try:
        # Download the file
        response = requests.get(file_url)
        response.raise_for_status()

        # Save to temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            tmp_file.write(response.content)
            tmp_path = tmp_file.name

        # Apply modifications
        modifier = ExcelModifier(tmp_path)
        output_path = modifier.apply_modifications(modifications)

        # Upload to storage (simplified - in practice, use S3 or similar)
        # For now, we'll return a local file URL
        file_url = f"/files/{os.path.basename(output_path)}"

        # Clean up original temp file
        os.unlink(tmp_path)

        return {
            "success": True,
            "file_url": file_url,
            "modifications_log": modifier.modifications_log,
            "file_id": file_id,
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/create-from-template")
async def create_from_template(
    template_id: str, customizations: Optional[Dict[str, Any]] = None
):
    """Create a new Excel file from a template"""
    try:
        # Load template (simplified - in practice, load from template library)
        template_path = f"templates/{template_id}.xlsx"

        if not os.path.exists(template_path):
            # Create a basic template
            wb = openpyxl.Workbook()
            ws = wb.active

            # Add template content based on template_id
            if template_id == "monthly_report":
                ws.title = "Monthly Report"
                headers = ["Date", "Revenue", "Expenses", "Profit"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

            elif template_id == "inventory":
                ws.title = "Inventory"
                headers = [
                    "Item Code",
                    "Description",
                    "Quantity",
                    "Unit Price",
                    "Total Value",
                ]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)

            # Save template
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"/tmp/template_{template_id}_{timestamp}.xlsx"
            wb.save(output_path)

        else:
            # Copy template
            wb = openpyxl.load_workbook(template_path)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"/tmp/template_{template_id}_{timestamp}.xlsx"
            wb.save(output_path)

        return {
            "success": True,
            "file_url": f"/files/{os.path.basename(output_path)}",
            "template_id": template_id,
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/create-from-ai")
async def create_from_ai(description: str, requirements: Optional[List[str]] = None):
    """Create a new Excel file based on AI interpretation"""
    try:
        # Use OpenAI to interpret requirements
        openai_service = OpenAIService()

        prompt = f"""
        Create an Excel file structure based on this description: {description}

        Requirements: {requirements or 'None specified'}

        Return a JSON structure with:
        - sheets: array of sheet definitions
        - each sheet should have: name, headers, sample_data (2-3 rows), formulas
        """

        ai_response = await openai_service.generate_completion(prompt)
        structure = eval(
            ai_response
        )  # In practice, use json.loads with proper error handling

        # Create Excel file from AI structure
        wb = openpyxl.Workbook()

        for idx, sheet_def in enumerate(structure.get("sheets", [])):
            if idx == 0:
                ws = wb.active
                ws.title = sheet_def["name"]
            else:
                ws = wb.create_sheet(sheet_def["name"])

            # Add headers
            headers = sheet_def.get("headers", [])
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Add sample data
            sample_data = sheet_def.get("sample_data", [])
            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            # Add formulas
            formulas = sheet_def.get("formulas", {})
            for cell_ref, formula in formulas.items():
                ws[cell_ref] = formula

        # Save file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"/tmp/ai_generated_{timestamp}.xlsx"
        wb.save(output_path)

        return {
            "success": True,
            "file_url": f"/files/{os.path.basename(output_path)}",
            "structure": structure,
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
