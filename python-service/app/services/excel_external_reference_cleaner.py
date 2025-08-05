import re
import os
import tempfile
import logging
from typing import Tuple, Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

logger = logging.getLogger(__name__)


class ExcelExternalReferenceCleaner:
    """Excel 파일에서 외부 참조를 제거하는 서비스"""

    def __init__(self):
        # 외부 참조 패턴 정의
        self.external_ref_patterns = [
            # file:/// 형식의 외부 참조
            re.compile(
                r"='?file:///[^']+\.xlsx?'?#?\$?([^!.]+)\.([A-Z]+\d+)", re.IGNORECASE
            ),
            # [파일명.xlsx]시트명 형식
            re.compile(r"='\[([^\]]+)\]([^'!]+)'!([A-Z]+\d+)", re.IGNORECASE),
            # 절대 경로가 포함된 참조
            re.compile(r"='?(/[^']+\.xlsx?)'?#?\$?([^!.]+)!([A-Z]+\d+)", re.IGNORECASE),
            re.compile(
                r"='?([A-Z]:[^']+\.xlsx?)'?#?\$?([^!.]+)!([A-Z]+\d+)", re.IGNORECASE
            ),
        ]

    def clean_external_references(
        self, file_path: str
    ) -> Tuple[bool, str, Dict[str, Any]]:
        """
        Excel 파일에서 외부 참조를 제거

        Returns:
            Tuple[bool, str, Dict]: (성공여부, 결과파일경로 또는 에러메시지, 메타데이터)
        """
        try:
            # 파일 존재 확인
            if not os.path.exists(file_path):
                return False, f"File not found: {file_path}", {}

            # Excel 파일 로드
            wb = load_workbook(file_path, data_only=False)
            has_external_refs = False
            cells_cleaned = 0
            sheets_affected = []

            # 모든 시트에서 외부 참조 검사 및 제거
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_modified = False

                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            # 수식인 경우에만 처리
                            if cell.value.startswith("="):
                                original_formula = cell.value
                                cleaned_formula = self._clean_formula(original_formula)

                                if original_formula != cleaned_formula:
                                    has_external_refs = True
                                    sheet_modified = True
                                    cells_cleaned += 1

                                    # 외부 참조를 값으로 대체 시도
                                    if cleaned_formula == "#REF!":
                                        # 외부 파일에서 값을 가져올 수 없는 경우
                                        # 다른 워크북에서 데이터 읽기 시도
                                        external_value = self._try_get_external_value(
                                            original_formula, file_path
                                        )
                                        if external_value is not None:
                                            cell.value = external_value
                                        else:
                                            # 값을 가져올 수 없으면 기본값 설정
                                            cell.value = 0
                                    else:
                                        cell.value = cleaned_formula

                                # 단순 외부 참조 데이터 (데이터만 있는 경우)
                                elif self._is_external_data_reference(original_formula):
                                    # 파일 경로에서 데이터 추출 시도
                                    file_match = re.search(
                                        r"file:///([^']+\.xlsx?)", original_formula
                                    )
                                    if file_match:
                                        external_file = file_match.group(1)
                                        if os.path.exists(external_file):
                                            try:
                                                wb_data = load_workbook(
                                                    external_file, data_only=True
                                                )
                                                # 시트와 셀 위치 추출
                                                cell_match = re.search(
                                                    r"#\$?([^!.]+)\.([A-Z]+\d+)",
                                                    original_formula,
                                                )
                                                if cell_match:
                                                    sheet_name_ref = cell_match.group(1)
                                                    cell_ref = cell_match.group(2)
                                                    if (
                                                        sheet_name_ref
                                                        in wb_data.sheetnames
                                                    ):
                                                        sheet_data = wb_data[
                                                            sheet_name_ref
                                                        ]
                                                        cell_data = sheet_data[cell_ref]
                                                        if cell_data.value is not None:
                                                            cell.value = cell_data.value
                                                            cells_cleaned += 1
                                                        else:
                                                            # If no value, replace with 0 or empty string
                                                            is_numeric = (
                                                                "SUM"
                                                                in original_formula
                                                                or "/"
                                                                in original_formula
                                                            )
                                                            cell.value = (
                                                                0 if is_numeric else ""
                                                            )
                                                            cells_cleaned += 1

                                                wb_data.close()

                                            except Exception as e:
                                                logger.warning(
                                                    f"Could not read external file {external_file}: {e}"
                                                )
                                                cell.value = 0
                                                cells_cleaned += 1
                                        else:
                                            # 외부 파일이 없으면 기본값
                                            cell.value = 0
                                            cells_cleaned += 1
                                            has_external_refs = True
                                            sheet_modified = True

                if sheet_modified:
                    sheets_affected.append(sheet_name)

            if has_external_refs:
                # 정리된 파일 저장
                with tempfile.NamedTemporaryFile(
                    suffix=".xlsx", delete=False
                ) as tmp_file:
                    output_path = tmp_file.name
                    wb.save(output_path)

                wb.close()

                metadata = {
                    "has_external_references": True,
                    "external_refs_cleaned": True,
                    "cells_cleaned": cells_cleaned,
                    "sheets_affected": sheets_affected,
                }

                return True, output_path, metadata
            else:
                wb.close()
                return (
                    True,
                    file_path,
                    {
                        "has_external_references": False,
                        "external_refs_cleaned": False,
                        "cells_cleaned": 0,
                    },
                )

        except InvalidFileException as e:
            logger.error(f"Invalid Excel file: {e}")
            return False, f"Invalid Excel file: {str(e)}", {}
        except Exception as e:
            logger.error(f"Error cleaning external references: {e}")
            return False, f"Error: {str(e)}", {}

    def _clean_formula(self, formula: str) -> str:
        """수식에서 외부 참조 제거"""
        cleaned = formula

        for pattern in self.external_ref_patterns:
            # 패턴에 맞는 외부 참조를 찾아서 제거
            matches = pattern.findall(cleaned)
            if matches:
                # 외부 참조를 #REF!로 대체
                cleaned = pattern.sub("#REF!", cleaned)

        return cleaned

    def _is_external_data_reference(self, formula: str) -> bool:
        """외부 데이터 참조인지 확인"""
        # file:/// 로 시작하는 경우
        if "file:///" in formula:
            return True
        # [파일명] 형식이 있는 경우
        if re.search(r"\[.+\]", formula):
            return True
        return False

    def _try_get_external_value(
        self, formula: str, current_file_path: str
    ) -> Optional[Any]:
        """외부 참조에서 실제 값을 가져오려고 시도"""
        # 이 메서드는 필요에 따라 구현
        # 현재는 None을 반환하여 기본값 사용
        return None


# Create singleton instance
excel_external_reference_cleaner = ExcelExternalReferenceCleaner()
