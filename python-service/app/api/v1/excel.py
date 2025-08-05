"""
Excel analysis API endpoints
"""

from fastapi import APIRouter, File, UploadFile, HTTPException, Depends
from pydantic import BaseModel, Field
from sqlalchemy.ext.asyncio import AsyncSession
import tempfile
import os
from typing import Dict, Any, List, Optional
import logging
from datetime import datetime

from app.core.database import get_db
from app.core.config import settings
from app.core.i18n_dependencies import get_i18n_context, I18nContext
from app.core.validators import FileValidator
from app.services.excel_analyzer import excel_analyzer
from app.services.openai_service import openai_service
from app.services.vector_search import vector_search_service
from app.services.excel_chart_analyzer import excel_chart_analyzer
from app.services.excel_pivot_analyzer import excel_pivot_analyzer
from app.services.template_selection_service import template_selection_service
from app.services.excel_auto_fixer import ExcelAutoFixer
from app.services.circular_reference_detector import CircularReferenceDetector
from app.services.fast_formula_fixer import FastFormulaFixer
from app.services.context import get_enhanced_context_manager
from app.core.responses import ResponseBuilder

logger = logging.getLogger(__name__)

router = APIRouter()


class AnalyzeRequest(BaseModel):
    user_query: Optional[str] = Field(None, max_length=1000)
    session_id: Optional[str] = Field(
        None, pattern=r"^[a-zA-Z0-9\-_]+$", max_length=128
    )


@router.post("/analyze")
async def analyze_excel_file(
    file: UploadFile = File(...),
    request: AnalyzeRequest = AnalyzeRequest(),
    db: AsyncSession = Depends(get_db),
    i18n: I18nContext = Depends(get_i18n_context),
) -> Dict[str, Any]:
    """
    Analyze an Excel file and optionally answer a specific query
    """
    # Enhanced file validation
    if not FileValidator.validate_filename(file.filename):
        error_message = i18n.get_error_message(
            "invalid_file_type", extensions=", ".join(settings.ALLOWED_EXTENSIONS)
        )
        raise HTTPException(status_code=400, detail=error_message)

    # Validate file size
    if not FileValidator.validate_file_size(file.size):
        error_message = i18n.get_error_message(
            "file_too_large", max_size=settings.MAX_UPLOAD_SIZE
        )
        raise HTTPException(status_code=400, detail=error_message)

    # Validate MIME type by reading first bytes
    content = await file.read()
    await file.seek(0)  # Reset file pointer

    if not FileValidator.validate_mime_type(file.filename, content[:8]):
        raise HTTPException(
            status_code=400,
            detail="Invalid file format. Please upload a valid Excel file.",
        )

    # Save uploaded file temporarily
    with tempfile.NamedTemporaryFile(
        delete=False, suffix=os.path.splitext(file.filename)[1]
    ) as tmp_file:
        content = await file.read()
        tmp_file.write(content)
        tmp_path = tmp_file.name

    try:
        # Generate unique file ID and save mapping
        from app.core.file_path_resolver import FilePathResolver

        file_id = await FilePathResolver.generate_file_id(file.filename)
        await FilePathResolver.save_file_mapping(
            file_id,
            tmp_path,
            {
                "original_filename": file.filename,
                "size": file.size,
                "session_id": request.session_id,
            },
        )

        # Analyze file structure using IntegratedErrorDetector
        logger.info(f"Analyzing file: {file.filename} (ID: {file_id})")
        from app.services.detection.integrated_error_detector import (
            IntegratedErrorDetector,
        )

        # Detect errors with IntegratedErrorDetector
        detector = IntegratedErrorDetector()
        detection_result = await detector.detect_all_errors(tmp_path)
        errors = detection_result["errors"]

        # Also get basic file analysis from excel_analyzer for compatibility
        analysis_result = await excel_analyzer.analyze_file(tmp_path)

        # Add detected errors to analysis result
        analysis_result["errors"] = [
            {
                "id": error.id,
                "type": error.type,
                "sheet": error.sheet,
                "cell": error.cell,
                "message": error.message,
                "severity": error.severity,
                "auto_fixable": error.is_auto_fixable,  # Changed from is_auto_fixable to auto_fixable for Rails compatibility
                "suggested_fix": error.suggested_fix,
            }
            for error in errors
        ]

        # Update summary with error count
        analysis_result["summary"]["total_errors"] = len(errors)
        analysis_result["summary"]["has_errors"] = len(errors) > 0

        # Get AI insights
        ai_analysis = await openai_service.analyze_excel_content(
            analysis_result, request.user_query
        )

        # 고급 기능 분석 추가
        advanced_analysis = {}

        # 차트 제안
        chart_suggestions = excel_chart_analyzer.suggest_optimal_charts(tmp_path)
        advanced_analysis["chart_suggestions"] = chart_suggestions

        # 피벗테이블 제안
        pivot_suggestions = excel_pivot_analyzer.suggest_optimal_pivots(tmp_path)
        advanced_analysis["pivot_suggestions"] = pivot_suggestions

        # 템플릿 추천 (사용자 쿼리가 있는 경우)
        if request.user_query:
            template_recommendations = (
                await template_selection_service.recommend_templates(
                    user_intent=request.user_query,
                    excel_file_path=tmp_path,
                    max_recommendations=3,
                )
            )
            advanced_analysis["template_recommendations"] = template_recommendations

        # Initialize workbook context if session_id is provided
        if request.session_id:
            context_manager = get_enhanced_context_manager()
            try:
                # 분석 결과 구조 변환 (detection_result와 analysis_result 통합)
                combined_result = {
                    **analysis_result,
                    "errors": detection_result["errors"],
                    "summary": {
                        **analysis_result.get("summary", {}),
                        **detection_result.get("summary", {}),
                    },
                }

                workbook_context = await context_manager.initialize_workbook_context(
                    session_id=request.session_id,
                    file_id=file_id,  # 고유 파일 ID 사용
                    file_name=file.filename,
                    analysis_result=combined_result,
                )

                # IntegratedErrorDetector 결과로 업데이트
                await context_manager.update_from_detector_result(
                    request.session_id, detection_result
                )

                logger.info(f"워크북 컨텍스트 초기화 완료: {file.filename}")

                # WebSocket으로 분석 완료 알림
                from app.api.v1.context_websocket import manager as ws_manager

                await ws_manager.send_to_session(
                    request.session_id,
                    {
                        "type": "analysis_complete",
                        "data": {
                            "file_id": file_id,
                            "file_name": file.filename,
                            "total_errors": len(errors),
                            "error_types": list(
                                set(
                                    e["type"]
                                    for e in errors
                                    if isinstance(e, dict) and "type" in e
                                )
                            ),
                            "summary": detection_result.get("summary", {}),
                        },
                        "timestamp": datetime.now().isoformat(),
                    },
                )

            except Exception as e:
                logger.error(f"워크북 컨텍스트 초기화 실패: {str(e)}")
                # 실패해도 계속 진행

        # Index content for future searches with enhanced summary
        content_summary = _create_enhanced_content_summary(
            analysis_result, detection_result
        )
        await vector_search_service.index_document(
            document_id=file_id,  # 파일 ID 사용
            document_type="excel_analysis",
            content=content_summary,
            metadata={
                "file_id": file_id,
                "filename": file.filename,
                "sheets": list(analysis_result["sheets"].keys()),
                "has_errors": analysis_result["summary"]["has_errors"],
                "error_count": len(errors),
                "error_types": list(
                    set(
                        e["type"] for e in errors if isinstance(e, dict) and "type" in e
                    )
                ),
                "session_id": request.session_id,
            },
            db=db,
        )

        # 표준 응답 생성
        response_data = {
            "file_id": file_id,
            "filename": file.filename,
            "file_analysis": analysis_result,
            "ai_insights": ai_analysis,
            "advanced_analysis": advanced_analysis,
            "user_query": request.user_query,
            "session_id": request.session_id,
            "language": i18n.language,
            "localized_labels": {
                "charts": i18n.get_text("excel.charts.title"),
                "pivot_tables": i18n.get_text("excel.pivot.title"),
                "templates": i18n.get_text("templates.title"),
            },
        }

        return ResponseBuilder.success(
            data=response_data, message=i18n.get_text("file.analysis_complete")
        )

    except Exception as e:
        logger.error(f"Error analyzing Excel file: {str(e)}")
        error_response = ResponseBuilder.from_exception(
            exception=e,
            context={"filename": file.filename},
            include_traceback=settings.DEBUG,
        )
        raise HTTPException(status_code=500, detail=error_response)
    finally:
        # Clean up temporary file
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


@router.post("/extract-formulas")
async def extract_formulas(file: UploadFile = File(...)) -> Dict[str, Any]:
    """
    Extract all formulas from an Excel file
    """
    # Validate file type
    if not any(file.filename.endswith(ext) for ext in settings.ALLOWED_EXTENSIONS):
        raise HTTPException(
            status_code=400,
            detail=f"Invalid file type. Allowed types: {settings.ALLOWED_EXTENSIONS}",
        )

    # Save uploaded file temporarily
    with tempfile.NamedTemporaryFile(
        delete=False, suffix=os.path.splitext(file.filename)[1]
    ) as tmp_file:
        content = await file.read()
        tmp_file.write(content)
        tmp_path = tmp_file.name

    try:
        # Extract formulas
        formulas = await excel_analyzer.extract_formulas(tmp_path)

        # Count total formulas
        total_formulas = sum(
            len(sheet_formulas) for sheet_formulas in formulas.values()
        )

        return ResponseBuilder.success(
            data={
                "filename": file.filename,
                "total_formulas": total_formulas,
                "formulas_by_sheet": formulas,
            }
        )

    except Exception as e:
        logger.error(f"Error extracting formulas: {str(e)}")
        error_response = ResponseBuilder.from_exception(
            exception=e, context={"filename": file.filename}
        )
        raise HTTPException(status_code=500, detail=error_response)
    finally:
        # Clean up temporary file
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


@router.post("/validate-formula")
async def validate_formula(
    formula: str, context: Dict[str, Any] = None
) -> Dict[str, Any]:
    """
    Validate and explain an Excel formula
    """
    if not formula.startswith("="):
        formula = "=" + formula

    # Get AI explanation
    prompt = f"Explain this Excel formula and check for any issues: {formula}"
    if context:
        prompt += f"\nContext: {context}"

    solution = await openai_service.generate_excel_solution(prompt, context)

    return ResponseBuilder.success(
        data={"formula": formula, "validation": solution, "context": context}
    )


@router.post("/auto-fix")
async def auto_fix_excel_file(
    file: UploadFile = File(...),
    fix_formulas: bool = True,
    fix_data_quality: bool = True,
    fix_structural: bool = True,
    fix_formatting: bool = True,
    save_fixed_file: bool = True,
    db: AsyncSession = Depends(get_db),
    i18n: I18nContext = Depends(get_i18n_context),
) -> Dict[str, Any]:
    """
    자동으로 Excel 파일의 모든 오류를 감지하고 수정
    """
    # 파일 타입 검증
    if not any(file.filename.endswith(ext) for ext in settings.ALLOWED_EXTENSIONS):
        error_message = i18n.get_error_message(
            "invalid_file_type", extensions=", ".join(settings.ALLOWED_EXTENSIONS)
        )
        raise HTTPException(status_code=400, detail=error_message)

    # 파일 크기 검증
    if file.size > settings.MAX_UPLOAD_SIZE:
        error_message = i18n.get_error_message(
            "file_too_large", max_size=settings.MAX_UPLOAD_SIZE
        )
        raise HTTPException(status_code=400, detail=error_message)

    # 임시 파일 저장
    with tempfile.NamedTemporaryFile(
        delete=False, suffix=os.path.splitext(file.filename)[1]
    ) as tmp_file:
        content = await file.read()
        tmp_file.write(content)
        tmp_path = tmp_file.name

    try:
        # 자동 수정 옵션 설정
        fix_options = {
            "fix_formulas": fix_formulas,
            "fix_data_quality": fix_data_quality,
            "fix_structural": fix_structural,
            "fix_formatting": fix_formatting,
            "save_fixed_file": save_fixed_file,
            "revalidate": True,
        }

        # 자동 수정 실행
        logger.info(f"Starting auto-fix for file: {file.filename}")
        auto_fixer = ExcelAutoFixer()
        fix_result = await auto_fixer.auto_fix_file(tmp_path, fix_options)

        # 결과 요약
        response_data = {
            "filename": file.filename,
            "summary": fix_result["summary"],
            "performance": fix_result["performance"],
            "details": {
                "original_errors": fix_result["original_errors"].get("summary", {}),
                "fixed_errors": {
                    "formulas": fix_result["fixed_errors"]
                    .get("formulas", {})
                    .get("fixed_formulas", []),
                    "data_quality": fix_result["fixed_errors"]
                    .get("data_quality", {})
                    .get("fixes_applied", []),
                    "structural": fix_result["fixed_errors"]
                    .get("structural", {})
                    .get("fixes_applied", []),
                    "formatting": fix_result["fixed_errors"]
                    .get("formatting", {})
                    .get("fixes_applied", []),
                },
                "remaining_errors": len(fix_result.get("unfixed_errors", [])),
            },
            "language": i18n.language,
        }

        # 수정된 파일 경로 추가
        if "fixed_file_path" in fix_result:
            response_data["fixed_file_path"] = fix_result["fixed_file_path"]

        # 벡터 DB에 인덱싱
        await vector_search_service.index_document(
            document_id=f"{file.filename}_autofix",
            document_type="excel_autofix",
            content=f"Auto-fixed Excel file: {file.filename}. Fixed {fix_result['summary']['total_errors_fixed']} errors.",
            metadata={
                "filename": file.filename,
                "total_errors_fixed": fix_result["summary"]["total_errors_fixed"],
                "fix_rate": fix_result["summary"]["fix_rate"],
            },
            db=db,
        )

        return ResponseBuilder.success(
            data=response_data, message=i18n.get_text("file.auto_fix_complete")
        )

    except Exception as e:
        logger.error(f"Error in auto-fix: {str(e)}")
        error_response = ResponseBuilder.from_exception(
            exception=e, context={"filename": file.filename}
        )
        raise HTTPException(status_code=500, detail=error_response)
    finally:
        # 임시 파일 정리
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


@router.post("/detect-errors")
async def detect_errors(
    file: UploadFile = File(...), session_id: str = None
) -> Dict[str, Any]:
    """
    Excel 파일의 오류 감지 - IntegratedErrorDetector 사용
    """
    # Validate file type
    if not any(file.filename.endswith(ext) for ext in settings.ALLOWED_EXTENSIONS):
        raise HTTPException(
            status_code=400,
            detail=f"Invalid file type. Allowed types: {settings.ALLOWED_EXTENSIONS}",
        )

    # Save uploaded file temporarily
    with tempfile.NamedTemporaryFile(
        delete=False, suffix=os.path.splitext(file.filename)[1]
    ) as tmp_file:
        content = await file.read()
        tmp_file.write(content)
        tmp_path = tmp_file.name

    try:
        # Use IntegratedErrorDetector for comprehensive error detection
        from app.services.detection.integrated_error_detector import (
            IntegratedErrorDetector,
        )

        logger.info(f"Detecting errors in file: {file.filename}")
        detector = IntegratedErrorDetector()
        detection_result = await detector.detect_all_errors(tmp_path)
        errors = detection_result["errors"]

        # Convert errors to dict format
        errors_dict = [
            {
                "id": error.id,
                "type": error.type,
                "sheet": error.sheet,
                "cell": error.cell,
                "message": error.message,
                "severity": error.severity,
                "auto_fixable": error.is_auto_fixable,  # Changed to auto_fixable for Rails compatibility
                "suggested_fix": error.suggested_fix,
                "formula": error.formula,
                "value": error.value,
                "confidence": error.confidence,
            }
            for error in errors
        ]

        # Generate summary
        summary = {
            "total_errors": len(errors),
            "critical_errors": sum(1 for e in errors if e.severity == "critical"),
            "high_errors": sum(1 for e in errors if e.severity == "high"),
            "medium_errors": sum(1 for e in errors if e.severity == "medium"),
            "low_errors": sum(1 for e in errors if e.severity == "low"),
            "auto_fixable": sum(1 for e in errors if e.is_auto_fixable),
            "has_errors": len(errors) > 0,
        }

        return ResponseBuilder.success(
            data={
                "filename": file.filename,
                "errors": errors_dict,
                "summary": summary,
                "session_id": session_id,
            },
            message="오류 감지 완료",
        )

    except Exception as e:
        logger.error(f"Error detecting Excel errors: {str(e)}")
        error_response = ResponseBuilder.from_exception(
            exception=e, context={"filename": file.filename}
        )
        raise HTTPException(status_code=500, detail=error_response)
    finally:
        # Clean up temporary file
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


def _create_enhanced_content_summary(
    analysis: Dict[str, Any], detection_result: Dict[str, Any]
) -> str:
    """IntegratedErrorDetector 결과를 포함한 향상된 컨텐츠 요약 생성"""
    parts = []
    summary = analysis.get("summary", {})

    # 파일 개요
    parts.append(
        f"Excel 파일 개요: {summary.get('total_sheets', 0)}개 시트, "
        f"{summary.get('total_rows', 0)}행, {summary.get('total_cells_with_data', 0)}개 데이터 셀"
    )

    # IntegratedErrorDetector 오류 요약
    if detection_result and "summary" in detection_result:
        det_summary = detection_result["summary"]
        parts.append("\n오류 감지 결과:")
        parts.append(f"- 총 오류: {det_summary.get('total_errors', 0)}개")
        parts.append(
            f"- 자동 수정 가능: {det_summary.get('auto_fixable', 0)}개 "
            f"({det_summary.get('auto_fixable_percentage', 0)}%)"
        )

        # 오류 타입별 분류
        if "by_type" in det_summary:
            parts.append("- 오류 타입:")
            for error_type, count in det_summary["by_type"].items():
                parts.append(f"  · {error_type}: {count}개")

        # 심각도별 분류
        if "by_severity" in det_summary:
            parts.append("- 심각도:")
            for severity, count in det_summary["by_severity"].items():
                parts.append(f"  · {severity}: {count}개")

    # 패턴 분석 결과 (있는 경우)
    if "pattern_analysis" in detection_result:
        patterns = detection_result["pattern_analysis"]
        if patterns.get("patterns"):
            parts.append("\n발견된 패턴:")
            for pattern in patterns["patterns"][:3]:  # 상위 3개 패턴
                parts.append(f"- {pattern.get('description', 'Unknown pattern')}")

    # 시트별 상세 정보
    parts.append("\n시트별 상세:")
    for sheet_name, sheet_data in analysis.get("sheets", {}).items():
        sheet_summary = sheet_data.get("summary", {})
        parts.append(f"\n[{sheet_name}]")
        parts.append(
            f"- 크기: {sheet_summary.get('rows', 0)}행 × {sheet_summary.get('columns', 0)}열"
        )
        parts.append(f"- 데이터 셀: {sheet_summary.get('non_empty_cells', 0)}개")

        # 시트별 오류
        if detection_result and "by_sheet" in detection_result.get("summary", {}):
            sheet_errors = detection_result["summary"]["by_sheet"].get(sheet_name, 0)
            if sheet_errors > 0:
                parts.append(f"- 오류: {sheet_errors}개")

        # 데이터 타입 분포
        if "data_types" in sheet_data:
            type_summary = []
            for dtype, count in sheet_data["data_types"].items():
                if count > 0:
                    type_summary.append(f"{dtype}: {count}")
            if type_summary:
                parts.append(f"- 데이터 타입: {', '.join(type_summary)}")

        # 수식 정보
        if sheet_data.get("formulas"):
            parts.append(f"- 수식: {len(sheet_data['formulas'])}개")
            # 주요 수식 예시
            for formula in sheet_data["formulas"][:3]:
                parts.append(f"  · {formula['cell']}: {formula['formula'][:30]}...")

    # AI가 이해하기 쉬운 맥락 정보 추가
    parts.append("\n맥락 정보:")
    parts.append(
        f"- 이 파일은 {summary.get('total_sheets', 0)}개의 워크시트로 구성되어 있습니다."
    )
    if (
        detection_result
        and detection_result.get("summary", {}).get("total_errors", 0) > 0
    ):
        parts.append(
            f"- {detection_result['summary']['total_errors']}개의 오류가 발견되었으며, "
            f"가장 많은 오류 타입은 '{detection_result['summary'].get('most_common_type', 'unknown')}'입니다."
        )
    if summary.get("has_charts"):
        parts.append("- 차트가 포함되어 있습니다.")
    if summary.get("has_pivot_tables"):
        parts.append("- 피벗 테이블이 포함되어 있습니다.")

    return "\n".join(parts)


def _create_content_summary(analysis: Dict[str, Any]) -> str:
    """Create a comprehensive searchable summary of Excel file content with detailed context"""
    parts = []
    summary = analysis.get("summary", {})

    # File overview
    parts.append(
        f"Excel 파일 개요: {summary.get('total_sheets', 0)}개 시트, "
        f"{summary.get('total_rows', 0)}행, {summary.get('total_cells_with_data', 0)}개 데이터 셀"
    )

    # Error summary if present
    if summary.get("has_errors"):
        parts.append(f"\n총 {summary.get('total_errors', 0)}개 오류 발견:")
        if "errors" in analysis:
            error_types = {}
            for error in analysis["errors"][:10]:  # First 10 errors for context
                error_types[error["type"]] = error_types.get(error["type"], 0) + 1
            for error_type, count in error_types.items():
                parts.append(f"- {error_type}: {count}개")

    # Sheet details
    parts.append("\n시트별 상세:")
    for sheet_name, sheet_data in analysis.get("sheets", {}).items():
        parts.append(f"\n[{sheet_name}]")

        # Column information
        if "columns" in sheet_data:
            cols = sheet_data["columns"][:15]  # More columns for better context
            parts.append(f"열: {', '.join(cols)}")
            if len(sheet_data["columns"]) > 15:
                parts.append(f"... 외 {len(sheet_data['columns']) - 15}개 열")

        # Data type analysis
        if "column_analysis" in sheet_data:
            data_types = {}
            for col_info in sheet_data["column_analysis"]:
                dtype = col_info.get("data_type", "unknown")
                data_types[dtype] = data_types.get(dtype, 0) + 1
            if data_types:
                parts.append(
                    "데이터 타입: "
                    + ", ".join([f"{k}({v})" for k, v in data_types.items()])
                )

        # Formulas
        if sheet_data.get("formula_count", 0) > 0:
            parts.append(f"수식 {sheet_data['formula_count']}개 포함")
            if "complex_formulas" in sheet_data:
                parts.append(
                    f"복잡한 수식: {', '.join(sheet_data['complex_formulas'][:3])}"
                )

        # Data quality
        if "data_quality_issues" in sheet_data:
            parts.append(
                f"데이터 품질 이슈: {len(sheet_data['data_quality_issues'])}개"
            )

        # Key metrics if identified
        if "key_metrics" in sheet_data:
            parts.append(f"주요 지표: {', '.join(sheet_data['key_metrics'][:5])}")

    # Business context hints (inferred from column names and data)
    business_keywords = _infer_business_context(analysis)
    if business_keywords:
        parts.append(f"\n비즈니스 컨텍스트: {', '.join(business_keywords)}")

    return " ".join(parts)


def _infer_business_context(analysis: Dict[str, Any]) -> List[str]:
    """Infer business context from column names and data patterns"""
    keywords = set()
    business_terms = {
        "매출": ["sales", "revenue", "매출", "수익"],
        "비용": ["cost", "expense", "비용", "지출"],
        "이익": ["profit", "margin", "이익", "마진"],
        "재고": ["inventory", "stock", "재고", "입고", "출고"],
        "고객": ["customer", "client", "고객", "거래처"],
        "날짜": ["date", "time", "날짜", "일자", "년", "월"],
    }

    # Check all column names across sheets
    for sheet_data in analysis.get("sheets", {}).values():
        for col in sheet_data.get("columns", []):
            col_lower = col.lower()
            for category, terms in business_terms.items():
                if any(term in col_lower for term in terms):
                    keywords.add(category)

    return list(keywords)


@router.post("/detect-circular-references")
async def detect_circular_references(
    file_path: str, i18n: I18nContext = Depends(get_i18n_context)
) -> Dict[str, Any]:
    """
    Detect circular references in Excel file using advanced graph analysis
    """
    try:
        import openpyxl

        # Load workbook
        workbook = openpyxl.load_workbook(file_path, data_only=False)

        # Initialize detector
        detector = CircularReferenceDetector()

        # Analyze workbook
        circular_chains = detector.analyze_workbook(workbook)

        # Convert to JSON-serializable format
        results = []
        for chain in circular_chains:
            results.append(
                {
                    "cells": chain.cells,
                    "chain_type": chain.chain_type,
                    "severity": chain.severity,
                    "description": chain.description,
                    "break_suggestions": chain.break_suggestions,
                }
            )

        return {
            "circular_references": results,
            "total_found": len(results),
            "message": i18n.get_message(
                "circular_references_detected", count=len(results)
            ),
        }

    except Exception as e:
        logger.error(f"Circular reference detection failed: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=i18n.get_error_message("analysis_failed", error=str(e)),
        )


@router.post("/fix-formula")
async def fix_formula(
    formula: str, error_type: str, i18n: I18nContext = Depends(get_i18n_context)
) -> Dict[str, Any]:
    """
    Fix a formula error using FastFormulaFixer
    """
    try:
        # Initialize fixer
        fixer = FastFormulaFixer()

        # Map error description to error type
        error_mapping = {
            "#DIV/0!": "#DIV/0!",
            "#N/A": "#N/A",
            "#NAME?": "#NAME?",
            "#REF!": "#REF!",
            "#VALUE!": "#VALUE!",
            "#NUM!": "#NUM!",
            "#NULL!": "#NULL!",
            "#SPILL!": "#SPILL!",
            "#CALC!": "#CALC!",
        }

        # Find matching error type
        detected_error = None
        for error, code in error_mapping.items():
            if error in error_type:
                detected_error = code
                break

        if not detected_error:
            detected_error = "#VALUE!"  # Default

        # Fix formula
        result = fixer.fix_formula(formula, detected_error)

        return {
            "original_formula": formula,
            "fixed_formula": result.fixed_formula,
            "confidence": result.confidence,
            "explanation": result.explanation,
            "error_type": detected_error,
        }

    except Exception as e:
        logger.error(f"Formula fix failed: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=i18n.get_error_message("formula_fix_failed", error=str(e)),
        )


@router.post("/update-cell")
async def update_cell(
    file_path: str,
    location: str,
    value: str,
    i18n: I18nContext = Depends(get_i18n_context),
) -> Dict[str, Any]:
    """
    Update a specific cell in Excel file
    """
    try:
        import openpyxl

        # Load workbook
        workbook = openpyxl.load_workbook(file_path)

        # Parse location (e.g., "Sheet1!A1" or "A1")
        if "!" in location:
            sheet_name, cell_ref = location.split("!", 1)
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
            cell_ref = location

        # Update cell
        sheet[cell_ref] = value

        # Save workbook
        workbook.save(file_path)
        workbook.close()

        return {
            "success": True,
            "location": location,
            "new_value": value,
            "message": i18n.get_message("cell_updated", location=location),
        }

    except Exception as e:
        logger.error(f"Cell update failed: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=i18n.get_error_message("cell_update_failed", error=str(e)),
        )


@router.post("/remove-duplicates")
async def remove_duplicates(
    file_path: str, i18n: I18nContext = Depends(get_i18n_context)
) -> Dict[str, Any]:
    """
    Remove duplicate rows from Excel file
    """
    try:
        import pandas as pd

        # Read Excel file
        df = pd.read_excel(file_path)

        # Count duplicates before removal
        duplicates_before = df.duplicated().sum()

        # Remove duplicates
        df_cleaned = df.drop_duplicates()

        # Save back to Excel
        df_cleaned.to_excel(file_path, index=False)

        return {
            "success": True,
            "duplicates_removed": duplicates_before,
            "rows_before": len(df),
            "rows_after": len(df_cleaned),
            "message": i18n.get_message("duplicates_removed", count=duplicates_before),
        }

    except Exception as e:
        logger.error(f"Duplicate removal failed: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=i18n.get_error_message("duplicate_removal_failed", error=str(e)),
        )


@router.post("/optimize-formulas")
async def optimize_formulas(
    file_path: str, i18n: I18nContext = Depends(get_i18n_context)
) -> Dict[str, Any]:
    """
    Optimize formulas in Excel file for better performance
    """
    try:
        import openpyxl

        # Load workbook
        workbook = openpyxl.load_workbook(file_path, data_only=False)

        # Initialize fixer for optimization
        fixer = FastFormulaFixer()

        optimized_count = 0
        optimization_details = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f" and cell.value:
                        original_formula = cell.value

                        # Check if formula can be optimized
                        if fixer._is_inefficient_formula(original_formula):
                            optimized = fixer._optimize_formula(original_formula)

                            if optimized != original_formula:
                                cell.value = optimized
                                optimized_count += 1

                                optimization_details.append(
                                    {
                                        "location": f"{sheet_name}!{cell.coordinate}",
                                        "original": original_formula,
                                        "optimized": optimized,
                                    }
                                )

        # Save workbook
        workbook.save(file_path)
        workbook.close()

        return {
            "success": True,
            "formulas_optimized": optimized_count,
            "details": optimization_details[:10],  # First 10 optimizations
            "message": i18n.get_message("formulas_optimized", count=optimized_count),
        }

    except Exception as e:
        logger.error(f"Formula optimization failed: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=i18n.get_error_message("optimization_failed", error=str(e)),
        )


class CellVerificationRequest(BaseModel):
    file_id: str
    cell_address: str


@router.post("/validate-and-report")
async def validate_and_report(
    file: UploadFile = File(...),
    session_id: str = None,
    user_id: str = None,
    db: AsyncSession = Depends(get_db),
    i18n: I18nContext = Depends(get_i18n_context),
) -> Dict[str, Any]:
    """
    사용자가 검증 버튼 클릭 시 Excel 파일을 전체 검증하고 AI 채팅에 보고
    IntegratedErrorDetector를 사용하여 포괄적인 오류 감지 수행
    """
    # 파일 타입 검증
    if not any(file.filename.endswith(ext) for ext in settings.ALLOWED_EXTENSIONS):
        error_message = i18n.get_error_message(
            "invalid_file_type", extensions=", ".join(settings.ALLOWED_EXTENSIONS)
        )
        raise HTTPException(status_code=400, detail=error_message)

    # 파일 크기 검증
    if file.size > settings.MAX_UPLOAD_SIZE:
        error_message = i18n.get_error_message(
            "file_too_large", max_size=settings.MAX_UPLOAD_SIZE
        )
        raise HTTPException(status_code=400, detail=error_message)

    # 임시 파일 저장
    with tempfile.NamedTemporaryFile(
        delete=False, suffix=os.path.splitext(file.filename)[1]
    ) as tmp_file:
        content = await file.read()
        tmp_file.write(content)
        tmp_path = tmp_file.name

    try:
        # IntegratedErrorDetector 사용하여 전체 검증
        from app.services.detection.integrated_error_detector import (
            IntegratedErrorDetector,
        )
        from app.services.error_classification_service import ErrorClassificationService
        from app.services.fix_recommendation_service import FixRecommendationService

        logger.info(f"Starting validation for file: {file.filename}")

        # 1. 오류 감지
        detector = IntegratedErrorDetector()
        detection_result = await detector.detect_all_errors(tmp_path)
        errors = detection_result["errors"]

        # 2. 오류 분류 및 우선순위 지정
        classifier = ErrorClassificationService()
        classified_errors = classifier.classify_errors(errors)
        prioritized_errors = classifier.prioritize_errors(errors)
        error_summary = classifier.get_error_summary(errors)

        # 3. 수정 제안 생성
        recommender = FixRecommendationService()
        recommendations = recommender.batch_recommendations(prioritized_errors)

        # 4. AI 보고서 생성
        validation_report = await _generate_validation_report(
            filename=file.filename,
            errors=prioritized_errors,
            error_summary=error_summary,
            classified_errors=classified_errors,
            recommendations=recommendations,
            i18n=i18n,
        )

        # 5. 결과 반환 (AI 채팅에 전달될 데이터)
        response = {
            "status": "success",
            "message": i18n.get_text("validation.complete"),
            "filename": file.filename,
            "validation_results": {
                "total_errors": len(errors),
                "error_summary": error_summary,
                "classified_errors": {
                    category: len(errors_list)
                    for category, errors_list in classified_errors.items()
                },
                "critical_errors": [
                    {
                        "id": error.id,
                        "type": error.type,
                        "location": f"{error.sheet}!{error.cell}",
                        "message": error.message,
                    }
                    for error in prioritized_errors[:5]  # 상위 5개 중요 오류
                    if error.severity == "critical"
                ],
                "auto_fixable_count": error_summary.get("auto_fixable", 0),
            },
            "ai_report": validation_report,
            "session_id": session_id,
            "user_id": user_id,
            "timestamp": datetime.utcnow().isoformat(),
        }

        # 6. 벡터 DB에 검증 결과 저장
        await vector_search_service.index_document(
            document_id=f"{file.filename}_validation_{session_id}",
            document_type="excel_validation",
            content=validation_report,
            metadata={
                "filename": file.filename,
                "total_errors": len(errors),
                "critical_errors": error_summary.get("critical_count", 0),
                "session_id": session_id,
                "user_id": user_id,
            },
            db=db,
        )

        # 7. AI 채팅에 보고서 전송
        if session_id and response.get("ai_report"):
            await _report_to_ai_chat(
                session_id=session_id,
                user_id=user_id,
                report=response["ai_report"],
                error_count=len(errors),
                filename=file.filename,
                db=db,
            )

        logger.info(
            f"Validation complete for {file.filename}: {len(errors)} errors found and reported to AI chat"
        )

        return response

    except Exception as e:
        logger.error(f"Error in validate-and-report: {str(e)}")
        error_message = i18n.get_error_message("validation_failed", error=str(e))
        raise HTTPException(status_code=500, detail=error_message)
    finally:
        # 임시 파일 정리
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


async def _generate_validation_report(
    filename: str,
    errors: List[Any],
    error_summary: Dict[str, Any],
    classified_errors: Dict[str, List[Any]],
    recommendations: Dict[str, Any],
    i18n: I18nContext,
) -> str:
    """
    AI가 채팅창에 보고할 검증 결과 생성
    """
    # OpenRouter를 사용하여 보고서 생성
    from app.services.openai_service import openai_service

    # 보고서 컨텍스트 준비
    context = {
        "filename": filename,
        "total_errors": len(errors),
        "error_summary": error_summary,
        "categories": list(classified_errors.keys()),
        "critical_errors": sum(1 for e in errors if e.severity == "critical"),
        "auto_fixable": error_summary.get("auto_fixable", 0),
        "language": i18n.language,
    }

    # 주요 오류 설명
    critical_errors_desc = []
    for error in errors[:10]:  # 상위 10개
        if error.severity in ["critical", "high"]:
            critical_errors_desc.append(
                f"- {error.type} at {error.sheet}!{error.cell}: {error.message}"
            )

    # AI에게 보고서 생성 요청
    prompt = f"""
    Excel 파일 '{filename}'의 검증이 완료되었습니다.
    사용자에게 친근하고 이해하기 쉬운 보고서를 작성해주세요.

    검증 결과:
    - 총 {context['total_errors']}개의 오류 발견
    - 중요 오류: {context['critical_errors']}개
    - 자동 수정 가능: {context['auto_fixable']}개

    오류 카테고리: {', '.join(context['categories'])}

    주요 오류:
    {chr(10).join(critical_errors_desc[:5])}

    보고서는 다음 구조로 작성해주세요:
    1. 전체 요약
    2. 주요 발견사항
    3. 권장 조치사항
    4. 다음 단계

    사용자가 쉽게 이해할 수 있도록 기술적인 용어는 피하고,
    구체적인 개선 방법을 제시해주세요.
    언어: {i18n.language}
    """

    try:
        report = await openai_service.generate_excel_solution(prompt, context)
        return report
    except Exception as e:
        # OpenRouter 실패 시 기본 보고서 생성
        logger.warning(f"AI report generation failed: {e}")
        return _generate_fallback_report(context, critical_errors_desc, i18n)


async def _report_to_ai_chat(
    session_id: str,
    user_id: str,
    report: str,
    error_count: int,
    filename: str,
    db: AsyncSession,
) -> None:
    """
    AI 채팅에 검증 보고서를 전송
    """
    try:
        # Rails API를 통해 AI 채팅에 메시지 추가
        import httpx

        rails_api_url = settings.RAILS_API_URL
        rails_api_key = settings.RAILS_INTERNAL_API_KEY

        async with httpx.AsyncClient() as client:
            # ChatMessage 생성 요청
            response = await client.post(
                f"{rails_api_url}/api/v1/ai_consultation/chat_messages",
                headers={
                    "Authorization": f"Bearer {rails_api_key}",
                    "Content-Type": "application/json",
                },
                json={
                    "session_id": session_id,
                    "user_id": user_id,
                    "role": "assistant",
                    "content": report,
                    "metadata": {
                        "type": "validation_report",
                        "filename": filename,
                        "error_count": error_count,
                        "generated_by": "excel_validator",
                    },
                },
            )

            if response.status_code != 200:
                logger.error(
                    f"Failed to send report to AI chat: {response.status_code} - {response.text}"
                )
            else:
                logger.info(
                    f"Successfully sent validation report to AI chat for session {session_id}"
                )

    except Exception as e:
        logger.error(f"Error sending report to AI chat: {str(e)}")
        # Don't raise exception - this is a non-critical failure


def _generate_fallback_report(
    context: Dict[str, Any], critical_errors: List[str], i18n: I18nContext
) -> str:
    """
    AI 서비스 실패 시 기본 보고서 생성
    """
    if i18n.language == "ko":
        return f"""
📊 Excel 파일 검증 결과

**파일명**: {context['filename']}

**전체 요약**
검증 결과 총 {context['total_errors']}개의 오류가 발견되었습니다.
이 중 {context['critical_errors']}개는 즉시 수정이 필요한 중요 오류이며,
{context['auto_fixable']}개는 자동으로 수정 가능합니다.

**주요 발견사항**
{chr(10).join(critical_errors[:5])}

**권장 조치사항**
1. 중요 오류부터 우선적으로 수정하세요
2. 자동 수정 기능을 활용하여 빠르게 해결 가능한 오류를 처리하세요
3. 수식 오류는 참조 범위를 확인하고 수정하세요
4. 데이터 유효성 검사를 추가하여 향후 오류를 방지하세요

**다음 단계**
- "자동 수정" 버튼을 클릭하여 수정 가능한 오류를 해결하세요
- 개별 오류를 클릭하여 상세 정보와 수정 방법을 확인하세요
- 수정 후 다시 검증하여 모든 오류가 해결되었는지 확인하세요
"""
    else:
        return f"""
📊 Excel File Validation Results

**Filename**: {context['filename']}

**Summary**
Validation found {context['total_errors']} total errors.
{context['critical_errors']} are critical errors requiring immediate attention,
and {context['auto_fixable']} can be automatically fixed.

**Key Findings**
{chr(10).join(critical_errors[:5])}

**Recommended Actions**
1. Priority fix critical errors first
2. Use auto-fix feature for quickly resolvable errors
3. Check and correct formula references
4. Add data validation to prevent future errors

**Next Steps**
- Click "Auto Fix" to resolve fixable errors
- Click individual errors for details and fix methods
- Re-validate after fixes to ensure all errors are resolved
"""


@router.post("/verify-cell-position")
async def verify_cell_position(
    request: CellVerificationRequest,
    db: AsyncSession = Depends(get_db),
    i18n: I18nContext = Depends(get_i18n_context),
) -> Dict[str, Any]:
    """
    Verify cell position by analyzing the same cell in Python
    Used for testing coordinate consistency between ExcelJS and Python
    """
    try:
        # TODO: 실제 파일 경로를 file_id로부터 가져오는 로직 구현
        # 현재는 테스트용으로 하드코딩
        if request.file_id == "36":
            file_path = "/Users/kevin/Desktop/사고조사.xlsx"
        else:
            # 기본 테스트 파일 사용
            file_path = "/Users/kevin/Desktop/사고조사.xlsx"

        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail=f"File not found: {file_path}")

        # openpyxl로 Excel 파일 읽기
        import openpyxl

        workbook = openpyxl.load_workbook(file_path, data_only=True)
        worksheet = workbook.active

        # 셀 주소를 좌표로 변환 (예: "C3" -> row=3, col=3)
        try:
            cell = worksheet[request.cell_address]
            python_address = cell.coordinate
            python_value = cell.value
            python_row = cell.row
            python_col = cell.column

            # 셀 타입 판단
            if python_value is None:
                cell_type = "empty"
            elif isinstance(python_value, (int, float)):
                cell_type = "number"
            elif isinstance(python_value, str):
                cell_type = "text"
            else:
                cell_type = "other"

            workbook.close()

            return {
                "success": True,
                "address": python_address,
                "value": python_value,
                "row": python_row,
                "column": python_col,
                "type": cell_type,
                "requested_address": request.cell_address,
                "match": request.cell_address == python_address,
            }

        except Exception as cell_error:
            workbook.close()
            return {
                "success": False,
                "error": f"Cell access error: {str(cell_error)}",
                "address": "ERROR",
                "value": None,
                "type": "error",
                "requested_address": request.cell_address,
                "match": False,
            }

    except Exception as e:
        logger.error(f"Cell position verification failed: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Verification failed: {str(e)}")
