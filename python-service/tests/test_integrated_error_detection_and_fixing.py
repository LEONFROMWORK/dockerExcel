"""
통합 테스트 시나리오 - 오류 감지 및 수정
WebSocket 진행 상황 보고 포함
"""

import pytest
from unittest.mock import Mock, AsyncMock
from app.services.detection.integrated_error_detector import IntegratedErrorDetector
from app.services.fixing.integrated_error_fixer import IntegratedErrorFixer
from app.websocket.progress_reporter import WebSocketProgressReporter
from app.core.interfaces import ExcelError, ExcelErrorType, ProcessingTier
import tempfile
import os
from openpyxl import Workbook


class TestIntegratedErrorDetectionAndFixing:
    """통합 오류 감지 및 수정 테스트"""

    @pytest.fixture
    def mock_websocket_manager(self):
        """Mock WebSocket Manager"""
        manager = Mock()
        manager.send_message = AsyncMock()
        return manager

    @pytest.fixture
    def progress_reporter(self, mock_websocket_manager):
        """WebSocket Progress Reporter"""
        reporter = WebSocketProgressReporter("test_session_123")
        reporter.manager = mock_websocket_manager
        return reporter

    @pytest.fixture
    def detector_with_progress(self, progress_reporter):
        """Progress Reporter가 포함된 Error Detector"""
        return IntegratedErrorDetector(progress_reporter=progress_reporter)

    @pytest.fixture
    def fixer_with_progress(self, progress_reporter):
        """Progress Reporter가 포함된 Error Fixer"""
        return IntegratedErrorFixer(progress_reporter=progress_reporter)

    @pytest.fixture
    def sample_excel_file(self):
        """오류가 포함된 샘플 Excel 파일 생성"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # 다양한 오류 시나리오 설정
        ws["A1"] = "=B1/C1"  # DIV/0 오류 (C1이 0)
        ws["B1"] = 10
        ws["C1"] = 0

        ws["A2"] = "=VLOOKUP(D2,E2:F10,2,FALSE)"  # #N/A 오류
        ws["D2"] = "NotFound"

        ws["A3"] = "=UNKNOWNFUNC()"  # #NAME? 오류

        ws["A4"] = "=Sheet2!A1"  # #REF! 오류 (Sheet2 없음)

        ws["A5"] = '=IF(A1>0,"Text",A1+B1)'  # #VALUE! 오류 가능성

        ws["A6"] = "=A7"  # 순환 참조
        ws["A7"] = "=A6"

        # 임시 파일로 저장
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            return tmp.name

    @pytest.mark.asyncio
    async def test_full_detection_and_fixing_workflow(
        self,
        detector_with_progress,
        fixer_with_progress,
        sample_excel_file,
        mock_websocket_manager,
    ):
        """전체 워크플로우 테스트: 감지 → 수정 → WebSocket 보고"""

        # 1. 오류 감지
        print("\n=== 1단계: 오류 감지 시작 ===")
        detection_result = await detector_with_progress.detect_all_errors(
            sample_excel_file
        )

        # WebSocket으로 진행 상황이 보고되었는지 확인
        assert mock_websocket_manager.send_message.called
        progress_calls = [
            call
            for call in mock_websocket_manager.send_message.call_args_list
            if call[0][1].get("type") == "progress"
        ]
        assert len(progress_calls) > 0

        # 감지 결과 검증
        assert detection_result["status"] == "success"
        assert len(detection_result["errors"]) >= 6  # 최소 6개 오류 예상

        errors = detection_result["errors"]
        error_types = {error["type"] for error in errors}

        print(f"감지된 오류 수: {len(errors)}")
        print(f"오류 타입: {error_types}")

        # 2. 오류별 수정 시도
        print("\n=== 2단계: 오류 수정 시작 ===")
        fixed_count = 0
        failed_count = 0

        for error_dict in errors:
            # dict를 ExcelError 객체로 변환
            error = ExcelError(
                id=error_dict["id"],
                type=error_dict["type"],
                severity=error_dict["severity"],
                message=error_dict["message"],
                cell=error_dict["cell"],
                sheet=error_dict["sheet"],
                formula=error_dict.get("formula"),
                value=error_dict.get("value"),
                is_auto_fixable=error_dict.get("is_auto_fixable", False),
            )

            if fixer_with_progress.can_fix(error):
                print(f"\n수정 시도: {error.type} at {error.cell}")
                fix_result = await fixer_with_progress.fix(error)

                if fix_result.success:
                    fixed_count += 1
                    print(f"✓ 수정 성공: {fix_result.fixed_formula}")
                else:
                    failed_count += 1
                    print(f"✗ 수정 실패: {fix_result.message}")
            else:
                print(f"\n수정 불가: {error.type} at {error.cell}")

        print(f"\n수정 결과: 성공 {fixed_count}, 실패 {failed_count}")

        # WebSocket 수정 보고 확인
        fix_reports = [
            call
            for call in mock_websocket_manager.send_message.call_args_list
            if "error_fixed" in str(call)
        ]
        assert len(fix_reports) >= fixed_count

        # 3. 일괄 수정 테스트
        print("\n=== 3단계: 일괄 수정 테스트 ===")

        # ExcelError 객체 리스트 생성
        error_objects = []
        for error_dict in errors[:3]:  # 처음 3개만 테스트
            error_objects.append(
                ExcelError(
                    id=error_dict["id"],
                    type=error_dict["type"],
                    severity=error_dict["severity"],
                    message=error_dict["message"],
                    cell=error_dict["cell"],
                    sheet=error_dict["sheet"],
                    formula=error_dict.get("formula"),
                    value=error_dict.get("value"),
                    is_auto_fixable=error_dict.get("is_auto_fixable", False),
                )
            )

        batch_result = await fixer_with_progress.fix_batch(
            error_objects, strategy="safe"
        )

        assert batch_result["total"] == 3
        assert batch_result["success"] >= 0
        assert batch_result["failed"] >= 0
        assert (
            batch_result["success"]
            + batch_result["failed"]
            + batch_result.get("skipped", 0)
            == 3
        )

        print(f"일괄 수정 결과: {batch_result}")

        # 정리
        os.unlink(sample_excel_file)

    @pytest.mark.asyncio
    async def test_websocket_progress_reporting(
        self, detector_with_progress, sample_excel_file, mock_websocket_manager
    ):
        """WebSocket 진행 상황 보고 테스트"""

        # 오류 감지 실행
        await detector_with_progress.detect_all_errors(sample_excel_file)

        # WebSocket 메시지 분석
        all_messages = mock_websocket_manager.send_message.call_args_list
        message_types = [call[0][1].get("type") for call in all_messages]

        # 필수 메시지 타입 확인
        assert "task_start" in message_types
        assert "progress" in message_types
        assert "task_complete" in message_types

        # 진행률 메시지 검증
        progress_messages = [
            call[0][1] for call in all_messages if call[0][1].get("type") == "progress"
        ]

        for msg in progress_messages:
            assert "current" in msg["data"]
            assert "total" in msg["data"]
            assert "percentage" in msg["data"]
            assert msg["data"]["percentage"] >= 0
            assert msg["data"]["percentage"] <= 100

        # 정리
        os.unlink(sample_excel_file)

    @pytest.mark.asyncio
    async def test_multi_cell_error_detection(self, detector_with_progress):
        """멀티 셀 오류 감지 테스트"""

        # 테스트용 셀 데이터
        cells = [
            {
                "sheet": "Sheet1",
                "address": "A1",
                "value": "#DIV/0!",
                "formula": "=B1/C1",
            },
            {
                "sheet": "Sheet1",
                "address": "A2",
                "value": "#N/A",
                "formula": "=VLOOKUP(D2,E2:F10,2,FALSE)",
            },
            {
                "sheet": "Sheet1",
                "address": "A3",
                "value": "#NAME?",
                "formula": "=UNKNOWNFUNC()",
            },
        ]

        # Mock workbook 생성
        wb = Workbook()
        wb.active

        # 임시 파일로 저장
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)

            # 멀티 셀 분석
            result = await detector_with_progress.detect_multi_cell_errors(
                tmp.name, cells
            )

            assert "individual_cells" in result
            assert "pattern_analysis" in result
            assert "cross_cell_issues" in result
            assert result["total_errors"] >= 0

            # 개별 셀 결과 확인
            for cell_result in result["individual_cells"]:
                assert "address" in cell_result
                assert "sheet" in cell_result
                assert "errors" in cell_result or "error" in cell_result

            os.unlink(tmp.name)

    @pytest.mark.asyncio
    async def test_error_fix_strategies(self, fixer_with_progress):
        """각 오류 타입별 수정 전략 테스트"""

        # DIV/0 오류
        div_zero_error = ExcelError(
            id="test_1",
            type=ExcelErrorType.DIV_ZERO,
            severity="high",
            message="Division by zero",
            cell="A1",
            sheet="Sheet1",
            formula="=B1/C1",
            is_auto_fixable=True,
        )

        result = await fixer_with_progress.fix(div_zero_error)
        assert result.success
        assert "IFERROR" in result.fixed_formula or "IF" in result.fixed_formula

        # #N/A 오류
        na_error = ExcelError(
            id="test_2",
            type=ExcelErrorType.NA,
            severity="medium",
            message="#N/A error",
            cell="A2",
            sheet="Sheet1",
            formula="=VLOOKUP(D2,E2:F10,2,FALSE)",
            is_auto_fixable=True,
        )

        result = await fixer_with_progress.fix(na_error)
        assert result.success
        assert "IFNA" in result.fixed_formula or "IFERROR" in result.fixed_formula

        # #NAME? 오류
        name_error = ExcelError(
            id="test_3",
            type=ExcelErrorType.NAME,
            severity="high",
            message="#NAME? error",
            cell="A3",
            sheet="Sheet1",
            formula="=UNKNOWNFUNC()",
            is_auto_fixable=True,
        )

        result = await fixer_with_progress.fix(name_error)
        # NAME 오류는 자동 수정이 어려울 수 있음
        if result.success:
            assert result.fixed_formula != name_error.formula

    @pytest.mark.asyncio
    async def test_caching_performance(self, detector_with_progress, sample_excel_file):
        """캐싱 성능 테스트"""
        import time

        # 첫 번째 실행 (캐시 없음)
        start_time = time.time()
        result1 = await detector_with_progress.detect_all_errors(sample_excel_file)
        first_run_time = time.time() - start_time

        # 두 번째 실행 (캐시 사용)
        start_time = time.time()
        result2 = await detector_with_progress.detect_all_errors(sample_excel_file)
        second_run_time = time.time() - start_time

        # 캐시가 작동하면 두 번째 실행이 더 빨라야 함
        assert second_run_time < first_run_time

        # 결과가 동일해야 함
        assert len(result1["errors"]) == len(result2["errors"])
        assert result2.get("tier_used") == ProcessingTier.CACHE.value

        # 정리
        detector_with_progress.clear_cache()
        os.unlink(sample_excel_file)


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
