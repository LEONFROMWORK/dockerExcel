#\!/usr/bin/env python3
import requests
import os

# Create a simple Excel file for testing
import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 'Test'
ws['B1'] = 'Excel'
ws['A2'] = 'File'
ws['B2'] = '2024'

# Save the file
test_file = 'test_excel.xlsx'
wb.save(test_file)

# Upload the file
url = 'http://localhost:3000/api/v1/excel_analysis/files'

with open(test_file, 'rb') as f:
    files = {'file': (test_file, f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
    response = requests.post(url, files=files)
    
    print(f"Status Code: {response.status_code}")
    print(f"Response: {response.json()}")
    
# Clean up
os.remove(test_file)
EOF < /dev/null