"""
Simple Excel analysis API endpoint
"""
from fastapi import APIRouter, File, UploadFile, HTTPException
import tempfile
import os
import openpyxl
import logging

logger = logging.getLogger(__name__)

router = APIRouter()

@router.post("/analyze")
async def analyze_excel_simple(file: UploadFile = File(...)):
    """
    Simple Excel analysis that returns cell errors
    """
    # Validate file type
    if not file.filename.endswith(('.xlsx', '.xls', '.xlsm')):
        raise HTTPException(status_code=400, detail="Invalid file type")
    
    # Save uploaded file temporarily
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        content = await file.read()
        tmp_file.write(content)
        tmp_path = tmp_file.name
    
    try:
        workbook = openpyxl.load_workbook(tmp_path, data_only=False)
        errors = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Check for errors in cells
            for row in sheet.iter_rows(max_row=100):  # First 100 rows
                for cell in row:
                    if cell.value is not None:
                        cell_value = str(cell.value)
                        
                        # Check for Excel error values
                        error_patterns = {
                            '#DIV/0!': 'Division by zero error',
                            '#N/A': 'Value not available error', 
                            '#NAME?': 'Unrecognized formula name',
                            '#NULL!': 'Null intersection error',
                            '#NUM!': 'Invalid numeric value',
                            '#REF!': 'Invalid cell reference',
                            '#VALUE!': 'Wrong value type'
                        }
                        
                        for error_code, description in error_patterns.items():
                            if error_code in cell_value:
                                errors.append({
                                    "cell": cell.coordinate,
                                    "sheet": sheet_name,
                                    "error_type": error_code,
                                    "message": description,
                                    "severity": "high",
                                    "auto_fixable": False
                                })
        
        workbook.close()
        
        return {
            "errors": errors,
            "summary": {
                "total_errors": len(errors),
                "has_errors": len(errors) > 0
            }
        }
        
    except Exception as e:
        logger.error(f"Error analyzing Excel file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Analysis failed: {str(e)}")
    finally:
        # Clean up temporary file
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)