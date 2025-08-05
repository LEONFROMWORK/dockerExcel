"""
Test to see exact response structure
"""
import requests
from openpyxl import Workbook
import tempfile
import os
import json

# Create test Excel file
wb = Workbook()
ws = wb.active
ws['A1'] = "Test"
ws['B1'] = 10
ws['C1'] = 0
ws['A2'] = "=B1/C1"

with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
    wb.save(tmp.name)
    test_file = tmp.name

try:
    url = "http://localhost:8000/api/v1/excel-error-analysis/analyze-upload"

    with open(test_file, 'rb') as f:
        files = {'file': ('test.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {'session_id': 'test-123'}

        response = requests.post(url, files=files, data=data)

    print("=== RAW RESPONSE ===")
    print(json.dumps(response.json(), indent=2))

finally:
    if os.path.exists(test_file):
        os.unlink(test_file)
