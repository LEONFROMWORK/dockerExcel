"""
Centralized style management for Excel generation
Provides consistent styling across all generated Excel files
"""

from typing import Dict, Any, Optional, List
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from dataclasses import dataclass
import json
from pathlib import Path
import logging

from .style_converters import StyleConverter

logger = logging.getLogger(__name__)


@dataclass
class ColorScheme:
    """Color scheme for Excel styling"""
    primary: str = "1F4788"
    secondary: str = "366092"
    accent: str = "4472C4"
    success: str = "70AD47"
    warning: str = "FFC000"
    error: str = "C00000"
    light_gray: str = "F2F2F2"
    medium_gray: str = "D9D9D9"
    dark_gray: str = "595959"


class StyleManager:
    """Manages Excel styles for consistent formatting"""
    
    def __init__(self, color_scheme: Optional[ColorScheme] = None):
        self.colors = color_scheme or ColorScheme()
        self._init_named_styles()
        self._init_style_presets()
    
    def _init_named_styles(self):
        """Initialize reusable named styles"""
        self.named_styles = {
            "title": NamedStyle(
                name="AI_Title",
                font=Font(bold=True, size=16, color="FFFFFF"),
                fill=PatternFill(start_color=self.colors.primary, end_color=self.colors.primary, fill_type="solid"),
                alignment=Alignment(horizontal="center", vertical="center")
            ),
            "header": NamedStyle(
                name="AI_Header",
                font=Font(bold=True, size=11, color="FFFFFF"),
                fill=PatternFill(start_color=self.colors.secondary, end_color=self.colors.secondary, fill_type="solid"),
                alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                border=Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            ),
            "subheader": NamedStyle(
                name="AI_Subheader",
                font=Font(bold=True, size=10),
                fill=PatternFill(start_color=self.colors.light_gray, end_color=self.colors.light_gray, fill_type="solid"),
                alignment=Alignment(horizontal="left", vertical="center")
            ),
            "currency": NamedStyle(
                name="AI_Currency",
                number_format="#,##0원",
                alignment=Alignment(horizontal="right")
            ),
            "percentage": NamedStyle(
                name="AI_Percentage",
                number_format="0.0%",
                alignment=Alignment(horizontal="center")
            ),
            "date": NamedStyle(
                name="AI_Date",
                number_format="YYYY-MM-DD",
                alignment=Alignment(horizontal="center")
            )
        }
    
    def _init_style_presets(self):
        """Initialize style presets from configuration file"""
        config_path = Path(__file__).parent / "configs" / "style_presets.json"
        
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                presets_data = json.load(f)
            
            # Use StyleConverter for better modularity
            converter = StyleConverter(self.colors)
            self.presets = converter.convert_all_styles(
                presets_data, self._get_thin_border
            )
        except Exception as e:
            logger.warning(f"Failed to load style presets from config: {e}")
            # Fallback to basic presets
            self.presets = self._get_fallback_presets()
    
    # _convert_json_to_styles method removed - now using StyleConverter
    
    def _get_fallback_presets(self) -> Dict[str, Dict[str, Any]]:
        """Get minimal fallback presets"""
        return {
            "default": {
                "title": {
                    "font": Font(bold=True, size=14),
                    "fill": PatternFill(start_color=self.colors.primary, end_color=self.colors.primary, fill_type="solid"),
                    "alignment": Alignment(horizontal="center", vertical="center")
                },
                "header": {
                    "font": Font(bold=True),
                    "fill": PatternFill(start_color=self.colors.light_gray, end_color=self.colors.light_gray, fill_type="solid"),
                    "border": self._get_thin_border()
                }
            }
        }
    
    def get_style(self, domain: str, style_type: str) -> Dict[str, Any]:
        """Get style dictionary for a specific domain and type"""
        if domain in self.presets and style_type in self.presets[domain]:
            return self.presets[domain][style_type]
        
        # Return default style if not found
        return self._get_default_style(style_type)
    
    def _get_default_style(self, style_type: str) -> Dict[str, Any]:
        """Get default style for a given type"""
        defaults = {
            "title": {
                "font": Font(bold=True, size=14),
                "fill": PatternFill(start_color=self.colors.primary, end_color=self.colors.primary, fill_type="solid"),
                "alignment": Alignment(horizontal="center", vertical="center")
            },
            "header": {
                "font": Font(bold=True),
                "fill": PatternFill(start_color=self.colors.light_gray, end_color=self.colors.light_gray, fill_type="solid"),
                "alignment": Alignment(horizontal="center", vertical="center"),
                "border": self._get_thin_border()
            },
            "data": {
                "font": Font(size=10),
                "alignment": Alignment(horizontal="left", vertical="center")
            }
        }
        return defaults.get(style_type, {})
    
    def _get_thin_border(self) -> Border:
        """Get thin border style"""
        thin = Side(style='thin')
        return Border(left=thin, right=thin, top=thin, bottom=thin)
    
    def get_number_format(self, data_type: str, locale: str = "ko") -> str:
        """Get number format based on data type and locale"""
        formats = {
            "ko": {
                "currency": "#,##0원",
                "percentage": "0.0%",
                "date": "YYYY-MM-DD",
                "datetime": "YYYY-MM-DD HH:MM",
                "number": "#,##0",
                "decimal": "#,##0.00"
            },
            "en": {
                "currency": "$#,##0.00",
                "percentage": "0.0%",
                "date": "MM/DD/YYYY",
                "datetime": "MM/DD/YYYY HH:MM",
                "number": "#,##0",
                "decimal": "#,##0.00"
            }
        }
        
        locale_formats = formats.get(locale, formats["ko"])
        return locale_formats.get(data_type, "General")
    
    def get_chart_colors(self, chart_type: str) -> List[str]:
        """Get color palette for charts"""
        palettes = {
            "default": ["4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5", "70AD47"],
            "financial": ["0B5394", "1C4587", "3D85C6", "6FA8DC", "9FC5E8", "CFE2F3"],
            "sales": ["70AD47", "548235", "375623", "A9D08E", "C6EFCE", "E2EFDA"],
            "warm": ["C00000", "E26B0A", "FFC000", "FFEB9C", "FFF2CC", "FCE4D6"]
        }
        return palettes.get(chart_type, palettes["default"])
    
    def create_gradient_fill(self, start_color: str, end_color: str) -> PatternFill:
        """Create gradient fill (Note: Excel doesn't support true gradients, this returns solid fill)"""
        # Excel doesn't support gradient fills in openpyxl, return solid fill with start color
        return PatternFill(start_color=start_color, end_color=end_color, fill_type="solid")