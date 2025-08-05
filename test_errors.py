import openpyxl
from openpyxl import Workbook

# 오류가 있는 Excel 파일 생성
wb = Workbook()
ws = wb.active

# 정상적인 데이터
ws['A1'] = '항목'
ws['B1'] = '수량'
ws['C1'] = '가격'
ws['D1'] = '합계'

ws['A2'] = '사과'
ws['B2'] = 10
ws['C2'] = 1000

# 오류 1: 잘못된 수식 (괄호 없음)
ws['D2'] = '=B2*C2'
ws['D3'] = '=SUM(B2:B5'  # 괄호가 닫히지 않음

# 오류 2: #REF! 오류
ws['E2'] = '=A10*B10'  # 존재하지 않는 셀 참조
ws['E3'] = '=#REF!'

# 오류 3: 순환 참조
ws['F2'] = '=F2+1'

# 오류 4: #DIV/0! 오류
ws['G2'] = '=B2/0'

# 파일 저장
wb.save('test_excel_errors.xlsx')
print("오류가 포함된 Excel 파일이 생성되었습니다: test_excel_errors.xlsx")
