"""
도메인별 특화 Excel 생성기
Domain-Specific Excel Generators with Industry Best Practices
"""

import json
import logging
from typing import Dict, List, Any, Optional
from datetime import datetime, timedelta
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule
from openpyxl.worksheet.datavalidation import DataValidation
import tempfile
import os

from .openai_service import openai_service

logger = logging.getLogger(__name__)


class DomainExcelGenerator:
    """도메인별 Excel 생성기 기본 클래스"""
    
    def __init__(self):
        self._create_named_styles()
    
    def _create_named_styles(self):
        """재사용 가능한 명명된 스타일 생성"""
        self.styles = {
            "title": NamedStyle(
                name="title",
                font=Font(bold=True, size=16, color="FFFFFF"),
                fill=PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid"),
                alignment=Alignment(horizontal="center", vertical="center")
            ),
            "header": NamedStyle(
                name="header",
                font=Font(bold=True, size=11, color="FFFFFF"),
                fill=PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
                alignment=Alignment(horizontal="center", vertical="center", wrap_text=True),
                border=Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            ),
            "currency": NamedStyle(
                name="currency",
                number_format="#,##0원",
                alignment=Alignment(horizontal="right")
            ),
            "percentage": NamedStyle(
                name="percentage",
                number_format="0.0%",
                alignment=Alignment(horizontal="center")
            ),
            "date": NamedStyle(
                name="date",
                number_format="YYYY-MM-DD",
                alignment=Alignment(horizontal="center")
            )
        }


class FinanceExcelGenerator(DomainExcelGenerator):
    """재무/회계 특화 Excel 생성기"""
    
    async def generate_financial_statements(
        self,
        company_name: str,
        period: str,
        data: Optional[Dict[str, Any]] = None
    ) -> str:
        """재무제표 생성"""
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # 1. 손익계산서
        income_stmt = wb.create_sheet("손익계산서")
        await self._create_income_statement(income_stmt, company_name, period, data)
        
        # 2. 재무상태표
        balance_sheet = wb.create_sheet("재무상태표")
        await self._create_balance_sheet(balance_sheet, company_name, period, data)
        
        # 3. 현금흐름표
        cash_flow = wb.create_sheet("현금흐름표")
        await self._create_cash_flow_statement(cash_flow, company_name, period, data)
        
        # 4. 재무비율 분석
        ratio_analysis = wb.create_sheet("재무비율분석")
        await self._create_ratio_analysis(ratio_analysis, data)
        
        # 5. 대시보드
        dashboard = wb.create_sheet("대시보드", 0)  # 첫 번째 시트로
        await self._create_financial_dashboard(dashboard, data)
        
        # 파일 저장
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"financial_statements_{company_name}_{timestamp}.xlsx"
        output_path = os.path.join(tempfile.gettempdir(), filename)
        wb.save(output_path)
        
        return output_path
    
    async def _create_income_statement(
        self,
        ws,
        company_name: str,
        period: str,
        data: Optional[Dict[str, Any]]
    ):
        """손익계산서 생성"""
        
        # 제목
        ws.merge_cells('A1:E1')
        ws['A1'] = f"{company_name} 손익계산서"
        ws['A1'].style = self.styles["title"]
        
        ws['A2'] = f"기간: {period}"
        ws.merge_cells('A2:E2')
        
        # 헤더
        headers = ['계정과목', '당기', '전기', '증감액', '증감률']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.style = self.styles["header"]
        
        # 손익계산서 항목
        items = [
            ('매출액', 1000000000, 900000000, 'revenue'),
            ('매출원가', 600000000, 550000000, 'expense'),
            ('매출총이익', 400000000, 350000000, 'subtotal'),
            ('판매비와관리비', 200000000, 180000000, 'expense'),
            ('영업이익', 200000000, 170000000, 'subtotal'),
            ('영업외수익', 10000000, 8000000, 'revenue'),
            ('영업외비용', 15000000, 12000000, 'expense'),
            ('법인세차감전순이익', 195000000, 166000000, 'subtotal'),
            ('법인세비용', 48750000, 41500000, 'expense'),
            ('당기순이익', 146250000, 124500000, 'total')
        ]
        
        row = 5
        for item_name, current, previous, item_type in items:
            ws.cell(row=row, column=1, value=item_name)
            
            # 당기
            current_cell = ws.cell(row=row, column=2, value=current)
            current_cell.style = self.styles["currency"]
            
            # 전기
            previous_cell = ws.cell(row=row, column=3, value=previous)
            previous_cell.style = self.styles["currency"]
            
            # 증감액
            change_cell = ws.cell(row=row, column=4)
            change_cell.value = f"=B{row}-C{row}"
            change_cell.style = self.styles["currency"]
            
            # 증감률
            rate_cell = ws.cell(row=row, column=5)
            rate_cell.value = f"=IF(C{row}=0,0,(B{row}-C{row})/C{row})"
            rate_cell.style = self.styles["percentage"]
            
            # 스타일링
            if item_type in ['subtotal', 'total']:
                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(bold=True)
                    if item_type == 'total':
                        cell.fill = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid")
            
            row += 1
        
        # 컬럼 너비 조정
        ws.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D']:
            ws.column_dimensions[col].width = 15
        ws.column_dimensions['E'].width = 12
    
    async def _create_balance_sheet(
        self,
        ws,
        company_name: str,
        period: str,
        data: Optional[Dict[str, Any]]
    ):
        """재무상태표 생성"""
        
        # 제목
        ws.merge_cells('A1:F1')
        ws['A1'] = f"{company_name} 재무상태표"
        ws['A1'].style = self.styles["title"]
        
        ws['A2'] = f"기준일: {period}"
        ws.merge_cells('A2:F2')
        
        # 자산 섹션
        ws['A4'] = "자 산"
        ws['A4'].font = Font(bold=True, size=14)
        ws.merge_cells('A4:F4')
        
        # 헤더
        headers = ['계정과목', '당기', '전기', '계정과목', '당기', '전기']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.style = self.styles["header"]
        
        # 자산 항목
        assets = [
            ('유동자산', 500000000, 450000000, 'header'),
            ('  현금및현금성자산', 100000000, 90000000, 'item'),
            ('  매출채권', 150000000, 140000000, 'item'),
            ('  재고자산', 200000000, 180000000, 'item'),
            ('  기타유동자산', 50000000, 40000000, 'item'),
            ('비유동자산', 800000000, 750000000, 'header'),
            ('  유형자산', 600000000, 580000000, 'item'),
            ('  무형자산', 150000000, 130000000, 'item'),
            ('  기타비유동자산', 50000000, 40000000, 'item'),
            ('자산총계', 1300000000, 1200000000, 'total')
        ]
        
        # 부채 및 자본 항목
        liabilities_equity = [
            ('유동부채', 300000000, 280000000, 'header'),
            ('  매입채무', 100000000, 95000000, 'item'),
            ('  단기차입금', 150000000, 140000000, 'item'),
            ('  기타유동부채', 50000000, 45000000, 'item'),
            ('비유동부채', 200000000, 220000000, 'header'),
            ('  장기차입금', 150000000, 170000000, 'item'),
            ('  기타비유동부채', 50000000, 50000000, 'item'),
            ('부채총계', 500000000, 500000000, 'total'),
            ('자본금', 300000000, 300000000, 'header'),
            ('자본잉여금', 200000000, 200000000, 'item'),
            ('이익잉여금', 300000000, 200000000, 'item'),
            ('자본총계', 800000000, 700000000, 'total'),
            ('부채와자본총계', 1300000000, 1200000000, 'total')
        ]
        
        # 데이터 입력
        row = 6
        for i, (item_name, current, previous, item_type) in enumerate(assets):
            ws.cell(row=row + i, column=1, value=item_name)
            ws.cell(row=row + i, column=2, value=current).style = self.styles["currency"]
            ws.cell(row=row + i, column=3, value=previous).style = self.styles["currency"]
            
            if item_type == 'total':
                for col in range(1, 4):
                    ws.cell(row=row + i, column=col).font = Font(bold=True)
                    ws.cell(row=row + i, column=col).fill = PatternFill(
                        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
                    )
        
        # 부채 및 자본 입력
        for i, (item_name, current, previous, item_type) in enumerate(liabilities_equity):
            ws.cell(row=row + i, column=4, value=item_name)
            ws.cell(row=row + i, column=5, value=current).style = self.styles["currency"]
            ws.cell(row=row + i, column=6, value=previous).style = self.styles["currency"]
            
            if item_type == 'total':
                for col in range(4, 7):
                    ws.cell(row=row + i, column=col).font = Font(bold=True)
                    ws.cell(row=row + i, column=col).fill = PatternFill(
                        start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
                    )
        
        # 컬럼 너비 조정
        for col in ['A', 'D']:
            ws.column_dimensions[col].width = 25
        for col in ['B', 'C', 'E', 'F']:
            ws.column_dimensions[col].width = 15
    
    async def _create_cash_flow_statement(
        self,
        ws,
        company_name: str,
        period: str,
        data: Optional[Dict[str, Any]]
    ):
        """현금흐름표 생성"""
        
        # 제목
        ws.merge_cells('A1:D1')
        ws['A1'] = f"{company_name} 현금흐름표"
        ws['A1'].style = self.styles["title"]
        
        # 현금흐름 항목
        cash_flows = [
            ('영업활동 현금흐름', '', '', 'section'),
            ('  당기순이익', 146250000, 124500000, 'item'),
            ('  감가상각비', 50000000, 45000000, 'item'),
            ('  운전자본 변동', -30000000, -25000000, 'item'),
            ('영업활동으로 인한 순현금흐름', 166250000, 144500000, 'subtotal'),
            ('', '', '', 'blank'),
            ('투자활동 현금흐름', '', '', 'section'),
            ('  유형자산 취득', -80000000, -70000000, 'item'),
            ('  유형자산 처분', 10000000, 5000000, 'item'),
            ('투자활동으로 인한 순현금흐름', -70000000, -65000000, 'subtotal'),
            ('', '', '', 'blank'),
            ('재무활동 현금흐름', '', '', 'section'),
            ('  차입금 증가', 50000000, 30000000, 'item'),
            ('  배당금 지급', -40000000, -35000000, 'item'),
            ('재무활동으로 인한 순현금흐름', 10000000, -5000000, 'subtotal'),
            ('', '', '', 'blank'),
            ('현금및현금성자산의 순증감', 106250000, 74500000, 'total'),
            ('기초 현금및현금성자산', 90000000, 15500000, 'item'),
            ('기말 현금및현금성자산', 196250000, 90000000, 'total')
        ]
        
        # 헤더
        headers = ['항목', '당기', '전기', '증감']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.style = self.styles["header"]
        
        # 데이터 입력
        row = 5
        for item_name, current, previous, item_type in cash_flows:
            if item_type != 'blank':
                ws.cell(row=row, column=1, value=item_name)
                
                if current != '':
                    ws.cell(row=row, column=2, value=current).style = self.styles["currency"]
                    ws.cell(row=row, column=3, value=previous).style = self.styles["currency"]
                    ws.cell(row=row, column=4).value = f"=B{row}-C{row}"
                    ws.cell(row=row, column=4).style = self.styles["currency"]
                
                # 스타일링
                if item_type in ['section', 'subtotal', 'total']:
                    for col in range(1, 5):
                        ws.cell(row=row, column=col).font = Font(bold=True)
                    
                    if item_type == 'total':
                        for col in range(1, 5):
                            ws.cell(row=row, column=col).fill = PatternFill(
                                start_color="E6B8B7", end_color="E6B8B7", fill_type="solid"
                            )
            
            row += 1
    
    async def _create_ratio_analysis(
        self,
        ws,
        data: Optional[Dict[str, Any]]
    ):
        """재무비율 분석"""
        
        # 제목
        ws.merge_cells('A1:F1')
        ws['A1'] = "재무비율 분석"
        ws['A1'].style = self.styles["title"]
        
        # 비율 카테고리
        ratios = [
            ('수익성 지표', '', '', '', '', 'header'),
            ('매출총이익률', 0.40, 0.39, '=B3/1000000000', '매출총이익/매출액', 'item'),
            ('영업이익률', 0.20, 0.19, '=B4/1000000000', '영업이익/매출액', 'item'),
            ('순이익률', 0.146, 0.138, '=B5/1000000000', '당기순이익/매출액', 'item'),
            ('ROE', 0.183, 0.178, '=B6/800000000', '당기순이익/자본총계', 'item'),
            ('ROA', 0.113, 0.104, '=B7/1300000000', '당기순이익/자산총계', 'item'),
            ('', '', '', '', '', 'blank'),
            ('안정성 지표', '', '', '', '', 'header'),
            ('유동비율', 1.67, 1.61, '=500000000/300000000', '유동자산/유동부채', 'item'),
            ('부채비율', 0.625, 0.714, '=500000000/800000000', '부채총계/자본총계', 'item'),
            ('', '', '', '', '', 'blank'),
            ('활동성 지표', '', '', '', '', 'header'),
            ('총자산회전율', 0.77, 0.75, '=1000000000/1300000000', '매출액/총자산', 'item'),
            ('매출채권회전율', 6.67, 6.43, '=1000000000/150000000', '매출액/매출채권', 'item')
        ]
        
        # 헤더
        headers = ['지표명', '당기', '전기', '계산식', '설명', '평가']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.style = self.styles["header"]
        
        # 데이터 입력
        row = 4
        for ratio_name, current, previous, formula, description, item_type in ratios:
            if item_type != 'blank':
                ws.cell(row=row, column=1, value=ratio_name)
                
                if item_type == 'item':
                    # 값
                    current_cell = ws.cell(row=row, column=2, value=current)
                    previous_cell = ws.cell(row=row, column=3, value=previous)
                    
                    # 백분율 형식
                    if ratio_name in ['매출총이익률', '영업이익률', '순이익률', 'ROE', 'ROA', '부채비율']:
                        current_cell.style = self.styles["percentage"]
                        previous_cell.style = self.styles["percentage"]
                    else:
                        current_cell.number_format = "0.00"
                        previous_cell.number_format = "0.00"
                    
                    # 계산식 및 설명
                    ws.cell(row=row, column=4, value=formula)
                    ws.cell(row=row, column=5, value=description)
                    
                    # 평가 (개선/유지/악화)
                    evaluation_cell = ws.cell(row=row, column=6)
                    if current > previous:
                        if ratio_name == '부채비율':  # 부채비율은 낮을수록 좋음
                            evaluation_cell.value = "악화"
                            evaluation_cell.font = Font(color="FF0000")
                        else:
                            evaluation_cell.value = "개선"
                            evaluation_cell.font = Font(color="008000")
                    elif current < previous:
                        if ratio_name == '부채비율':
                            evaluation_cell.value = "개선"
                            evaluation_cell.font = Font(color="008000")
                        else:
                            evaluation_cell.value = "악화"
                            evaluation_cell.font = Font(color="FF0000")
                    else:
                        evaluation_cell.value = "유지"
                
                # 스타일링
                if item_type == 'header':
                    for col in range(1, 7):
                        ws.cell(row=row, column=col).font = Font(bold=True, size=12)
                        ws.cell(row=row, column=col).fill = PatternFill(
                            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
                        )
            
            row += 1
        
        # 컬럼 너비 조정
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 10
    
    async def _create_financial_dashboard(
        self,
        ws,
        data: Optional[Dict[str, Any]]
    ):
        """재무 대시보드"""
        
        # 제목
        ws.merge_cells('A1:H1')
        ws['A1'] = "재무 대시보드"
        ws['A1'].style = self.styles["title"]
        
        # KPI 섹션
        ws['A3'] = "주요 재무 지표"
        ws['A3'].font = Font(bold=True, size=14)
        
        # KPI 박스들
        kpis = [
            ('매출액', 1000000000, '전기 대비 11.1% ↑'),
            ('영업이익', 200000000, '전기 대비 17.6% ↑'),
            ('당기순이익', 146250000, '전기 대비 17.5% ↑'),
            ('ROE', 0.183, '18.3%')
        ]
        
        for i, (kpi_name, value, change) in enumerate(kpis):
            col = i * 2 + 1
            
            # KPI 이름
            ws.cell(row=5, column=col, value=kpi_name)
            ws.cell(row=5, column=col).font = Font(bold=True, size=12)
            
            # KPI 값
            value_cell = ws.cell(row=6, column=col, value=value)
            if isinstance(value, (int, float)) and value < 1:
                value_cell.style = self.styles["percentage"]
            elif isinstance(value, (int, float)):
                value_cell.style = self.styles["currency"]
            value_cell.font = Font(size=16, bold=True, color="2F5597")
            
            # 변화
            ws.cell(row=7, column=col, value=change)
            ws.cell(row=7, column=col).font = Font(size=10, color="008000")
        
        # 차트 공간 예약
        ws['A10'] = "매출 및 이익 추이"
        ws['A10'].font = Font(bold=True, size=12)
        
        # 차트 데이터 준비
        chart_data = [
            ['구분', '전전기', '전기', '당기'],
            ['매출액', 850000000, 900000000, 1000000000],
            ['영업이익', 150000000, 170000000, 200000000],
            ['당기순이익', 110000000, 124500000, 146250000]
        ]
        
        for row_idx, row_data in enumerate(chart_data, 12):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 12:  # 헤더
                    cell.font = Font(bold=True)
                elif col_idx > 1:  # 숫자 데이터
                    cell.style = self.styles["currency"]
        
        # 막대 차트 생성
        chart = BarChart()
        chart.title = "매출 및 이익 추이"
        chart.y_axis.title = "금액 (원)"
        chart.x_axis.title = "구분"
        
        data = Reference(ws, min_col=2, min_row=12, max_row=15, max_col=4)
        categories = Reference(ws, min_col=1, min_row=13, max_row=15)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.width = 15
        chart.height = 10
        
        ws.add_chart(chart, "A17")
        
        # 재무비율 요약
        ws['E10'] = "주요 재무비율"
        ws['E10'].font = Font(bold=True, size=12)
        
        ratio_summary = [
            ['비율', '당기', '전기', '업계평균'],
            ['영업이익률', 0.20, 0.19, 0.15],
            ['순이익률', 0.146, 0.138, 0.10],
            ['부채비율', 0.625, 0.714, 0.80],
            ['유동비율', 1.67, 1.61, 1.50]
        ]
        
        for row_idx, row_data in enumerate(ratio_summary, 12):
            for col_idx, value in enumerate(row_data, 5):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 12:  # 헤더
                    cell.font = Font(bold=True)
                elif col_idx > 5 and isinstance(value, (int, float)):  # 비율 데이터
                    if row_data[0] in ['영업이익률', '순이익률', '부채비율']:
                        cell.style = self.styles["percentage"]
                    else:
                        cell.number_format = "0.00"
                    
                    # 조건부 서식 (업계평균과 비교)
                    if col_idx == 6:  # 당기 컬럼
                        industry_avg = row_data[4]
                        if row_data[0] == '부채비율':  # 낮을수록 좋음
                            if value < industry_avg:
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            else:
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        else:  # 높을수록 좋음
                            if value > industry_avg:
                                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            else:
                                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")


class HRExcelGenerator(DomainExcelGenerator):
    """인사관리 특화 Excel 생성기"""
    
    async def generate_hr_management_system(
        self,
        company_name: str,
        employee_count: int,
        data: Optional[Dict[str, Any]] = None
    ) -> str:
        """인사관리 시스템 생성"""
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # 1. 직원 명부
        employee_roster = wb.create_sheet("직원명부")
        await self._create_employee_roster(employee_roster, employee_count)
        
        # 2. 급여 관리
        payroll = wb.create_sheet("급여관리")
        await self._create_payroll_sheet(payroll)
        
        # 3. 근태 관리
        attendance = wb.create_sheet("근태관리")
        await self._create_attendance_sheet(attendance)
        
        # 4. 휴가 관리
        leave_mgmt = wb.create_sheet("휴가관리")
        await self._create_leave_management(leave_mgmt)
        
        # 5. HR 대시보드
        hr_dashboard = wb.create_sheet("HR대시보드", 0)
        await self._create_hr_dashboard(hr_dashboard, employee_count)
        
        # 파일 저장
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"hr_management_{company_name}_{timestamp}.xlsx"
        output_path = os.path.join(tempfile.gettempdir(), filename)
        wb.save(output_path)
        
        return output_path
    
    async def _create_employee_roster(self, ws, employee_count: int):
        """직원 명부 생성"""
        
        # 제목
        ws.merge_cells('A1:K1')
        ws['A1'] = "직원 명부"
        ws['A1'].style = self.styles["title"]
        
        # 헤더
        headers = [
            '사번', '성명', '부서', '직급', '입사일', '생년월일',
            '연락처', '이메일', '주소', '비상연락처', '상태'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.style = self.styles["header"]
        
        # 샘플 데이터 생성
        departments = ['경영지원', '영업', '개발', '마케팅', '생산', '품질관리']
        positions = ['사원', '대리', '과장', '차장', '부장', '이사']
        
        for i in range(min(employee_count, 50)):  # 최대 50명까지 샘플
            row = i + 4
            
            # 사번
            ws.cell(row=row, column=1, value=f"EMP{2024000 + i:04d}")
            
            # 성명
            ws.cell(row=row, column=2, value=f"직원{i+1}")
            
            # 부서
            ws.cell(row=row, column=3, value=np.random.choice(departments))
            
            # 직급
            ws.cell(row=row, column=4, value=np.random.choice(positions))
            
            # 입사일
            hire_date = datetime.now() - timedelta(days=np.random.randint(30, 3650))
            ws.cell(row=row, column=5, value=hire_date).style = self.styles["date"]
            
            # 생년월일
            birth_date = datetime.now() - timedelta(days=np.random.randint(7300, 18250))
            ws.cell(row=row, column=6, value=birth_date).style = self.styles["date"]
            
            # 연락처
            ws.cell(row=row, column=7, value=f"010-{np.random.randint(1000, 9999)}-{np.random.randint(1000, 9999)}")
            
            # 이메일
            ws.cell(row=row, column=8, value=f"employee{i+1}@company.com")
            
            # 주소
            ws.cell(row=row, column=9, value=f"서울시 {np.random.choice(['강남구', '서초구', '송파구', '마포구'])}")
            
            # 비상연락처
            ws.cell(row=row, column=10, value=f"010-{np.random.randint(1000, 9999)}-{np.random.randint(1000, 9999)}")
            
            # 상태
            ws.cell(row=row, column=11, value=np.random.choice(['재직', '재직', '재직', '재직', '휴직']))
        
        # 테이블 스타일 적용
        if employee_count > 0:
            table_range = f"A3:K{min(employee_count + 3, 53)}"
            table = Table(displayName="EmployeeRoster", ref=table_range)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            ws.add_table(table)
        
        # 컬럼 너비 조정
        column_widths = [10, 10, 12, 10, 12, 12, 15, 25, 25, 15, 10]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    async def _create_payroll_sheet(self, ws):
        """급여 관리 시트"""
        
        # 제목
        ws.merge_cells('A1:L1')
        ws['A1'] = f"{datetime.now().strftime('%Y년 %m월')} 급여 명세"
        ws['A1'].style = self.styles["title"]
        
        # 헤더
        headers = [
            '사번', '성명', '부서', '직급', '기본급', '수당',
            '총지급액', '소득세', '4대보험', '공제합계', '실지급액', '비고'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.style = self.styles["header"]
        
        # 샘플 급여 데이터
        base_salaries = {
            '사원': 3000000,
            '대리': 3500000,
            '과장': 4500000,
            '차장': 5500000,
            '부장': 6500000,
            '이사': 8000000
        }
        
        for i in range(20):  # 20명 샘플
            row = i + 4
            
            position = np.random.choice(list(base_salaries.keys()))
            base_salary = base_salaries[position]
            
            # 기본 정보
            ws.cell(row=row, column=1, value=f"EMP{2024000 + i:04d}")
            ws.cell(row=row, column=2, value=f"직원{i+1}")
            ws.cell(row=row, column=3, value=np.random.choice(['경영지원', '영업', '개발', '마케팅']))
            ws.cell(row=row, column=4, value=position)
            
            # 급여 계산
            ws.cell(row=row, column=5, value=base_salary).style = self.styles["currency"]
            
            # 수당 (기본급의 0~20%)
            allowance = int(base_salary * np.random.uniform(0, 0.2))
            ws.cell(row=row, column=6, value=allowance).style = self.styles["currency"]
            
            # 총지급액
            total_pay_cell = ws.cell(row=row, column=7)
            total_pay_cell.value = f"=E{row}+F{row}"
            total_pay_cell.style = self.styles["currency"]
            
            # 소득세 (간이세액표 기준 약 5~10%)
            tax_cell = ws.cell(row=row, column=8)
            tax_cell.value = f"=G{row}*0.08"
            tax_cell.style = self.styles["currency"]
            
            # 4대보험 (약 9%)
            insurance_cell = ws.cell(row=row, column=9)
            insurance_cell.value = f"=G{row}*0.09"
            insurance_cell.style = self.styles["currency"]
            
            # 공제합계
            deduction_cell = ws.cell(row=row, column=10)
            deduction_cell.value = f"=H{row}+I{row}"
            deduction_cell.style = self.styles["currency"]
            
            # 실지급액
            net_pay_cell = ws.cell(row=row, column=11)
            net_pay_cell.value = f"=G{row}-J{row}"
            net_pay_cell.style = self.styles["currency"]
            net_pay_cell.font = Font(bold=True)
        
        # 합계 행
        total_row = 24
        ws.cell(row=total_row, column=4, value="합계").font = Font(bold=True)
        
        for col in [5, 6, 7, 8, 9, 10, 11]:
            cell = ws.cell(row=total_row, column=col)
            cell.value = f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}23)"
            cell.style = self.styles["currency"]
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
        
        # 컬럼 너비 조정
        column_widths = [10, 10, 12, 10, 12, 12, 12, 12, 12, 12, 12, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    async def _create_attendance_sheet(self, ws):
        """근태 관리 시트"""
        
        # 제목
        ws.merge_cells('A1:AF1')
        month = datetime.now().strftime('%Y년 %m월')
        ws['A1'] = f"{month} 근태 관리"
        ws['A1'].style = self.styles["title"]
        
        # 날짜 헤더 생성
        headers = ['사번', '성명', '부서']
        
        # 현재 월의 일수
        current_date = datetime.now()
        if current_date.month == 12:
            next_month = current_date.replace(year=current_date.year + 1, month=1)
        else:
            next_month = current_date.replace(month=current_date.month + 1)
        
        days_in_month = (next_month - current_date.replace(day=1)).days
        
        for day in range(1, days_in_month + 1):
            headers.append(str(day))
        
        headers.extend(['출근일수', '결근', '지각', '조퇴'])
        
        # 헤더 입력
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.style = self.styles["header"]
            
            # 주말 표시
            if col > 3 and col <= 3 + days_in_month:
                day = int(header)
                date = current_date.replace(day=day)
                if date.weekday() >= 5:  # 토요일, 일요일
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # 샘플 근태 데이터
        attendance_codes = ['○', '○', '○', '○', '○', '○', '●', '△', '□', '휴']
        
        for i in range(15):  # 15명 샘플
            row = i + 4
            
            # 기본 정보
            ws.cell(row=row, column=1, value=f"EMP{2024000 + i:04d}")
            ws.cell(row=row, column=2, value=f"직원{i+1}")
            ws.cell(row=row, column=3, value=np.random.choice(['경영지원', '영업', '개발', '마케팅']))
            
            # 일별 근태
            attendance_record = []
            for day_col in range(4, 4 + days_in_month):
                day = day_col - 3
                date = current_date.replace(day=day)
                
                # 주말은 기본적으로 휴무
                if date.weekday() >= 5:
                    code = '휴'
                else:
                    code = np.random.choice(attendance_codes, p=[0.85, 0.02, 0.02, 0.01, 0.01, 0.01, 0.02, 0.02, 0.02, 0.02])
                
                cell = ws.cell(row=row, column=day_col, value=code)
                cell.alignment = Alignment(horizontal="center")
                
                # 색상 코딩
                if code == '●':  # 결근
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif code == '△':  # 지각
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif code == '□':  # 조퇴
                    cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                
                attendance_record.append(code)
            
            # 집계
            work_days = attendance_record.count('○')
            absent_days = attendance_record.count('●')
            late_days = attendance_record.count('△')
            early_leave_days = attendance_record.count('□')
            
            ws.cell(row=row, column=4 + days_in_month, value=work_days)
            ws.cell(row=row, column=5 + days_in_month, value=absent_days)
            ws.cell(row=row, column=6 + days_in_month, value=late_days)
            ws.cell(row=row, column=7 + days_in_month, value=early_leave_days)
        
        # 범례
        legend_row = 20
        ws.cell(row=legend_row, column=1, value="범례:")
        legends = [
            ('○', '정상출근'), ('●', '결근'), ('△', '지각'),
            ('□', '조퇴'), ('휴', '휴무/휴가')
        ]
        
        for i, (code, desc) in enumerate(legends):
            ws.cell(row=legend_row, column=3 + i*2, value=code)
            ws.cell(row=legend_row, column=4 + i*2, value=desc)
        
        # 컬럼 너비 조정
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        
        for col in range(4, 4 + days_in_month):
            ws.column_dimensions[get_column_letter(col)].width = 3
    
    async def _create_leave_management(self, ws):
        """휴가 관리 시트"""
        
        # 제목
        ws.merge_cells('A1:J1')
        ws['A1'] = f"{datetime.now().year}년 휴가 관리"
        ws['A1'].style = self.styles["title"]
        
        # 헤더
        headers = [
            '사번', '성명', '부서', '입사일', '근속연수',
            '연차발생', '사용', '잔여', '대체휴가', '비고'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.style = self.styles["header"]
        
        # 샘플 휴가 데이터
        for i in range(20):  # 20명 샘플
            row = i + 4
            
            # 기본 정보
            ws.cell(row=row, column=1, value=f"EMP{2024000 + i:04d}")
            ws.cell(row=row, column=2, value=f"직원{i+1}")
            ws.cell(row=row, column=3, value=np.random.choice(['경영지원', '영업', '개발', '마케팅']))
            
            # 입사일 및 근속연수
            years_of_service = np.random.randint(0, 20)
            hire_date = datetime.now() - timedelta(days=years_of_service * 365)
            ws.cell(row=row, column=4, value=hire_date).style = self.styles["date"]
            ws.cell(row=row, column=5, value=years_of_service)
            
            # 연차 계산 (근속연수에 따라)
            if years_of_service < 1:
                annual_leave = int(years_of_service * 12 * 15 / 12)  # 월차
            elif years_of_service < 3:
                annual_leave = 15
            else:
                annual_leave = min(15 + (years_of_service - 2) // 2, 25)
            
            ws.cell(row=row, column=6, value=annual_leave)
            
            # 사용 연차 (랜덤)
            used_leave = np.random.randint(0, annual_leave + 1)
            ws.cell(row=row, column=7, value=used_leave)
            
            # 잔여 연차
            remaining_cell = ws.cell(row=row, column=8)
            remaining_cell.value = f"=F{row}-G{row}"
            remaining_cell.font = Font(bold=True)
            
            # 조건부 서식 (잔여 연차가 3일 이하면 빨간색)
            if annual_leave - used_leave <= 3:
                remaining_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # 대체휴가
            comp_leave = np.random.randint(0, 5)
            ws.cell(row=row, column=9, value=comp_leave)
        
        # 요약 통계
        summary_row = 26
        ws.cell(row=summary_row, column=1, value="요약").font = Font(bold=True, size=12)
        
        summary_items = [
            ('총 직원수', '=COUNTA(A4:A23)', '명'),
            ('평균 근속연수', '=AVERAGE(E4:E23)', '년'),
            ('총 연차 발생', '=SUM(F4:F23)', '일'),
            ('총 연차 사용', '=SUM(G4:G23)', '일'),
            ('평균 사용률', '=G28/G27', '%')
        ]
        
        for i, (label, formula, unit) in enumerate(summary_items):
            ws.cell(row=summary_row + i + 1, column=1, value=label)
            value_cell = ws.cell(row=summary_row + i + 1, column=2, value=formula)
            ws.cell(row=summary_row + i + 1, column=3, value=unit)
            
            if unit == '%':
                value_cell.style = self.styles["percentage"]
            elif unit == '년':
                value_cell.number_format = "0.0"
    
    async def _create_hr_dashboard(self, ws, employee_count: int):
        """HR 대시보드"""
        
        # 제목
        ws.merge_cells('A1:H1')
        ws['A1'] = "HR 대시보드"
        ws['A1'].style = self.styles["title"]
        
        # KPI 섹션
        ws['A3'] = "인사 현황"
        ws['A3'].font = Font(bold=True, size=14)
        
        # 직원 현황
        kpis = [
            ('총 직원수', employee_count, '명'),
            ('신규 입사', np.random.randint(5, 15), '명'),
            ('퇴사자', np.random.randint(0, 5), '명'),
            ('평균 근속연수', round(np.random.uniform(3, 8), 1), '년')
        ]
        
        for i, (kpi_name, value, unit) in enumerate(kpis):
            row = 5 + i * 3
            
            ws.cell(row=row, column=1, value=kpi_name).font = Font(bold=True)
            value_cell = ws.cell(row=row + 1, column=1, value=value)
            value_cell.font = Font(size=20, bold=True, color="2F5597")
            ws.cell(row=row + 1, column=2, value=unit)
        
        # 부서별 인원 현황
        ws['D3'] = "부서별 인원 현황"
        ws['D3'].font = Font(bold=True, size=12)
        
        departments = ['경영지원', '영업', '개발', '마케팅', '생산', '품질관리']
        dept_headers = ['부서', '인원', '비율']
        
        for col, header in enumerate(dept_headers, 4):
            ws.cell(row=5, column=col, value=header).font = Font(bold=True)
        
        total_employees = 0
        for i, dept in enumerate(departments):
            count = np.random.randint(5, 30)
            total_employees += count
            
            ws.cell(row=6 + i, column=4, value=dept)
            ws.cell(row=6 + i, column=5, value=count)
            
            # 비율은 나중에 계산
        
        # 비율 계산
        for i in range(len(departments)):
            ratio_cell = ws.cell(row=6 + i, column=6)
            ratio_cell.value = f"=E{6+i}/{total_employees}"
            ratio_cell.style = self.styles["percentage"]
        
        # 차트 생성 - 부서별 인원
        pie_chart = PieChart()
        pie_chart.title = "부서별 인원 분포"
        
        labels = Reference(ws, min_col=4, min_row=6, max_row=6 + len(departments) - 1)
        data = Reference(ws, min_col=5, min_row=6, max_row=6 + len(departments) - 1)
        
        pie_chart.add_data(data)
        pie_chart.set_categories(labels)
        pie_chart.width = 10
        pie_chart.height = 8
        
        ws.add_chart(pie_chart, "D15")
        
        # 근태 현황 요약
        ws['A18'] = "이번 달 근태 현황"
        ws['A18'].font = Font(bold=True, size=12)
        
        attendance_summary = [
            ['구분', '인원', '비율'],
            ['정상 출근', int(employee_count * 0.92), ''],
            ['지각', int(employee_count * 0.03), ''],
            ['조퇴', int(employee_count * 0.02), ''],
            ['결근', int(employee_count * 0.01), ''],
            ['휴가', int(employee_count * 0.02), '']
        ]
        
        for row_idx, row_data in enumerate(attendance_summary, 20):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 20:
                    cell.font = Font(bold=True)
                elif col_idx == 3 and row_idx > 20:
                    cell.value = f"=B{row_idx}/SUM(B21:B25)"
                    cell.style = self.styles["percentage"]
        
        # 연차 사용률
        ws['E18'] = "연차 사용 현황"
        ws['E18'].font = Font(bold=True, size=12)
        
        leave_data = [
            ['구분', '일수'],
            ['총 발생 연차', employee_count * 15],
            ['사용 연차', int(employee_count * 15 * 0.6)],
            ['잔여 연차', '=F21-F22'],
            ['평균 사용률', '=F22/F21']
        ]
        
        for row_idx, row_data in enumerate(leave_data, 20):
            for col_idx, value in enumerate(row_data, 5):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 20:
                    cell.font = Font(bold=True)
                elif row_idx == 24 and col_idx == 6:
                    cell.style = self.styles["percentage"]


class SalesExcelGenerator(DomainExcelGenerator):
    """영업/판매 특화 Excel 생성기"""
    
    async def generate_sales_dashboard(
        self,
        company_name: str,
        period: str,
        data: Optional[Dict[str, Any]] = None
    ) -> str:
        """영업 대시보드 생성"""
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # 1. 영업 대시보드
        dashboard = wb.create_sheet("영업대시보드")
        await self._create_sales_main_dashboard(dashboard, company_name, period)
        
        # 2. 일일 매출 보고
        daily_sales = wb.create_sheet("일일매출")
        await self._create_daily_sales_report(daily_sales)
        
        # 3. 고객별 매출
        customer_sales = wb.create_sheet("고객별매출")
        await self._create_customer_sales_analysis(customer_sales)
        
        # 4. 제품별 판매
        product_sales = wb.create_sheet("제품별판매")
        await self._create_product_sales_analysis(product_sales)
        
        # 5. 영업사원 성과
        sales_performance = wb.create_sheet("영업성과")
        await self._create_sales_performance_sheet(sales_performance)
        
        # 첫 번째 시트를 대시보드로 이동
        wb._sheets = [dashboard] + [s for s in wb._sheets if s != dashboard]
        
        # 파일 저장
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"sales_dashboard_{company_name}_{timestamp}.xlsx"
        output_path = os.path.join(tempfile.gettempdir(), filename)
        wb.save(output_path)
        
        return output_path
    
    async def _create_sales_main_dashboard(self, ws, company_name: str, period: str):
        """메인 영업 대시보드"""
        
        # 제목
        ws.merge_cells('A1:J1')
        ws['A1'] = f"{company_name} 영업 대시보드 - {period}"
        ws['A1'].style = self.styles["title"]
        
        # 주요 지표
        ws['A3'] = "핵심 성과 지표 (KPI)"
        ws['A3'].font = Font(bold=True, size=14)
        
        # KPI 데이터
        monthly_target = 500000000
        monthly_actual = 450000000
        
        kpis = [
            ('월 매출 목표', monthly_target, '원'),
            ('월 매출 실적', monthly_actual, '원'),
            ('달성률', monthly_actual / monthly_target, '%'),
            ('전월 대비', 0.08, '%')
        ]
        
        for i, (kpi_name, value, unit) in enumerate(kpis):
            col = (i % 2) * 4 + 1
            row = 5 + (i // 2) * 3
            
            ws.cell(row=row, column=col, value=kpi_name).font = Font(bold=True, size=11)
            value_cell = ws.cell(row=row, column=col + 1, value=value)
            value_cell.font = Font(size=16, bold=True, color="2F5597")
            
            if unit == '%':
                value_cell.style = self.styles["percentage"]
            elif unit == '원':
                value_cell.style = self.styles["currency"]
            
            ws.cell(row=row, column=col + 2, value=unit)
        
        # 월별 매출 추이
        ws['A12'] = "월별 매출 추이"
        ws['A12'].font = Font(bold=True, size=12)
        
        # 차트 데이터
        months = ['1월', '2월', '3월', '4월', '5월', '6월']
        sales_data = [380000000, 420000000, 400000000, 440000000, 420000000, 450000000]
        targets = [400000000] * 6
        
        chart_data = [['월', '실적', '목표']]
        for i, month in enumerate(months):
            chart_data.append([month, sales_data[i], targets[i]])
        
        for row_idx, row_data in enumerate(chart_data, 14):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 14:
                    cell.font = Font(bold=True)
                elif col_idx > 1:
                    cell.style = self.styles["currency"]
        
        # 라인 차트 생성
        line_chart = LineChart()
        line_chart.title = "월별 매출 실적 vs 목표"
        line_chart.y_axis.title = "매출액 (원)"
        line_chart.x_axis.title = "월"
        
        data = Reference(ws, min_col=2, min_row=14, max_row=20, max_col=3)
        categories = Reference(ws, min_col=1, min_row=15, max_row=20)
        
        line_chart.add_data(data, titles_from_data=True)
        line_chart.set_categories(categories)
        line_chart.width = 15
        line_chart.height = 10
        
        ws.add_chart(line_chart, "E14")
        
        # 판매 퍼널
        ws['A25'] = "판매 퍼널 분석"
        ws['A25'].font = Font(bold=True, size=12)
        
        funnel_data = [
            ['단계', '건수', '전환율'],
            ['잠재고객', 500, 1.0],
            ['상담', 300, 0.6],
            ['제안', 150, 0.5],
            ['협상', 75, 0.5],
            ['계약', 30, 0.4]
        ]
        
        for row_idx, row_data in enumerate(funnel_data, 27):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 27:
                    cell.font = Font(bold=True)
                elif col_idx == 3 and row_idx > 27:
                    cell.style = self.styles["percentage"]
        
        # 영업팀 성과 요약
        ws['E25'] = "영업팀 성과 TOP 5"
        ws['E25'].font = Font(bold=True, size=12)
        
        team_performance = [
            ['영업사원', '매출액', '달성률'],
            ['김영업', 95000000, 1.19],
            ['이판매', 88000000, 1.10],
            ['박세일', 82000000, 1.03],
            ['최고객', 78000000, 0.98],
            ['정성과', 75000000, 0.94]
        ]
        
        for row_idx, row_data in enumerate(team_performance, 27):
            for col_idx, value in enumerate(row_data, 5):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 27:
                    cell.font = Font(bold=True)
                elif col_idx == 6:
                    cell.style = self.styles["currency"]
                elif col_idx == 7:
                    cell.style = self.styles["percentage"]
                    
                    # 달성률에 따른 색상
                    if row_idx > 27:
                        if value >= 1.0:
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    async def _create_daily_sales_report(self, ws):
        """일일 매출 보고서"""
        
        # 제목
        ws.merge_cells('A1:I1')
        today = datetime.now().strftime('%Y년 %m월 %d일')
        ws['A1'] = f"{today} 일일 매출 보고"
        ws['A1'].style = self.styles["title"]
        
        # 일일 요약
        ws['A3'] = "일일 매출 요약"
        ws['A3'].font = Font(bold=True, size=12)
        
        daily_summary = [
            ('일일 목표', 16666667, '원'),
            ('일일 실적', 18500000, '원'),
            ('달성률', 1.11, '%'),
            ('전일 대비', 0.05, '%')
        ]
        
        for i, (label, value, unit) in enumerate(daily_summary):
            ws.cell(row=5 + i, column=1, value=label).font = Font(bold=True)
            value_cell = ws.cell(row=5 + i, column=2, value=value)
            
            if unit == '%':
                value_cell.style = self.styles["percentage"]
            else:
                value_cell.style = self.styles["currency"]
            
            ws.cell(row=5 + i, column=3, value=unit)
        
        # 시간대별 매출
        ws['E3'] = "시간대별 매출"
        ws['E3'].font = Font(bold=True, size=12)
        
        hourly_data = [
            ['시간', '매출액', '거래수'],
            ['09:00', 1500000, 15],
            ['10:00', 2200000, 22],
            ['11:00', 2800000, 28],
            ['12:00', 1800000, 18],
            ['13:00', 2100000, 21],
            ['14:00', 2500000, 25],
            ['15:00', 2300000, 23],
            ['16:00', 1900000, 19],
            ['17:00', 1400000, 14]
        ]
        
        for row_idx, row_data in enumerate(hourly_data, 5):
            for col_idx, value in enumerate(row_data, 5):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 5:
                    cell.font = Font(bold=True)
                elif col_idx == 6:
                    cell.style = self.styles["currency"]
        
        # 일일 거래 상세
        ws['A11'] = "주요 거래 내역"
        ws['A11'].font = Font(bold=True, size=12)
        
        headers = ['시간', '고객명', '담당자', '제품', '수량', '단가', '매출액', '결제방법', '상태']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=13, column=col, value=header)
            cell.style = self.styles["header"]
        
        # 샘플 거래 데이터
        transactions = []
        customers = ['A기업', 'B상사', 'C무역', 'D유통', 'E마트']
        products = ['제품A', '제품B', '제품C', '서비스A', '서비스B']
        sales_reps = ['김영업', '이판매', '박세일', '최고객', '정성과']
        payment_methods = ['현금', '카드', '계좌이체', '어음']
        
        for i in range(20):
            time = f"{9 + i//2:02d}:{(i%2)*30:02d}"
            customer = np.random.choice(customers)
            rep = np.random.choice(sales_reps)
            product = np.random.choice(products)
            quantity = np.random.randint(1, 100)
            unit_price = np.random.randint(10000, 100000)
            amount = quantity * unit_price
            payment = np.random.choice(payment_methods)
            status = np.random.choice(['완료', '완료', '완료', '진행중', '보류'])
            
            transactions.append([time, customer, rep, product, quantity, unit_price, amount, payment, status])
        
        for row_idx, transaction in enumerate(transactions, 14):
            for col_idx, value in enumerate(transaction, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                if col_idx in [5, 6, 7]:  # 수량, 단가, 매출액
                    if col_idx == 5:
                        cell.number_format = "#,##0"
                    else:
                        cell.style = self.styles["currency"]
                
                if col_idx == 9:  # 상태
                    if value == '완료':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value == '보류':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    async def _create_customer_sales_analysis(self, ws):
        """고객별 매출 분석"""
        
        # 제목
        ws.merge_cells('A1:H1')
        ws['A1'] = "고객별 매출 분석"
        ws['A1'].style = self.styles["title"]
        
        # 고객 등급별 요약
        ws['A3'] = "고객 등급별 매출"
        ws['A3'].font = Font(bold=True, size=12)
        
        grade_summary = [
            ['등급', '고객수', '매출액', '비중', '평균구매액'],
            ['VIP', 20, 150000000, 0.33, 7500000],
            ['Gold', 50, 180000000, 0.40, 3600000],
            ['Silver', 100, 100000000, 0.22, 1000000],
            ['Bronze', 200, 20000000, 0.05, 100000]
        ]
        
        for row_idx, row_data in enumerate(grade_summary, 5):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 5:
                    cell.font = Font(bold=True)
                elif col_idx == 3 or col_idx == 5:
                    cell.style = self.styles["currency"]
                elif col_idx == 4:
                    cell.style = self.styles["percentage"]
        
        # 고객별 상세 데이터
        ws['A12'] = "주요 고객 매출 TOP 20"
        ws['A12'].font = Font(bold=True, size=12)
        
        customer_headers = [
            '순위', '고객명', '등급', '당월매출', '전월매출', '증감률',
            '누적매출', '거래횟수'
        ]
        
        for col, header in enumerate(customer_headers, 1):
            cell = ws.cell(row=14, column=col, value=header)
            cell.style = self.styles["header"]
        
        # 샘플 고객 데이터
        for i in range(20):
            rank = i + 1
            customer_name = f"{chr(65 + i%5)}기업"
            grade = ['VIP', 'VIP', 'Gold', 'Gold', 'Silver'][min(i//4, 4)]
            current_sales = np.random.randint(5000000, 30000000)
            prev_sales = np.random.randint(4000000, 25000000)
            change_rate = (current_sales - prev_sales) / prev_sales
            total_sales = current_sales * np.random.randint(5, 20)
            transactions = np.random.randint(10, 100)
            
            row = 15 + i
            ws.cell(row=row, column=1, value=rank)
            ws.cell(row=row, column=2, value=customer_name)
            ws.cell(row=row, column=3, value=grade)
            ws.cell(row=row, column=4, value=current_sales).style = self.styles["currency"]
            ws.cell(row=row, column=5, value=prev_sales).style = self.styles["currency"]
            
            change_cell = ws.cell(row=row, column=6, value=change_rate)
            change_cell.style = self.styles["percentage"]
            if change_rate > 0:
                change_cell.font = Font(color="008000")
            elif change_rate < 0:
                change_cell.font = Font(color="FF0000")
            
            ws.cell(row=row, column=7, value=total_sales).style = self.styles["currency"]
            ws.cell(row=row, column=8, value=transactions)
    
    async def _create_product_sales_analysis(self, ws):
        """제품별 판매 분석"""
        
        # 제목
        ws.merge_cells('A1:I1')
        ws['A1'] = "제품별 판매 분석"
        ws['A1'].style = self.styles["title"]
        
        # 카테고리별 요약
        ws['A3'] = "카테고리별 판매 현황"
        ws['A3'].font = Font(bold=True, size=12)
        
        category_data = [
            ['카테고리', '판매수량', '매출액', '비중', '평균단가'],
            ['제품군A', 1200, 180000000, 0.40, 150000],
            ['제품군B', 800, 120000000, 0.27, 150000],
            ['제품군C', 1500, 90000000, 0.20, 60000],
            ['서비스', 200, 60000000, 0.13, 300000]
        ]
        
        for row_idx, row_data in enumerate(category_data, 5):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 5:
                    cell.font = Font(bold=True)
                elif col_idx == 2:
                    cell.number_format = "#,##0"
                elif col_idx in [3, 5]:
                    cell.style = self.styles["currency"]
                elif col_idx == 4:
                    cell.style = self.styles["percentage"]
        
        # 파이 차트 - 카테고리별 매출 비중
        pie_chart = PieChart()
        pie_chart.title = "카테고리별 매출 비중"
        
        labels = Reference(ws, min_col=1, min_row=6, max_row=9)
        data = Reference(ws, min_col=3, min_row=6, max_row=9)
        
        pie_chart.add_data(data)
        pie_chart.set_categories(labels)
        pie_chart.width = 10
        pie_chart.height = 8
        
        ws.add_chart(pie_chart, "F5")
        
        # 제품별 상세
        ws['A13'] = "제품별 판매 순위"
        ws['A13'].font = Font(bold=True, size=12)
        
        product_headers = [
            '순위', '제품코드', '제품명', '카테고리', '판매수량',
            '매출액', '재고', '회전율', '수익률'
        ]
        
        for col, header in enumerate(product_headers, 1):
            cell = ws.cell(row=15, column=col, value=header)
            cell.style = self.styles["header"]
        
        # 샘플 제품 데이터
        products = []
        for i in range(30):
            products.append({
                'rank': i + 1,
                'code': f"PRD{1000 + i:04d}",
                'name': f"제품{chr(65 + i%26)}",
                'category': ['제품군A', '제품군B', '제품군C', '서비스'][i % 4],
                'quantity': np.random.randint(50, 500),
                'sales': np.random.randint(5000000, 20000000),
                'stock': np.random.randint(0, 200),
                'turnover': np.random.uniform(2, 8),
                'profit_rate': np.random.uniform(0.1, 0.4)
            })
        
        for i, product in enumerate(products):
            row = 16 + i
            ws.cell(row=row, column=1, value=product['rank'])
            ws.cell(row=row, column=2, value=product['code'])
            ws.cell(row=row, column=3, value=product['name'])
            ws.cell(row=row, column=4, value=product['category'])
            ws.cell(row=row, column=5, value=product['quantity']).number_format = "#,##0"
            ws.cell(row=row, column=6, value=product['sales']).style = self.styles["currency"]
            ws.cell(row=row, column=7, value=product['stock']).number_format = "#,##0"
            ws.cell(row=row, column=8, value=product['turnover']).number_format = "0.0"
            ws.cell(row=row, column=9, value=product['profit_rate']).style = self.styles["percentage"]
            
            # 재고 부족 경고
            if product['stock'] < 20:
                ws.cell(row=row, column=7).fill = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )
    
    async def _create_sales_performance_sheet(self, ws):
        """영업사원 성과 시트"""
        
        # 제목
        ws.merge_cells('A1:J1')
        ws['A1'] = "영업사원 성과 분석"
        ws['A1'].style = self.styles["title"]
        
        # 영업팀 성과 요약
        ws['A3'] = "영업팀 성과 대시보드"
        ws['A3'].font = Font(bold=True, size=12)
        
        # 성과 데이터
        headers = [
            '순위', '사원명', '팀', '월목표', '월실적', '달성률',
            '신규고객', '거래건수', '평균단가', '인센티브'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.style = self.styles["header"]
        
        # 샘플 영업사원 데이터
        teams = ['영업1팀', '영업2팀', '영업3팀']
        sales_people = []
        
        for i in range(15):
            target = np.random.randint(50000000, 100000000)
            actual = int(target * np.random.uniform(0.7, 1.3))
            achievement = actual / target
            
            sales_people.append({
                'rank': i + 1,
                'name': f"영업사원{i+1}",
                'team': np.random.choice(teams),
                'target': target,
                'actual': actual,
                'achievement': achievement,
                'new_customers': np.random.randint(5, 20),
                'transactions': np.random.randint(50, 200),
                'avg_price': actual // np.random.randint(50, 150),
                'incentive': max(0, int((actual - target) * 0.01))
            })
        
        # 데이터 입력
        for i, person in enumerate(sales_people):
            row = 6 + i
            
            ws.cell(row=row, column=1, value=person['rank'])
            ws.cell(row=row, column=2, value=person['name'])
            ws.cell(row=row, column=3, value=person['team'])
            ws.cell(row=row, column=4, value=person['target']).style = self.styles["currency"]
            ws.cell(row=row, column=5, value=person['actual']).style = self.styles["currency"]
            
            achievement_cell = ws.cell(row=row, column=6, value=person['achievement'])
            achievement_cell.style = self.styles["percentage"]
            
            # 달성률에 따른 색상
            if person['achievement'] >= 1.1:
                achievement_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                achievement_cell.font = Font(bold=True, color="008000")
            elif person['achievement'] >= 1.0:
                achievement_cell.fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
            elif person['achievement'] >= 0.9:
                achievement_cell.fill = PatternFill(start_color="FFF4E5", end_color="FFF4E5", fill_type="solid")
            else:
                achievement_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                achievement_cell.font = Font(color="FF0000")
            
            ws.cell(row=row, column=7, value=person['new_customers'])
            ws.cell(row=row, column=8, value=person['transactions'])
            ws.cell(row=row, column=9, value=person['avg_price']).style = self.styles["currency"]
            ws.cell(row=row, column=10, value=person['incentive']).style = self.styles["currency"]
        
        # 팀별 성과 요약
        ws['A24'] = "팀별 성과 비교"
        ws['A24'].font = Font(bold=True, size=12)
        
        team_headers = ['팀명', '인원', '목표합계', '실적합계', '평균달성률']
        
        for col, header in enumerate(team_headers, 1):
            cell = ws.cell(row=26, column=col, value=header)
            cell.font = Font(bold=True)
        
        # 팀별 집계 (실제로는 위 데이터를 집계해야 하지만 샘플로 작성)
        team_data = [
            ['영업1팀', 5, 400000000, 420000000, 1.05],
            ['영업2팀', 5, 350000000, 330000000, 0.94],
            ['영업3팀', 5, 380000000, 400000000, 1.05]
        ]
        
        for i, team in enumerate(team_data):
            row = 27 + i
            for col, value in enumerate(team, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if col in [3, 4]:
                    cell.style = self.styles["currency"]
                elif col == 5:
                    cell.style = self.styles["percentage"]
        
        # 막대 차트 - 팀별 성과
        bar_chart = BarChart()
        bar_chart.title = "팀별 실적 vs 목표"
        bar_chart.y_axis.title = "금액 (원)"
        
        teams = Reference(ws, min_col=1, min_row=27, max_row=29)
        targets = Reference(ws, min_col=3, min_row=27, max_row=29)
        actuals = Reference(ws, min_col=4, min_row=27, max_row=29)
        
        bar_chart.add_data(targets, titles_from_data=False)
        bar_chart.add_data(actuals, titles_from_data=False)
        bar_chart.set_categories(teams)
        bar_chart.series[0].title = "목표"
        bar_chart.series[1].title = "실적"
        bar_chart.width = 12
        bar_chart.height = 8
        
        ws.add_chart(bar_chart, "F26")


class InventoryExcelGenerator(DomainExcelGenerator):
    """재고 관리 특화 Excel 생성기"""
    
    async def generate_inventory_system(
        self,
        company_name: str,
        data: Optional[Dict[str, Any]] = None
    ) -> str:
        """재고 관리 시스템 생성"""
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # 1. 재고 현황
        inventory_status = wb.create_sheet("재고현황")
        await self._create_inventory_status_sheet(inventory_status, data)
        
        # 2. 입출고 내역
        transactions = wb.create_sheet("입출고내역")
        await self._create_transaction_history_sheet(transactions, data)
        
        # 3. ABC 분석
        abc_analysis = wb.create_sheet("ABC분석")
        await self._create_abc_analysis_sheet(abc_analysis, data)
        
        # 4. 재주문 알림
        reorder_alert = wb.create_sheet("재주문알림")
        await self._create_reorder_alert_sheet(reorder_alert, data)
        
        # 5. 재고 대시보드
        dashboard = wb.create_sheet("대시보드", 0)
        await self._create_inventory_dashboard(dashboard, data)
        
        # 파일 저장
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"inventory_management_{company_name}_{timestamp}.xlsx"
        output_path = os.path.join(tempfile.gettempdir(), filename)
        wb.save(output_path)
        
        return output_path
    
    async def _create_inventory_status_sheet(self, ws, data: Optional[Dict[str, Any]]):
        """재고 현황 시트 생성"""
        # 제목
        ws.merge_cells('A1:N1')
        ws['A1'] = "재고 현황"
        ws['A1'].style = self.styles["title"]
        
        # 헤더
        headers = [
            '품목코드', '품목명', '카테고리', '현재재고', '단위',
            '안전재고', '재주문점', '최대재고', '단가', '재고금액',
            '위치', '재고상태', '최종입고일', '재고회전율'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.style = self.styles["header"]
        
        # 샘플 재고 데이터
        sample_items = [
            ['ITM001', '노트북 스탠드', '사무용품', 150, '개', 50, 75, 300, 25000, 150*25000, 'A-01-01', '정상', datetime.now(), 8.5],
            ['ITM002', '무선 마우스', '전자기기', 45, '개', 30, 40, 150, 35000, 45*35000, 'B-02-03', '재주문필요', datetime.now(), 12.3],
            ['ITM003', 'A4 용지', '사무용품', 280, '박스', 100, 150, 500, 15000, 280*15000, 'C-01-02', '정상', datetime.now(), 24.5],
            ['ITM004', 'USB 케이블', '전자기기', 12, '개', 20, 25, 80, 8000, 12*8000, 'B-03-05', '재주문필요', datetime.now(), 15.2],
            ['ITM005', '화이트보드', '사무용품', 320, '개', 50, 100, 200, 45000, 320*45000, 'D-01-01', '초과재고', datetime.now(), 3.2],
        ]
        
        for row_idx, item in enumerate(sample_items, 4):
            for col_idx, value in enumerate(item, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # 숫자 형식 지정
                if col_idx in [9, 10]:  # 단가, 재고금액
                    cell.style = self.styles["currency"]
                elif col_idx == 13:  # 최종입고일
                    cell.style = self.styles["date"]
                elif col_idx == 14:  # 재고회전율
                    cell.number_format = "0.0"
                
                # 재고상태에 따른 조건부 서식
                if col_idx == 12:  # 재고상태
                    if value == '재주문필요':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006")
                    elif value == '초과재고':
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(color="9C5700")
        
        # 합계 행
        total_row = 9
        ws.cell(row=total_row, column=3, value="합계").font = Font(bold=True)
        ws.cell(row=total_row, column=10).value = f"=SUM(J4:J8)"
        ws.cell(row=total_row, column=10).style = self.styles["currency"]
        ws.cell(row=total_row, column=10).font = Font(bold=True)
        
        # 조건부 서식 규칙 추가
        # 재고 수준이 재주문점 이하인 경우 강조
        for row in range(4, 9):
            ws.cell(row=row, column=4).conditional_formatting.add(
                f'D{row}',
                ColorScaleRule(
                    start_type='num', start_value=0, start_color='FF0000',
                    mid_type='num', mid_value=50, mid_color='FFFF00',
                    end_type='num', end_value=100, end_color='00FF00'
                )
            )
        
        # 컬럼 너비 조정
        column_widths = [10, 20, 12, 10, 8, 10, 10, 10, 12, 15, 10, 12, 12, 10]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    async def _create_transaction_history_sheet(self, ws, data: Optional[Dict[str, Any]]):
        """입출고 내역 시트"""
        # 제목
        ws.merge_cells('A1:K1')
        ws['A1'] = "입출고 내역"
        ws['A1'].style = self.styles["title"]
        
        # 헤더
        headers = [
            '거래번호', '거래일시', '구분', '품목코드', '품목명',
            '수량', '단가', '금액', '거래처', '담당자', '비고'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.style = self.styles["header"]
        
        # 샘플 거래 데이터
        transactions = []
        items = ['노트북 스탠드', '무선 마우스', 'A4 용지', 'USB 케이블', '화이트보드']
        suppliers = ['(주)오피스월드', '(주)테크서플라이', '(주)페이퍼랜드', '(주)케이블테크']
        
        for i in range(30):
            trans_type = np.random.choice(['입고', '출고'], p=[0.4, 0.6])
            item_idx = np.random.randint(0, len(items))
            quantity = np.random.randint(10, 100)
            unit_price = [25000, 35000, 15000, 8000, 45000][item_idx]
            
            transaction = [
                f'TRX{2024000 + i:06d}',
                datetime.now() - timedelta(days=np.random.randint(0, 30)),
                trans_type,
                f'ITM{item_idx + 1:03d}',
                items[item_idx],
                quantity,
                unit_price,
                quantity * unit_price,
                np.random.choice(suppliers) if trans_type == '입고' else f'부서{np.random.randint(1, 6)}',
                f'담당자{np.random.randint(1, 10)}',
                np.random.choice(['정기입고', '긴급입고', '판매출고', '내부사용', '반품'])
            ]
            transactions.append(transaction)
        
        # 최신 거래순으로 정렬
        transactions.sort(key=lambda x: x[1], reverse=True)
        
        for row_idx, transaction in enumerate(transactions, 4):
            for col_idx, value in enumerate(transaction, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # 형식 지정
                if col_idx == 2:  # 거래일시
                    cell.style = self.styles["date"]
                elif col_idx in [7, 8]:  # 단가, 금액
                    cell.style = self.styles["currency"]
                
                # 입출고 구분에 따른 색상
                if col_idx == 3:  # 구분
                    if value == '입고':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(color="006100")
                    else:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(color="9C0006")
    
    async def _create_abc_analysis_sheet(self, ws, data: Optional[Dict[str, Any]]):
        """ABC 분석 시트"""
        # 제목
        ws.merge_cells('A1:H1')
        ws['A1'] = "ABC 분석"
        ws['A1'].style = self.styles["title"]
        
        # 설명
        ws['A3'] = "ABC 분석을 통한 재고 중요도 관리"
        ws['A3'].font = Font(italic=True)
        
        # 헤더
        headers = [
            '순위', '품목코드', '품목명', '연간사용금액', '누적금액',
            '구성비', '누적비율', 'ABC등급'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.style = self.styles["header"]
        
        # ABC 분석 데이터
        abc_items = [
            [1, 'ITM003', 'A4 용지', 180000000, 180000000, 0.36, 0.36, 'A'],
            [2, 'ITM002', '무선 마우스', 140000000, 320000000, 0.28, 0.64, 'A'],
            [3, 'ITM005', '화이트보드', 80000000, 400000000, 0.16, 0.80, 'B'],
            [4, 'ITM001', '노트북 스탠드', 60000000, 460000000, 0.12, 0.92, 'B'],
            [5, 'ITM004', 'USB 케이블', 40000000, 500000000, 0.08, 1.00, 'C'],
        ]
        
        for row_idx, item in enumerate(abc_items, 6):
            for col_idx, value in enumerate(item, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # 형식 지정
                if col_idx in [4, 5]:  # 금액
                    cell.style = self.styles["currency"]
                elif col_idx in [6, 7]:  # 비율
                    cell.style = self.styles["percentage"]
                
                # ABC 등급에 따른 색상
                if col_idx == 8:  # ABC등급
                    if value == 'A':
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif value == 'B':
                        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        
        # ABC 등급별 관리 전략
        ws['A13'] = "ABC 등급별 관리 전략"
        ws['A13'].font = Font(bold=True, size=12)
        
        strategies = [
            ['등급', '비중', '관리전략', '재고정책'],
            ['A', '70%', '중점관리', '일일 점검, 정확한 수요예측'],
            ['B', '20%', '일반관리', '주간 점검, 적정 재고 유지'],
            ['C', '10%', '간소관리', '월간 점검, 최소 재고 유지']
        ]
        
        for row_idx, strategy in enumerate(strategies, 15):
            for col_idx, value in enumerate(strategy, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 15:
                    cell.font = Font(bold=True)
    
    async def _create_reorder_alert_sheet(self, ws, data: Optional[Dict[str, Any]]):
        """재주문 알림 시트"""
        # 제목
        ws.merge_cells('A1:J1')
        ws['A1'] = "재주문 알림"
        ws['A1'].style = self.styles["title"]
        ws['A1'].fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        
        # 헤더
        headers = [
            '긴급도', '품목코드', '품목명', '현재재고', '재주문점',
            '부족수량', '권장주문량', '공급업체', '리드타임', '예상입고일'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.style = self.styles["header"]
        
        # 재주문 필요 품목
        reorder_items = [
            ['긴급', 'ITM002', '무선 마우스', 45, 40, 5, 105, '(주)테크서플라이', 3, datetime.now() + timedelta(days=3)],
            ['긴급', 'ITM004', 'USB 케이블', 12, 25, 13, 80, '(주)케이블테크', 2, datetime.now() + timedelta(days=2)],
            ['높음', 'ITM001', '노트북 스탠드', 80, 75, 0, 120, '(주)오피스월드', 5, datetime.now() + timedelta(days=5)],
        ]
        
        for row_idx, item in enumerate(reorder_items, 4):
            for col_idx, value in enumerate(item, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # 형식 지정
                if col_idx == 10:  # 예상입고일
                    cell.style = self.styles["date"]
                
                # 긴급도에 따른 색상
                if col_idx == 1:  # 긴급도
                    if value == '긴급':
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                    elif value == '높음':
                        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # 주문 승인 섹션
        ws['A10'] = "주문 승인"
        ws['A10'].font = Font(bold=True, size=12)
        
        approval_headers = ['선택', '품목코드', '품목명', '주문수량', '단가', '주문금액', '승인']
        for col, header in enumerate(approval_headers, 1):
            cell = ws.cell(row=12, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    async def _create_inventory_dashboard(self, ws, data: Optional[Dict[str, Any]]):
        """재고 대시보드"""
        # 제목
        ws.merge_cells('A1:J1')
        ws['A1'] = "재고 관리 대시보드"
        ws['A1'].style = self.styles["title"]
        
        # KPI 섹션
        ws['A3'] = "핵심 성과 지표"
        ws['A3'].font = Font(bold=True, size=14)
        
        # KPI 카드들
        kpis = [
            ('총 재고 가치', 24200000, '원', 'currency'),
            ('평균 재고 회전율', 12.7, '회', 'number'),
            ('재고 정확도', 0.985, '%', 'percentage'),
            ('재주문 필요', 3, '개', 'number')
        ]
        
        for i, (kpi_name, value, unit, format_type) in enumerate(kpis):
            row = 5 + (i // 2) * 3
            col = 1 + (i % 2) * 4
            
            ws.cell(row=row, column=col, value=kpi_name).font = Font(bold=True, size=11)
            value_cell = ws.cell(row=row + 1, column=col, value=value)
            value_cell.font = Font(size=20, bold=True, color="2F5597")
            
            if format_type == 'currency':
                value_cell.style = self.styles["currency"]
            elif format_type == 'percentage':
                value_cell.style = self.styles["percentage"]
            else:
                value_cell.number_format = "0.0"
            
            ws.cell(row=row + 1, column=col + 1, value=unit)
        
        # 재고 현황 요약
        ws['A13'] = "재고 상태별 현황"
        ws['A13'].font = Font(bold=True, size=12)
        
        status_summary = [
            ['상태', '품목수', '재고가치', '비중'],
            ['정상', 12, 18500000, 0.764],
            ['재주문필요', 3, 1700000, 0.070],
            ['초과재고', 2, 4000000, 0.165],
            ['합계', 17, 24200000, 1.000]
        ]
        
        for row_idx, row_data in enumerate(status_summary, 15):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 15:  # 헤더
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                elif col_idx == 3:  # 재고가치
                    cell.style = self.styles["currency"]
                elif col_idx == 4:  # 비중
                    cell.style = self.styles["percentage"]
                
                if row_idx == 19:  # 합계 행
                    cell.font = Font(bold=True)
        
        # ABC 분석 요약
        ws['F13'] = "ABC 등급별 현황"
        ws['F13'].font = Font(bold=True, size=12)
        
        abc_summary = [
            ['등급', '품목수', '재고가치', '비중'],
            ['A', 2, 16000000, 0.661],
            ['B', 2, 5600000, 0.231],
            ['C', 13, 2600000, 0.107],
            ['합계', 17, 24200000, 1.000]
        ]
        
        for row_idx, row_data in enumerate(abc_summary, 15):
            for col_idx, value in enumerate(row_data, 6):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 15:  # 헤더
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                elif col_idx == 8:  # 재고가치
                    cell.style = self.styles["currency"]
                elif col_idx == 9:  # 비중
                    cell.style = self.styles["percentage"]
                
                if row_idx == 19:  # 합계 행
                    cell.font = Font(bold=True)
        
        # 차트 위치 표시
        ws['A22'] = "[월별 재고 추이 차트]"
        ws['F22'] = "[재고 회전율 차트]"
        ws['A28'] = "[ABC 분석 파레토 차트]"
        ws['F28'] = "[재고 상태 분포 차트]"


class ProjectExcelGenerator(DomainExcelGenerator):
    """프로젝트 관리 특화 Excel 생성기"""
    
    async def generate_project_management(
        self,
        project_name: str,
        project_duration: int = 90,
        data: Optional[Dict[str, Any]] = None
    ) -> str:
        """프로젝트 관리 시스템 생성"""
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # 1. 프로젝트 개요
        overview = wb.create_sheet("프로젝트개요")
        await self._create_project_overview_sheet(overview, project_name, project_duration, data)
        
        # 2. 작업 일정 (간트 차트)
        gantt = wb.create_sheet("작업일정")
        await self._create_gantt_chart_sheet(gantt, project_duration, data)
        
        # 3. 리소스 관리
        resources = wb.create_sheet("리소스관리")
        await self._create_resource_management_sheet(resources, data)
        
        # 4. 위험 관리
        risks = wb.create_sheet("위험관리")
        await self._create_risk_management_sheet(risks, data)
        
        # 5. 프로젝트 대시보드
        dashboard = wb.create_sheet("대시보드", 0)
        await self._create_project_dashboard(dashboard, project_name, data)
        
        # 파일 저장
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"project_management_{project_name.replace(' ', '_')}_{timestamp}.xlsx"
        output_path = os.path.join(tempfile.gettempdir(), filename)
        wb.save(output_path)
        
        return output_path
    
    async def _create_project_overview_sheet(self, ws, project_name: str, duration: int, data: Optional[Dict[str, Any]]):
        """프로젝트 개요 시트"""
        # 제목
        ws.merge_cells('A1:F1')
        ws['A1'] = f"{project_name} - 프로젝트 개요"
        ws['A1'].style = self.styles["title"]
        
        # 프로젝트 기본 정보
        ws['A3'] = "프로젝트 정보"
        ws['A3'].font = Font(bold=True, size=14)
        
        project_info = [
            ['프로젝트명', project_name],
            ['프로젝트 관리자', data.get('pm', '홍길동')],
            ['시작일', datetime.now().strftime('%Y-%m-%d')],
            ['종료일(예정)', (datetime.now() + timedelta(days=duration)).strftime('%Y-%m-%d')],
            ['프로젝트 기간', f'{duration}일'],
            ['전체 예산', '₩150,000,000'],
            ['현재 상태', '진행중'],
            ['진행률', '32%']
        ]
        
        for row_idx, (label, value) in enumerate(project_info, 5):
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)
        
        # 프로젝트 목표
        ws['D3'] = "프로젝트 목표"
        ws['D3'].font = Font(bold=True, size=14)
        
        goals = data.get('goals', [
            '고품질 제품 개발',
            '일정 내 완료',
            '예산 내 실행',
            '고객 만족도 향상'
        ])
        
        for idx, goal in enumerate(goals, 5):
            ws.cell(row=idx, column=4, value=f"{idx-4}. {goal}")
        
        # 주요 마일스톤
        ws['A15'] = "주요 마일스톤"
        ws['A15'].font = Font(bold=True, size=14)
        
        milestone_headers = ['마일스톤', '목표일', '상태', '진행률']
        for col, header in enumerate(milestone_headers, 1):
            cell = ws.cell(row=17, column=col, value=header)
            cell.style = self.styles["header"]
        
        milestones = [
            ['요구사항 분석', (datetime.now() + timedelta(days=14)).strftime('%Y-%m-%d'), '진행중', 0.8],
            ['설계 완료', (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d'), '진행중', 0.3],
            ['개발 완료', (datetime.now() + timedelta(days=75)).strftime('%Y-%m-%d'), '예정', 0],
            ['테스트 완료', (datetime.now() + timedelta(days=85)).strftime('%Y-%m-%d'), '예정', 0],
            ['배포', (datetime.now() + timedelta(days=90)).strftime('%Y-%m-%d'), '예정', 0]
        ]
        
        for row_idx, milestone in enumerate(milestones, 18):
            for col_idx, value in enumerate(milestone, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 4:  # 진행률
                    cell.style = self.styles["percentage"]
                    
                    # 진행률에 따른 색상
                    if value >= 0.8:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value >= 0.5:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif value > 0:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # 팀 구성
        ws['F15'] = "프로젝트 팀"
        ws['F15'].font = Font(bold=True, size=14)
        
        team_headers = ['역할', '담당자', '참여율']
        for col, header in enumerate(team_headers, 6):
            cell = ws.cell(row=17, column=col, value=header)
            cell.style = self.styles["header"]
        
        team = [
            ['프로젝트 관리자', '홍길동', '100%'],
            ['기술 리드', '김개발', '100%'],
            ['개발자', '이코딩', '100%'],
            ['개발자', '박프로', '80%'],
            ['디자이너', '최디자인', '60%'],
            ['QA 엔지니어', '정테스트', '50%']
        ]
        
        for row_idx, member in enumerate(team, 18):
            for col_idx, value in enumerate(member, 6):
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    async def _create_gantt_chart_sheet(self, ws, duration: int, data: Optional[Dict[str, Any]]):
        """간트 차트 시트"""
        # 제목
        ws.merge_cells('A1:Z1')
        ws['A1'] = "프로젝트 간트 차트"
        ws['A1'].style = self.styles["title"]
        
        # 작업 정보 헤더
        task_headers = ['ID', '작업명', '담당자', '시작일', '종료일', '기간', '진행률', '선행작업']
        for col, header in enumerate(task_headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.style = self.styles["header"]
        
        # 날짜 헤더 (주 단위)
        start_date = datetime.now()
        weeks = duration // 7 + 1
        
        for week in range(weeks):
            col = 9 + week
            date = start_date + timedelta(weeks=week)
            cell = ws.cell(row=3, column=col, value=f"W{week+1}\n{date.strftime('%m/%d')}")
            cell.style = self.styles["header"]
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 작업 데이터
        tasks = [
            ['1', '프로젝트 착수', '홍길동', 0, 3, 3, 1.0, ''],
            ['2', '요구사항 분석', '김분석', 3, 14, 11, 0.8, '1'],
            ['3', '시스템 설계', '이설계', 10, 21, 11, 0.5, '2'],
            ['4', 'UI/UX 디자인', '최디자인', 14, 28, 14, 0.3, '2'],
            ['5', '데이터베이스 설계', '박DB', 21, 28, 7, 0.1, '3'],
            ['6', 'API 개발', '김개발', 28, 56, 28, 0, '3,5'],
            ['7', '프론트엔드 개발', '이코딩', 28, 63, 35, 0, '4'],
            ['8', '통합 테스트', '정테스트', 63, 77, 14, 0, '6,7'],
            ['9', '사용자 교육', '최교육', 77, 84, 7, 0, '8'],
            ['10', '배포', '홍길동', 84, 90, 6, 0, '8,9']
        ]
        
        for row_idx, task in enumerate(tasks, 4):
            # 작업 정보 입력
            for col_idx in range(len(task_headers)):
                value = task[col_idx] if col_idx < len(task) else ''
                cell = ws.cell(row=row_idx, column=col_idx + 1, value=value)
                
                if col_idx == 4:  # 종료일
                    cell.value = (start_date + timedelta(days=task[4])).strftime('%Y-%m-%d')
                elif col_idx == 3:  # 시작일
                    cell.value = (start_date + timedelta(days=task[3])).strftime('%Y-%m-%d')
                elif col_idx == 6:  # 진행률
                    cell.style = self.styles["percentage"]
            
            # 간트 차트 바 그리기
            start_week = task[3] // 7
            end_week = task[4] // 7
            progress = task[6]
            
            for week in range(start_week, end_week + 1):
                col = 9 + week
                cell = ws.cell(row=row_idx, column=col)
                
                # 진행 상태에 따른 색상
                if progress == 1.0:
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                elif progress > 0:
                    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        
        # 현재 날짜 표시 (세로선)
        current_week = 0  # 현재 주
        for row in range(3, 14):
            cell = ws.cell(row=row, column=9 + current_week)
            cell.border = Border(right=Side(style='thick', color='FF0000'))
        
        # 컬럼 너비 조정
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 12
        
        for col in range(9, 9 + weeks):
            ws.column_dimensions[get_column_letter(col)].width = 5
    
    async def _create_resource_management_sheet(self, ws, data: Optional[Dict[str, Any]]):
        """리소스 관리 시트"""
        # 제목
        ws.merge_cells('A1:I1')
        ws['A1'] = "리소스 관리"
        ws['A1'].style = self.styles["title"]
        
        # 리소스 할당 현황
        ws['A3'] = "팀원별 작업 할당 현황"
        ws['A3'].font = Font(bold=True, size=14)
        
        # 헤더
        headers = ['팀원', '역할', '총 가용시간', '할당시간', '잔여시간', '활용률', '현재작업', '작업수', '상태']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.style = self.styles["header"]
        
        # 팀원 데이터
        team_members = [
            ['홍길동', '프로젝트 관리자', 720, 720, 0, 1.0, '전체 관리', 2, '과부하'],
            ['김개발', '기술 리드', 720, 640, 80, 0.89, 'API 개발', 3, '정상'],
            ['이코딩', '개발자', 720, 560, 160, 0.78, '프론트엔드 개발', 2, '정상'],
            ['박프로', '개발자', 576, 480, 96, 0.83, '백엔드 개발', 2, '정상'],
            ['최디자인', '디자이너', 432, 336, 96, 0.78, 'UI 디자인', 1, '정상'],
            ['정테스트', 'QA 엔지니어', 360, 224, 136, 0.62, '테스트 계획', 1, '여유']
        ]
        
        for row_idx, member in enumerate(team_members, 6):
            for col_idx, value in enumerate(member, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                if col_idx in [3, 4, 5]:  # 시간
                    if col_idx == 5:  # 잔여시간
                        if value < 50:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif col_idx == 6:  # 활용률
                    cell.style = self.styles["percentage"]
                    if value > 0.9:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif value < 0.7:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif col_idx == 9:  # 상태
                    if value == '과부하':
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(color="FFFFFF")
                    elif value == '여유':
                        cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        
        # 리소스 차트 (월별 할당)
        ws['A14'] = "월별 리소스 할당 계획"
        ws['A14'].font = Font(bold=True, size=14)
        
        month_headers = ['팀원', '1월', '2월', '3월', '4월']
        for col, header in enumerate(month_headers, 1):
            cell = ws.cell(row=16, column=col, value=header)
            cell.style = self.styles["header"]
        
        # 월별 할당률 데이터
        monthly_allocation = [
            ['홍길동', 1.0, 1.0, 1.0, 0.5],
            ['김개발', 0.8, 1.0, 1.0, 0.8],
            ['이코딩', 0.6, 0.8, 1.0, 0.9],
            ['박프로', 0.7, 0.9, 0.9, 0.7],
            ['최디자인', 0.8, 0.6, 0.4, 0.2],
            ['정테스트', 0.2, 0.4, 0.8, 1.0]
        ]
        
        for row_idx, allocation in enumerate(monthly_allocation, 17):
            for col_idx, value in enumerate(allocation, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx > 1:
                    cell.style = self.styles["percentage"]
                    
                    # 데이터 바 형식
                    rule = DataBarRule(
                        start_type='num', start_value=0,
                        end_type='num', end_value=1,
                        color="638EC6"
                    )
                    ws.conditional_formatting.add(f'{get_column_letter(col_idx)}{row_idx}', rule)
    
    async def _create_risk_management_sheet(self, ws, data: Optional[Dict[str, Any]]):
        """위험 관리 시트"""
        # 제목
        ws.merge_cells('A1:J1')
        ws['A1'] = "위험 관리 등록부"
        ws['A1'].style = self.styles["title"]
        
        # 위험 등록부
        ws['A3'] = "식별된 위험 요소"
        ws['A3'].font = Font(bold=True, size=14)
        
        # 헤더
        headers = [
            'ID', '위험요소', '카테고리', '발생확률', '영향도',
            '위험점수', '대응전략', '담당자', '대응계획', '상태'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.style = self.styles["header"]
        
        # 위험 데이터
        risks = [
            ['R001', '핵심 개발자 이탈', '인적자원', '중', '높음', 6, '완화', '홍길동', '백업 인력 확보 및 지식 전수', '모니터링'],
            ['R002', '요구사항 변경', '범위', '높음', '중', 6, '수용', '홍길동', '변경관리 프로세스 수립', '대응중'],
            ['R003', '신기술 도입 실패', '기술', '중', '높음', 6, '완화', '김개발', '파일럿 테스트 실시', '대응완료'],
            ['R004', '예산 초과', '비용', '낮음', '높음', 3, '회피', '홍길동', '비용 절감 방안 수립', '모니터링'],
            ['R005', '일정 지연', '일정', '중', '중', 4, '완화', '홍길동', '버퍼 일정 확보', '모니터링'],
            ['R006', '품질 기준 미달', '품질', '낮음', '높음', 3, '완화', '정테스트', 'QA 프로세스 강화', '대응중']
        ]
        
        for row_idx, risk in enumerate(risks, 6):
            for col_idx, value in enumerate(risk, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # 위험점수에 따른 색상
                if col_idx == 6:  # 위험점수
                    if value >= 6:
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(color="FFFFFF")
                    elif value >= 4:
                        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    else:
                        cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
                
                # 상태에 따른 색상
                elif col_idx == 10:  # 상태
                    if value == '대응중':
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif value == '대응완료':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        
        # 위험 매트릭스
        ws['A14'] = "위험 평가 매트릭스"
        ws['A14'].font = Font(bold=True, size=14)
        
        # 매트릭스 헤더
        ws['B16'] = "발생확률 →"
        ws['A17'] = "영향도"
        ws['A18'] = "↓"
        
        prob_labels = ['낮음', '중', '높음']
        impact_labels = ['높음', '중', '낮음']
        
        # 확률 라벨
        for i, label in enumerate(prob_labels):
            ws.cell(row=16, column=3+i, value=label).alignment = Alignment(horizontal="center")
        
        # 영향도 라벨
        for i, label in enumerate(impact_labels):
            ws.cell(row=17+i, column=2, value=label).alignment = Alignment(horizontal="right")
        
        # 매트릭스 값
        matrix_values = [
            [3, 6, 9],  # 높음
            [2, 4, 6],  # 중
            [1, 2, 3]   # 낮음
        ]
        
        for row_idx, row_values in enumerate(matrix_values):
            for col_idx, value in enumerate(row_values):
                cell = ws.cell(row=17+row_idx, column=3+col_idx, value=value)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True, size=14)
                
                # 점수에 따른 색상
                if value >= 6:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True, size=14)
                elif value >= 4:
                    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        
        # 대응 전략 설명
        ws['G16'] = "대응 전략"
        ws['G16'].font = Font(bold=True)
        
        strategies = [
            ['회피', '위험을 제거하거나 프로젝트 범위에서 제외'],
            ['완화', '위험의 확률이나 영향을 감소시킴'],
            ['전가', '위험을 제3자에게 이전 (보험, 아웃소싱 등)'],
            ['수용', '위험을 인지하고 발생 시 대응 계획 수립']
        ]
        
        for i, (strategy, description) in enumerate(strategies):
            ws.cell(row=17+i, column=7, value=strategy).font = Font(bold=True)
            ws.cell(row=17+i, column=8, value=description)
    
    async def _create_project_dashboard(self, ws, project_name: str, data: Optional[Dict[str, Any]]):
        """프로젝트 대시보드"""
        # 제목
        ws.merge_cells('A1:J1')
        ws['A1'] = f"{project_name} - 프로젝트 대시보드"
        ws['A1'].style = self.styles["title"]
        
        # 전체 현황
        ws['A3'] = "프로젝트 현황"
        ws['A3'].font = Font(bold=True, size=14)
        
        # KPI 카드
        kpis = [
            ('전체 진행률', 0.32, '32%'),
            ('예산 사용률', 0.28, '₩42,000,000 / ₩150,000,000'),
            ('일정 준수율', 0.95, '19/20 작업 정상 진행'),
            ('위험 지수', 4.2, '평균 위험도')
        ]
        
        for i, (kpi_name, value, description) in enumerate(kpis):
            row = 5 + (i // 2) * 4
            col = 1 + (i % 2) * 5
            
            ws.cell(row=row, column=col, value=kpi_name).font = Font(bold=True, size=12)
            value_cell = ws.cell(row=row+1, column=col, value=value)
            
            if kpi_name in ['전체 진행률', '예산 사용률', '일정 준수율']:
                value_cell.style = self.styles["percentage"]
                value_cell.font = Font(size=24, bold=True, color="2F5597")
            else:
                value_cell.font = Font(size=24, bold=True, color="2F5597")
                value_cell.number_format = "0.0"
            
            ws.cell(row=row+2, column=col, value=description).font = Font(size=10, italic=True)
        
        # 마일스톤 진행 현황
        ws['A15'] = "마일스톤 진행 현황"
        ws['A15'].font = Font(bold=True, size=12)
        
        milestones = [
            ['마일스톤', '진행률', '상태'],
            ['요구사항 분석', 0.8, '진행중'],
            ['설계', 0.3, '진행중'],
            ['개발', 0.0, '예정'],
            ['테스트', 0.0, '예정'],
            ['배포', 0.0, '예정']
        ]
        
        for row_idx, milestone in enumerate(milestones, 17):
            for col_idx, value in enumerate(milestone, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 17:  # 헤더
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                elif col_idx == 2 and row_idx > 17:  # 진행률
                    cell.style = self.styles["percentage"]
                    
                    # 진행률 데이터 바
                    rule = DataBarRule(
                        start_type='num', start_value=0,
                        end_type='num', end_value=1,
                        color="70AD47"
                    )
                    ws.conditional_formatting.add(f'B{row_idx}', rule)
        
        # 주요 이슈
        ws['E15'] = "주요 이슈 및 결정사항"
        ws['E15'].font = Font(bold=True, size=12)
        
        issues = [
            ['구분', '내용', '담당자', '기한'],
            ['이슈', 'API 설계 변경 필요', '김개발', '1/25'],
            ['결정필요', '프레임워크 선택', '이코딩', '1/20'],
            ['위험', '일정 지연 우려', '홍길동', '즉시']
        ]
        
        for row_idx, issue in enumerate(issues, 17):
            for col_idx, value in enumerate(issue, 5):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 17:  # 헤더
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                elif col_idx == 5 and row_idx > 17:  # 구분
                    if value == '이슈':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    elif value == '위험':
                        cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                        cell.font = Font(color="FFFFFF")
        
        # 차트 위치 표시
        ws['A25'] = "[번다운 차트]"
        ws['E25'] = "[예산 사용 추이]"
        ws['A32'] = "[리소스 활용률 차트]"
        ws['E32'] = "[위험도 추이 차트]"
        
        # 컬럼 너비 조정
        for col in range(1, 11):
            ws.column_dimensions[get_column_letter(col)].width = 15


# 도메인별 생성기 인스턴스
finance_generator = FinanceExcelGenerator()
hr_generator = HRExcelGenerator()
sales_generator = SalesExcelGenerator()
inventory_generator = InventoryExcelGenerator()
project_generator = ProjectExcelGenerator()


# 통합 도메인 생성기
class DomainSpecificGenerator:
    """도메인별 Excel 생성 통합 관리자"""
    
    def __init__(self):
        self.generators = {
            'finance': finance_generator,
            'hr': hr_generator,
            'sales': sales_generator,
            'inventory': inventory_generator,
            'project': project_generator
        }
    
    async def generate(
        self,
        domain: str,
        template_type: str,
        **kwargs
    ) -> str:
        """도메인별 Excel 생성"""
        
        generator = self.generators.get(domain)
        if not generator:
            raise ValueError(f"Unsupported domain: {domain}")
        
        # 도메인별 템플릿 타입에 따른 생성
        if domain == 'finance':
            if template_type == 'statements':
                return await generator.generate_financial_statements(**kwargs)
        
        elif domain == 'hr':
            if template_type == 'management':
                return await generator.generate_hr_management_system(**kwargs)
        
        elif domain == 'sales':
            if template_type == 'dashboard':
                return await generator.generate_sales_dashboard(**kwargs)
        
        elif domain == 'inventory':
            if template_type == 'system':
                return await generator.generate_inventory_system(**kwargs)
        
        elif domain == 'project':
            if template_type == 'management':
                return await generator.generate_project_management(**kwargs)
        
        raise ValueError(f"Unsupported template type: {template_type} for domain: {domain}")


# 전역 도메인 생성기 인스턴스
domain_generator = DomainSpecificGenerator()