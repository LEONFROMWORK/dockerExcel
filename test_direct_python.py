"""
Direct test of Python error detection API
"""
import requests
from openpyxl import Workbook
import tempfile
import os

# Create test Excel file
wb = Workbook()
ws = wb.active
ws.title = "TestSheet"

# Add test data
ws['A1'] = "Test Data"
ws['B1'] = 10
ws['C1'] = 0
ws['A2'] = "=B1/C1"  # DIV/0 error

# Save to temp file
with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
    wb.save(tmp.name)
    test_file = tmp.name

print(f"Created test file: {test_file}")

try:
    # Call Python service directly
    url = "http://localhost:8000/api/v1/excel-error-analysis/analyze-upload"

    with open(test_file, 'rb') as f:
        files = {'file': ('test_direct.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {'session_id': 'test-session-123'}

        response = requests.post(url, files=files, data=data)

    print(f"\nResponse status: {response.status_code}")
    if response.status_code == 200:
        result = response.json()
        print(f"\nSuccess: {result.get('success')}")
        print(f"Message: {result.get('message')}")

        if 'data' in result:
            data = result['data']
            print(f"\nTotal errors: {data.get('total_errors', 0)}")
            print(f"Summary: {data.get('summary', {})}")

            if 'errors' in data:
                print(f"\nErrors found: {len(data['errors'])}")
                for i, error in enumerate(data['errors'][:5]):
                    print(f"\nError {i+1}:")
                    print(f"  Type: {error.get('type')}")
                    print(f"  Cell: {error.get('sheet')}!{error.get('cell')}")
                    print(f"  Message: {error.get('message')}")
    else:
        print(f"Error: {response.text}")

finally:
    # Clean up
    if os.path.exists(test_file):
        os.unlink(test_file)
        print(f"\nDeleted test file: {test_file}")
