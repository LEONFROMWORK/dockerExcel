"""
Base generator class for AI Excel generation
Provides common functionality for all domain-specific generators
"""

from abc import ABC, abstractmethod
from typing import Dict, List, Any, Optional
import logging
import pandas as pd
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime

logger = logging.getLogger(__name__)


class BaseExcelGenerator(ABC):
    """Abstract base class for Excel generators"""
    
    def __init__(self):
        self.workbook = None
        self.metadata = {}
        self.context = {}
        
    @abstractmethod
    async def generate_structure(self, request: str, context: Dict[str, Any]) -> Dict[str, Any]:
        """Generate Excel structure based on request and context"""
        pass
    
    @abstractmethod
    async def generate_data(self, structure: Dict[str, Any], context: Dict[str, Any]) -> Dict[str, pd.DataFrame]:
        """Generate data for each sheet based on structure"""
        pass
    
    @abstractmethod
    async def apply_formulas(self, worksheet: Worksheet, structure: Dict[str, Any]) -> None:
        """Apply formulas to the worksheet"""
        pass
    
    @abstractmethod
    async def create_visualizations(self, worksheet: Worksheet, structure: Dict[str, Any]) -> None:
        """Create charts and visualizations"""
        pass
    
    async def generate(self, request: str, context: Optional[Dict[str, Any]] = None) -> str:
        """Main generation method"""
        try:
            # Initialize context
            self.context = context or {}
            
            # Generate structure
            structure = await self.generate_structure(request, self.context)
            
            # Generate data
            data = await self.generate_data(structure, self.context)
            
            # Create workbook
            self.workbook = openpyxl.Workbook()
            self.workbook.remove(self.workbook.active)
            
            # Create sheets
            for sheet_config in structure.get('sheets', []):
                await self._create_sheet(sheet_config, data)
            
            # Save and return path
            file_path = self._save_workbook(structure)
            return file_path
            
        except Exception as e:
            logger.error(f"Generation failed: {str(e)}")
            raise
    
    async def _create_sheet(self, sheet_config: Dict[str, Any], data: Dict[str, pd.DataFrame]) -> None:
        """Create a single sheet with data, formulas, and visualizations"""
        sheet_name = sheet_config.get('name', 'Sheet1')
        worksheet = self.workbook.create_sheet(sheet_name)
        
        # Add data if available
        if sheet_name in data:
            df = data[sheet_name]
            await self._write_dataframe(worksheet, df, sheet_config)
        
        # Apply formulas
        await self.apply_formulas(worksheet, sheet_config)
        
        # Create visualizations
        await self.create_visualizations(worksheet, sheet_config)
    
    async def _write_dataframe(self, worksheet: Worksheet, df: pd.DataFrame, config: Dict[str, Any]) -> None:
        """Write DataFrame to worksheet with formatting"""
        # Write headers
        for col_idx, column in enumerate(df.columns, 1):
            worksheet.cell(row=1, column=col_idx, value=str(column))
        
        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
    
    def _save_workbook(self, structure: Dict[str, Any]) -> str:
        """Save workbook to file"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        domain = structure.get('domain', 'general')
        filename = f"ai_generated_{domain}_{timestamp}.xlsx"
        
        import tempfile
        import os
        output_path = os.path.join(tempfile.gettempdir(), filename)
        self.workbook.save(output_path)
        
        return output_path
    
    def set_metadata(self, key: str, value: Any) -> None:
        """Set metadata for the generation"""
        self.metadata[key] = value
    
    def get_metadata(self, key: str, default: Any = None) -> Any:
        """Get metadata value"""
        return self.metadata.get(key, default)