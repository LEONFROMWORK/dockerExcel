from fastapi import APIRouter, HTTPException
from typing import Dict, Any, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tempfile
import os
from datetime import datetime

router = APIRouter()

@router.post("/create-from-structure")
async def create_excel_from_structure(excel_structure: Dict[str, Any]):
    """Create an Excel file from a structured data format"""
    try:
        # Create a new workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        sheets_data = excel_structure.get("sheets", [])
        if not sheets_data:
            raise ValueError("No sheets data provided")
        
        for sheet_info in sheets_data:
            sheet_name = sheet_info.get("name", "Sheet")
            sheet_data = sheet_info.get("data", [])
            
            # Create sheet
            ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet names max 31 chars
            
            # Write data
            for row_idx, row_data in enumerate(sheet_data, 1):
                if isinstance(row_data, list):
                    for col_idx, cell_value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                        
                        # Apply formatting for header row
                        if row_idx == 1:
                            cell.font = Font(bold=True, size=12)
                            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add borders to all cells with data
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        cell.border = thin_border
        
        # Save to temporary file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"converted_{timestamp}.xlsx"
        output_path = os.path.join(tempfile.gettempdir(), output_filename)
        
        wb.save(output_path)
        wb.close()
        
        # Return file path (in production, upload to S3 and return URL)
        return {
            "success": True,
            "file_url": f"/tmp/uploads/{output_filename}",
            "filename": output_filename,
            "sheets_count": len(sheets_data)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to create Excel: {str(e)}")

@router.post("/enhance-with-ocr")
async def enhance_excel_with_ocr(
    excel_structure: Dict[str, Any],
    ocr_results: List[Dict[str, Any]]
):
    """Enhance Excel structure with OCR results for better accuracy"""
    try:
        enhanced_structure = excel_structure.copy()
        
        # Merge OCR results with AI-extracted data
        for ocr_result in ocr_results:
            if ocr_result.get("type") == "table":
                # Find matching sheet or create new one
                sheet_name = ocr_result.get("sheet_name", "OCR Data")
                ocr_data = ocr_result.get("data", [])
                
                # Add to sheets
                sheet_found = False
                for sheet in enhanced_structure.get("sheets", []):
                    if sheet["name"] == sheet_name:
                        # Merge data
                        sheet["data"].extend(ocr_data)
                        sheet_found = True
                        break
                
                if not sheet_found:
                    enhanced_structure.setdefault("sheets", []).append({
                        "name": sheet_name,
                        "data": ocr_data
                    })
        
        return {
            "success": True,
            "enhanced_structure": enhanced_structure,
            "ocr_sheets_added": len([r for r in ocr_results if r.get("type") == "table"])
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to enhance with OCR: {str(e)}")