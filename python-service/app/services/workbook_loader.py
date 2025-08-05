"""
Workbook Loader Implementation
워크북 로더 구현 - 순환 참조 방지를 위한 분리
"""

from typing import Any, Optional
import openpyxl
import logging
from app.core.interfaces import IWorkbookLoader
from app.core.exceptions import WorkbookLoadError

logger = logging.getLogger(__name__)


class OpenpyxlWorkbookLoader(IWorkbookLoader):
    """Openpyxl을 사용한 워크북 로더 구현"""

    async def load_workbook(self, file_path: str) -> Any:
        """워크북을 로드합니다"""
        try:
            # data_only=True: 수식 대신 계산된 값을 로드
            # keep_vba=True: VBA 코드 유지
            workbook = openpyxl.load_workbook(
                file_path, data_only=False, keep_vba=True  # 수식을 유지
            )
            logger.info(f"워크북 로드 성공: {file_path}")
            return workbook

        except FileNotFoundError:
            raise WorkbookLoadError(
                f"파일을 찾을 수 없습니다: {file_path}", code="FILE_NOT_FOUND"
            )
        except Exception as e:
            logger.error(f"워크북 로드 실패: {str(e)}")
            raise WorkbookLoadError(
                f"워크북 로드 중 오류 발생: {str(e)}",
                code="LOAD_ERROR",
                details={"file_path": file_path},
            )

    async def save_workbook(self, workbook: Any, file_path: str) -> bool:
        """워크북을 저장합니다"""
        try:
            workbook.save(file_path)
            logger.info(f"워크북 저장 성공: {file_path}")
            return True

        except Exception as e:
            logger.error(f"워크북 저장 실패: {str(e)}")
            raise WorkbookLoadError(
                f"워크북 저장 중 오류 발생: {str(e)}",
                code="SAVE_ERROR",
                details={"file_path": file_path},
            )

    async def get_cell_value(self, workbook: Any, sheet: str, cell: str) -> Any:
        """셀 값을 가져옵니다"""
        try:
            worksheet = workbook[sheet] if sheet in workbook.sheetnames else None
            if not worksheet:
                raise WorkbookLoadError(
                    f"시트를 찾을 수 없습니다: {sheet}", code="SHEET_NOT_FOUND"
                )

            cell_obj = worksheet[cell]
            return cell_obj.value

        except Exception as e:
            logger.error(f"셀 값 읽기 실패: {sheet}!{cell} - {str(e)}")
            return None

    async def set_cell_value(
        self, workbook: Any, sheet: str, cell: str, value: Any
    ) -> bool:
        """셀 값을 설정합니다"""
        try:
            worksheet = workbook[sheet] if sheet in workbook.sheetnames else None
            if not worksheet:
                raise WorkbookLoadError(
                    f"시트를 찾을 수 없습니다: {sheet}", code="SHEET_NOT_FOUND"
                )

            worksheet[cell] = value
            return True

        except Exception as e:
            logger.error(f"셀 값 설정 실패: {sheet}!{cell} - {str(e)}")
            return False

    def get_formula(self, workbook: Any, sheet: str, cell: str) -> Optional[str]:
        """셀의 수식을 가져옵니다"""
        try:
            worksheet = workbook[sheet]
            cell_obj = worksheet[cell]

            # 수식이 있는 경우 반환
            if (
                hasattr(cell_obj, "value")
                and isinstance(cell_obj.value, str)
                and cell_obj.value.startswith("=")
            ):
                return cell_obj.value

            return None

        except Exception as e:
            logger.error(f"수식 읽기 실패: {sheet}!{cell} - {str(e)}")
            return None
