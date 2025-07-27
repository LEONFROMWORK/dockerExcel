"""
Excel 비교 분석 엔진
원하는 결과와 실제 결과를 비교하여 차이점을 분석
"""

from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime
import logging
from dataclasses import dataclass
from enum import Enum
import json
import openpyxl
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ComparisonType(Enum):
    """비교 유형"""
    VALUE = "value"  # 값 비교
    FORMULA = "formula"  # 수식 비교
    FORMAT = "format"  # 서식 비교
    STRUCTURE = "structure"  # 구조 비교
    ALL = "all"  # 전체 비교


class DifferenceType(Enum):
    """차이점 유형"""
    VALUE_MISMATCH = "value_mismatch"
    FORMULA_DIFFERENT = "formula_different"
    FORMAT_DIFFERENT = "format_different"
    CELL_MISSING = "cell_missing"
    CELL_EXTRA = "cell_extra"
    TYPE_MISMATCH = "type_mismatch"
    PRECISION_DIFFERENCE = "precision_difference"


@dataclass
class CellDifference:
    """셀 차이점 정보"""
    sheet: str
    cell: str
    difference_type: DifferenceType
    expected_value: Any
    actual_value: Any
    description: str
    severity: str  # low, medium, high
    suggestion: Optional[str] = None


@dataclass
class ComparisonResult:
    """비교 분석 결과"""
    total_cells_compared: int
    differences_found: int
    match_percentage: float
    differences: List[CellDifference]
    summary: Dict[str, Any]
    execution_time: float


class ComparisonEngine:
    """Excel 파일 비교 분석 엔진"""
    
    def __init__(self):
        self.tolerance = 1e-10  # 부동소수점 비교 허용 오차
        self.ignore_hidden = True  # 숨겨진 행/열 무시 여부
        self.case_sensitive = False  # 대소문자 구분 여부
        
    async def compare_files(
        self,
        expected_file: str,
        actual_file: str,
        comparison_type: ComparisonType = ComparisonType.ALL,
        sheets_to_compare: Optional[List[str]] = None
    ) -> ComparisonResult:
        """두 Excel 파일을 비교 분석"""
        
        start_time = datetime.now()
        
        try:
            # 워크북 로드
            wb_expected = openpyxl.load_workbook(expected_file, data_only=True)
            wb_actual = openpyxl.load_workbook(actual_file, data_only=True)
            
            # 비교할 시트 결정
            if sheets_to_compare:
                sheets = sheets_to_compare
            else:
                sheets = self._get_common_sheets(wb_expected, wb_actual)
            
            # 차이점 수집
            differences = []
            total_cells = 0
            
            for sheet_name in sheets:
                if sheet_name in wb_expected.sheetnames and sheet_name in wb_actual.sheetnames:
                    sheet_differences, cells_compared = await self._compare_sheets(
                        wb_expected[sheet_name],
                        wb_actual[sheet_name],
                        sheet_name,
                        comparison_type
                    )
                    differences.extend(sheet_differences)
                    total_cells += cells_compared
            
            # 결과 요약
            execution_time = (datetime.now() - start_time).total_seconds()
            match_percentage = ((total_cells - len(differences)) / total_cells * 100) if total_cells > 0 else 0
            
            summary = self._create_summary(differences)
            
            return ComparisonResult(
                total_cells_compared=total_cells,
                differences_found=len(differences),
                match_percentage=match_percentage,
                differences=differences,
                summary=summary,
                execution_time=execution_time
            )
            
        except Exception as e:
            logger.error(f"파일 비교 중 오류: {str(e)}")
            raise
    
    async def _compare_sheets(
        self,
        sheet_expected: Any,
        sheet_actual: Any,
        sheet_name: str,
        comparison_type: ComparisonType
    ) -> Tuple[List[CellDifference], int]:
        """시트 비교"""
        
        differences = []
        cells_compared = 0
        
        # 사용된 범위 결정
        max_row = max(sheet_expected.max_row, sheet_actual.max_row)
        max_col = max(sheet_expected.max_column, sheet_actual.max_column)
        
        for row in range(1, max_row + 1):
            # 숨겨진 행 건너뛰기
            if self.ignore_hidden and (
                sheet_expected.row_dimensions[row].hidden or
                sheet_actual.row_dimensions[row].hidden
            ):
                continue
                
            for col in range(1, max_col + 1):
                # 숨겨진 열 건너뛰기
                col_letter = get_column_letter(col)
                if self.ignore_hidden and (
                    sheet_expected.column_dimensions[col_letter].hidden or
                    sheet_actual.column_dimensions[col_letter].hidden
                ):
                    continue
                
                cell_expected = sheet_expected.cell(row=row, column=col)
                cell_actual = sheet_actual.cell(row=row, column=col)
                cell_address = f"{col_letter}{row}"
                
                cells_compared += 1
                
                # 비교 수행
                cell_differences = await self._compare_cells(
                    cell_expected,
                    cell_actual,
                    sheet_name,
                    cell_address,
                    comparison_type
                )
                
                differences.extend(cell_differences)
        
        return differences, cells_compared
    
    async def _compare_cells(
        self,
        cell_expected: Any,
        cell_actual: Any,
        sheet_name: str,
        cell_address: str,
        comparison_type: ComparisonType
    ) -> List[CellDifference]:
        """개별 셀 비교"""
        
        differences = []
        
        # 값 비교
        if comparison_type in [ComparisonType.VALUE, ComparisonType.ALL]:
            value_diff = self._compare_values(
                cell_expected.value,
                cell_actual.value,
                sheet_name,
                cell_address
            )
            if value_diff:
                differences.append(value_diff)
        
        # 수식 비교
        if comparison_type in [ComparisonType.FORMULA, ComparisonType.ALL]:
            if hasattr(cell_expected, 'formula') and hasattr(cell_actual, 'formula'):
                formula_diff = self._compare_formulas(
                    cell_expected,
                    cell_actual,
                    sheet_name,
                    cell_address
                )
                if formula_diff:
                    differences.append(formula_diff)
        
        # 서식 비교
        if comparison_type in [ComparisonType.FORMAT, ComparisonType.ALL]:
            format_diff = self._compare_formats(
                cell_expected,
                cell_actual,
                sheet_name,
                cell_address
            )
            if format_diff:
                differences.append(format_diff)
        
        return differences
    
    def _compare_values(
        self,
        expected: Any,
        actual: Any,
        sheet: str,
        cell: str
    ) -> Optional[CellDifference]:
        """값 비교"""
        
        # 둘 다 None인 경우
        if expected is None and actual is None:
            return None
        
        # 하나만 None인 경우
        if expected is None or actual is None:
            return CellDifference(
                sheet=sheet,
                cell=cell,
                difference_type=DifferenceType.CELL_MISSING if actual is None else DifferenceType.CELL_EXTRA,
                expected_value=expected,
                actual_value=actual,
                description=f"셀이 {'누락됨' if actual is None else '추가됨'}",
                severity="high"
            )
        
        # 타입 비교
        if type(expected) != type(actual):
            # 숫자와 문자열 변환 가능한 경우 체크
            try:
                if isinstance(expected, (int, float)) and isinstance(actual, str):
                    actual_num = float(actual)
                    if self._numbers_equal(expected, actual_num):
                        return None
                elif isinstance(expected, str) and isinstance(actual, (int, float)):
                    expected_num = float(expected)
                    if self._numbers_equal(expected_num, actual):
                        return None
            except (ValueError, TypeError) as e:
                logger.debug(f"타입 변환 실패: {e}")
            
            return CellDifference(
                sheet=sheet,
                cell=cell,
                difference_type=DifferenceType.TYPE_MISMATCH,
                expected_value=expected,
                actual_value=actual,
                description=f"타입 불일치: {type(expected).__name__} vs {type(actual).__name__}",
                severity="medium"
            )
        
        # 숫자 비교
        if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
            if not self._numbers_equal(expected, actual):
                return CellDifference(
                    sheet=sheet,
                    cell=cell,
                    difference_type=DifferenceType.VALUE_MISMATCH,
                    expected_value=expected,
                    actual_value=actual,
                    description=f"숫자 값 차이: {abs(expected - actual)}",
                    severity="high" if abs(expected - actual) > 1 else "medium",
                    suggestion=f"값을 {expected}로 수정"
                )
        
        # 문자열 비교
        elif isinstance(expected, str) and isinstance(actual, str):
            if self.case_sensitive:
                if expected != actual:
                    return CellDifference(
                        sheet=sheet,
                        cell=cell,
                        difference_type=DifferenceType.VALUE_MISMATCH,
                        expected_value=expected,
                        actual_value=actual,
                        description="문자열 불일치",
                        severity="medium",
                        suggestion=f"값을 '{expected}'로 수정"
                    )
            else:
                if expected.lower() != actual.lower():
                    return CellDifference(
                        sheet=sheet,
                        cell=cell,
                        difference_type=DifferenceType.VALUE_MISMATCH,
                        expected_value=expected,
                        actual_value=actual,
                        description="문자열 불일치 (대소문자 무시)",
                        severity="low" if expected.lower() == actual.lower() else "medium",
                        suggestion=f"값을 '{expected}'로 수정"
                    )
        
        # 기타 타입 비교
        elif expected != actual:
            return CellDifference(
                sheet=sheet,
                cell=cell,
                difference_type=DifferenceType.VALUE_MISMATCH,
                expected_value=expected,
                actual_value=actual,
                description="값 불일치",
                severity="medium"
            )
        
        return None
    
    def _numbers_equal(self, a: float, b: float) -> bool:
        """부동소수점 수 비교"""
        if abs(a - b) < self.tolerance:
            return True
        # 상대 오차 확인
        if a != 0:
            return abs((a - b) / a) < self.tolerance
        return False
    
    def _compare_formulas(
        self,
        cell_expected: Any,
        cell_actual: Any,
        sheet: str,
        cell: str
    ) -> Optional[CellDifference]:
        """수식 비교"""
        
        formula_expected = cell_expected.formula if hasattr(cell_expected, 'formula') else None
        formula_actual = cell_actual.formula if hasattr(cell_actual, 'formula') else None
        
        if formula_expected != formula_actual:
            if formula_expected and not formula_actual:
                return CellDifference(
                    sheet=sheet,
                    cell=cell,
                    difference_type=DifferenceType.FORMULA_DIFFERENT,
                    expected_value=f"={formula_expected}",
                    actual_value=cell_actual.value,
                    description="수식이 값으로 변경됨",
                    severity="high",
                    suggestion=f"수식 '{formula_expected}' 복원"
                )
            elif not formula_expected and formula_actual:
                return CellDifference(
                    sheet=sheet,
                    cell=cell,
                    difference_type=DifferenceType.FORMULA_DIFFERENT,
                    expected_value=cell_expected.value,
                    actual_value=f"={formula_actual}",
                    description="값이 수식으로 변경됨",
                    severity="medium"
                )
            else:
                return CellDifference(
                    sheet=sheet,
                    cell=cell,
                    difference_type=DifferenceType.FORMULA_DIFFERENT,
                    expected_value=f"={formula_expected}",
                    actual_value=f"={formula_actual}",
                    description="수식이 다름",
                    severity="high",
                    suggestion=f"수식을 '{formula_expected}'로 수정"
                )
        
        return None
    
    def _compare_formats(
        self,
        cell_expected: Any,
        cell_actual: Any,
        sheet: str,
        cell: str
    ) -> Optional[CellDifference]:
        """서식 비교"""
        
        # 간단한 서식 비교 (폰트, 색상 등)
        format_differences = []
        
        # 폰트 비교
        if hasattr(cell_expected, 'font') and hasattr(cell_actual, 'font'):
            if cell_expected.font.bold != cell_actual.font.bold:
                format_differences.append("굵기")
            if cell_expected.font.italic != cell_actual.font.italic:
                format_differences.append("기울임")
            if cell_expected.font.size != cell_actual.font.size:
                format_differences.append("크기")
        
        # 숫자 형식 비교
        if hasattr(cell_expected, 'number_format') and hasattr(cell_actual, 'number_format'):
            if cell_expected.number_format != cell_actual.number_format:
                format_differences.append("숫자 형식")
        
        if format_differences:
            return CellDifference(
                sheet=sheet,
                cell=cell,
                difference_type=DifferenceType.FORMAT_DIFFERENT,
                expected_value="원본 서식",
                actual_value="변경된 서식",
                description=f"서식 차이: {', '.join(format_differences)}",
                severity="low"
            )
        
        return None
    
    def _get_common_sheets(self, wb1: Any, wb2: Any) -> List[str]:
        """공통 시트 이름 추출"""
        sheets1 = set(wb1.sheetnames)
        sheets2 = set(wb2.sheetnames)
        return list(sheets1.intersection(sheets2))
    
    def _create_summary(self, differences: List[CellDifference]) -> Dict[str, Any]:
        """차이점 요약 생성"""
        
        summary = {
            "by_type": {},
            "by_severity": {"low": 0, "medium": 0, "high": 0},
            "by_sheet": {},
            "top_issues": []
        }
        
        # 타입별 집계
        for diff in differences:
            diff_type = diff.difference_type.value
            summary["by_type"][diff_type] = summary["by_type"].get(diff_type, 0) + 1
            
            # 심각도별 집계
            summary["by_severity"][diff.severity] += 1
            
            # 시트별 집계
            if diff.sheet not in summary["by_sheet"]:
                summary["by_sheet"][diff.sheet] = 0
            summary["by_sheet"][diff.sheet] += 1
        
        # 주요 이슈 추출
        high_severity_diffs = [d for d in differences if d.severity == "high"]
        summary["top_issues"] = [
            {
                "cell": f"{d.sheet}!{d.cell}",
                "issue": d.description,
                "suggestion": d.suggestion
            }
            for d in high_severity_diffs[:5]
        ]
        
        return summary
    
    async def generate_comparison_report(
        self,
        result: ComparisonResult,
        output_format: str = "excel"
    ) -> str:
        """비교 분석 보고서 생성"""
        
        if output_format == "excel":
            return await self._generate_excel_report(result)
        elif output_format == "json":
            return await self._generate_json_report(result)
        else:
            raise ValueError(f"지원하지 않는 출력 형식: {output_format}")
    
    async def _generate_excel_report(self, result: ComparisonResult) -> str:
        """Excel 형식 보고서 생성"""
        
        wb = openpyxl.Workbook()
        
        # 요약 시트
        ws_summary = wb.active
        ws_summary.title = "요약"
        
        # 요약 정보 작성
        ws_summary["A1"] = "Excel 비교 분석 보고서"
        ws_summary["A3"] = "전체 셀 수:"
        ws_summary["B3"] = result.total_cells_compared
        ws_summary["A4"] = "차이점 발견:"
        ws_summary["B4"] = result.differences_found
        ws_summary["A5"] = "일치율:"
        ws_summary["B5"] = f"{result.match_percentage:.2f}%"
        ws_summary["A6"] = "실행 시간:"
        ws_summary["B6"] = f"{result.execution_time:.2f}초"
        
        # 차이점 상세 시트
        ws_details = wb.create_sheet("차이점 상세")
        headers = ["시트", "셀", "유형", "예상값", "실제값", "설명", "심각도", "제안사항"]
        ws_details.append(headers)
        
        for diff in result.differences:
            ws_details.append([
                diff.sheet,
                diff.cell,
                diff.difference_type.value,
                str(diff.expected_value),
                str(diff.actual_value),
                diff.description,
                diff.severity,
                diff.suggestion or ""
            ])
        
        # 파일 저장
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"comparison_report_{timestamp}.xlsx"
        wb.save(filename)
        
        return filename
    
    async def _generate_json_report(self, result: ComparisonResult) -> str:
        """JSON 형식 보고서 생성"""
        
        report_data = {
            "timestamp": datetime.now().isoformat(),
            "summary": {
                "total_cells_compared": result.total_cells_compared,
                "differences_found": result.differences_found,
                "match_percentage": result.match_percentage,
                "execution_time": result.execution_time
            },
            "differences": [
                {
                    "sheet": diff.sheet,
                    "cell": diff.cell,
                    "type": diff.difference_type.value,
                    "expected": str(diff.expected_value),
                    "actual": str(diff.actual_value),
                    "description": diff.description,
                    "severity": diff.severity,
                    "suggestion": diff.suggestion
                }
                for diff in result.differences
            ],
            "analysis": result.summary
        }
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"comparison_report_{timestamp}.json"
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, ensure_ascii=False, indent=2)
        
        return filename