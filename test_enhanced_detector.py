"""
Test enhanced formula detector
"""
import asyncio
import sys
import os
from openpyxl import Workbook
import tempfile

# Add project to path
sys.path.insert(0, '/Users/kevin/excel-unified/python-service')

from app.services.detection.strategies.enhanced_formula_detector import EnhancedFormulaDetector
from app.services.workbook_loader import OpenpyxlWorkbookLoader

async def test_enhanced_detector():
    # Create workbook with errors
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"

    # Add test data
    ws['A1'] = "Test Data"
    ws['B1'] = 10
    ws['C1'] = 0

    # Add various error formulas
    ws['A2'] = "=B1/C1"          # DIV/0 error
    ws['A3'] = "=VLOOKUP(\"X\",E1:F10,2,FALSE)"  # #N/A error
    ws['A4'] = "=UNKNOWNFUNC()"  # #NAME? error
    ws['A5'] = "=Sheet2!A1"      # #REF! error (no Sheet2)
    ws['A6'] = "=A1+B1"          # #VALUE! error (text + number)
    ws['A7'] = "=A8"             # Circular reference
    ws['A8'] = "=A7"             # Circular reference

    # Save to temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        temp_file = tmp.name

    print(f"Created test file: {temp_file}")

    # Load workbook
    loader = OpenpyxlWorkbookLoader()
    workbook = await loader.load_workbook(temp_file)

    # Test enhanced detector
    print("\n=== Testing EnhancedFormulaDetector ===")
    detector = EnhancedFormulaDetector()
    errors = await detector.detect(workbook)

    print(f"\nFound {len(errors)} errors:")
    for i, error in enumerate(errors):
        print(f"\n{i+1}. {error.type} at {error.sheet}!{error.cell}")
        print(f"   Message: {error.message}")
        print(f"   Formula: {error.formula}")
        print(f"   Severity: {error.severity}")
        print(f"   Auto-fixable: {error.is_auto_fixable}")

    # Clean up
    os.unlink(temp_file)
    print(f"\nDeleted test file: {temp_file}")

if __name__ == "__main__":
    asyncio.run(test_enhanced_detector())
