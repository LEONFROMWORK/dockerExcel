"""
Excel Error Detection Service
포괄적인 Excel 파일 오류 감지 및 분석 시스템
"""

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import pandas as pd
import re
from datetime import datetime
import logging
from dataclasses import dataclass
from typing import Any, List, Dict, Optional, Union
from .circular_reference_detector import CircularReferenceDetector, CircularReferenceChain

@dataclass
class ExcelError:
    """Excel 오류 정보를 담는 데이터 클래스"""
    error_type: str
    severity: str  # critical, high, medium, low
    location: str  # Sheet!A1
    description: str
    current_value: Any
    suggested_fix: str
    auto_fixable: bool
    fix_confidence: float  # 0-1

class ExcelErrorDetector:
    """Excel 파일의 다양한 오류를 감지하는 메인 클래스"""

    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.errors_found = []
        self.formula_errors = [
            '#DIV/0!', '#N/A', '#NAME?', '#NULL!', '#NUM!',
            '#REF!', '#VALUE!', '#GETTING_DATA', '#SPILL!', '#CALC!'
        ]

    def analyze_file(self, file_path: str) -> Dict[str, Any]:
        """Excel 파일을 전체적으로 분석하여 모든 오류를 감지"""

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False, keep_vba=True)

            analysis_result = {
                'file_path': file_path,
                'analysis_timestamp': datetime.now().isoformat(),
                'errors': [],
                'summary': {
                    'total_errors': 0,
                    'critical_errors': 0,
                    'auto_fixable_errors': 0,
                    'sheets_analyzed': len(workbook.sheetnames)
                },
                'recommendations': []
            }

            # 각 시트별로 오류 분석
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_errors = self._analyze_sheet(sheet, sheet_name)
                analysis_result['errors'].extend(sheet_errors)

            # 워크북 전체 수준의 오류 분석
            workbook_errors = self._analyze_workbook_structure(workbook)
            analysis_result['errors'].extend(workbook_errors)

            # 순환 참조 검사
            circular_refs = self._detect_circular_references(workbook)
            analysis_result['errors'].extend(circular_refs)

            # 요약 통계 계산
            analysis_result['summary'] = self._calculate_summary(analysis_result['errors'])

            # 수정 권장사항 생성
            analysis_result['recommendations'] = self._generate_recommendations(analysis_result['errors'])

            return analysis_result

        except Exception as e:
            self.logger.error(f"파일 분석 중 오류 발생: {str(e)}")
            return {
                'error': f'파일 분석 실패: {str(e)}',
                'file_path': file_path
            }

    def _analyze_sheet(self, sheet, sheet_name: str) -> List[ExcelError]:
        """개별 시트 분석"""

        errors = []

        # 1. 수식 오류 검사
        formula_errors = self._detect_formula_errors(sheet, sheet_name)
        errors.extend(formula_errors)

        # 2. 데이터 품질 문제 검사
        data_quality_errors = self._detect_data_quality_issues(sheet, sheet_name)
        errors.extend(data_quality_errors)

        # 3. 구조적 문제 검사
        structural_errors = self._detect_structural_issues(sheet, sheet_name)
        errors.extend(structural_errors)

        # 4. 서식 문제 검사
        formatting_errors = self._detect_formatting_issues(sheet, sheet_name)
        errors.extend(formatting_errors)

        return errors

    def _detect_formula_errors(self, sheet, sheet_name: str) -> List[ExcelError]:
        """수식 관련 오류 감지"""

        errors = []

        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # 수식 셀
                    # 1. 수식 오류 값 검사
                    if cell.value and str(cell.value).startswith('#'):
                        error_type = str(cell.value)
                        errors.append(ExcelError(
                            error_type='formula_error',
                            severity='high',
                            location=f"{sheet_name}!{cell.coordinate}",
                            description=f"수식 오류: {error_type}",
                            current_value=cell.value,
                            suggested_fix=self._suggest_formula_fix(cell, error_type),
                            auto_fixable=self._is_formula_auto_fixable(error_type),
                            fix_confidence=self._calculate_fix_confidence(error_type)
                        ))

                    # 2. 깨진 참조 검사
                    if cell.value and '#REF!' in str(cell.value):
                        errors.append(ExcelError(
                            error_type='broken_reference',
                            severity='critical',
                            location=f"{sheet_name}!{cell.coordinate}",
                            description="깨진 셀 참조",
                            current_value=cell.value,
                            suggested_fix="참조되는 셀이나 범위를 복구하거나 수식을 수정해야 합니다",
                            auto_fixable=False,
                            fix_confidence=0.3
                        ))

                    # 3. 비효율적인 수식 패턴 검사
                    inefficient_patterns = self._detect_inefficient_formulas(cell)
                    for pattern in inefficient_patterns:
                        errors.append(ExcelError(
                            error_type='inefficient_formula',
                            severity='medium',
                            location=f"{sheet_name}!{cell.coordinate}",
                            description=f"비효율적인 수식 패턴: {pattern['issue']}",
                            current_value=cell.value,
                            suggested_fix=pattern['suggestion'],
                            auto_fixable=True,
                            fix_confidence=0.8
                        ))

        return errors

    def _detect_data_quality_issues(self, sheet, sheet_name: str) -> List[ExcelError]:
        """데이터 품질 문제 감지"""

        errors = []

        # 데이터를 DataFrame으로 변환하여 분석
        try:
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(row)

            if not data:
                return errors

            df = pd.DataFrame(data[1:], columns=data[0] if data[0] else None)

            # 1. 중복 데이터 검사
            duplicates = df.duplicated().sum()
            if duplicates > 0:
                errors.append(ExcelError(
                    error_type='duplicate_data',
                    severity='medium',
                    location=f"{sheet_name}!전체",
                    description=f"{duplicates}개의 중복 행 발견",
                    current_value=f"{duplicates} 중복 행",
                    suggested_fix="중복 행을 제거하고 고유 데이터만 유지",
                    auto_fixable=True,
                    fix_confidence=0.9
                ))

            # 2. 빈 셀/누락 데이터 검사
            for col_idx, column in enumerate(df.columns):
                if column is not None:
                    missing_count = df[column].isna().sum()
                    if missing_count > len(df) * 0.5:  # 50% 이상 누락
                        col_letter = get_column_letter(col_idx + 1)
                        errors.append(ExcelError(
                            error_type='missing_data',
                            severity='medium',
                            location=f"{sheet_name}!{col_letter}:{col_letter}",
                            description=f"컬럼 '{column}'에서 {missing_count}개 값 누락 ({missing_count/len(df)*100:.1f}%)",
                            current_value=f"{missing_count} 누락값",
                            suggested_fix="누락된 값을 기본값으로 채우거나 해당 행을 제거",
                            auto_fixable=True,
                            fix_confidence=0.7
                        ))

            # 3. 데이터 타입 불일치 검사
            type_errors = self._detect_data_type_inconsistencies(df, sheet_name)
            errors.extend(type_errors)

        except Exception as e:
            self.logger.warning(f"데이터 품질 검사 중 오류: {str(e)}")

        return errors

    def _detect_structural_issues(self, sheet, sheet_name: str) -> List[ExcelError]:
        """구조적 문제 감지"""

        errors = []

        # 1. 병합된 셀 검사
        merged_ranges = list(sheet.merged_cells.ranges)
        if merged_ranges:
            for merged_range in merged_ranges:
                errors.append(ExcelError(
                    error_type='merged_cells',
                    severity='medium',
                    location=f"{sheet_name}!{merged_range}",
                    description="병합된 셀이 데이터 처리를 방해할 수 있음",
                    current_value=f"병합 범위: {merged_range}",
                    suggested_fix="병합을 해제하고 데이터를 적절히 분배",
                    auto_fixable=True,
                    fix_confidence=0.8
                ))

        # 2. 빈 행/열 검사
        empty_rows = self._find_empty_rows(sheet)
        if len(empty_rows) > 10:  # 10개 이상의 빈 행
            errors.append(ExcelError(
                error_type='excessive_empty_rows',
                severity='low',
                location=f"{sheet_name}!전체",
                description=f"{len(empty_rows)}개의 빈 행이 파일 크기를 증가시킴",
                current_value=f"{len(empty_rows)} 빈 행",
                suggested_fix="불필요한 빈 행을 제거하여 파일 최적화",
                auto_fixable=True,
                fix_confidence=0.9
            ))

        # 3. 테이블 헤더 문제 검사
        header_issues = self._detect_header_issues(sheet, sheet_name)
        errors.extend(header_issues)

        return errors

    def _detect_formatting_issues(self, sheet, sheet_name: str) -> List[ExcelError]:
        """서식 관련 문제 감지"""

        errors = []

        # 1. 일관성 없는 날짜 형식 검사
        date_format_issues = self._detect_date_format_inconsistencies(sheet, sheet_name)
        errors.extend(date_format_issues)

        # 2. 숫자 형식 문제 검사
        number_format_issues = self._detect_number_format_issues(sheet, sheet_name)
        errors.extend(number_format_issues)

        return errors

    def _detect_circular_references(self, workbook) -> List[ExcelError]:
        """고급 순환 참조 감지 시스템을 사용한 순환 참조 검사"""

        errors = []
        
        # 고급 순환 참조 감지기 사용
        detector = CircularReferenceDetector()
        circular_chains = detector.analyze_workbook(workbook)
        
        # 순환 참조 체인을 ExcelError 형식으로 변환
        for chain in circular_chains:
            # 첫 번째 셀을 대표 위치로 사용
            primary_location = chain.cells[0]
            
            # 가장 신뢰도 높은 수정 제안 선택
            primary_suggestion = chain.break_suggestions[0] if chain.break_suggestions else {
                'description': '순환 참조를 수동으로 검토하고 수정하세요'
            }
            
            # 자동 수정 가능 여부 결정 (간단한 직접 순환 참조만)
            auto_fixable = (chain.chain_type == 'direct' and 
                          len(chain.cells) == 2 and 
                          primary_suggestion.get('impact') == 'low')
            
            errors.append(ExcelError(
                error_type='circular_reference',
                severity=chain.severity,
                location=primary_location,
                description=chain.description,
                current_value=f"순환 체인: {' → '.join(chain.cells)}",
                suggested_fix=primary_suggestion['description'],
                auto_fixable=auto_fixable,
                fix_confidence=0.8 if auto_fixable else 0.3
            ))
            
            # 추가 제안사항을 별도 오류로 추가 (정보 제공용)
            if len(chain.break_suggestions) > 1:
                for i, suggestion in enumerate(chain.break_suggestions[1:3], 1):  # 상위 3개만
                    errors.append(ExcelError(
                        error_type='circular_reference_suggestion',
                        severity='low',
                        location=primary_location,
                        description=f"대안 수정 방법 {i}: {suggestion['description']}",
                        current_value=f"영향도: {suggestion.get('impact', 'unknown')}",
                        suggested_fix=suggestion['description'],
                        auto_fixable=False,
                        fix_confidence=0.5
                    ))

        return errors

    def _suggest_formula_fix(self, cell, error_type: str) -> str:
        """수식 오류에 대한 수정 제안"""

        suggestions = {
            '#DIV/0!': "분모가 0이 되지 않도록 IF 함수나 IFERROR 함수를 사용하세요",
            '#N/A': "VLOOKUP이나 MATCH 함수에서 값을 찾을 수 없습니다. 데이터 범위나 검색값을 확인하세요",
            '#NAME?': "함수명이나 참조명이 잘못되었습니다. 철자를 확인하세요",
            '#REF!': "참조하는 셀이나 범위가 삭제되었습니다. 올바른 참조로 수정하세요",
            '#VALUE!': "수식에 잘못된 데이터 타입이 사용되었습니다. 인수를 확인하세요",
            '#NUM!': "숫자가 너무 크거나 잘못된 숫자 연산입니다",
            '#NULL!': "범위 연산자(공백)가 잘못 사용되었습니다"
        }

        return suggestions.get(error_type, "수식을 검토하고 수정이 필요합니다")

    def _is_formula_auto_fixable(self, error_type: str) -> bool:
        """수식 오류가 자동 수정 가능한지 판단 (확장됨)"""

        auto_fixable_errors = [
            '#DIV/0!',      # 0으로 나누기 - IFERROR로 처리
            '#N/A',         # 값 찾기 실패 - IFERROR로 처리
            '#NAME?',       # 함수명 오타 - 자동 교정
            '#NULL!',       # 범위 연산자 오류 - 문법 수정
            '#VALUE!',      # 타입 오류 - VALUE() 함수로 변환
            '#NUM!',        # 숫자 범위 오류 - 조건부 처리
            '#SPILL!',      # 동적 배열 충돌 - 범위 조정
            '#REF!',        # 참조 오류 - 제한적 자동 수정
        ]
        return error_type in auto_fixable_errors

    def _calculate_fix_confidence(self, error_type: str) -> float:
        """수정 신뢰도 계산"""

        confidence_map = {
            '#DIV/0!': 0.9,
            '#N/A': 0.7,
            '#NAME?': 0.6,
            '#REF!': 0.3,
            '#VALUE!': 0.5,
            '#NUM!': 0.6,
            '#NULL!': 0.8
        }

        return confidence_map.get(error_type, 0.5)

    def _detect_inefficient_formulas(self, cell) -> List[Dict[str, str]]:
        """비효율적인 수식 패턴 감지"""

        issues = []

        if not cell.value:
            return issues

        formula = str(cell.value).upper()

        # VLOOKUP 대신 XLOOKUP 사용 권장
        if 'VLOOKUP' in formula:
            issues.append({
                'issue': 'VLOOKUP 사용',
                'suggestion': 'XLOOKUP 함수를 사용하면 더 효율적이고 유연합니다'
            })

        # 중첩된 IF 함수
        if formula.count('IF(') > 3:
            issues.append({
                'issue': '과도한 IF 중첩',
                'suggestion': 'SWITCH 함수나 룩업 테이블 사용을 고려하세요'
            })

        # 비효율적인 범위 참조
        if ':' in formula and ('1048576' in formula or '16384' in formula):
            issues.append({
                'issue': '전체 열/행 참조',
                'suggestion': '실제 데이터 범위만 참조하여 성능을 개선하세요'
            })

        return issues

    def _analyze_workbook_structure(self, workbook) -> List[ExcelError]:
        """워크북 전체 구조 분석"""

        errors = []

        # 1. 과도한 시트 수 검사
        if len(workbook.sheetnames) > 20:
            errors.append(ExcelError(
                error_type='excessive_sheets',
                severity='low',
                location='워크북 전체',
                description=f"시트가 {len(workbook.sheetnames)}개로 너무 많음",
                current_value=f"{len(workbook.sheetnames)} 시트",
                suggested_fix="사용하지 않는 시트를 제거하거나 별도 파일로 분리",
                auto_fixable=False,
                fix_confidence=0.3
            ))

        # 2. 빈 시트 검사
        empty_sheets = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            if sheet.max_row == 1 and sheet.max_column == 1:
                empty_sheets.append(sheet_name)

        if empty_sheets:
            errors.append(ExcelError(
                error_type='empty_sheets',
                severity='low',
                location=f"시트: {', '.join(empty_sheets)}",
                description=f"{len(empty_sheets)}개의 빈 시트 발견",
                current_value=empty_sheets,
                suggested_fix="빈 시트를 제거하여 파일 정리",
                auto_fixable=True,
                fix_confidence=0.9
            ))

        return errors

    def _find_empty_rows(self, sheet) -> List[int]:
        """빈 행 찾기"""

        empty_rows = []
        for row_num in range(1, sheet.max_row + 1):
            row_cells = [cell.value for cell in sheet[row_num]]
            if all(cell is None for cell in row_cells):
                empty_rows.append(row_num)

        return empty_rows

    def _detect_header_issues(self, sheet, sheet_name: str) -> List[ExcelError]:
        """헤더 관련 문제 감지"""

        errors = []

        if sheet.max_row == 0:
            return errors

        # 첫 번째 행을 헤더로 가정
        header_row = [cell.value for cell in sheet[1]]

        # 1. 중복 헤더 검사
        non_empty_headers = [h for h in header_row if h is not None]
        if len(non_empty_headers) != len(set(non_empty_headers)):
            errors.append(ExcelError(
                error_type='duplicate_headers',
                severity='medium',
                location=f"{sheet_name}!1:1",
                description="중복된 헤더명 발견",
                current_value=header_row,
                suggested_fix="헤더명을 고유하게 수정",
                auto_fixable=True,
                fix_confidence=0.8
            ))

        # 2. 빈 헤더 검사
        empty_header_count = header_row.count(None)
        if empty_header_count > 0:
            errors.append(ExcelError(
                error_type='empty_headers',
                severity='medium',
                location=f"{sheet_name}!1:1",
                description=f"{empty_header_count}개의 빈 헤더 발견",
                current_value=header_row,
                suggested_fix="빈 헤더에 적절한 이름을 지정",
                auto_fixable=True,
                fix_confidence=0.7
            ))

        return errors

    def _detect_data_type_inconsistencies(self, df: pd.DataFrame, sheet_name: str) -> List[ExcelError]:
        """데이터 타입 불일치 감지"""

        errors = []

        for col_idx, column in enumerate(df.columns):
            if column is None:
                continue

            col_data = df[column].dropna()
            if len(col_data) == 0:
                continue

            # 숫자로 보이는 데이터가 텍스트로 저장된 경우
            numeric_pattern = re.compile(r'^-?\d+\.?\d*$')
            text_numbers = []

            for idx, value in col_data.items():
                if isinstance(value, str) and numeric_pattern.match(value.strip()):
                    text_numbers.append(idx + 2)  # Excel 행 번호 (헤더 제외)

            if len(text_numbers) > len(col_data) * 0.3:  # 30% 이상이 텍스트 숫자
                col_letter = get_column_letter(col_idx + 1)
                errors.append(ExcelError(
                    error_type='text_stored_as_number',
                    severity='medium',
                    location=f"{sheet_name}!{col_letter}:{col_letter}",
                    description=f"컬럼 '{column}'에서 숫자가 텍스트로 저장됨",
                    current_value=f"{len(text_numbers)}개 텍스트 숫자",
                    suggested_fix="텍스트로 저장된 숫자를 숫자 형식으로 변환",
                    auto_fixable=True,
                    fix_confidence=0.85
                ))

        return errors

    def _detect_date_format_inconsistencies(self, sheet, sheet_name: str) -> List[ExcelError]:
        """날짜 형식 불일치 감지"""

        errors = []

        # 날짜 패턴들
        date_patterns = [
            r'\d{4}-\d{2}-\d{2}',  # 2023-01-15
            r'\d{2}/\d{2}/\d{4}',  # 01/15/2023
            r'\d{2}-\d{2}-\d{4}',  # 15-01-2023
            r'\d{4}/\d{2}/\d{2}',  # 2023/01/15
        ]

        date_formats_found = set()
        date_cells = []

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    for pattern in date_patterns:
                        if re.match(pattern, str(cell.value).strip()):
                            date_formats_found.add(pattern)
                            date_cells.append(cell.coordinate)
                            break

        # 2개 이상의 다른 날짜 형식이 발견된 경우
        if len(date_formats_found) > 1:
            errors.append(ExcelError(
                error_type='inconsistent_date_format',
                severity='medium',
                location=f"{sheet_name}!다중 위치",
                description=f"{len(date_formats_found)}가지 다른 날짜 형식 발견",
                current_value=list(date_formats_found),
                suggested_fix="일관된 날짜 형식으로 표준화",
                auto_fixable=True,
                fix_confidence=0.8
            ))

        return errors

    def _detect_number_format_issues(self, sheet, sheet_name: str) -> List[ExcelError]:
        """숫자 형식 문제 감지"""

        errors = []

        # 통화 형식 불일치 검사
        currency_symbols = set()
        currency_cells = []

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    # 통화 기호 찾기
                    currency_pattern = r'[₩$€£¥]'
                    if re.search(currency_pattern, str(cell.value)):
                        symbols = re.findall(currency_pattern, str(cell.value))
                        currency_symbols.update(symbols)
                        currency_cells.append(cell.coordinate)

        if len(currency_symbols) > 1:
            errors.append(ExcelError(
                error_type='mixed_currency_symbols',
                severity='medium',
                location=f"{sheet_name}!다중 위치",
                description=f"다양한 통화 기호 사용: {', '.join(currency_symbols)}",
                current_value=list(currency_symbols),
                suggested_fix="일관된 통화 기호로 표준화",
                auto_fixable=True,
                fix_confidence=0.7
            ))

        return errors

    def _calculate_summary(self, errors: List[ExcelError]) -> Dict[str, Any]:
        """오류 요약 통계 계산"""

        summary = {
            'total_errors': len(errors),
            'critical_errors': 0,
            'high_errors': 0,
            'medium_errors': 0,
            'low_errors': 0,
            'auto_fixable_errors': 0,
            'error_types': {}
        }

        for error in errors:
            # 심각도별 카운트
            if error.severity == 'critical':
                summary['critical_errors'] += 1
            elif error.severity == 'high':
                summary['high_errors'] += 1
            elif error.severity == 'medium':
                summary['medium_errors'] += 1
            elif error.severity == 'low':
                summary['low_errors'] += 1

            # 자동 수정 가능 오류
            if error.auto_fixable:
                summary['auto_fixable_errors'] += 1

            # 오류 타입별 카운트
            error_type = error.error_type
            if error_type not in summary['error_types']:
                summary['error_types'][error_type] = 0
            summary['error_types'][error_type] += 1

        return summary

    def _generate_recommendations(self, errors: List[ExcelError]) -> List[str]:
        """수정 권장사항 생성"""

        recommendations = []

        # 심각도별 권장사항
        critical_count = sum(1 for e in errors if e.severity == 'critical')
        if critical_count > 0:
            recommendations.append(f"🚨 {critical_count}개의 심각한 오류를 즉시 수정해야 합니다")

        # 자동 수정 가능한 오류
        auto_fixable = sum(1 for e in errors if e.auto_fixable)
        if auto_fixable > 0:
            recommendations.append(f"✅ {auto_fixable}개의 오류는 자동으로 수정할 수 있습니다")

        # 특정 오류 타입별 권장사항
        error_types = {}
        for error in errors:
            if error.error_type not in error_types:
                error_types[error.error_type] = 0
            error_types[error.error_type] += 1

        if error_types.get('formula_error', 0) > 5:
            recommendations.append("🔧 수식 오류가 많습니다. 수식 검토 및 IFERROR 함수 사용을 고려하세요")

        if error_types.get('duplicate_data', 0) > 0:
            recommendations.append("🗂️ 중복 데이터를 제거하여 데이터 품질을 향상시키세요")

        if error_types.get('merged_cells', 0) > 0:
            recommendations.append("📊 병합된 셀을 해제하여 데이터 분석 도구와의 호환성을 개선하세요")

        return recommendations
