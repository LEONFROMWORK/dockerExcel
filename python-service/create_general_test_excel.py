#!/usr/bin/env python3
"""
일반적인 Excel 파일 생성 - 기본 기능 테스트용
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference
from datetime import date, datetime, timedelta
import random

# 워크북 생성
wb = Workbook()
ws = wb.active
ws.title = "판매 데이터"

# 헤더 스타일 정의
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="366092")
header_alignment = Alignment(horizontal="center", vertical="center")

# 테두리 스타일
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 1. 판매 데이터 시트
# 헤더 추가
headers = ["날짜", "제품명", "카테고리", "수량", "단가", "총액", "할인율", "최종금액"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 샘플 데이터
products = [
    ("노트북", "전자제품", 45000),
    ("마우스", "액세서리", 25000),
    ("키보드", "액세서리", 35000),
    ("모니터", "전자제품", 180000),
    ("USB 케이블", "액세서리", 8000),
    ("웹캠", "전자제품", 65000),
    ("스피커", "전자제품", 55000),
    ("헤드셋", "액세서리", 45000)
]

# 데이터 입력
start_date = date(2024, 1, 1)
for row in range(2, 52):  # 50개 행
    # 날짜
    ws.cell(row=row, column=1, value=start_date).number_format = 'yyyy-mm-dd'
    start_date = start_date + timedelta(days=7)
    
    # 제품 정보
    product = random.choice(products)
    ws.cell(row=row, column=2, value=product[0])  # 제품명
    ws.cell(row=row, column=3, value=product[1])  # 카테고리
    
    # 수량
    quantity = random.randint(1, 20)
    ws.cell(row=row, column=4, value=quantity)
    
    # 단가
    unit_price = product[2]
    ws.cell(row=row, column=5, value=unit_price).number_format = '₩#,##0'
    
    # 총액 (수식)
    ws.cell(row=row, column=6, value=f'=D{row}*E{row}').number_format = '₩#,##0'
    
    # 할인율
    discount = random.choice([0, 0.05, 0.1, 0.15, 0.2])
    ws.cell(row=row, column=7, value=discount).number_format = '0%'
    
    # 최종금액 (수식)
    ws.cell(row=row, column=8, value=f'=F{row}*(1-G{row})').number_format = '₩#,##0'
    
    # 테두리 적용
    for col in range(1, 9):
        ws.cell(row=row, column=col).border = thin_border

# 열 너비 조정
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 12

# 요약 행 추가
summary_row = 53
ws.cell(row=summary_row, column=2, value="합계").font = Font(bold=True)
ws.cell(row=summary_row, column=4, value=f'=SUM(D2:D51)').font = Font(bold=True)
ws.cell(row=summary_row, column=6, value=f'=SUM(F2:F51)').font = Font(bold=True)
ws.cell(row=summary_row, column=6).number_format = '₩#,##0'
ws.cell(row=summary_row, column=8, value=f'=SUM(H2:H51)').font = Font(bold=True)
ws.cell(row=summary_row, column=8).number_format = '₩#,##0'

# 2. 요약 시트 생성
ws2 = wb.create_sheet("요약")

# 카테고리별 매출 요약
ws2['A1'] = "카테고리별 매출 요약"
ws2['A1'].font = Font(bold=True, size=14)
ws2.merge_cells('A1:C1')

ws2['A3'] = "카테고리"
ws2['B3'] = "총 수량"
ws2['C3'] = "총 매출"

# 헤더 스타일 적용
for col in range(1, 4):
    cell = ws2.cell(row=3, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

# 카테고리별 데이터 (SUMIF 수식 사용)
categories = ["전자제품", "액세서리"]
for idx, category in enumerate(categories, 4):
    ws2.cell(row=idx, column=1, value=category)
    ws2.cell(row=idx, column=2, value=f"=SUMIF('판매 데이터'!C:C,A{idx},'판매 데이터'!D:D)")
    ws2.cell(row=idx, column=3, value=f"=SUMIF('판매 데이터'!C:C,A{idx},'판매 데이터'!H:H)")
    ws2.cell(row=idx, column=3).number_format = '₩#,##0'

# 3. 차트 추가
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "카테고리별 매출"
chart.y_axis.title = '매출액'
chart.x_axis.title = '카테고리'

data = Reference(ws2, min_col=3, min_row=3, max_row=5, max_col=3)
cats = Reference(ws2, min_col=1, min_row=4, max_row=5)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.shape = 4
ws2.add_chart(chart, "E3")

# 4. 셀 병합 예제
ws2['A10'] = "월별 실적 보고서"
ws2['A10'].font = Font(bold=True, size=16)
ws2['A10'].alignment = Alignment(horizontal="center", vertical="center")
ws2.merge_cells('A10:H12')

# 병합된 셀에 테두리 적용
thick_border = Border(
    left=Side(style='thick'),
    right=Side(style='thick'),
    top=Side(style='thick'),
    bottom=Side(style='thick')
)
for row in range(10, 13):
    for col in range(1, 9):
        ws2.cell(row=row, column=col).border = thick_border

# 5. 고정 창 설정
ws.freeze_panes = 'A2'  # 첫 번째 행 고정
ws2.freeze_panes = 'A4'  # 세 번째 행까지 고정

# 파일 저장
wb.save("general_test.xlsx")
print("일반 테스트 파일 생성 완료: general_test.xlsx")
print("\n포함된 기능:")
print("- 텍스트, 숫자, 날짜 데이터")
print("- 수식 (SUM, SUMIF, 기본 산술)")
print("- 서식 (글꼴, 색상, 테두리)")
print("- 숫자 형식 (통화, 백분율, 날짜)")
print("- 셀 병합")
print("- 차트")
print("- 고정 창")
print("- 여러 시트")