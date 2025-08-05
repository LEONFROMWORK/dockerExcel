"""
Table builder - specialized for Excel table operations
"""

from typing import Dict, Optional, Tuple
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import pandas as pd

from .configs import TableConfig, ValidationConfig


class TableBuilder:
    """Specialized builder for Excel tables and data operations"""

    def create_table(self, worksheet: Worksheet, config: TableConfig) -> Table:
        """Create a formatted table in the worksheet"""
        table = Table(
            displayName=config.table_name, ref=f"{config.start_cell}:{config.end_cell}"
        )
        table.tableStyleInfo = TableStyleInfo(
            name=config.style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
        return table

    def create_table_legacy(
        self,
        worksheet: Worksheet,
        start_cell: str,
        end_cell: str,
        table_name: str,
        style: str = "TableStyleMedium9",
    ) -> Table:
        """Legacy method for backward compatibility"""
        config = TableConfig(start_cell, end_cell, table_name, style)
        return self.create_table(worksheet, config)

    def write_dataframe(
        self,
        worksheet: Worksheet,
        df: pd.DataFrame,
        start_row: int = 1,
        start_col: int = 1,
    ) -> Tuple[int, int]:
        """Write DataFrame to worksheet"""
        # Write headers
        for col_idx, column in enumerate(df.columns):
            worksheet.cell(row=start_row, column=start_col + col_idx, value=str(column))

        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False)):
            for col_idx, value in enumerate(row):
                worksheet.cell(
                    row=start_row + row_idx + 1, column=start_col + col_idx, value=value
                )

        # Return end position
        end_row = start_row + len(df)
        end_col = start_col + len(df.columns) - 1
        return end_row, end_col

    def add_data_validation(
        self, worksheet: Worksheet, config: ValidationConfig
    ) -> None:
        """Add data validation to cells"""
        dv = DataValidation(
            type=config.validation_type,
            formula1=config.formula1,
            formula2=config.formula2,
            allow_blank=config.allow_blank,
        )
        dv.error = config.error_message
        dv.errorTitle = config.error_title
        dv.prompt = config.prompt_message
        dv.promptTitle = config.prompt_title

        worksheet.add_data_validation(dv)
        dv.add(config.cell_range)

    def add_data_validation_legacy(
        self,
        worksheet: Worksheet,
        cell_range: str,
        validation_type: str,
        formula1: str,
        formula2: Optional[str] = None,
        allow_blank: bool = True,
    ) -> None:
        """Legacy method for backward compatibility"""
        config = ValidationConfig(
            cell_range=cell_range,
            validation_type=validation_type,
            formula1=formula1,
            formula2=formula2,
            allow_blank=allow_blank,
        )
        self.add_data_validation(worksheet, config)

    def create_summary_row(
        self,
        worksheet: Worksheet,
        row_number: int,
        label: str,
        formulas: Dict[int, str],
    ) -> None:
        """Create a summary row with formulas"""
        # Add label
        worksheet.cell(row=row_number, column=1, value=label)

        # Add formulas
        for col, formula in formulas.items():
            worksheet.cell(row=row_number, column=col, value=formula)

    def auto_adjust_columns(
        self, worksheet: Worksheet, min_width: float = 8.0, max_width: float = 50.0
    ) -> None:
        """Automatically adjust column widths based on content"""
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            adjusted_width = min(max(length + 2, min_width), max_width)
            column_letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[column_letter].width = adjusted_width
