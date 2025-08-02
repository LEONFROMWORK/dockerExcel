#!/usr/bin/env python3
"""
RevoGrid 테스트용 샘플 Excel 파일 생성
다양한 포맷팅과 기능을 포함한 데모 파일
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def create_sample_excel():
    """샘플 Excel 파일 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = "샘플 데이터"
    
    # 헤더 행 생성
    headers = [
        "제품명", "카테고리", "가격", "재고", "판매량", 
        "총매출", "등급", "출시일", "상태", "비고"
    ]
    
    # 헤더 스타일
    header_font = Font(name="맑은 고딕", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_border = Border(
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin"),
        left=Side(border_style="thin"),
        right=Side(border_style="thin")
    )
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 헤더 작성
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = header_alignment
    
    # 샘플 데이터
    sample_data = [
        ["갤럭시 스마트폰", "전자제품", 850000, 50, 120, "=C2*E2", "A+", "2024-01-15", "판매중", "인기상품"],
        ["아이폰", "전자제품", 1200000, 30, 85, "=C3*E3", "A", "2024-02-01", "판매중", ""],
        ["노트북", "컴퓨터", 1500000, 25, 60, "=C4*E4", "A", "2024-01-20", "판매중", "고성능"],
        ["무선 이어폰", "오디오", 150000, 100, 200, "=C5*E5", "B+", "2023-12-10", "판매중", "베스트셀러"],
        ["태블릿", "전자제품", 600000, 40, 75, "=C6*E6", "B", "2024-01-05", "재고부족", ""],
        ["스마트워치", "웨어러블", 300000, 60, 95, "=C7*E7", "B+", "2023-11-20", "판매중", "건강관리"],
        ["게임 콘솔", "게임", 500000, 20, 45, "=C8*E8", "A-", "2024-02-15", "예약판매", "신제품"],
        ["키보드", "컴퓨터", 80000, 150, 180, "=C9*E9", "C+", "2023-10-01", "판매중", "기계식"],
        ["마우스", "컴퓨터", 45000, 200, 250, "=C10*E10", "C", "2023-09-15", "판매중", "게이밍"],
        ["모니터", "컴퓨터", 350000, 35, 65, "=C11*E11", "B", "2024-01-10", "판매중", "4K 지원"]
    ]
    
    # 데이터 행 스타일
    data_font = Font(name="맑은 고딕", size=10)
    data_border = Border(
        top=Side(border_style="thin", color="D0D0D0"),
        bottom=Side(border_style="thin", color="D0D0D0"),
        left=Side(border_style="thin", color="D0D0D0"),
        right=Side(border_style="thin", color="D0D0D0")
    )
    
    # 데이터 작성
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = data_border
            
            # 컬럼별 특별 스타일
            if col_idx == 3:  # 가격 컬럼
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 4 or col_idx == 5:  # 재고, 판매량 컬럼
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 6:  # 총매출 컬럼
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
                cell.font = Font(name="맑은 고딕", size=10, bold=True)
            elif col_idx == 7:  # 등급 컬럼
                cell.alignment = Alignment(horizontal="center")
                # 등급별 색상
                if value == "A+":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif value == "A":
                    cell.fill = PatternFill(start_color="D4F1D4", end_color="D4F1D4", fill_type="solid")
                elif value.startswith("B"):
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            elif col_idx == 8:  # 출시일 컬럼
                cell.number_format = 'YYYY-MM-DD'
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 9:  # 상태 컬럼
                cell.alignment = Alignment(horizontal="center")
                # 상태별 색상
                if value == "판매중":
                    cell.font = Font(name="맑은 고딕", size=10, color="008000")
                elif value == "재고부족":
                    cell.font = Font(name="맑은 고딕", size=10, color="FF8C00")
                elif value == "예약판매":
                    cell.font = Font(name="맑은 고딕", size=10, color="0066CC")
    
    # 컬럼 너비 조정
    column_widths = [15, 12, 12, 8, 10, 15, 8, 12, 12, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # 행 높이 조정
    ws.row_dimensions[1].height = 25  # 헤더 행
    for row in range(2, len(sample_data) + 2):
        ws.row_dimensions[row].height = 20
    
    # 병합 셀 추가 (제목용)
    ws.insert_rows(1)
    ws.merge_cells('A1:J1')
    title_cell = ws['A1']
    title_cell.value = "🛍️ 전자제품 판매 현황 대시보드"
    title_cell.font = Font(name="맑은 고딕", size=16, bold=True, color="1F4E79")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    ws.row_dimensions[1].height = 35
    
    # 두 번째 시트 추가 (차트 데이터용)
    ws2 = wb.create_sheet("차트 데이터")
    
    # 차트용 데이터
    chart_headers = ["월", "매출액", "목표", "달성률"]
    months = ["1월", "2월", "3월", "4월", "5월", "6월"]
    sales = [1200, 1350, 1100, 1450, 1600, 1750]
    targets = [1300, 1400, 1200, 1500, 1550, 1700]
    
    # 차트 헤더
    for col, header in enumerate(chart_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    
    # 차트 데이터
    for row, (month, sale, target) in enumerate(zip(months, sales, targets), 2):
        ws2.cell(row=row, column=1, value=month)
        ws2.cell(row=row, column=2, value=sale)
        ws2.cell(row=row, column=3, value=target)
        # 달성률 계산
        achievement = f"=B{row}/C{row}"
        ws2.cell(row=row, column=4, value=achievement)
        ws2.cell(row=row, column=4).number_format = '0.0%'
    
    # 파일 저장
    filename = "sample_excel_demo.xlsx"
    wb.save(filename)
    print(f"✅ 샘플 Excel 파일 생성 완료: {filename}")
    return filename

if __name__ == "__main__":
    create_sample_excel()