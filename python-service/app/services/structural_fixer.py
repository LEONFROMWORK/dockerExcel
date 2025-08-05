"""
Excel 구조적 문제 수정 도구
Excel Structural Problem Fixer - 병합 셀, 빈 행/열, 시트 구조 등 수정
"""

import logging
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from typing import Dict, List, Any, Optional, Tuple
import re

logger = logging.getLogger(__name__)


class StructuralFixer:
    """Excel 파일의 구조적 문제를 수정하는 클래스"""

    def __init__(self):
        self.fixes_applied = []
        self.warnings = []

    def fix_structural_issues(
        self, workbook: Workbook, issues: List[Dict]
    ) -> Dict[str, Any]:
        """구조적 문제들을 종합적으로 수정"""

        results = {
            "merged_cells_fixed": 0,
            "empty_rows_removed": 0,
            "empty_columns_removed": 0,
            "sheets_cleaned": 0,
            "tables_normalized": 0,
            "headers_fixed": 0,
            "formatting_cleaned": 0,
            "fixes_applied": [],
            "warnings": [],
        }

        try:
            # 1. 병합된 셀 처리
            merged_fixes = self._fix_merged_cells(workbook, issues)
            results["merged_cells_fixed"] = merged_fixes["count"]
            results["fixes_applied"].extend(merged_fixes["details"])

            # 2. 빈 행/열 제거
            empty_fixes = self._remove_empty_rows_columns(workbook, issues)
            results["empty_rows_removed"] = empty_fixes["rows_removed"]
            results["empty_columns_removed"] = empty_fixes["columns_removed"]
            results["fixes_applied"].extend(empty_fixes["details"])

            # 3. 시트 정리
            sheet_fixes = self._clean_sheets(workbook, issues)
            results["sheets_cleaned"] = sheet_fixes["count"]
            results["fixes_applied"].extend(sheet_fixes["details"])

            # 4. 테이블 구조 정규화
            table_fixes = self._normalize_table_structures(workbook, issues)
            results["tables_normalized"] = table_fixes["count"]
            results["fixes_applied"].extend(table_fixes["details"])

            # 5. 헤더 행 수정
            header_fixes = self._fix_headers(workbook, issues)
            results["headers_fixed"] = header_fixes["count"]
            results["fixes_applied"].extend(header_fixes["details"])

            # 6. 불필요한 서식 정리
            format_fixes = self._clean_formatting(workbook, issues)
            results["formatting_cleaned"] = format_fixes["count"]
            results["fixes_applied"].extend(format_fixes["details"])

            results["warnings"] = self.warnings

            logger.info(
                f"구조적 수정 완료: {len(results['fixes_applied'])}개 수정사항 적용"
            )

        except Exception as e:
            logger.error(f"구조적 수정 중 오류: {str(e)}")
            results["error"] = str(e)

        return results

    def _fix_merged_cells(
        self, workbook: Workbook, issues: List[Dict]
    ) -> Dict[str, Any]:
        """병합된 셀 수정"""

        fixes = {"count": 0, "details": []}

        for worksheet in workbook.worksheets:
            merged_ranges = list(worksheet.merged_cells.ranges)

            for merged_range in merged_ranges:
                try:
                    # 병합된 셀의 값 추출
                    top_left_cell = worksheet[merged_range.coord.split(":")[0]]
                    merged_value = top_left_cell.value

                    # 병합 해제
                    worksheet.unmerge_cells(merged_range.coord)

                    # 전략적 값 배치
                    if merged_value is not None:
                        # 병합된 영역의 첫 번째 셀에만 값 유지
                        top_left_cell.value = merged_value

                        # 나머지 셀들은 비워둠
                        min_row, min_col, max_row, max_col = merged_range.bounds
                        for row in range(min_row, max_row + 1):
                            for col in range(min_col, max_col + 1):
                                if row == min_row and col == min_col:
                                    continue  # 첫 번째 셀은 건드리지 않음
                                worksheet.cell(row=row, column=col).value = None

                    fixes["count"] += 1
                    fixes["details"].append(
                        {
                            "type": "merged_cell_unmerged",
                            "location": f"{worksheet.title}!{merged_range.coord}",
                            "action": f"병합 해제: {merged_range.coord}",
                            "value_preserved": merged_value is not None,
                        }
                    )

                except Exception as e:
                    self.warnings.append(
                        f"병합 셀 수정 실패 {merged_range.coord}: {str(e)}"
                    )

        return fixes

    def _remove_empty_rows_columns(
        self, workbook: Workbook, issues: List[Dict]
    ) -> Dict[str, Any]:
        """빈 행과 열 제거"""

        fixes = {"rows_removed": 0, "columns_removed": 0, "details": []}

        for worksheet in workbook.worksheets:
            try:
                # 빈 행 제거
                rows_to_delete = []
                for row_num in range(worksheet.max_row, 0, -1):
                    row_cells = worksheet[row_num]
                    if all(
                        cell.value is None or str(cell.value).strip() == ""
                        for cell in row_cells
                    ):
                        # 연속된 빈 행이 5개 이상인 경우에만 제거
                        empty_count = 0
                        for check_row in range(
                            row_num, min(row_num + 5, worksheet.max_row + 1)
                        ):
                            check_cells = worksheet[check_row]
                            if all(
                                cell.value is None or str(cell.value).strip() == ""
                                for cell in check_cells
                            ):
                                empty_count += 1
                            else:
                                break

                        if empty_count >= 3:  # 3개 이상 연속 빈 행
                            rows_to_delete.append(row_num)

                # 행 삭제 실행
                for row_num in rows_to_delete:
                    worksheet.delete_rows(row_num)
                    fixes["rows_removed"] += 1
                    fixes["details"].append(
                        {
                            "type": "empty_row_removed",
                            "location": f"{worksheet.title}!행{row_num}",
                            "action": f"빈 행 제거: 행 {row_num}",
                        }
                    )

                # 빈 열 제거
                cols_to_delete = []
                for col_num in range(worksheet.max_column, 0, -1):
                    col_letter = get_column_letter(col_num)
                    col_cells = [
                        worksheet[f"{col_letter}{row}"]
                        for row in range(1, worksheet.max_row + 1)
                    ]

                    if all(
                        cell.value is None or str(cell.value).strip() == ""
                        for cell in col_cells
                    ):
                        # 연속된 빈 열이 3개 이상인 경우에만 제거
                        empty_count = 0
                        for check_col in range(
                            col_num, min(col_num + 3, worksheet.max_column + 1)
                        ):
                            check_letter = get_column_letter(check_col)
                            check_cells = [
                                worksheet[f"{check_letter}{row}"]
                                for row in range(1, worksheet.max_row + 1)
                            ]
                            if all(
                                cell.value is None or str(cell.value).strip() == ""
                                for cell in check_cells
                            ):
                                empty_count += 1
                            else:
                                break

                        if empty_count >= 2:  # 2개 이상 연속 빈 열
                            cols_to_delete.append(col_num)

                # 열 삭제 실행
                for col_num in cols_to_delete:
                    worksheet.delete_cols(col_num)
                    fixes["columns_removed"] += 1
                    fixes["details"].append(
                        {
                            "type": "empty_column_removed",
                            "location": f"{worksheet.title}!열{get_column_letter(col_num)}",
                            "action": f"빈 열 제거: 열 {get_column_letter(col_num)}",
                        }
                    )

            except Exception as e:
                self.warnings.append(
                    f"시트 {worksheet.title} 빈 행/열 제거 실패: {str(e)}"
                )

        return fixes

    def _clean_sheets(self, workbook: Workbook, issues: List[Dict]) -> Dict[str, Any]:
        """빈 시트 및 불필요한 시트 정리"""

        fixes = {"count": 0, "details": []}

        sheets_to_remove = []

        for worksheet in workbook.worksheets:
            try:
                # 시트가 완전히 비어있는지 확인
                has_data = False
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value is not None and str(cell.value).strip() != "":
                            has_data = True
                            break
                    if has_data:
                        break

                # 빈 시트이고 시트가 2개 이상인 경우에만 제거
                if not has_data and len(workbook.worksheets) > 1:
                    # 기본 시트명이거나 의미없는 이름인 경우
                    if (
                        worksheet.title.lower()
                        in [
                            "sheet",
                            "sheet1",
                            "sheet2",
                            "sheet3",
                            "시트1",
                            "시트2",
                            "시트3",
                        ]
                        or re.match(r"^sheet\d+$", worksheet.title.lower())
                        or re.match(r"^시트\d+$", worksheet.title)
                    ):

                        sheets_to_remove.append(worksheet.title)

            except Exception as e:
                self.warnings.append(f"시트 {worksheet.title} 정리 검사 실패: {str(e)}")

        # 시트 제거 실행
        for sheet_name in sheets_to_remove:
            try:
                workbook.remove(workbook[sheet_name])
                fixes["count"] += 1
                fixes["details"].append(
                    {
                        "type": "empty_sheet_removed",
                        "location": f"시트: {sheet_name}",
                        "action": f"빈 시트 제거: {sheet_name}",
                    }
                )
            except Exception as e:
                self.warnings.append(f"시트 {sheet_name} 제거 실패: {str(e)}")

        return fixes

    def _normalize_table_structures(
        self, workbook: Workbook, issues: List[Dict]
    ) -> Dict[str, Any]:
        """테이블 구조 정규화"""

        fixes = {"count": 0, "details": []}

        for worksheet in workbook.worksheets:
            try:
                # 데이터 범위 식별
                data_range = self._identify_data_range(worksheet)

                if data_range:
                    min_row, min_col, max_row, max_col = data_range

                    # 헤더 행 정규화
                    header_fixed = self._normalize_headers(
                        worksheet, min_row, min_col, max_col
                    )
                    if header_fixed:
                        fixes["count"] += 1
                        fixes["details"].append(
                            {
                                "type": "table_headers_normalized",
                                "location": f"{worksheet.title}!{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{min_row}",
                                "action": "테이블 헤더 정규화",
                            }
                        )

                    # 데이터 타입 일관성 확인 및 수정
                    type_fixes = self._fix_column_data_types(
                        worksheet, min_row + 1, min_col, max_row, max_col
                    )
                    fixes["count"] += type_fixes
                    if type_fixes > 0:
                        fixes["details"].append(
                            {
                                "type": "column_types_normalized",
                                "location": f"{worksheet.title}!데이터 영역",
                                "action": f"{type_fixes}개 열의 데이터 타입 정규화",
                            }
                        )

            except Exception as e:
                self.warnings.append(
                    f"시트 {worksheet.title} 테이블 정규화 실패: {str(e)}"
                )

        return fixes

    def _identify_data_range(self, worksheet) -> Optional[Tuple[int, int, int, int]]:
        """워크시트에서 실제 데이터 범위 식별"""

        min_row = min_col = float("inf")
        max_row = max_col = 0

        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None and str(cell.value).strip() != "":
                    min_row = min(min_row, cell.row)
                    max_row = max(max_row, cell.row)
                    min_col = min(min_col, cell.column)
                    max_col = max(max_col, cell.column)

        if min_row == float("inf"):
            return None

        return (min_row, min_col, max_row, max_col)

    def _normalize_headers(
        self, worksheet, header_row: int, min_col: int, max_col: int
    ) -> bool:
        """헤더 행 정규화"""

        fixed = False

        for col in range(min_col, max_col + 1):
            cell = worksheet.cell(row=header_row, column=col)

            if cell.value is None or str(cell.value).strip() == "":
                # 빈 헤더에 기본값 제공
                cell.value = f"Column_{get_column_letter(col)}"
                fixed = True
            else:
                # 헤더 텍스트 정리
                original_value = str(cell.value).strip()
                clean_value = re.sub(r"\s+", " ", original_value)  # 여러 공백을 하나로
                clean_value = re.sub(r"[^\w\s가-힣]", "", clean_value)  # 특수문자 제거

                if clean_value != original_value:
                    cell.value = clean_value
                    fixed = True

        return fixed

    def _fix_column_data_types(
        self, worksheet, start_row: int, min_col: int, max_row: int, max_col: int
    ) -> int:
        """열별 데이터 타입 일관성 수정"""

        fixes_count = 0

        for col in range(min_col, max_col + 1):
            column_values = []

            # 열의 데이터 수집
            for row in range(start_row, max_row + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.value is not None:
                    column_values.append((row, cell.value))

            if not column_values:
                continue

            # 데이터 타입 분석
            numeric_count = 0
            date_count = 0
            text_count = 0

            for _, value in column_values:
                if isinstance(value, (int, float)):
                    numeric_count += 1
                elif self._is_date_like(str(value)):
                    date_count += 1
                else:
                    text_count += 1

            # 주요 타입 결정
            total_count = len(column_values)
            if numeric_count / total_count > 0.7:
                # 숫자형 열로 정규화
                fixed = self._normalize_numeric_column(worksheet, col, column_values)
                if fixed:
                    fixes_count += 1
            elif date_count / total_count > 0.7:
                # 날짜형 열로 정규화
                fixed = self._normalize_date_column(worksheet, col, column_values)
                if fixed:
                    fixes_count += 1

        return fixes_count

    def _is_date_like(self, text: str) -> bool:
        """텍스트가 날짜 형식인지 확인"""

        date_patterns = [
            r"\d{4}-\d{2}-\d{2}",
            r"\d{2}/\d{2}/\d{4}",
            r"\d{4}/\d{2}/\d{2}",
            r"\d{2}-\d{2}-\d{4}",
        ]

        for pattern in date_patterns:
            if re.match(pattern, text.strip()):
                return True

        return False

    def _normalize_numeric_column(
        self, worksheet, col: int, values: List[Tuple[int, Any]]
    ) -> bool:
        """숫자형 열 정규화"""

        fixed = False

        for row, value in values:
            if not isinstance(value, (int, float)):
                # 텍스트를 숫자로 변환 시도
                try:
                    clean_text = re.sub(r"[^\d.-]", "", str(value))
                    if clean_text:
                        numeric_value = float(clean_text)
                        if numeric_value.is_integer():
                            numeric_value = int(numeric_value)
                        worksheet.cell(row=row, column=col).value = numeric_value
                        fixed = True
                except Exception:
                    pass  # 변환 실패 시 원본 유지

        return fixed

    def _normalize_date_column(
        self, worksheet, col: int, values: List[Tuple[int, Any]]
    ) -> bool:
        """날짜형 열 정규화"""

        fixed = False

        for row, value in values:
            if isinstance(value, str) and self._is_date_like(value):
                try:
                    from datetime import datetime

                    # 다양한 날짜 형식 파싱 시도
                    date_formats = ["%Y-%m-%d", "%m/%d/%Y", "%Y/%m/%d", "%m-%d-%Y"]

                    for fmt in date_formats:
                        try:
                            date_obj = datetime.strptime(value.strip(), fmt)
                            worksheet.cell(row=row, column=col).value = date_obj
                            fixed = True
                            break
                        except Exception:
                            continue

                except Exception:
                    pass  # 변환 실패 시 원본 유지

        return fixed

    def _fix_headers(self, workbook: Workbook, issues: List[Dict]) -> Dict[str, Any]:
        """헤더 행 수정"""

        fixes = {"count": 0, "details": []}

        for worksheet in workbook.worksheets:
            try:
                # 첫 번째 행이 헤더인지 확인
                first_row = list(worksheet[1])
                if any(cell.value is not None for cell in first_row):
                    header_fixed = False

                    for cell in first_row:
                        if cell.value is None:
                            cell.value = f"Header_{get_column_letter(cell.column)}"
                            header_fixed = True

                    if header_fixed:
                        fixes["count"] += 1
                        fixes["details"].append(
                            {
                                "type": "headers_fixed",
                                "location": f"{worksheet.title}!1:1",
                                "action": "헤더 행 수정 완료",
                            }
                        )

            except Exception as e:
                self.warnings.append(f"시트 {worksheet.title} 헤더 수정 실패: {str(e)}")

        return fixes

    def _clean_formatting(
        self, workbook: Workbook, issues: List[Dict]
    ) -> Dict[str, Any]:
        """불필요한 서식 정리"""

        fixes = {"count": 0, "details": []}

        try:
            # 전체 워크북에 대해 기본 서식 정리
            for worksheet in workbook.worksheets:
                formatting_cleaned = False

                # 빈 셀의 서식 제거
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value is None or str(cell.value).strip() == "":
                            if (
                                cell.font.name != "Calibri"
                                or cell.font.size != 11
                                or cell.fill.start_color.index != 0
                            ):

                                # 기본 서식으로 재설정
                                from openpyxl.styles import Font, PatternFill

                                cell.font = Font(name="Calibri", size=11)
                                cell.fill = PatternFill()
                                formatting_cleaned = True

                if formatting_cleaned:
                    fixes["count"] += 1
                    fixes["details"].append(
                        {
                            "type": "formatting_cleaned",
                            "location": f"시트: {worksheet.title}",
                            "action": "불필요한 서식 정리 완료",
                        }
                    )

        except Exception as e:
            self.warnings.append(f"서식 정리 실패: {str(e)}")

        return fixes
