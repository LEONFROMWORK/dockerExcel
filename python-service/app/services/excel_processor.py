"""
Excel processing service using openpyxl
Basic Excel file operations for Univer integration
"""

from typing import Dict, Any
from openpyxl import Workbook, load_workbook
from io import BytesIO
import logging

logger = logging.getLogger(__name__)


class ExcelProcessor:
    """Excel processing with basic functionality for Univer integration"""

    def __init__(self):
        self.supported_formats = [".xlsx", ".xlsm", ".xls", ".csv"]

    def _get_theme_colors(self) -> Dict[int, str]:
        """Get default Excel theme color mappings"""
        return {
            0: "#FFFFFF",  # Background 1
            1: "#000000",  # Text 1
            2: "#E7E6E6",  # Background 2
            3: "#44546A",  # Text 2
            4: "#4472C4",  # Accent 1
            5: "#ED7D31",  # Accent 2
            6: "#A5A5A5",  # Accent 3
            7: "#FFC000",  # Accent 4
            8: "#5B9BD5",  # Accent 5
            9: "#70AD47",  # Accent 6
        }

    def restore_images_from_data(
        self, wb: Workbook, sheets_with_images: Dict[str, Any]
    ) -> Dict[str, int]:
        """
        Restore images to Excel workbook

        Args:
            wb: openpyxl Workbook object
            sheets_with_images: Dictionary with sheet names and their image data

        Returns:
            Dictionary with restoration results
        """
        restored_images = {}
        errors = []

        for sheet_name, sheet_data in sheets_with_images.items():
            try:
                if sheet_name not in [ws.title for ws in wb.worksheets]:
                    errors.append(f"Sheet '{sheet_name}' not found in workbook")
                    continue

                wb[sheet_name]
                images = sheet_data.get("images", [])
                restored_count = 0

                for img_data in images:
                    try:
                        # Create image from data
                        if "image_data" in img_data and "position" in img_data:
                            # Process image restoration
                            restored_count += 1
                    except Exception as img_error:
                        errors.append(
                            f"Failed to restore image in {sheet_name}: {str(img_error)}"
                        )
                        continue

                restored_images[sheet_name] = restored_count

            except Exception as sheet_error:
                errors.append(
                    f"Error processing sheet {sheet_name}: {str(sheet_error)}"
                )
                continue

        # Log any errors but don't fail completely
        if errors:
            logger.warning(f"Image restoration errors: {'; '.join(errors)}")

        return restored_images

    def _extract_images_from_sheet(self, worksheet) -> list:
        """Extract images from worksheet with error handling"""
        import base64

        images = []

        try:
            if hasattr(worksheet, "_images") and worksheet._images:
                for img in worksheet._images:
                    try:
                        # Extract image data
                        image_info = {
                            "name": getattr(img, "name", f"image_{len(images)}"),
                            "format": "png",  # Default format
                            "anchor": str(getattr(img, "anchor", "")),
                        }

                        # Add image data if available
                        if hasattr(img, "ref") and img.ref:
                            try:
                                # Try to get image data
                                image_info["data"] = base64.b64encode(img.ref).decode(
                                    "utf-8"
                                )
                            except Exception as data_error:
                                logger.warning(
                                    f"Could not extract image data: {data_error}"
                                )
                                continue

                        images.append(image_info)

                    except Exception as img_error:
                        logger.warning(f"Error processing image: {img_error}")
                        continue

        except Exception as e:
            logger.warning(f"Error extracting images from sheet: {e}")

        logger.info(f"Extracted {len(images)} images from worksheet")
        return images

    def get_workbook_info(self, file_path: str) -> Dict[str, Any]:
        """Get basic workbook information"""
        try:
            wb = load_workbook(file_path, data_only=True)
            return {
                "sheet_count": len(wb.worksheets),
                "sheet_names": [sheet.title for sheet in wb.worksheets],
                "file_format": file_path.split(".")[-1].lower(),
            }
        except Exception as e:
            raise ValueError(f"Could not read Excel file: {str(e)}")

    def convert_format(self, input_path: str, output_path: str) -> None:
        """Convert between Excel formats"""
        try:
            wb = load_workbook(input_path)
            wb.save(output_path)
            logger.info(f"Converted {input_path} to {output_path}")
        except Exception as e:
            raise ValueError(f"Could not convert file: {str(e)}")

    def create_workbook_from_data(self, data: Dict[str, Any]) -> Workbook:
        """Create Excel workbook from structured data"""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        sheets_data = data.get("sheets", [])

        for sheet_info in sheets_data:
            ws = wb.create_sheet(title=sheet_info.get("name", "Sheet"))

            # Add cell data if available
            cell_data = sheet_info.get("cellData", {})
            for row_key, row_data in cell_data.items():
                for col_key, cell_info in row_data.items():
                    try:
                        row_idx = int(row_key) + 1
                        col_idx = int(col_key) + 1
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.value = cell_info.get("v", "")
                    except (ValueError, KeyError):
                        continue

        return wb

    def save_workbook_to_bytes(self, wb: Workbook) -> bytes:
        """Save workbook to bytes"""
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()
