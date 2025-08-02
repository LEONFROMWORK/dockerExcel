#!/usr/bin/env python3
"""
테스트용 Excel 파일 생성 스크립트
다양한 기능을 포함한 Excel 파일을 생성합니다.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def create_test_excel():
    """테스트용 Excel 파일 생성"""
    
    # 워크북 생성
    wb = Workbook()
    
    # 첫 번째 시트 (데이터 시트)
    ws1 = wb.active
    ws1.title = "매출 데이터"
    
    # 헤더 데이터
    headers = ["월", "제품A", "제품B", "제품C", "총매출"]
    months = ["1월", "2월", "3월", "4월", "5월", "6월"]
    product_a = [120, 135, 150, 165, 180, 195]
    product_b = [80, 85, 90, 95, 100, 105]  
    product_c = [200, 210, 220, 230, 240, 250]
    
    # 데이터 입력
    for col, header in enumerate(headers, 1):
        ws1.cell(1, col, header)
    
    for row, month in enumerate(months, 2):
        ws1.cell(row, 1, month)
        ws1.cell(row, 2, product_a[row-2])
        ws1.cell(row, 3, product_b[row-2])
        ws1.cell(row, 4, product_c[row-2])
        # 수식으로 총매출 계산
        ws1.cell(row, 5, f"=B{row}+C{row}+D{row}")
    
    # 스타일 적용
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for col in range(1, 6):
        cell = ws1.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # 테두리 추가
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(1, 8):
        for col in range(1, 6):
            ws1.cell(row, col).border = thin_border
    
    # 차트 생성
    chart = LineChart()
    chart.title = "월별 매출 추이"
    chart.y_axis.title = "매출액"
    chart.x_axis.title = "월"
    
    # 데이터 범위 설정
    data = Reference(ws1, min_col=2, min_row=1, max_col=4, max_row=7)
    cats = Reference(ws1, min_col=1, min_row=2, max_row=7)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    # 차트 위치 설정
    ws1.add_chart(chart, "G2")
    
    # 두 번째 시트 (요약 시트)
    ws2 = wb.create_sheet("요약")
    
    # 요약 데이터
    ws2["A1"] = "매출 요약"
    ws2["A1"].font = Font(size=16, bold=True)
    
    ws2["A3"] = "총 매출:"
    ws2["B3"] = f"=SUM('{ws1.title}'!E2:E7)"
    
    ws2["A4"] = "평균 매출:"
    ws2["B4"] = f"=AVERAGE('{ws1.title}'!E2:E7)"
    
    ws2["A5"] = "최대 매출:"
    ws2["B5"] = f"=MAX('{ws1.title}'!E2:E7)"
    
    # 조건부 서식 (간단한 색상 적용)
    ws2["B3"].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    ws2["B4"].fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")  
    ws2["B5"].fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    
    # 파일 저장
    output_path = "/Users/kevin/excel-unified/test_sample.xlsx"
    wb.save(output_path)
    
    print(f"✅ 테스트 Excel 파일이 생성되었습니다: {output_path}")
    print("📊 포함된 기능:")
    print("  - 다중 시트 (매출 데이터, 요약)")
    print("  - 수식 (SUM, AVERAGE, MAX)")
    print("  - 차트 (라인 차트)")
    print("  - 셀 서식 (폰트, 색상, 정렬)")
    print("  - 테두리")
    print("  - 조건부 서식")
    
    return output_path

if __name__ == "__main__":
    create_test_excel()