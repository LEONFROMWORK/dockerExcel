"""
Comprehensive Error Detection System
Implements parallel analysis with basic tools, AI, and pattern matching
"""

import asyncio
from typing import List, Dict, Any, Optional
from dataclasses import dataclass
from enum import Enum
import logging
from datetime import datetime
import openpyxl
import json

from ..services.excel_analyzer import excel_analyzer
from ..services.openai_service import openai_service
from ..services.advanced_vba_analyzer import AdvancedVBAAnalyzer

logger = logging.getLogger(__name__)


class ErrorCategory(Enum):
    """Error categories for comprehensive detection"""

    SYNTAX = "syntax"  # 구문 오류 (ExcelJS/openpyxl)
    FORMULA = "formula"  # 수식 오류
    REFERENCE = "reference"  # 참조 오류
    LOGIC = "logic"  # 비즈니스 로직 오류
    SEMANTIC = "semantic"  # 의미적 오류
    PATTERN = "pattern"  # 패턴 오류
    PERFORMANCE = "performance"  # 성능 문제
    DATA_QUALITY = "data_quality"  # 데이터 품질
    STRUCTURAL = "structural"  # 구조적 문제


@dataclass
class ErrorReport:
    """Represents a detected error with all details"""

    category: ErrorCategory
    severity: str  # critical, high, medium, low
    location: Dict[str, Any]
    description: str
    detection_method: str  # basic_tool, ai_analysis, pattern_matching
    suggested_fix: Optional[str] = None
    confidence: float = 1.0
    error_id: Optional[str] = None
    metadata: Optional[Dict[str, Any]] = None


class ComprehensiveErrorDetector:
    """
    Comprehensive error detection system that runs all analysis methods in parallel
    Ensures AI analysis always runs regardless of basic tool results
    """

    def __init__(self):
        self.basic_analyzer = excel_analyzer
        self.vba_analyzer = AdvancedVBAAnalyzer()

    async def analyze_excel_file(
        self, file_path: str, options: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """
        Analyze Excel file comprehensively using all available methods

        Args:
            file_path: Path to the Excel file
            options: Optional configuration for analysis

        Returns:
            Comprehensive analysis results with all detected errors
        """
        options = options or {}

        # Run all analysis methods in parallel
        analysis_tasks = []

        # Always run basic analysis
        if options.get("include_basic", True):
            analysis_tasks.append(self._run_basic_analysis(file_path))

        # Always run AI analysis (doesn't depend on basic results)
        if options.get("include_ai", True):
            analysis_tasks.append(self._run_ai_analysis(file_path))

        # Always run pattern analysis
        if options.get("include_patterns", True):
            analysis_tasks.append(self._run_pattern_analysis(file_path))

        # Run VBA analysis if applicable
        if options.get("include_vba", True):
            analysis_tasks.append(self._run_vba_analysis(file_path))

        # Execute all tasks in parallel
        results = await asyncio.gather(*analysis_tasks, return_exceptions=True)

        # Process results
        all_errors = []
        ai_insights = []

        for idx, result in enumerate(results):
            if isinstance(result, Exception):
                logger.error(f"Analysis task {idx} failed: {str(result)}")
                continue

            if isinstance(result, list):
                all_errors.extend(result)
            elif isinstance(result, dict) and "errors" in result:
                all_errors.extend(result["errors"])
                if "insights" in result:
                    ai_insights.extend(result["insights"])

        # Deduplicate and sort errors
        unique_errors = self._deduplicate_errors(all_errors)
        sorted_errors = sorted(
            unique_errors,
            key=lambda x: self._get_severity_score(x.severity),
            reverse=True,
        )

        # Generate comprehensive report
        return {
            "file_path": file_path,
            "analysis_date": datetime.now().isoformat(),
            "total_errors": len(sorted_errors),
            "errors": [self._error_to_dict(e) for e in sorted_errors],
            "summary": self._generate_summary(sorted_errors),
            "ai_insights": ai_insights,
            "fix_recommendations": self._generate_fix_recommendations(sorted_errors),
        }

    async def _run_basic_analysis(self, file_path: str) -> List[ErrorReport]:
        """Run basic analysis using ExcelJS/openpyxl"""
        errors = []

        try:
            # Use existing excel analyzer
            result = await self.basic_analyzer.analyze_file(file_path)

            # Convert formula errors
            for sheet_name, sheet_data in result.get("sheets", {}).items():
                for error in sheet_data.get("errors", []):
                    errors.append(
                        ErrorReport(
                            category=ErrorCategory.FORMULA,
                            severity="high" if "#REF!" in error["error"] else "medium",
                            location={"sheet": sheet_name, "cell": error["cell"]},
                            description=error["description"],
                            detection_method="basic_tool",
                            suggested_fix=self._get_basic_fix(error["error"]),
                            confidence=0.95,
                        )
                    )

            # Check for data quality issues
            for sheet_name, sheet_data in result.get("sheets", {}).items():
                if "data_quality_issues" in sheet_data:
                    for issue in sheet_data["data_quality_issues"]:
                        errors.append(
                            ErrorReport(
                                category=ErrorCategory.DATA_QUALITY,
                                severity="low",
                                location={"sheet": sheet_name},
                                description=f"{issue['type']}: {issue.get('percentage', 0)}%",
                                detection_method="basic_tool",
                                metadata=issue,
                            )
                        )

        except Exception as e:
            logger.error(f"Basic analysis failed: {str(e)}")

        return errors

    async def _run_ai_analysis(self, file_path: str) -> Dict[str, Any]:
        """
        Run AI-based analysis for complex errors
        This ALWAYS runs regardless of basic tool results
        """
        errors = []
        insights = []

        try:
            # First, extract workbook data
            workbook_data = await self._extract_workbook_data_for_ai(file_path)

            # Create comprehensive AI prompt
            ai_prompt = self._build_ai_analysis_prompt(workbook_data)

            # Get AI analysis
            messages = [
                {
                    "role": "system",
                    "content": """You are an expert Excel auditor specializing in detecting complex errors
                    that automated tools might miss. Focus on business logic, semantic errors, and data relationships.""",
                },
                {"role": "user", "content": ai_prompt},
            ]

            ai_response = await openai_service.chat_completion(
                messages=messages, temperature=0.3, max_tokens=2000
            )

            # Parse AI response
            ai_errors = self._parse_ai_response(ai_response)

            for error in ai_errors:
                errors.append(
                    ErrorReport(
                        category=self._map_ai_category(error.get("type", "unknown")),
                        severity=error.get("severity", "medium"),
                        location=error.get("location", {}),
                        description=error.get("description", ""),
                        detection_method="ai_analysis",
                        suggested_fix=error.get("fix"),
                        confidence=error.get("confidence", 0.8),
                        metadata=error.get("metadata", {}),
                    )
                )

            # Extract insights
            if isinstance(ai_response, str) and "INSIGHTS:" in ai_response:
                insights_text = ai_response.split("INSIGHTS:")[1].strip()
                insights = [
                    line.strip() for line in insights_text.split("\n") if line.strip()
                ]

        except Exception as e:
            logger.error(f"AI analysis failed: {str(e)}")

        return {"errors": errors, "insights": insights}

    async def _run_pattern_analysis(self, file_path: str) -> List[ErrorReport]:
        """Run pattern-based error detection"""
        errors = []

        try:
            # Load workbook data
            import openpyxl

            workbook = openpyxl.load_workbook(file_path, data_only=True)

            for sheet in workbook.worksheets:
                # Check for consecutive blanks
                blank_errors = self._check_consecutive_blanks(sheet)
                errors.extend(blank_errors)

                # Check for inconsistent formulas
                formula_errors = self._check_formula_consistency(sheet)
                errors.extend(formula_errors)

                # Check for hardcoded values where formulas expected
                hardcoded_errors = self._check_hardcoded_values(sheet)
                errors.extend(hardcoded_errors)

                # Check for circular references
                circular_errors = self._check_circular_references(sheet)
                errors.extend(circular_errors)

        except Exception as e:
            logger.error(f"Pattern analysis failed: {str(e)}")

        return errors

    async def _run_vba_analysis(self, file_path: str) -> List[ErrorReport]:
        """Run VBA analysis if file contains macros"""
        errors = []

        try:
            vba_result = await self.vba_analyzer.analyze_file(file_path)

            if vba_result.get("has_vba"):
                for error in vba_result.get("errors", []):
                    errors.append(
                        ErrorReport(
                            category=(
                                ErrorCategory.SYNTAX
                                if error["category"] == "syntax"
                                else ErrorCategory.LOGIC
                            ),
                            severity=error["severity"],
                            location={
                                "module": error["module_name"],
                                "line": error.get("line_number"),
                            },
                            description=error["description"],
                            detection_method="vba_analyzer",
                            suggested_fix=error["fix_suggestion"],
                            confidence=error["confidence"],
                        )
                    )

        except Exception as e:
            logger.error(f"VBA analysis failed: {str(e)}")

        return errors

    def _build_ai_analysis_prompt(self, workbook_data: Dict[str, Any]) -> str:
        """Build comprehensive prompt for AI analysis"""
        prompt = f"""
        Analyze this Excel workbook for complex errors that basic tools might miss.

        Workbook Overview:
        {json.dumps(workbook_data.get('summary', {}), indent=2)}

        Focus on detecting:

        1. **Business Logic Errors**:
           - Revenue < Costs but Profit > 0
           - Sum of parts != Total
           - Percentage calculations > 100% or < 0%
           - Time series data with impossible jumps

        2. **Semantic Errors**:
           - Unit mismatches (mixing 원, 천원, 백만원)
           - Date format inconsistencies
           - Category classification errors
           - Currency mixing

        3. **Data Quality Issues**:
           - Abnormal outliers in context
           - Missing required data for calculations
           - Duplicate entries that shouldn't exist
           - Inconsistent data patterns

        4. **Structural Problems**:
           - Inconsistent data layout across sheets
           - Wrong aggregation ranges
           - Hidden circular reference risks
           - Formula ranges not updating with data

        5. **Relationship Errors**:
           - Parent-child relationships that don't add up
           - Cross-sheet references to wrong cells
           - Lookup values that don't exist
           - Index mismatches

        Sample Data from Sheets:
        {json.dumps(workbook_data.get('sample_data', {}), indent=2)}

        Return a JSON array of errors found, each with:
        {{
            "type": "logic|semantic|quality|structural|relationship",
            "severity": "critical|high|medium|low",
            "location": {{"sheet": "name", "cell": "A1" or "range": "A1:B10"}},
            "description": "Clear description of the error",
            "confidence": 0.0-1.0,
            "fix": "Suggested fix",
            "metadata": {{additional context}}
        }}

        INSIGHTS:
        List any general observations about the workbook quality
        """

        return prompt

    async def _extract_workbook_data_for_ai(self, file_path: str) -> Dict[str, Any]:
        """Extract relevant workbook data for AI analysis"""
        import pandas as pd

        data = {"summary": {}, "sample_data": {}, "formulas": {}, "relationships": []}

        try:
            # Get basic file info
            basic_info = await self.basic_analyzer.analyze_file(file_path)
            data["summary"] = basic_info.get("summary", {})

            # Load with pandas for data sampling
            excel_file = pd.ExcelFile(file_path)

            for sheet_name in excel_file.sheet_names[:5]:  # Limit to first 5 sheets
                df = pd.read_excel(excel_file, sheet_name=sheet_name, nrows=20)

                # Sample data
                data["sample_data"][sheet_name] = {
                    "shape": df.shape,
                    "columns": list(df.columns),
                    "head": df.head(5).to_dict("records"),
                    "dtypes": df.dtypes.astype(str).to_dict(),
                    "null_counts": df.isnull().sum().to_dict(),
                }

                # Detect numeric columns
                numeric_cols = df.select_dtypes(include=["number"]).columns
                if len(numeric_cols) > 0:
                    data["sample_data"][sheet_name]["numeric_summary"] = {
                        col: {
                            "min": (
                                float(df[col].min())
                                if pd.notna(df[col].min())
                                else None
                            ),
                            "max": (
                                float(df[col].max())
                                if pd.notna(df[col].max())
                                else None
                            ),
                            "mean": (
                                float(df[col].mean())
                                if pd.notna(df[col].mean())
                                else None
                            ),
                        }
                        for col in numeric_cols
                    }

        except Exception as e:
            logger.error(f"Error extracting workbook data: {str(e)}")

        return data

    def _check_consecutive_blanks(self, sheet) -> List[ErrorReport]:
        """Check for suspicious consecutive blank cells"""
        errors = []

        # Check for blank rows in the middle of data
        last_data_row = 0
        blank_count = 0

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if any(cell is not None for cell in row):
                if blank_count > 3 and last_data_row > 0:
                    errors.append(
                        ErrorReport(
                            category=ErrorCategory.PATTERN,
                            severity="low",
                            location={
                                "sheet": sheet.title,
                                "range": f"A{last_data_row+1}:A{row_idx-1}",
                            },
                            description=f"{blank_count} consecutive blank rows detected",
                            detection_method="pattern_matching",
                            suggested_fix="Consider removing unnecessary blank rows",
                        )
                    )
                last_data_row = row_idx
                blank_count = 0
            else:
                blank_count += 1

        return errors

    def _check_formula_consistency(self, sheet) -> List[ErrorReport]:
        """Check for inconsistent formula patterns"""
        errors = []

        # Group cells by column to check formula consistency

        for col_idx in range(1, sheet.max_column + 1):
            col_formulas = []
            for row_idx in range(
                2, min(sheet.max_row + 1, 100)
            ):  # Sample first 100 rows
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.formula:
                    col_formulas.append((row_idx, cell.formula))

            if len(col_formulas) > 5:  # Need enough formulas to detect pattern
                # Check if formulas follow a pattern
                base_formula = col_formulas[0][1]
                inconsistent = []

                for row_idx, formula in col_formulas[1:]:
                    # Simple check - formulas should have similar structure
                    if not self._formulas_similar(
                        base_formula, formula, row_idx - col_formulas[0][0]
                    ):
                        inconsistent.append(row_idx)

                if inconsistent:
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    errors.append(
                        ErrorReport(
                            category=ErrorCategory.PATTERN,
                            severity="medium",
                            location={
                                "sheet": sheet.title,
                                "cells": [
                                    f"{col_letter}{row}" for row in inconsistent[:5]
                                ],
                            },
                            description=f"Inconsistent formula pattern in column {col_letter}",
                            detection_method="pattern_matching",
                            suggested_fix="Review formulas for consistency",
                        )
                    )

        return errors

    def _check_hardcoded_values(self, sheet) -> List[ErrorReport]:
        """Check for hardcoded values where formulas might be expected"""
        errors = []

        # Look for numeric values in cells surrounded by formulas
        for row_idx in range(2, min(sheet.max_row + 1, 100)):
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)

                if isinstance(cell.value, (int, float)) and not cell.formula:
                    # Check if surrounding cells have formulas
                    surrounding_formulas = 0
                    for dr in [-1, 0, 1]:
                        for dc in [-1, 0, 1]:
                            if dr == 0 and dc == 0:
                                continue
                            adj_cell = sheet.cell(row=row_idx + dr, column=col_idx + dc)
                            if adj_cell.formula:
                                surrounding_formulas += 1

                    if (
                        surrounding_formulas >= 4
                    ):  # Most surrounding cells have formulas
                        errors.append(
                            ErrorReport(
                                category=ErrorCategory.PATTERN,
                                severity="low",
                                location={
                                    "sheet": sheet.title,
                                    "cell": f"{openpyxl.utils.get_column_letter(col_idx)}{row_idx}",
                                },
                                description="Hardcoded value surrounded by formulas",
                                detection_method="pattern_matching",
                                suggested_fix="Consider using a formula instead of hardcoded value",
                                confidence=0.7,
                            )
                        )

        return errors

    def _check_circular_references(self, sheet) -> List[ErrorReport]:
        """Basic circular reference detection"""
        errors = []

        # This is a simplified check - full circular reference detection requires
        # building a dependency graph
        formula_refs = {}

        for row in sheet.iter_rows():
            for cell in row:
                if cell.formula:
                    # Extract cell references from formula
                    import re

                    refs = re.findall(r"[A-Z]+\d+", cell.formula)
                    formula_refs[cell.coordinate] = refs

        # Check for direct circular references
        for cell_addr, refs in formula_refs.items():
            if cell_addr in refs:
                errors.append(
                    ErrorReport(
                        category=ErrorCategory.STRUCTURAL,
                        severity="critical",
                        location={"sheet": sheet.title, "cell": cell_addr},
                        description="Direct circular reference detected",
                        detection_method="pattern_matching",
                        suggested_fix="Remove self-reference from formula",
                    )
                )

        return errors

    def _formulas_similar(self, formula1: str, formula2: str, row_diff: int) -> bool:
        """Check if two formulas follow the same pattern with row adjustment"""
        # Simple heuristic - in practice, would need more sophisticated parsing
        import re

        # Replace row numbers with placeholders
        pattern1 = re.sub(r"\d+", "X", formula1)
        pattern2 = re.sub(r"\d+", "X", formula2)

        return pattern1 == pattern2

    def _parse_ai_response(self, response: str) -> List[Dict[str, Any]]:
        """Parse AI response to extract errors"""
        try:
            # Try to extract JSON from response
            import re

            json_match = re.search(r"\[.*\]", response, re.DOTALL)
            if json_match:
                return json.loads(json_match.group())

            # Fallback: try to parse as direct JSON
            if response.strip().startswith("["):
                return json.loads(response)

        except Exception as e:
            logger.error(f"Failed to parse AI response: {str(e)}")

        return []

    def _map_ai_category(self, ai_type: str) -> ErrorCategory:
        """Map AI error types to our categories"""
        mapping = {
            "logic": ErrorCategory.LOGIC,
            "semantic": ErrorCategory.SEMANTIC,
            "quality": ErrorCategory.DATA_QUALITY,
            "structural": ErrorCategory.STRUCTURAL,
            "relationship": ErrorCategory.REFERENCE,
            "formula": ErrorCategory.FORMULA,
        }
        return mapping.get(ai_type.lower(), ErrorCategory.SEMANTIC)

    def _deduplicate_errors(self, errors: List[ErrorReport]) -> List[ErrorReport]:
        """Remove duplicate errors based on location and description"""
        seen = set()
        unique_errors = []

        for error in errors:
            # Create a key for deduplication
            location_key = json.dumps(error.location, sort_keys=True)
            desc_key = error.description[:50]  # First 50 chars
            key = (error.category.value, location_key, desc_key)

            if key not in seen:
                seen.add(key)
                # Assign unique error ID
                error.error_id = f"{error.category.value}_{len(unique_errors)+1:04d}"
                unique_errors.append(error)

        return unique_errors

    def _get_severity_score(self, severity: str) -> int:
        """Convert severity to numeric score for sorting"""
        scores = {"critical": 4, "high": 3, "medium": 2, "low": 1}
        return scores.get(severity.lower(), 0)

    def _generate_summary(self, errors: List[ErrorReport]) -> Dict[str, Any]:
        """Generate error summary statistics"""
        summary = {
            "total": len(errors),
            "by_category": {},
            "by_severity": {},
            "by_detection_method": {},
            "critical_count": 0,
        }

        for error in errors:
            # By category
            cat = error.category.value
            summary["by_category"][cat] = summary["by_category"].get(cat, 0) + 1

            # By severity
            sev = error.severity
            summary["by_severity"][sev] = summary["by_severity"].get(sev, 0) + 1

            # By detection method
            method = error.detection_method
            summary["by_detection_method"][method] = (
                summary["by_detection_method"].get(method, 0) + 1
            )

            # Count critical
            if error.severity == "critical":
                summary["critical_count"] += 1

        return summary

    def _generate_fix_recommendations(
        self, errors: List[ErrorReport]
    ) -> List[Dict[str, Any]]:
        """Generate prioritized fix recommendations"""
        recommendations = []

        # Group errors by category and severity
        error_groups = {}
        for error in errors:
            key = (error.category.value, error.severity)
            if key not in error_groups:
                error_groups[key] = []
            error_groups[key].append(error)

        # Sort groups by priority
        sorted_groups = sorted(
            error_groups.items(),
            key=lambda x: (self._get_severity_score(x[0][1]), len(x[1])),
            reverse=True,
        )

        # Generate recommendations for top groups
        for (category, severity), group_errors in sorted_groups[:5]:
            if len(group_errors) > 0:
                sample_error = group_errors[0]
                recommendations.append(
                    {
                        "priority": severity,
                        "category": category,
                        "count": len(group_errors),
                        "description": f"Fix {len(group_errors)} {category} errors",
                        "sample_location": sample_error.location,
                        "suggested_approach": sample_error.suggested_fix
                        or "Manual review required",
                        "estimated_impact": self._estimate_impact(
                            category, severity, len(group_errors)
                        ),
                    }
                )

        return recommendations

    def _estimate_impact(self, category: str, severity: str, count: int) -> str:
        """Estimate the impact of fixing certain errors"""
        if severity == "critical":
            return "Very High - May prevent calculation errors"
        elif severity == "high" and count > 5:
            return "High - Significant improvement in reliability"
        elif category in ["logic", "semantic"]:
            return "High - Improves business logic accuracy"
        elif category == "performance" and count > 10:
            return "Medium - Noticeable performance improvement"
        else:
            return "Low to Medium - General quality improvement"

    def _error_to_dict(self, error: ErrorReport) -> Dict[str, Any]:
        """Convert ErrorReport to dictionary"""
        return {
            "id": error.error_id,
            "category": error.category.value,
            "severity": error.severity,
            "location": error.location,
            "description": error.description,
            "detection_method": error.detection_method,
            "suggested_fix": error.suggested_fix,
            "confidence": error.confidence,
            "metadata": error.metadata,
        }

    def _get_basic_fix(self, error_type: str) -> Optional[str]:
        """Get basic fix suggestion for common Excel errors"""
        fixes = {
            "#DIV/0!": "Use IFERROR or IF to handle division by zero",
            "#N/A": "Check VLOOKUP/MATCH reference or use IFERROR",
            "#NAME?": "Check formula name spelling or add quotes to text",
            "#NULL!": "Check cell range references",
            "#NUM!": "Check numeric arguments are within valid range",
            "#REF!": "Update cell references - cells may have been deleted",
            "#VALUE!": "Check data types in formula arguments",
        }
        return fixes.get(error_type)
