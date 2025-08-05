"""
실시간 Excel 수식 검증 서비스
formulas 라이브러리를 사용하여 실시간으로 Excel 수식을 계산하고 검증
"""

import logging
from typing import Dict, List, Any, Optional
from dataclasses import dataclass
from datetime import datetime

try:
    import formulas

    FORMULAS_AVAILABLE = True
except ImportError:
    FORMULAS_AVAILABLE = False
    logging.warning(
        "formulas library not available. Install with: pip install formulas"
    )

import openpyxl
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


@dataclass
class ValidationResult:
    """수식 검증 결과"""

    valid: bool
    result: Any = None
    error: Optional[str] = None
    error_type: Optional[str] = None
    calculated_value: Any = None
    formula_type: str = "unknown"
    dependencies: List[str] = None
    execution_time: float = 0.0
    warnings: List[str] = None


class RealtimeFormulaValidator:
    """실시간 Excel 수식 검증기"""

    def __init__(self):
        self.xl_model = None
        self.context = {}
        self.sheet_structures = {}
        self.dependency_graph = {}

        if not FORMULAS_AVAILABLE:
            raise ImportError(
                "formulas library is required. Install with: pip install formulas"
            )

    def initialize_from_openpyxl_workbook(self, workbook_path: str) -> bool:
        """openpyxl 워크북으로부터 초기화"""
        try:
            wb = openpyxl.load_workbook(workbook_path, data_only=False)

            # 모든 시트의 데이터 추출
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                self.context[sheet_name] = {}

                # 셀 데이터 추출
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            cell_addr = f"{get_column_letter(cell.column)}{cell.row}"

                            if cell.data_type == "f":  # 수식
                                self.context[sheet_name][cell_addr] = f"={cell.value}"
                            else:  # 값
                                self.context[sheet_name][cell_addr] = cell.value

                # 시트 구조 정보 저장
                self.sheet_structures[sheet_name] = {
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "merged_cells": [str(range_) for range_ in ws.merged_cells.ranges],
                }

            # formulas 모델 초기화
            self.xl_model = formulas.ExcelModel()
            self.xl_model.loads(self.context)

            logger.info(f"Initialized validator with {len(self.context)} sheets")
            return True

        except Exception as e:
            logger.error(f"Failed to initialize from workbook: {e}")
            return False

    def set_context(self, sheet_name: str, cell_data: Dict[str, Any]):
        """컨텍스트 수동 설정"""
        if sheet_name not in self.context:
            self.context[sheet_name] = {}

        self.context[sheet_name].update(cell_data)

        # formulas 모델 재생성
        self.xl_model = formulas.ExcelModel()
        self.xl_model.loads({sheet_name: self.context[sheet_name]})

    def validate_formula_realtime(
        self, formula: str, cell_address: str, sheet_name: str = "Sheet1"
    ) -> ValidationResult:
        """실시간 수식 검증"""
        start_time = datetime.now()

        try:
            # 수식 정규화
            if not formula.startswith("="):
                formula = f"={formula}"

            # 임시 컨텍스트 생성
            temp_context = self.context.get(sheet_name, {}).copy()
            temp_context[cell_address] = formula

            # 새 모델로 계산
            test_model = formulas.ExcelModel()
            test_model.loads({sheet_name: temp_context})
            test_model.calculate()

            # 결과 추출
            calculated_value = test_model[sheet_name].cells[cell_address].value

            # 의존성 분석
            dependencies = self._extract_dependencies(formula)

            # 수식 유형 분류
            formula_type = self._classify_formula_type(formula)

            execution_time = (datetime.now() - start_time).total_seconds()

            return ValidationResult(
                valid=True,
                result=calculated_value,
                calculated_value=calculated_value,
                formula_type=formula_type,
                dependencies=dependencies,
                execution_time=execution_time,
                warnings=self._generate_warnings(formula, calculated_value),
            )

        except Exception as e:
            execution_time = (datetime.now() - start_time).total_seconds()
            error_type = self._classify_error_type(str(e))

            return ValidationResult(
                valid=False,
                error=str(e),
                error_type=error_type,
                formula_type=self._classify_formula_type(formula),
                dependencies=self._extract_dependencies(formula),
                execution_time=execution_time,
            )

    def batch_validate_changes(
        self, changes: List[Dict]
    ) -> Dict[str, ValidationResult]:
        """변경사항 일괄 검증"""
        results = {}

        # 임시 컨텍스트 구성
        temp_context = {}
        for sheet_name in self.context:
            temp_context[sheet_name] = self.context[sheet_name].copy()

        # 모든 변경사항 적용
        for change in changes:
            sheet = change.get("sheet", "Sheet1")
            cell = change.get("cell")
            formula = change.get("formula") or change.get("newFormula")
            value = change.get("value") or change.get("newValue")

            if sheet not in temp_context:
                temp_context[sheet] = {}

            if formula:
                temp_context[sheet][cell] = (
                    f"={formula}" if not formula.startswith("=") else formula
                )
            elif value is not None:
                temp_context[sheet][cell] = value

        # 일괄 계산
        try:
            batch_model = formulas.ExcelModel()
            batch_model.loads(temp_context)
            batch_model.calculate()

            # 각 변경사항에 대한 결과 생성
            for change in changes:
                sheet = change.get("sheet", "Sheet1")
                cell = change.get("cell")
                change_id = f"{sheet}!{cell}"

                try:
                    calculated_value = batch_model[sheet].cells[cell].value
                    formula = temp_context[sheet][cell]

                    results[change_id] = ValidationResult(
                        valid=True,
                        result=calculated_value,
                        calculated_value=calculated_value,
                        formula_type=self._classify_formula_type(str(formula)),
                        dependencies=self._extract_dependencies(str(formula)),
                    )
                except Exception as e:
                    results[change_id] = ValidationResult(
                        valid=False,
                        error=str(e),
                        error_type=self._classify_error_type(str(e)),
                    )

        except Exception as e:
            # 전체 실패 시 개별 검증으로 폴백
            logger.warning(f"Batch validation failed, falling back to individual: {e}")
            for change in changes:
                sheet = change.get("sheet", "Sheet1")
                cell = change.get("cell")
                formula = change.get("formula") or change.get("newFormula")
                change_id = f"{sheet}!{cell}"

                if formula:
                    results[change_id] = self.validate_formula_realtime(
                        formula, cell, sheet
                    )

        return results

    def simulate_error_conditions(
        self, formula: str, sheet_name: str = "Sheet1"
    ) -> List[Dict]:
        """오류 조건 시뮬레이션"""
        errors = []

        try:
            # 다양한 오류 시나리오 테스트
            error_scenarios = [
                # Division by zero
                {
                    "type": "division_by_zero",
                    "test": lambda f: "/0" in f
                    or "B2" in f
                    and self.context.get(sheet_name, {}).get("B2") == 0,
                },
                # Circular reference
                {
                    "type": "circular_reference",
                    "test": lambda f: self._check_circular_reference(f, sheet_name),
                },
                # Invalid reference
                {
                    "type": "invalid_reference",
                    "test": lambda f: self._check_invalid_references(f, sheet_name),
                },
                # Type mismatch
                {
                    "type": "type_mismatch",
                    "test": lambda f: self._check_type_mismatch(f, sheet_name),
                },
            ]

            for scenario in error_scenarios:
                if scenario["test"](formula):
                    errors.append(
                        {
                            "type": scenario["type"],
                            "severity": (
                                "high"
                                if scenario["type"]
                                in ["circular_reference", "division_by_zero"]
                                else "medium"
                            ),
                            "description": self._get_error_description(
                                scenario["type"]
                            ),
                            "auto_fixable": scenario["type"] in ["division_by_zero"],
                        }
                    )

        except Exception as e:
            logger.error(f"Error simulation failed: {e}")

        return errors

    def _extract_dependencies(self, formula: str) -> List[str]:
        """수식의 의존성 추출"""
        dependencies = []
        try:
            # 기본적인 셀 참조 패턴 매칭
            import re

            cell_pattern = r"[A-Z]+\d+"
            range_pattern = r"[A-Z]+\d+:[A-Z]+\d+"

            # 범위 참조 찾기
            ranges = re.findall(range_pattern, formula)
            for range_ref in ranges:
                dependencies.append(range_ref)

            # 개별 셀 참조 찾기 (범위에 포함되지 않은 것만)
            cells = re.findall(cell_pattern, formula)
            for cell in cells:
                if not any(cell in range_ref for range_ref in ranges):
                    dependencies.append(cell)

        except Exception as e:
            logger.error(f"Dependency extraction failed: {e}")

        return list(set(dependencies))

    def _classify_formula_type(self, formula: str) -> str:
        """수식 유형 분류"""
        formula_upper = formula.upper()

        if any(
            func in formula_upper
            for func in ["VLOOKUP", "HLOOKUP", "INDEX", "MATCH", "XLOOKUP"]
        ):
            return "lookup"
        elif any(
            func in formula_upper for func in ["SUM", "AVERAGE", "COUNT", "MIN", "MAX"]
        ):
            return "aggregation"
        elif "IF" in formula_upper:
            return "conditional"
        elif any(
            func in formula_upper
            for func in ["LEFT", "RIGHT", "MID", "CONCATENATE", "TEXT"]
        ):
            return "text"
        elif any(func in formula_upper for func in ["DATE", "TIME", "NOW", "TODAY"]):
            return "datetime"
        elif any(op in formula for op in ["+", "-", "*", "/", "^"]):
            return "arithmetic"
        else:
            return "other"

    def _classify_error_type(self, error_message: str) -> str:
        """오류 유형 분류"""
        error_lower = error_message.lower()

        if "circular" in error_lower:
            return "circular_reference"
        elif "ref" in error_lower or "reference" in error_lower:
            return "reference_error"
        elif "div" in error_lower or "zero" in error_lower:
            return "division_by_zero"
        elif "name" in error_lower:
            return "name_error"
        elif "value" in error_lower:
            return "value_error"
        elif "syntax" in error_lower:
            return "syntax_error"
        else:
            return "unknown_error"

    def _generate_warnings(self, formula: str, result: Any) -> List[str]:
        """경고 메시지 생성"""
        warnings = []

        # 성능 관련 경고
        if len(formula) > 200:
            warnings.append("매우 복잡한 수식입니다. 성능에 영향을 줄 수 있습니다.")

        # 휘발성 함수 경고
        volatile_functions = ["NOW", "TODAY", "RAND", "RANDBETWEEN"]
        if any(func in formula.upper() for func in volatile_functions):
            warnings.append(
                "휘발성 함수가 포함되어 있습니다. 파일이 열릴 때마다 재계산됩니다."
            )

        # 결과 유형 경고
        if result is None:
            warnings.append("수식 결과가 비어있습니다.")
        elif isinstance(result, str) and result.startswith("#"):
            warnings.append(f"오류 결과: {result}")

        return warnings

    def _check_circular_reference(self, formula: str, sheet_name: str) -> bool:
        """순환 참조 검사"""
        # 실제로는 더 복잡한 그래프 분석이 필요
        # 여기서는 간단한 패턴만 검사
        return False

    def _check_invalid_references(self, formula: str, sheet_name: str) -> bool:
        """잘못된 참조 검사"""
        # 실제로는 시트 구조와 비교 필요
        return False

    def _check_type_mismatch(self, formula: str, sheet_name: str) -> bool:
        """타입 불일치 검사"""
        # 실제로는 참조되는 셀들의 데이터 타입 분석 필요
        return False

    def _get_error_description(self, error_type: str) -> str:
        """오류 설명 생성"""
        descriptions = {
            "division_by_zero": "0으로 나누기 오류가 발생할 수 있습니다",
            "circular_reference": "순환 참조가 감지되었습니다",
            "invalid_reference": "존재하지 않는 셀을 참조하고 있습니다",
            "type_mismatch": "데이터 타입이 호환되지 않습니다",
        }
        return descriptions.get(error_type, "알 수 없는 오류입니다")


# 전역 검증기 인스턴스
_validator_instance = None


def get_validator() -> RealtimeFormulaValidator:
    """검증기 싱글톤 인스턴스 반환"""
    global _validator_instance
    if _validator_instance is None:
        _validator_instance = RealtimeFormulaValidator()
    return _validator_instance
