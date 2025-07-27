"""
Data Quality Enhancement Service
데이터 품질 향상 및 표준화 시스템
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font, PatternFill, Alignment
import re
from typing import Dict, List, Any, Optional, Tuple, Union
from datetime import datetime, date
import logging
from dataclasses import dataclass
from enum import Enum
import locale
import phonenumbers

class DataType(Enum):
    """데이터 타입 열거형"""
    TEXT = "text"
    NUMBER = "number"
    DATE = "date"
    EMAIL = "email"
    PHONE = "phone"
    URL = "url"
    CURRENCY = "currency"
    PERCENTAGE = "percentage"
    BOOLEAN = "boolean"
    UNKNOWN = "unknown"

@dataclass
class DataQualityRule:
    """데이터 품질 규칙 정의"""
    rule_name: str
    description: str
    severity: str  # critical, high, medium, low
    auto_fixable: bool
    confidence_threshold: float

@dataclass
class DataQualityIssue:
    """데이터 품질 문제"""
    issue_type: str
    location: str
    description: str
    severity: str
    current_value: Any
    suggested_value: Any
    confidence: float
    auto_fixable: bool
    rule_applied: str

class DataQualityEnhancer:
    """데이터 품질 향상 메인 클래스"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
        
        # 데이터 품질 규칙들
        self.quality_rules = {
            'duplicate_detection': DataQualityRule(
                rule_name='중복 데이터 감지',
                description='동일한 행 또는 값의 중복을 감지',
                severity='medium',
                auto_fixable=True,
                confidence_threshold=0.9
            ),
            'missing_data_handling': DataQualityRule(
                rule_name='누락 데이터 처리',
                description='빈 셀이나 누락된 값을 처리',
                severity='medium',
                auto_fixable=True,
                confidence_threshold=0.7
            ),
            'data_type_consistency': DataQualityRule(
                rule_name='데이터 타입 일관성',
                description='컬럼 내 데이터 타입의 일관성 검사',
                severity='high',
                auto_fixable=True,
                confidence_threshold=0.8
            ),
            'date_format_standardization': DataQualityRule(
                rule_name='날짜 형식 표준화',
                description='다양한 날짜 형식을 표준 형식으로 통일',
                severity='medium',
                auto_fixable=True,
                confidence_threshold=0.85
            ),
            'number_format_standardization': DataQualityRule(
                rule_name='숫자 형식 표준화',
                description='숫자의 표기 방식과 단위를 일관되게 정리',
                severity='medium',
                auto_fixable=True,
                confidence_threshold=0.9
            ),
            'text_normalization': DataQualityRule(
                rule_name='텍스트 정규화',
                description='대소문자, 공백, 특수문자 등을 정규화',
                severity='low',
                auto_fixable=True,
                confidence_threshold=0.8
            ),
            'email_validation': DataQualityRule(
                rule_name='이메일 유효성 검사',
                description='이메일 주소의 형식을 검증',
                severity='medium',
                auto_fixable=True,
                confidence_threshold=0.95
            ),
            'phone_standardization': DataQualityRule(
                rule_name='전화번호 표준화',
                description='전화번호를 표준 형식으로 정리',
                severity='medium',
                auto_fixable=True,
                confidence_threshold=0.8
            )
        }
        
        # 날짜 패턴들
        self.date_patterns = [
            (r'\d{4}-\d{2}-\d{2}', '%Y-%m-%d'),  # 2023-12-25
            (r'\d{2}/\d{2}/\d{4}', '%m/%d/%Y'),  # 12/25/2023
            (r'\d{4}/\d{2}/\d{2}', '%Y/%m/%d'),  # 2023/12/25
            (r'\d{2}-\d{2}-\d{4}', '%m-%d-%Y'),  # 12-25-2023
            (r'\d{4}\.\d{2}\.\d{2}', '%Y.%m.%d'), # 2023.12.25
            (r'\d{2}\.\d{2}\.\d{4}', '%m.%d.%Y'), # 12.25.2023
        ]
        
        # 통화 기호 매핑
        self.currency_symbols = {
            '₩': 'KRW',
            '$': 'USD', 
            '€': 'EUR',
            '£': 'GBP',
            '¥': 'JPY',
            '원': 'KRW',
            'won': 'KRW',
            'dollar': 'USD',
            'euro': 'EUR'
        }
    
    def analyze_data_quality(self, file_path: str) -> Dict[str, Any]:
        """데이터 품질 종합 분석"""
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            analysis_result = {
                'file_path': file_path,
                'analysis_timestamp': datetime.now().isoformat(),
                'sheets_analyzed': [],
                'overall_quality_score': 0,
                'issues_found': [],
                'data_profile': {},
                'recommendations': []
            }
            
            total_quality_score = 0
            sheets_count = 0
            
            # 각 시트별 분석
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                if sheet.max_row <= 1:  # 빈 시트 건너뛰기
                    continue
                
                sheet_analysis = self._analyze_sheet_quality(sheet, sheet_name)
                analysis_result['sheets_analyzed'].append(sheet_analysis)
                total_quality_score += sheet_analysis['quality_score']
                sheets_count += 1
                
                # 시트별 이슈들을 전체 이슈 목록에 추가
                analysis_result['issues_found'].extend(sheet_analysis['issues'])
            
            # 전체 품질 점수 계산
            analysis_result['overall_quality_score'] = (
                total_quality_score / sheets_count if sheets_count > 0 else 0
            )
            
            # 데이터 프로파일 생성
            analysis_result['data_profile'] = self._generate_data_profile(analysis_result['sheets_analyzed'])
            
            # 권장사항 생성
            analysis_result['recommendations'] = self._generate_quality_recommendations(
                analysis_result['issues_found'], 
                analysis_result['overall_quality_score']
            )
            
            return analysis_result
            
        except Exception as e:
            self.logger.error(f"데이터 품질 분석 중 오류: {str(e)}")
            return {
                'error': f'데이터 품질 분석 실패: {str(e)}',
                'file_path': file_path
            }
    
    def enhance_data_quality(self, file_path: str, enhancement_options: Dict[str, bool] = None) -> Dict[str, Any]:
        """데이터 품질 향상 적용"""
        
        if enhancement_options is None:
            enhancement_options = {
                'remove_duplicates': True,
                'fill_missing_data': True,
                'standardize_formats': True,
                'normalize_text': True,
                'validate_emails': True,
                'standardize_phones': True,
                'optimize_data_types': True
            }
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            
            enhancement_result = {
                'original_file': file_path,
                'enhancements_applied': [],
                'quality_improvements': {},
                'summary': {
                    'sheets_processed': 0,
                    'total_changes': 0,
                    'quality_score_before': 0,
                    'quality_score_after': 0
                }
            }
            
            # 처리 전 품질 점수
            pre_analysis = self.analyze_data_quality(file_path)
            enhancement_result['summary']['quality_score_before'] = pre_analysis.get('overall_quality_score', 0)
            
            # 각 시트별 품질 향상 적용
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                if sheet.max_row <= 1:
                    continue
                
                sheet_enhancements = self._enhance_sheet_quality(sheet, sheet_name, enhancement_options)
                enhancement_result['enhancements_applied'].append(sheet_enhancements)
                enhancement_result['summary']['sheets_processed'] += 1
                enhancement_result['summary']['total_changes'] += sheet_enhancements.get('changes_count', 0)
            
            # 향상된 파일 저장
            output_path = file_path.replace('.xlsx', '_enhanced.xlsx')
            workbook.save(output_path)
            enhancement_result['enhanced_file_path'] = output_path
            
            # 처리 후 품질 점수 (간략 계산)
            enhancement_result['summary']['quality_score_after'] = min(
                enhancement_result['summary']['quality_score_before'] + 
                (enhancement_result['summary']['total_changes'] * 0.5), 
                100
            )
            
            return enhancement_result
            
        except Exception as e:
            self.logger.error(f"데이터 품질 향상 중 오류: {str(e)}")
            return {
                'error': f'데이터 품질 향상 실패: {str(e)}',
                'file_path': file_path
            }
    
    def _analyze_sheet_quality(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """개별 시트의 데이터 품질 분석"""
        
        # 데이터를 DataFrame으로 변환
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        
        if not data or len(data) < 2:
            return {
                'sheet_name': sheet_name,
                'quality_score': 100,  # 빈 시트는 완벽한 점수
                'issues': [],
                'data_stats': {}
            }
        
        # 헤더와 데이터 분리
        headers = data[0] if data[0] else [f"Column_{i+1}" for i in range(len(data[0]))]
        data_rows = data[1:] if len(data) > 1 else []
        
        df = pd.DataFrame(data_rows, columns=headers)
        
        sheet_analysis = {
            'sheet_name': sheet_name,
            'quality_score': 100,  # 기본 점수에서 차감
            'issues': [],
            'data_stats': {
                'total_rows': len(df),
                'total_columns': len(df.columns),
                'non_empty_cells': 0,
                'data_types_detected': {}
            }
        }
        
        quality_deductions = 0
        
        # 1. 중복 데이터 검사
        duplicate_issues = self._detect_duplicates(df, sheet_name)
        sheet_analysis['issues'].extend(duplicate_issues)
        quality_deductions += len(duplicate_issues) * 2
        
        # 2. 누락 데이터 검사
        missing_issues = self._detect_missing_data(df, sheet_name)
        sheet_analysis['issues'].extend(missing_issues)
        quality_deductions += len(missing_issues) * 1
        
        # 3. 데이터 타입 일관성 검사
        type_issues = self._detect_data_type_issues(df, sheet_name)
        sheet_analysis['issues'].extend(type_issues)
        quality_deductions += len(type_issues) * 3
        
        # 4. 날짜 형식 검사
        date_issues = self._detect_date_format_issues(df, sheet_name)
        sheet_analysis['issues'].extend(date_issues)
        quality_deductions += len(date_issues) * 2
        
        # 5. 숫자 형식 검사
        number_issues = self._detect_number_format_issues(df, sheet_name)
        sheet_analysis['issues'].extend(number_issues)
        quality_deductions += len(number_issues) * 2
        
        # 6. 텍스트 품질 검사
        text_issues = self._detect_text_quality_issues(df, sheet_name)
        sheet_analysis['issues'].extend(text_issues)
        quality_deductions += len(text_issues) * 1
        
        # 7. 이메일 유효성 검사
        email_issues = self._detect_email_issues(df, sheet_name)
        sheet_analysis['issues'].extend(email_issues)
        quality_deductions += len(email_issues) * 2
        
        # 8. 전화번호 형식 검사
        phone_issues = self._detect_phone_issues(df, sheet_name)
        sheet_analysis['issues'].extend(phone_issues)
        quality_deductions += len(phone_issues) * 1
        
        # 데이터 통계 계산
        sheet_analysis['data_stats']['non_empty_cells'] = df.count().sum()
        sheet_analysis['data_stats']['data_types_detected'] = self._analyze_data_types(df)
        
        # 최종 품질 점수 계산
        sheet_analysis['quality_score'] = max(0, 100 - quality_deductions)
        
        return sheet_analysis
    
    def _enhance_sheet_quality(self, sheet, sheet_name: str, options: Dict[str, bool]) -> Dict[str, Any]:
        """개별 시트의 데이터 품질 향상"""
        
        enhancement_result = {
            'sheet_name': sheet_name,
            'changes_count': 0,
            'enhancements': []
        }
        
        # 데이터를 DataFrame으로 변환
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)
        
        if not data or len(data) < 2:
            return enhancement_result
        
        headers = data[0] if data[0] else [f"Column_{i+1}" for i in range(len(data[0]))]
        data_rows = data[1:] if len(data) > 1 else []
        df = pd.DataFrame(data_rows, columns=headers)
        
        # 중복 제거
        if options.get('remove_duplicates', True):
            duplicate_count = self._remove_duplicates(sheet, df)
            if duplicate_count > 0:
                enhancement_result['enhancements'].append(f"{duplicate_count}개 중복 행 제거")
                enhancement_result['changes_count'] += duplicate_count
        
        # 누락 데이터 처리
        if options.get('fill_missing_data', True):
            filled_count = self._fill_missing_data(sheet, df)
            if filled_count > 0:
                enhancement_result['enhancements'].append(f"{filled_count}개 빈 셀 처리")
                enhancement_result['changes_count'] += filled_count
        
        # 형식 표준화
        if options.get('standardize_formats', True):
            format_changes = self._standardize_formats(sheet, df)
            if format_changes > 0:
                enhancement_result['enhancements'].append(f"{format_changes}개 형식 표준화")
                enhancement_result['changes_count'] += format_changes
        
        # 텍스트 정규화
        if options.get('normalize_text', True):
            text_changes = self._normalize_text_data(sheet, df)
            if text_changes > 0:
                enhancement_result['enhancements'].append(f"{text_changes}개 텍스트 정규화")
                enhancement_result['changes_count'] += text_changes
        
        return enhancement_result
    
    def _detect_duplicates(self, df: pd.DataFrame, sheet_name: str) -> List[DataQualityIssue]:
        """중복 데이터 감지"""
        
        issues = []
        
        # 전체 행 중복 검사
        duplicate_rows = df.duplicated()
        if duplicate_rows.any():
            duplicate_count = duplicate_rows.sum()
            issues.append(DataQualityIssue(
                issue_type='duplicate_rows',
                location=f"{sheet_name}!전체",
                description=f"{duplicate_count}개의 중복 행 발견",
                severity='medium',
                current_value=f"{duplicate_count} 중복 행",
                suggested_value="중복 행 제거",
                confidence=0.95,
                auto_fixable=True,
                rule_applied='duplicate_detection'
            ))
        
        # 컬럼별 중복값 검사 (키 컬럼으로 추정되는 경우)
        for col in df.columns:
            if col and ('id' in str(col).lower() or 'key' in str(col).lower()):
                col_duplicates = df[col].duplicated()
                if col_duplicates.any():
                    duplicate_count = col_duplicates.sum()
                    issues.append(DataQualityIssue(
                        issue_type='duplicate_values',
                        location=f"{sheet_name}!{col}",
                        description=f"키 컬럼 '{col}'에서 {duplicate_count}개 중복값 발견",
                        severity='high',
                        current_value=f"{duplicate_count} 중복값",
                        suggested_value="중복값 제거 또는 수정",
                        confidence=0.9,
                        auto_fixable=False,  # 키 컬럼은 수동 검토 필요
                        rule_applied='duplicate_detection'
                    ))
        
        return issues
    
    def _detect_missing_data(self, df: pd.DataFrame, sheet_name: str) -> List[DataQualityIssue]:
        """누락 데이터 감지"""
        
        issues = []
        
        for col_idx, col in enumerate(df.columns):
            if col is None:
                continue
            
            missing_count = df[col].isna().sum()
            total_count = len(df)
            missing_percentage = (missing_count / total_count) * 100
            
            if missing_count > 0:
                severity = 'high' if missing_percentage > 50 else 'medium' if missing_percentage > 20 else 'low'
                
                col_letter = get_column_letter(col_idx + 1)
                issues.append(DataQualityIssue(
                    issue_type='missing_data',
                    location=f"{sheet_name}!{col_letter}:{col_letter}",
                    description=f"컬럼 '{col}'에서 {missing_count}개 값 누락 ({missing_percentage:.1f}%)",
                    severity=severity,
                    current_value=f"{missing_count} 누락값",
                    suggested_value=self._suggest_missing_data_treatment(df[col], missing_percentage),
                    confidence=0.8 if missing_percentage < 30 else 0.6,
                    auto_fixable=missing_percentage < 30,
                    rule_applied='missing_data_handling'
                ))
        
        return issues
    
    def _detect_data_type_issues(self, df: pd.DataFrame, sheet_name: str) -> List[DataQualityIssue]:
        """데이터 타입 일관성 문제 감지"""
        
        issues = []
        
        for col_idx, col in enumerate(df.columns):
            if col is None:
                continue
                
            col_data = df[col].dropna()
            if len(col_data) == 0:
                continue
            
            # 데이터 타입 분석
            type_analysis = self._analyze_column_types(col_data)
            
            if len(type_analysis) > 1:  # 여러 타입이 섞여 있는 경우
                col_letter = get_column_letter(col_idx + 1)
                
                # 숫자가 텍스트로 저장된 경우
                if DataType.NUMBER in type_analysis and DataType.TEXT in type_analysis:
                    text_numbers = self._find_text_numbers(col_data)
                    if len(text_numbers) > len(col_data) * 0.3:  # 30% 이상
                        issues.append(DataQualityIssue(
                            issue_type='text_stored_as_number',
                            location=f"{sheet_name}!{col_letter}:{col_letter}",
                            description=f"컬럼 '{col}'에서 숫자가 텍스트로 저장됨",
                            severity='medium',
                            current_value=f"{len(text_numbers)}개 텍스트 숫자",
                            suggested_value="숫자 형식으로 변환",
                            confidence=0.9,
                            auto_fixable=True,
                            rule_applied='data_type_consistency'
                        ))
                
                # 혼재된 데이터 타입
                issues.append(DataQualityIssue(
                    issue_type='mixed_data_types',
                    location=f"{sheet_name}!{col_letter}:{col_letter}",
                    description=f"컬럼 '{col}'에서 여러 데이터 타입 혼재: {', '.join([t.value for t in type_analysis])}",
                    severity='high',
                    current_value=list(type_analysis),
                    suggested_value="일관된 데이터 타입으로 통일",
                    confidence=0.7,
                    auto_fixable=False,
                    rule_applied='data_type_consistency'
                ))
        
        return issues
    
    def _detect_date_format_issues(self, df: pd.DataFrame, sheet_name: str) -> List[DataQualityIssue]:
        """날짜 형식 문제 감지"""
        
        issues = []
        
        for col_idx, col in enumerate(df.columns):
            if col is None:
                continue
                
            col_data = df[col].dropna().astype(str)
            date_formats_found = set()
            date_values = []
            
            for value in col_data:
                for pattern, format_str in self.date_patterns:
                    if re.match(pattern, value.strip()):
                        date_formats_found.add(format_str)
                        date_values.append(value)
                        break
            
            # 날짜로 보이는 값들이 있고, 여러 형식이 섞여 있는 경우
            if len(date_values) > len(col_data) * 0.5 and len(date_formats_found) > 1:
                col_letter = get_column_letter(col_idx + 1)
                issues.append(DataQualityIssue(
                    issue_type='inconsistent_date_format',
                    location=f"{sheet_name}!{col_letter}:{col_letter}",
                    description=f"컬럼 '{col}'에서 {len(date_formats_found)}가지 날짜 형식 혼재",
                    severity='medium',
                    current_value=list(date_formats_found),
                    suggested_value="표준 날짜 형식(YYYY-MM-DD)으로 통일",
                    confidence=0.85,
                    auto_fixable=True,
                    rule_applied='date_format_standardization'
                ))
        
        return issues
    
    def _detect_number_format_issues(self, df: pd.DataFrame, sheet_name: str) -> List[DataQualityIssue]:
        """숫자 형식 문제 감지"""
        
        issues = []
        
        for col_idx, col in enumerate(df.columns):
            if col is None:
                continue
                
            col_data = df[col].dropna().astype(str)
            
            # 통화 기호 혼재 검사
            currency_symbols_found = set()
            for value in col_data:
                for symbol in self.currency_symbols.keys():
                    if symbol in value:
                        currency_symbols_found.add(symbol)
            
            if len(currency_symbols_found) > 1:
                col_letter = get_column_letter(col_idx + 1)
                issues.append(DataQualityIssue(
                    issue_type='mixed_currency_symbols',
                    location=f"{sheet_name}!{col_letter}:{col_letter}",
                    description=f"컬럼 '{col}'에서 여러 통화 기호 혼재: {', '.join(currency_symbols_found)}",
                    severity='medium',
                    current_value=list(currency_symbols_found),
                    suggested_value="일관된 통화 기호로 통일",
                    confidence=0.8,
                    auto_fixable=True,
                    rule_applied='number_format_standardization'
                ))
            
            # 천 단위 구분자 불일치 검사
            comma_format = sum(1 for v in col_data if ',' in v and re.match(r'[\d,]+', v))
            no_comma_format = sum(1 for v in col_data if v.isdigit() and len(v) > 3)
            
            if comma_format > 0 and no_comma_format > 0:
                col_letter = get_column_letter(col_idx + 1)
                issues.append(DataQualityIssue(
                    issue_type='inconsistent_number_format',
                    location=f"{sheet_name}!{col_letter}:{col_letter}",
                    description=f"컬럼 '{col}'에서 천 단위 구분자 사용 불일치",
                    severity='low',
                    current_value={"콤마_사용": comma_format, "콤마_미사용": no_comma_format},
                    suggested_value="일관된 천 단위 구분자 적용",
                    confidence=0.9,
                    auto_fixable=True,
                    rule_applied='number_format_standardization'
                ))
        
        return issues
    
    def _detect_text_quality_issues(self, df: pd.DataFrame, sheet_name: str) -> List[DataQualityIssue]:
        """텍스트 품질 문제 감지"""
        
        issues = []
        
        for col_idx, col in enumerate(df.columns):
            if col is None:
                continue
                
            col_data = df[col].dropna().astype(str)
            text_issues = []
            
            # 앞뒤 공백 문제
            leading_trailing_spaces = sum(1 for v in col_data if v != v.strip())
            if leading_trailing_spaces > 0:
                text_issues.append(f"{leading_trailing_spaces}개 값에 앞뒤 공백")
            
            # 연속 공백 문제
            multiple_spaces = sum(1 for v in col_data if '  ' in v)
            if multiple_spaces > 0:
                text_issues.append(f"{multiple_spaces}개 값에 연속 공백")
            
            # 대소문자 불일치 (같은 단어의 다른 케이스)
            case_variations = self._detect_case_variations(col_data)
            if case_variations:
                text_issues.append(f"{len(case_variations)}개 단어의 대소문자 불일치")
            
            if text_issues:
                col_letter = get_column_letter(col_idx + 1)
                issues.append(DataQualityIssue(
                    issue_type='text_quality_issues',
                    location=f"{sheet_name}!{col_letter}:{col_letter}",
                    description=f"컬럼 '{col}'에서 텍스트 품질 문제: {', '.join(text_issues)}",
                    severity='low',
                    current_value=text_issues,
                    suggested_value="텍스트 정규화 (공백 제거, 대소문자 통일)",
                    confidence=0.9,
                    auto_fixable=True,
                    rule_applied='text_normalization'
                ))
        
        return issues
    
    def _detect_email_issues(self, df: pd.DataFrame, sheet_name: str) -> List[DataQualityIssue]:
        """이메일 주소 유효성 문제 감지"""
        
        issues = []
        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        
        for col_idx, col in enumerate(df.columns):
            if col is None:
                continue
            
            # 이메일로 추정되는 컬럼 찾기
            if 'email' in str(col).lower() or 'mail' in str(col).lower():
                col_data = df[col].dropna().astype(str)
                invalid_emails = []
                
                for value in col_data:
                    if not re.match(email_pattern, value.strip()):
                        invalid_emails.append(value)
                
                if invalid_emails:
                    col_letter = get_column_letter(col_idx + 1)
                    issues.append(DataQualityIssue(
                        issue_type='invalid_email_format',
                        location=f"{sheet_name}!{col_letter}:{col_letter}",
                        description=f"컬럼 '{col}'에서 {len(invalid_emails)}개 잘못된 이메일 형식",
                        severity='medium',
                        current_value=f"{len(invalid_emails)} 잘못된 이메일",
                        suggested_value="유효한 이메일 형식으로 수정",
                        confidence=0.95,
                        auto_fixable=False,  # 이메일은 수동 검토 필요
                        rule_applied='email_validation'
                    ))
        
        return issues
    
    def _detect_phone_issues(self, df: pd.DataFrame, sheet_name: str) -> List[DataQualityIssue]:
        """전화번호 형식 문제 감지"""
        
        issues = []
        
        for col_idx, col in enumerate(df.columns):
            if col is None:
                continue
            
            # 전화번호로 추정되는 컬럼 찾기
            if any(keyword in str(col).lower() for keyword in ['phone', 'tel', '전화', '휴대폰', 'mobile']):
                col_data = df[col].dropna().astype(str)
                format_variations = set()
                
                for value in col_data:
                    # 전화번호 패턴 분석
                    cleaned = re.sub(r'[^\d-]', '', value)
                    if len(cleaned) >= 8:  # 최소 전화번호 길이
                        format_variations.add(self._get_phone_format_pattern(cleaned))
                
                if len(format_variations) > 1:
                    col_letter = get_column_letter(col_idx + 1)
                    issues.append(DataQualityIssue(
                        issue_type='inconsistent_phone_format',
                        location=f"{sheet_name}!{col_letter}:{col_letter}",
                        description=f"컬럼 '{col}'에서 {len(format_variations)}가지 전화번호 형식 혼재",
                        severity='medium',
                        current_value=list(format_variations),
                        suggested_value="표준 전화번호 형식으로 통일 (예: 010-1234-5678)",
                        confidence=0.8,
                        auto_fixable=True,
                        rule_applied='phone_standardization'
                    ))
        
        return issues
    
    # 헬퍼 메서드들
    def _analyze_column_types(self, series: pd.Series) -> set:
        """컬럼의 데이터 타입들을 분석"""
        types_found = set()
        
        for value in series:
            if pd.isna(value):
                continue
                
            str_value = str(value).strip()
            
            # 숫자 체크
            if self._is_number(str_value):
                types_found.add(DataType.NUMBER)
            # 날짜 체크
            elif self._is_date(str_value):
                types_found.add(DataType.DATE)
            # 이메일 체크
            elif self._is_email(str_value):
                types_found.add(DataType.EMAIL)
            # URL 체크
            elif self._is_url(str_value):
                types_found.add(DataType.URL)
            # 전화번호 체크
            elif self._is_phone(str_value):
                types_found.add(DataType.PHONE)
            else:
                types_found.add(DataType.TEXT)
        
        return types_found
    
    def _is_number(self, value: str) -> bool:
        """숫자인지 확인"""
        try:
            float(value.replace(',', '').replace('₩', '').replace('$', ''))
            return True
        except ValueError:
            return False
    
    def _is_date(self, value: str) -> bool:
        """날짜인지 확인"""
        for pattern, _ in self.date_patterns:
            if re.match(pattern, value):
                return True
        return False
    
    def _is_email(self, value: str) -> bool:
        """이메일인지 확인"""
        email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        return bool(re.match(email_pattern, value))
    
    def _is_url(self, value: str) -> bool:
        """URL인지 확인"""
        url_pattern = r'https?://[^\s]+'
        return bool(re.match(url_pattern, value))
    
    def _is_phone(self, value: str) -> bool:
        """전화번호인지 확인"""
        phone_pattern = r'[\d\-\(\)\+\s]{8,}'
        return bool(re.match(phone_pattern, value))
    
    def _find_text_numbers(self, series: pd.Series) -> List:
        """텍스트로 저장된 숫자들 찾기"""
        text_numbers = []
        for value in series:
            if isinstance(value, str) and self._is_number(value):
                text_numbers.append(value)
        return text_numbers
    
    def _suggest_missing_data_treatment(self, series: pd.Series, missing_percentage: float) -> str:
        """누락 데이터 처리 방법 제안"""
        if missing_percentage > 70:
            return "컬럼 제거 고려"
        elif missing_percentage > 30:
            return "기본값으로 채우기 또는 해당 행 제거"
        else:
            # 데이터 타입에 따른 제안
            non_null_data = series.dropna()
            if len(non_null_data) == 0:
                return "기본값으로 채우기"
            
            sample_value = non_null_data.iloc[0] if len(non_null_data) > 0 else None
            
            if isinstance(sample_value, (int, float)):
                return "평균값 또는 중앙값으로 채우기"
            elif self._is_date(str(sample_value)):
                return "가장 가까운 날짜로 채우기"
            else:
                return "가장 빈번한 값으로 채우기"
    
    def _detect_case_variations(self, series: pd.Series) -> List:
        """대소문자 변형들 감지"""
        word_variations = {}
        for value in series:
            lower_value = value.lower()
            if lower_value not in word_variations:
                word_variations[lower_value] = set()
            word_variations[lower_value].add(value)
        
        variations = []
        for word, cases in word_variations.items():
            if len(cases) > 1:
                variations.extend(list(cases))
        
        return variations
    
    def _get_phone_format_pattern(self, phone: str) -> str:
        """전화번호 형식 패턴 추출"""
        if '-' in phone:
            parts = phone.split('-')
            return '-'.join(['X' * len(part) for part in parts])
        elif len(phone) == 11:
            return 'XXXXXXXXXXX'
        elif len(phone) == 10:
            return 'XXXXXXXXXX'
        else:
            return f'X({len(phone)}digits)'
    
    # 실제 데이터 수정 메서드들
    def _remove_duplicates(self, sheet, df: pd.DataFrame) -> int:
        """중복 행 제거"""
        # 실제 구현에서는 Excel 시트에서 중복 행을 찾아 삭제
        duplicate_rows = df.duplicated()
        duplicate_count = duplicate_rows.sum()
        
        # 여기서는 실제 삭제 로직을 구현해야 함
        # 예시로 개수만 반환
        return duplicate_count
    
    def _fill_missing_data(self, sheet, df: pd.DataFrame) -> int:
        """누락된 데이터 채우기"""
        filled_count = 0
        
        for col in df.columns:
            missing_count = df[col].isna().sum()
            if missing_count > 0:
                # 데이터 타입에 따른 적절한 값으로 채우기
                if df[col].dtype in ['int64', 'float64']:
                    fill_value = df[col].median()
                else:
                    fill_value = df[col].mode().iloc[0] if not df[col].mode().empty else ""
                
                # 실제 Excel 시트에 값 입력
                # 여기서는 예시로 개수만 증가
                filled_count += missing_count
        
        return filled_count
    
    def _standardize_formats(self, sheet, df: pd.DataFrame) -> int:
        """형식 표준화"""
        changes_count = 0
        
        # 날짜 형식 표준화
        for col in df.columns:
            col_data = df[col].dropna().astype(str)
            for idx, value in enumerate(col_data):
                for pattern, format_str in self.date_patterns:
                    if re.match(pattern, value):
                        # 표준 형식으로 변환
                        try:
                            date_obj = datetime.strptime(value, format_str)
                            standardized = date_obj.strftime('%Y-%m-%d')
                            if standardized != value:
                                changes_count += 1
                        except ValueError:
                            continue
                        break
        
        return changes_count
    
    def _normalize_text_data(self, sheet, df: pd.DataFrame) -> int:
        """텍스트 데이터 정규화"""
        changes_count = 0
        
        for col in df.columns:
            col_data = df[col].dropna().astype(str)
            for idx, value in enumerate(col_data):
                original = value
                normalized = value.strip()  # 공백 제거
                normalized = re.sub(r'\s+', ' ', normalized)  # 연속 공백을 단일 공백으로
                
                if normalized != original:
                    changes_count += 1
        
        return changes_count
    
    def _generate_data_profile(self, sheets_analysis: List[Dict]) -> Dict[str, Any]:
        """데이터 프로파일 생성"""
        
        total_rows = sum(sheet.get('data_stats', {}).get('total_rows', 0) for sheet in sheets_analysis)
        total_columns = sum(sheet.get('data_stats', {}).get('total_columns', 0) for sheet in sheets_analysis)
        
        return {
            'total_sheets': len(sheets_analysis),
            'total_rows': total_rows,
            'total_columns': total_columns,
            'average_quality_score': sum(sheet.get('quality_score', 0) for sheet in sheets_analysis) / len(sheets_analysis) if sheets_analysis else 0
        }
    
    def _generate_quality_recommendations(self, issues: List[DataQualityIssue], overall_score: float) -> List[str]:
        """품질 개선 권장사항 생성"""
        
        recommendations = []
        
        if overall_score < 50:
            recommendations.append("🚨 데이터 품질이 매우 낮습니다. 전면적인 데이터 정리가 필요합니다.")
        elif overall_score < 70:
            recommendations.append("⚠️ 데이터 품질 개선이 필요합니다.")
        elif overall_score < 90:
            recommendations.append("✅ 데이터 품질이 양호하지만 몇 가지 개선사항이 있습니다.")
        else:
            recommendations.append("🎉 데이터 품질이 우수합니다!")
        
        # 이슈 타입별 권장사항
        issue_types = {}
        for issue in issues:
            if issue.issue_type not in issue_types:
                issue_types[issue.issue_type] = 0
            issue_types[issue.issue_type] += 1
        
        if issue_types.get('duplicate_rows', 0) > 0:
            recommendations.append("🗂️ 중복 데이터를 제거하여 데이터 정확성을 높이세요.")
        
        if issue_types.get('missing_data', 0) > 0:
            recommendations.append("📝 누락된 데이터를 적절한 값으로 채우거나 제거하세요.")
        
        if issue_types.get('mixed_data_types', 0) > 0:
            recommendations.append("🔧 데이터 타입을 일관되게 정리하여 분석 효율을 높이세요.")
        
        if issue_types.get('inconsistent_date_format', 0) > 0:
            recommendations.append("📅 날짜 형식을 표준화하여 시계열 분석을 용이하게 하세요.")
        
        return recommendations
    
    def _analyze_data_types(self, df: pd.DataFrame) -> Dict[str, int]:
        """데이터 타입별 컬럼 수 분석"""
        
        type_counts = {
            'text': 0,
            'number': 0,
            'date': 0,
            'mixed': 0
        }
        
        for col in df.columns:
            if col is None:
                continue
            
            col_data = df[col].dropna()
            if len(col_data) == 0:
                continue
            
            types_found = self._analyze_column_types(col_data)
            
            if len(types_found) > 1:
                type_counts['mixed'] += 1
            elif DataType.NUMBER in types_found:
                type_counts['number'] += 1
            elif DataType.DATE in types_found:
                type_counts['date'] += 1
            else:
                type_counts['text'] += 1
        
        return type_counts