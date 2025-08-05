import openpyxl
from openpyxl.utils import get_column_letter


def analyze_excel_file(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)

    results = {"sheets": [], "formulas": [], "errors": [], "data_summary": {}}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info = {
            "name": sheet_name,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "cells_with_data": 0,
            "cells_with_formulas": 0,
            "cells_with_errors": 0,
        }

        # 시트의 모든 셀 분석
        for row in range(1, min(ws.max_row + 1, 100)):  # 처음 100행만
            for col in range(1, min(ws.max_column + 1, 20)):  # 처음 20열만
                cell = ws.cell(row=row, column=col)
                cell_addr = f"{get_column_letter(col)}{row}"

                # 값이 있는 셀
                if cell.value is not None:
                    sheet_info["cells_with_data"] += 1

                    # 오류 값 체크
                    if isinstance(cell.value, str) and cell.value.startswith("#"):
                        sheet_info["cells_with_errors"] += 1
                        results["errors"].append(
                            {
                                "sheet": sheet_name,
                                "cell": cell_addr,
                                "error": cell.value,
                                "row": row,
                                "column": col,
                            }
                        )

                # 수식이 있는 셀 (read_only=False로 열어야 함)
                wb_formula = openpyxl.load_workbook(filepath, data_only=False)
                ws_formula = wb_formula[sheet_name]
                cell_formula = ws_formula.cell(row=row, column=col)

                if (
                    cell_formula.value
                    and isinstance(cell_formula.value, str)
                    and cell_formula.value.startswith("=")
                ):
                    sheet_info["cells_with_formulas"] += 1
                    results["formulas"].append(
                        {
                            "sheet": sheet_name,
                            "cell": cell_addr,
                            "formula": cell_formula.value,
                            "value": cell.value,
                        }
                    )

        results["sheets"].append(sheet_info)
        wb_formula.close()

    wb.close()
    return results


# 파일 분석
filepath = "/Users/kevin/Downloads/7777777.xlsx"
analysis = analyze_excel_file(filepath)

print("=== Excel 파일 분석 결과 ===")
print(f"\n시트 개수: {len(analysis['sheets'])}")

for sheet in analysis["sheets"]:
    print(f"\n시트: {sheet['name']}")
    print(f"  - 크기: {sheet['max_row']}행 x {sheet['max_column']}열")
    print(f"  - 데이터 셀: {sheet['cells_with_data']}")
    print(f"  - 수식 셀: {sheet['cells_with_formulas']}")
    print(f"  - 오류 셀: {sheet['cells_with_errors']}")

print(f"\n=== 발견된 오류 ({len(analysis['errors'])}개) ===")
for error in analysis["errors"]:
    print(f"  {error['sheet']} - {error['cell']}: {error['error']}")

print("\n=== 수식 분석 (일부) ===")
for formula in analysis["formulas"][:10]:  # 처음 10개만
    print(
        f"  {formula['sheet']} - {formula['cell']}: {formula['formula']} = {formula['value']}"
    )
