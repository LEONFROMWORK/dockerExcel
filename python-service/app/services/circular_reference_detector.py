"""
Advanced Circular Reference Detection Service
복잡한 순환 참조 패턴을 감지하고 분석하는 고급 시스템
"""

from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
import re
from typing import List, Dict, Set
from dataclasses import dataclass
import networkx as nx
import logging


@dataclass
class CircularReferenceChain:
    """순환 참조 체인 정보"""

    cells: List[str]  # 순환 참조에 포함된 셀들
    chain_type: str  # direct, indirect, multi-sheet
    severity: str  # critical, high, medium
    description: str
    break_suggestions: List[Dict[str, str]]  # 순환 참조를 끊을 수 있는 제안들


class CircularReferenceDetector:
    """고급 순환 참조 감지 시스템"""

    def __init__(self):
        self.logger = logging.getLogger(__name__)
        self.dependency_graph = nx.DiGraph()
        self.cell_formulas = {}
        self.sheet_references = {}

    def analyze_workbook(self, workbook) -> List[CircularReferenceChain]:
        """워크북 전체의 순환 참조를 분석"""

        # 1단계: 모든 셀의 의존성 그래프 구축
        self._build_dependency_graph(workbook)

        # 2단계: 순환 참조 감지
        circular_refs = self._detect_all_circular_references()

        # 3단계: 순환 참조 체인 분석 및 수정 제안 생성
        circular_chains = self._analyze_circular_chains(circular_refs)

        return circular_chains

    def _build_dependency_graph(self, workbook):
        """모든 셀의 의존성 그래프를 구축"""

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f" and cell.value:
                        cell_ref = f"{sheet_name}!{cell.coordinate}"
                        self.cell_formulas[cell_ref] = cell.value

                        # 수식에서 참조하는 셀들 추출
                        referenced_cells = self._extract_cell_references(
                            cell.value, sheet_name, workbook
                        )

                        # 의존성 그래프에 엣지 추가
                        for ref_cell in referenced_cells:
                            self.dependency_graph.add_edge(ref_cell, cell_ref)

    def _extract_cell_references(
        self, formula: str, current_sheet: str, workbook
    ) -> Set[str]:
        """수식에서 참조하는 모든 셀 추출 (다른 시트 포함)"""

        references = set()

        # 일반 셀 참조 패턴 (A1, $A$1, A:A, 1:1 등)
        cell_pattern = r"(?:(?:\'[^\']+\'|[A-Za-z_]\w*)!)?(?:\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?|\$?[A-Z]+:\$?[A-Z]+|\d+:\d+)"

        # 명명된 범위 패턴
        named_range_pattern = r"[A-Za-z_]\w*(?![!:\(])"

        # 함수 이름 제외
        function_names = {
            "SUM",
            "AVERAGE",
            "COUNT",
            "MAX",
            "MIN",
            "IF",
            "VLOOKUP",
            "HLOOKUP",
            "INDEX",
            "MATCH",
            "SUMIF",
            "COUNTIF",
            "AND",
            "OR",
            "NOT",
            "IFERROR",
            "CONCATENATE",
            "LEFT",
            "RIGHT",
            "MID",
            "LEN",
            "TRIM",
            "DATE",
            "NOW",
        }

        # 셀 참조 추출
        for match in re.finditer(cell_pattern, formula):
            ref = match.group()

            # 시트 이름이 포함된 경우
            if "!" in ref:
                sheet_ref, cell_ref = ref.split("!", 1)
                sheet_ref = sheet_ref.strip("'")
            else:
                sheet_ref = current_sheet
                cell_ref = ref

            # 범위 참조인 경우 개별 셀로 확장
            if ":" in cell_ref:
                for cell in self._expand_range(cell_ref, sheet_ref):
                    references.add(f"{sheet_ref}!{cell}")
            else:
                references.add(f"{sheet_ref}!{cell_ref}")

        # 명명된 범위 처리
        for match in re.finditer(named_range_pattern, formula):
            name = match.group()
            if name.upper() not in function_names:
                # 실제 명명된 범위인지 확인하고 참조 추가
                if (
                    hasattr(workbook, "defined_names")
                    and name in workbook.defined_names
                ):
                    named_range = workbook.defined_names[name]
                    # 명명된 범위가 참조하는 셀들 추가
                    for dest in named_range.destinations:
                        sheet_name, cell_range = dest
                        if ":" in cell_range:
                            for cell in self._expand_range(cell_range, sheet_name):
                                references.add(f"{sheet_name}!{cell}")
                        else:
                            references.add(f"{sheet_name}!{cell_range}")

        return references

    def _expand_range(self, range_ref: str, sheet_name: str) -> List[str]:
        """범위 참조를 개별 셀로 확장 (메모리 효율성을 위해 제한적으로)"""

        cells = []
        try:
            if ":" in range_ref:
                start, end = range_ref.split(":")

                # 전체 열/행 참조는 처리하지 않음
                if start.isalpha() or end.isalpha() or start.isdigit() or end.isdigit():
                    return []

                start_col, start_row = coordinate_from_string(start)
                end_col, end_row = coordinate_from_string(end)

                # 너무 큰 범위는 제한
                if (
                    column_index_from_string(end_col)
                    - column_index_from_string(start_col)
                    > 100
                    or end_row - start_row > 100
                ):
                    return []

                for row in range(start_row, end_row + 1):
                    for col_idx in range(
                        column_index_from_string(start_col),
                        column_index_from_string(end_col) + 1,
                    ):
                        cells.append(f"{get_column_letter(col_idx)}{row}")
        except (KeyError, IndexError, AttributeError):
            pass

        return cells

    def _detect_all_circular_references(self) -> List[List[str]]:
        """모든 순환 참조 감지"""

        try:
            # NetworkX의 순환 감지 알고리즘 사용
            cycles = list(nx.simple_cycles(self.dependency_graph))
            return cycles
        except (KeyError, IndexError, AttributeError):
            # 대안: 수동 DFS 기반 순환 감지
            return self._manual_cycle_detection()

    def _manual_cycle_detection(self) -> List[List[str]]:
        """수동 DFS 기반 순환 참조 감지"""

        cycles = []
        visited = set()
        rec_stack = set()

        def dfs(node, path):
            visited.add(node)
            rec_stack.add(node)
            path.append(node)

            for neighbor in self.dependency_graph.neighbors(node):
                if neighbor not in visited:
                    if dfs(neighbor, path.copy()):
                        return True
                elif neighbor in rec_stack:
                    # 순환 참조 발견
                    cycle_start = path.index(neighbor)
                    cycle = path[cycle_start:] + [neighbor]
                    cycles.append(cycle)

            rec_stack.remove(node)
            return False

        for node in self.dependency_graph.nodes():
            if node not in visited:
                dfs(node, [])

        return cycles

    def _analyze_circular_chains(
        self, cycles: List[List[str]]
    ) -> List[CircularReferenceChain]:
        """순환 참조 체인을 분석하고 수정 제안 생성"""

        circular_chains = []

        for cycle in cycles:
            # 순환 참조 타입 판단
            chain_type = self._determine_chain_type(cycle)

            # 심각도 평가
            severity = self._evaluate_severity(cycle, chain_type)

            # 설명 생성
            description = self._generate_description(cycle, chain_type)

            # 수정 제안 생성
            break_suggestions = self._generate_break_suggestions(cycle)

            circular_chains.append(
                CircularReferenceChain(
                    cells=cycle,
                    chain_type=chain_type,
                    severity=severity,
                    description=description,
                    break_suggestions=break_suggestions,
                )
            )

        return circular_chains

    def _determine_chain_type(self, cycle: List[str]) -> str:
        """순환 참조 체인의 타입 결정"""

        sheets = set(cell.split("!")[0] for cell in cycle)

        if len(sheets) > 1:
            return "multi-sheet"
        elif len(cycle) == 2:
            return "direct"
        else:
            return "indirect"

    def _evaluate_severity(self, cycle: List[str], chain_type: str) -> str:
        """순환 참조의 심각도 평가"""

        if chain_type == "multi-sheet" or len(cycle) > 5:
            return "critical"
        elif len(cycle) > 3:
            return "high"
        else:
            return "medium"

    def _generate_description(self, cycle: List[str], chain_type: str) -> str:
        """순환 참조에 대한 설명 생성"""

        if chain_type == "direct":
            return f"{cycle[0]}과 {cycle[1]}이 서로를 참조하고 있습니다."
        elif chain_type == "multi-sheet":
            sheets = set(cell.split("!")[0] for cell in cycle)
            return f"여러 시트({', '.join(sheets)})에 걸친 순환 참조가 발견되었습니다."
        else:
            return (
                f"{len(cycle)}개의 셀이 순환 참조 체인을 형성하고 있습니다: "
                + " → ".join(cycle)
                + " → "
                + cycle[0]
            )

    def _generate_break_suggestions(self, cycle: List[str]) -> List[Dict[str, str]]:
        """순환 참조를 끊기 위한 제안 생성"""

        suggestions = []

        # 각 연결점에서 순환을 끊는 제안
        for i, cell in enumerate(cycle):
            next_cell = cycle[(i + 1) % len(cycle)]

            suggestions.append(
                {
                    "action": "remove_reference",
                    "target_cell": cell,
                    "remove_reference_to": next_cell,
                    "description": f"{cell}에서 {next_cell}에 대한 참조를 제거하면 순환이 해결됩니다.",
                    "impact": self._evaluate_break_impact(cell, next_cell),
                }
            )

        # 가장 영향이 적은 제안을 우선순위로 정렬
        suggestions.sort(key=lambda x: x["impact"])

        # 추가 제안: 중간 계산 셀 사용
        suggestions.append(
            {
                "action": "use_intermediate_cell",
                "description": "중간 계산용 보조 셀을 만들어 순환 참조를 해결할 수 있습니다.",
                "impact": "low",
            }
        )

        return suggestions[:3]  # 상위 3개 제안만 반환

    def _evaluate_break_impact(self, source: str, target: str) -> str:
        """참조를 끊었을 때의 영향도 평가"""

        # 해당 셀을 참조하는 다른 셀의 수 확인
        dependents = len(list(self.dependency_graph.predecessors(source)))

        if dependents > 10:
            return "high"
        elif dependents > 5:
            return "medium"
        else:
            return "low"
