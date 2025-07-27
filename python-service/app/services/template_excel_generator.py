"""
템플릿 기반 Excel 생성 엔진
Template-based Excel Generation Engine
"""

import json
import logging
import os
import tempfile
from typing import Dict, List, Any, Optional
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import numpy as np

from .template_selection_service import template_selection_service
from .excel_chart_analyzer import excel_chart_analyzer
from .excel_pivot_analyzer import excel_pivot_analyzer

logger = logging.getLogger(__name__)


class TemplateExcelGenerator:
    """템플릿 기반 Excel 생성기"""
    
    def __init__(self):
        self.templates_metadata = template_selection_service.templates_metadata
        
        # 스타일 정의
        self.styles = {
            "header": {
                "font": Font(bold=True, color="FFFFFF", size=12),
                "fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
                "alignment": Alignment(horizontal="center", vertical="center")
            },
            "subheader": {
                "font": Font(bold=True, size=11),
                "fill": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
                "alignment": Alignment(horizontal="center", vertical="center")
            },
            "data": {
                "font": Font(size=10),
                "alignment": Alignment(horizontal="right", vertical="center")
            },
            "currency": {
                "font": Font(size=10),
                "alignment": Alignment(horizontal="right", vertical="center"),
                "number_format": "#,##0"
            },
            "percentage": {
                "font": Font(size=10),
                "alignment": Alignment(horizontal="right", vertical="center"),
                "number_format": "0.0%"
            },
            "border": Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    
    async def generate_from_template(
        self,
        template_id: str,
        user_data: Dict[str, Any] = None,
        customization: Dict[str, Any] = None
    ) -> Dict[str, Any]:
        """템플릿에서 Excel 파일 생성"""
        
        try:
            logger.info(f"템플릿 기반 Excel 생성 시작: {template_id}")
            
            # 1. 템플릿 메타데이터 로드
            template_info = template_selection_service.get_template_by_id(template_id)
            
            if template_info["status"] != "success":
                return template_info
            
            template = template_info["template"]
            
            # 2. 템플릿별 생성 함수 호출
            generator_method = getattr(
                self, 
                f"_generate_{template_id}", 
                self._generate_generic_template
            )
            
            # 3. Excel 파일 생성
            result = await generator_method(template, user_data, customization)
            
            if result["status"] == "success":
                logger.info(f"템플릿 기반 Excel 생성 완료: {result['output_file']}")
            
            return result
            
        except Exception as e:
            logger.error(f"템플릿 기반 Excel 생성 실패: {str(e)}")
            return {
                "status": "error",
                "error": str(e)
            }
    
    async def _generate_quarterly_financial_report(
        self,
        template: Dict[str, Any],
        user_data: Dict[str, Any] = None,
        customization: Dict[str, Any] = None
    ) -> Dict[str, Any]:
        """분기별 재무보고서 생성"""
        
        try:
            # 새 워크북 생성
            wb = openpyxl.Workbook()
            
            # 기본 시트 제거 후 새 시트 생성
            wb.remove(wb.active)
            
            # 1. 요약 시트
            summary_ws = wb.create_sheet("재무요약")
            self._create_financial_summary_sheet(summary_ws, user_data)
            
            # 2. 상세 손익계산서
            income_ws = wb.create_sheet("손익계산서") 
            self._create_income_statement_sheet(income_ws, user_data)
            
            # 3. 차트 시트
            chart_ws = wb.create_sheet("재무분석")
            self._create_financial_charts_sheet(chart_ws, user_data)
            
            # 파일 저장
            output_path = self._save_workbook(wb, "quarterly_financial_report")
            
            return {
                "status": "success",
                "template_id": "quarterly_financial_report",
                "output_file": output_path,
                "sheets_created": ["재무요약", "손익계산서", "재무분석"],
                "features_applied": [
                    "자동 합계 계산",
                    "전기 대비 증감률", 
                    "재무 차트",
                    "조건부 서식"
                ]
            }
            
        except Exception as e:
            logger.error(f"분기별 재무보고서 생성 실패: {str(e)}")
            return {"status": "error", "error": str(e)}
    
    def _create_financial_summary_sheet(self, ws, user_data: Dict[str, Any] = None):
        """재무요약 시트 생성"""
        
        # 제목
        ws['A1'] = '분기별 재무 요약'
        ws.merge_cells('A1:E1')
        self._apply_style(ws['A1'], self.styles["header"])
        
        # 헤더
        headers = ['항목', '당분기', '전분기', '증감액', '증감률']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            self._apply_style(cell, self.styles["subheader"])
        
        # 데이터 (샘플 또는 사용자 데이터)
        if user_data and 'financial_data' in user_data:
            financial_data = user_data['financial_data']
        else:
            financial_data = self._generate_sample_financial_data()
        
        row = 3
        for item in financial_data:
            ws.cell(row=row, column=1, value=item['category'])
            ws.cell(row=row, column=2, value=item['current'])
            ws.cell(row=row, column=3, value=item['previous'])
            
            # 증감액 계산
            change = item['current'] - item['previous']
            ws.cell(row=row, column=4, value=change)
            
            # 증감률 계산
            if item['previous'] != 0:
                change_rate = change / item['previous']
                ws.cell(row=row, column=5, value=change_rate)
                self._apply_style(ws.cell(row=row, column=5), self.styles["percentage"])
            
            # 스타일 적용
            for col in range(2, 5):
                self._apply_style(ws.cell(row=row, column=col), self.styles["currency"])
            
            row += 1
        
        # 열 너비 조정
        column_widths = [20, 15, 15, 15, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_income_statement_sheet(self, ws, user_data: Dict[str, Any] = None):
        """손익계산서 시트 생성"""
        
        # 제목
        ws['A1'] = '손익계산서'
        ws['E1'] = '(단위: 천원)'
        ws.merge_cells('A1:D1')
        self._apply_style(ws['A1'], self.styles["header"])
        
        # 헤더
        headers = ['계정과목', '당분기', '전분기', '전년동기', '비고']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            self._apply_style(cell, self.styles["subheader"])
        
        # 손익계산서 항목
        income_items = [
            '매출액', '매출원가', '매출총이익', '판매관리비', 
            '영업이익', '영업외수익', '영업외비용', '법인세비용', '당기순이익'
        ]
        
        # 데이터 입력
        if user_data and 'income_data' in user_data:
            income_data = user_data['income_data']
        else:
            income_data = self._generate_sample_income_data()
        
        row = 3
        for item in income_items:
            ws.cell(row=row, column=1, value=item)
            
            if item in income_data:
                data = income_data[item]
                ws.cell(row=row, column=2, value=data.get('current', 0))
                ws.cell(row=row, column=3, value=data.get('previous', 0))
                ws.cell(row=row, column=4, value=data.get('last_year', 0))
            
            # 핵심 항목 강조
            if item in ['매출총이익', '영업이익', '당기순이익']:
                for col in range(1, 6):
                    cell = ws.cell(row=row, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            
            row += 1
        
        # 열 너비 조정
        column_widths = [20, 15, 15, 15, 20]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_financial_charts_sheet(self, ws, user_data: Dict[str, Any] = None):
        """재무분석 차트 시트 생성"""
        
        # 제목
        ws['A1'] = '재무 분석 차트'
        ws.merge_cells('A1:F1')
        self._apply_style(ws['A1'], self.styles["header"])
        
        # 차트용 데이터 준비
        chart_data = self._prepare_chart_data(user_data)
        
        # 1. 매출 추이 차트
        self._create_revenue_trend_chart(ws, chart_data, 'A3')
        
        # 2. 수익성 분석 차트  
        self._create_profitability_chart(ws, chart_data, 'A20')
        
        # 3. 비용 구성 파이 차트
        self._create_cost_composition_chart(ws, chart_data, 'H3')
    
    def _create_revenue_trend_chart(self, ws, data, position):
        """매출 추이 라인 차트 생성"""
        
        # 데이터 테이블 생성
        periods = ['1분기', '2분기', '3분기', '4분기']
        revenues = data.get('quarterly_revenue', [100000, 120000, 110000, 150000])
        
        # 데이터 입력
        ws[position] = '분기별 매출 추이'
        start_row = int(position[1:]) + 1
        
        for i, period in enumerate(periods):
            ws.cell(row=start_row + i, column=1, value=period)
            ws.cell(row=start_row + i, column=2, value=revenues[i])
        
        # 라인 차트 생성
        chart = LineChart()
        chart.title = "분기별 매출 추이"
        chart.y_axis.title = "매출액 (천원)"
        chart.x_axis.title = "분기"
        
        data_ref = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(periods) - 1)
        categories = Reference(ws, min_col=1, min_row=start_row, max_row=start_row + len(periods) - 1)
        
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(categories)
        
        # 차트를 시트에 추가
        chart_position = f'D{start_row}'
        ws.add_chart(chart, chart_position)
    
    def _create_profitability_chart(self, ws, data, position):
        """수익성 분석 막대 차트 생성"""
        
        # 수익성 지표
        profitability_items = ['매출총이익률', '영업이익률', '순이익률']
        rates = data.get('profitability_rates', [0.25, 0.15, 0.10])
        
        start_row = int(position[1:]) + 1
        
        # 데이터 입력
        ws[position] = '수익성 분석'
        for i, (item, rate) in enumerate(zip(profitability_items, rates)):
            ws.cell(row=start_row + i, column=1, value=item)
            ws.cell(row=start_row + i, column=2, value=rate)
        
        # 막대 차트 생성
        chart = BarChart()
        chart.title = "수익성 지표"
        chart.y_axis.title = "비율"
        chart.x_axis.title = "지표"
        
        data_ref = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(rates) - 1)
        categories = Reference(ws, min_col=1, min_row=start_row, max_row=start_row + len(rates) - 1)
        
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(categories)
        
        chart_position = f'D{start_row}'
        ws.add_chart(chart, chart_position)
    
    def _create_cost_composition_chart(self, ws, data, position):
        """비용 구성 파이 차트 생성"""
        
        # 비용 구성
        cost_items = ['재료비', '인건비', '관리비', '기타']
        costs = data.get('cost_composition', [40000, 30000, 20000, 10000])
        
        start_row = int(position[1:]) + 1
        
        # 데이터 입력
        ws[position] = '비용 구성'
        for i, (item, cost) in enumerate(zip(cost_items, costs)):
            ws.cell(row=start_row + i, column=1, value=item)
            ws.cell(row=start_row + i, column=2, value=cost)
        
        # 파이 차트 생성
        chart = PieChart()
        chart.title = "비용 구성"
        
        data_ref = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(costs) - 1)
        categories = Reference(ws, min_col=1, min_row=start_row, max_row=start_row + len(costs) - 1)
        
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(categories)
        
        chart_position = f'J{start_row}'
        ws.add_chart(chart, chart_position)
    
    async def _generate_sales_performance_dashboard(
        self,
        template: Dict[str, Any],
        user_data: Dict[str, Any] = None,
        customization: Dict[str, Any] = None
    ) -> Dict[str, Any]:
        """영업 성과 대시보드 생성"""
        
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            # 1. 대시보드 메인 시트
            dashboard_ws = wb.create_sheet("영업대시보드")
            self._create_sales_dashboard_sheet(dashboard_ws, user_data)
            
            # 2. 상세 데이터 시트
            detail_ws = wb.create_sheet("상세데이터")
            self._create_sales_detail_sheet(detail_ws, user_data)
            
            output_path = self._save_workbook(wb, "sales_performance_dashboard")
            
            return {
                "status": "success",
                "template_id": "sales_performance_dashboard",
                "output_file": output_path,
                "sheets_created": ["영업대시보드", "상세데이터"],
                "features_applied": [
                    "실시간 성과 추적",
                    "목표 대비 실적",
                    "순위 및 랭킹",
                    "성과 시각화"
                ]
            }
            
        except Exception as e:
            logger.error(f"영업 성과 대시보드 생성 실패: {str(e)}")
            return {"status": "error", "error": str(e)}
    
    def _create_sales_dashboard_sheet(self, ws, user_data: Dict[str, Any] = None):
        """영업 대시보드 메인 시트 생성"""
        
        # 제목
        ws['A1'] = '영업 성과 대시보드'
        ws.merge_cells('A1:H1')
        self._apply_style(ws['A1'], self.styles["header"])
        
        # KPI 섹션
        ws['A3'] = 'Key Performance Indicators'
        self._apply_style(ws['A3'], self.styles["subheader"])
        
        # 샘플 KPI 데이터
        kpi_data = user_data.get('kpi_data', {}) if user_data else {}
        kpis = [
            ('총 매출', kpi_data.get('total_sales', 500000000), '원'),
            ('목표 달성률', kpi_data.get('target_achievement', 0.85), '%'),
            ('신규 고객', kpi_data.get('new_customers', 150), '명'),
            ('평균 거래액', kpi_data.get('avg_deal', 3333333), '원')
        ]
        
        # KPI 박스 생성
        for i, (label, value, unit) in enumerate(kpis):
            col = i * 2 + 1
            ws.cell(row=4, column=col, value=label)
            ws.cell(row=5, column=col, value=value)
            
            # KPI 값 스타일링
            kpi_cell = ws.cell(row=5, column=col)
            kpi_cell.font = Font(size=16, bold=True, color="2F5597")
            
            if unit == '%':
                kpi_cell.number_format = "0%"
            elif unit == '원':
                kpi_cell.number_format = "#,##0"
        
        # 영업사원별 성과 테이블
        ws['A8'] = '영업사원별 성과'
        self._apply_style(ws['A8'], self.styles["subheader"])
        
        # 테이블 헤더
        sales_headers = ['이름', '목표', '실적', '달성률', '순위']
        for col, header in enumerate(sales_headers, 1):
            cell = ws.cell(row=9, column=col, value=header)
            self._apply_style(cell, self.styles["subheader"])
        
        # 영업사원 데이터
        sales_data = user_data.get('sales_team_data', []) if user_data else self._generate_sample_sales_data()
        
        for i, person in enumerate(sales_data, 10):
            ws.cell(row=i, column=1, value=person['name'])
            ws.cell(row=i, column=2, value=person['target'])
            ws.cell(row=i, column=3, value=person['actual'])
            
            # 달성률 계산
            achievement = person['actual'] / person['target'] if person['target'] > 0 else 0
            achievement_cell = ws.cell(row=i, column=4, value=achievement)
            achievement_cell.number_format = "0%"
            
            # 달성률에 따른 조건부 서식
            if achievement >= 1.0:
                achievement_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif achievement >= 0.8:
                achievement_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                achievement_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            ws.cell(row=i, column=5, value=person['rank'])
        
        # 월별 매출 추이 차트 생성
        self._create_monthly_sales_chart(ws, user_data)
    
    def _create_monthly_sales_chart(self, ws, user_data: Dict[str, Any] = None):
        """월별 매출 추이 차트 생성"""
        
        # 차트 데이터 위치
        chart_start_row = 9
        chart_start_col = 7
        
        # 월별 데이터
        months = ['1월', '2월', '3월', '4월', '5월', '6월']
        monthly_sales = user_data.get('monthly_sales', [50000, 60000, 55000, 70000, 80000, 75000]) if user_data else [50000, 60000, 55000, 70000, 80000, 75000]
        
        # 데이터 테이블 생성
        ws.cell(row=chart_start_row, column=chart_start_col, value='월')
        ws.cell(row=chart_start_row, column=chart_start_col + 1, value='매출')
        
        for i, (month, sales) in enumerate(zip(months, monthly_sales)):
            ws.cell(row=chart_start_row + 1 + i, column=chart_start_col, value=month)
            ws.cell(row=chart_start_row + 1 + i, column=chart_start_col + 1, value=sales)
        
        # 라인 차트 생성
        chart = LineChart()
        chart.title = "월별 매출 추이"
        chart.y_axis.title = "매출 (천원)"
        chart.x_axis.title = "월"
        
        data_ref = Reference(ws, min_col=chart_start_col + 1, min_row=chart_start_row + 1, max_row=chart_start_row + len(months))
        categories = Reference(ws, min_col=chart_start_col, min_row=chart_start_row + 1, max_row=chart_start_row + len(months))
        
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(categories)
        
        # 차트 위치 설정
        ws.add_chart(chart, f'{get_column_letter(chart_start_col + 3)}{chart_start_row}')
    
    def _create_sales_detail_sheet(self, ws, user_data: Dict[str, Any] = None):
        """영업 상세 데이터 시트 생성"""
        
        # 제목
        ws['A1'] = '영업 상세 데이터'
        ws.merge_cells('A1:H1')
        self._apply_style(ws['A1'], self.styles["header"])
        
        # 헤더
        headers = ['날짜', '영업사원', '고객명', '제품', '수량', '단가', '매출액', '상태']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            self._apply_style(cell, self.styles["subheader"])
        
        # 상세 데이터 (샘플 또는 사용자 데이터)
        detail_data = user_data.get('sales_detail', []) if user_data else self._generate_sample_sales_detail()
        
        for i, record in enumerate(detail_data, 3):
            ws.cell(row=i, column=1, value=record.get('date', datetime.now().strftime('%Y-%m-%d')))
            ws.cell(row=i, column=2, value=record.get('salesperson', ''))
            ws.cell(row=i, column=3, value=record.get('customer', ''))
            ws.cell(row=i, column=4, value=record.get('product', ''))
            ws.cell(row=i, column=5, value=record.get('quantity', 0))
            ws.cell(row=i, column=6, value=record.get('unit_price', 0))
            ws.cell(row=i, column=7, value=record.get('amount', 0))
            ws.cell(row=i, column=8, value=record.get('status', ''))
        
        # 테이블 스타일 적용
        table_range = f"A2:{get_column_letter(len(headers))}{len(detail_data) + 2}"
        table = Table(displayName="SalesData", ref=table_range)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9", 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=True
        )
        ws.add_table(table)
    
    async def _generate_generic_template(
        self,
        template: Dict[str, Any],
        user_data: Dict[str, Any] = None,
        customization: Dict[str, Any] = None
    ) -> Dict[str, Any]:
        """범용 템플릿 생성"""
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = template.get("name", "Generic Template")
            
            # 기본 헤더 생성
            ws['A1'] = template.get("name", "템플릿")
            ws.merge_cells('A1:D1')
            self._apply_style(ws['A1'], self.styles["header"])
            
            # 템플릿 설명
            ws['A3'] = template.get("description", "")
            
            # 기본 데이터 구조 생성
            if user_data:
                self._populate_generic_data(ws, user_data)
            else:
                self._create_generic_structure(ws, template)
            
            output_path = self._save_workbook(wb, template.get("id", "generic_template"))
            
            return {
                "status": "success",
                "template_id": template.get("id", "generic"),
                "output_file": output_path,
                "sheets_created": [ws.title],
                "features_applied": ["기본 구조", "스타일링"]
            }
            
        except Exception as e:
            logger.error(f"범용 템플릿 생성 실패: {str(e)}")
            return {"status": "error", "error": str(e)}
    
    def _populate_generic_data(self, ws, user_data: Dict[str, Any]):
        """범용 데이터 입력"""
        
        # 사용자 데이터가 pandas DataFrame인 경우
        if isinstance(user_data.get('data'), pd.DataFrame):
            df = user_data['data']
            
            # 헤더 입력
            for col, header in enumerate(df.columns, 1):
                cell = ws.cell(row=5, column=col, value=header)
                self._apply_style(cell, self.styles["subheader"])
            
            # 데이터 입력
            for row_idx, row in enumerate(df.itertuples(index=False), 6):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 딕셔너리 형태의 데이터인 경우
        elif isinstance(user_data.get('data'), list):
            data_list = user_data['data']
            if data_list:
                # 첫 번째 레코드에서 키 추출
                headers = list(data_list[0].keys()) if isinstance(data_list[0], dict) else []
                
                # 헤더 입력
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=5, column=col, value=header)
                    self._apply_style(cell, self.styles["subheader"])
                
                # 데이터 입력
                for row_idx, record in enumerate(data_list, 6):
                    if isinstance(record, dict):
                        for col_idx, header in enumerate(headers, 1):
                            value = record.get(header, "")
                            ws.cell(row=row_idx, column=col_idx, value=value)
    
    def _create_generic_structure(self, ws, template: Dict[str, Any]):
        """범용 구조 생성"""
        
        # 데이터 요구사항 기반 구조 생성
        requirements = template.get("data_requirements", {})
        
        # 기본 헤더 생성
        headers = ["항목", "값", "비고"]
        
        # 숫자 컬럼이 필요한 경우
        numeric_cols = requirements.get("numeric_columns", {})
        if isinstance(numeric_cols, dict):
            min_numeric = numeric_cols.get("min", 0)
            for i in range(min_numeric):
                headers.append(f"수치{i+1}")
        
        # 날짜 컬럼이 필요한 경우
        if requirements.get("date_column", {}).get("required", False):
            headers.append("날짜")
        
        # 헤더 입력
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            self._apply_style(cell, self.styles["subheader"])
        
        # 샘플 데이터 생성
        min_rows = requirements.get("min_rows", 5)
        for row in range(6, 6 + min_rows):
            ws.cell(row=row, column=1, value=f"항목 {row-5}")
            for col in range(2, len(headers) + 1):
                if "수치" in headers[col-1]:
                    ws.cell(row=row, column=col, value=np.random.randint(1, 100))
                elif "날짜" in headers[col-1]:
                    ws.cell(row=row, column=col, value=datetime.now().strftime('%Y-%m-%d'))
                else:
                    ws.cell(row=row, column=col, value="")
    
    def _generate_sample_financial_data(self) -> List[Dict[str, Any]]:
        """샘플 재무 데이터 생성"""
        
        return [
            {'category': '매출액', 'current': 500000, 'previous': 450000},
            {'category': '매출원가', 'current': 300000, 'previous': 280000},
            {'category': '판관비', 'current': 100000, 'previous': 95000},
            {'category': '영업이익', 'current': 100000, 'previous': 75000},
            {'category': '당기순이익', 'current': 80000, 'previous': 60000}
        ]
    
    def _generate_sample_income_data(self) -> Dict[str, Dict[str, int]]:
        """샘플 손익계산서 데이터 생성"""
        
        return {
            '매출액': {'current': 500000, 'previous': 450000, 'last_year': 420000},
            '매출원가': {'current': 300000, 'previous': 280000, 'last_year': 270000},
            '매출총이익': {'current': 200000, 'previous': 170000, 'last_year': 150000},
            '판매관리비': {'current': 100000, 'previous': 95000, 'last_year': 90000},
            '영업이익': {'current': 100000, 'previous': 75000, 'last_year': 60000},
            '영업외수익': {'current': 5000, 'previous': 3000, 'last_year': 2000},
            '영업외비용': {'current': 3000, 'previous': 2000, 'last_year': 1500},
            '법인세비용': {'current': 20000, 'previous': 15000, 'last_year': 12000},
            '당기순이익': {'current': 82000, 'previous': 61000, 'last_year': 48500}
        }
    
    def _generate_sample_sales_data(self) -> List[Dict[str, Any]]:
        """샘플 영업 데이터 생성"""
        
        return [
            {'name': '김철수', 'target': 100000, 'actual': 120000, 'rank': 1},
            {'name': '이영희', 'target': 90000, 'actual': 95000, 'rank': 2},
            {'name': '박민수', 'target': 80000, 'actual': 75000, 'rank': 3},
            {'name': '정수진', 'target': 70000, 'actual': 85000, 'rank': 4},
            {'name': '최동훈', 'target': 60000, 'actual': 55000, 'rank': 5}
        ]
    
    def _generate_sample_sales_detail(self) -> List[Dict[str, Any]]:
        """샘플 영업 상세 데이터 생성"""
        
        detail_data = []
        salespeople = ['김철수', '이영희', '박민수', '정수진', '최동훈']
        products = ['제품A', '제품B', '제품C']
        
        for i in range(20):
            detail_data.append({
                'date': (datetime.now() - timedelta(days=i)).strftime('%Y-%m-%d'),
                'salesperson': np.random.choice(salespeople),
                'customer': f'고객{i+1}',
                'product': np.random.choice(products),
                'quantity': np.random.randint(1, 10),
                'unit_price': np.random.randint(10000, 50000),
                'amount': np.random.randint(50000, 500000),
                'status': np.random.choice(['완료', '진행중', '대기'])
            })
        
        return detail_data
    
    def _prepare_chart_data(self, user_data: Dict[str, Any] = None) -> Dict[str, Any]:
        """차트용 데이터 준비"""
        
        if user_data and 'chart_data' in user_data:
            return user_data['chart_data']
        
        return {
            'quarterly_revenue': [100000, 120000, 110000, 150000],
            'profitability_rates': [0.25, 0.15, 0.10],
            'cost_composition': [40000, 30000, 20000, 10000]
        }
    
    def _apply_style(self, cell, style_dict: Dict[str, Any]):
        """셀에 스타일 적용"""
        
        if 'font' in style_dict:
            cell.font = style_dict['font']
        if 'fill' in style_dict:
            cell.fill = style_dict['fill']
        if 'alignment' in style_dict:
            cell.alignment = style_dict['alignment']
        if 'number_format' in style_dict:
            cell.number_format = style_dict['number_format']
        if 'border' in style_dict:
            cell.border = style_dict['border']
    
    def _save_workbook(self, workbook: openpyxl.Workbook, filename_prefix: str) -> str:
        """워크북 저장"""
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{filename_prefix}_{timestamp}.xlsx"
        output_path = os.path.join(tempfile.gettempdir(), filename)
        
        workbook.save(output_path)
        logger.info(f"Excel 파일 저장 완료: {output_path}")
        
        return output_path


# 전역 템플릿 Excel 생성기 인스턴스
template_excel_generator = TemplateExcelGenerator()