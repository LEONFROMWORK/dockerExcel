#!/usr/bin/env python3
"""
Simple FastAPI server for Excel analysis - for testing purposes
"""
from fastapi import FastAPI, File, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
import tempfile
import os
import json
import openpyxl
import pandas as pd
import uvicorn
import cv2
import numpy as np
import pytesseract
from PIL import Image
import base64

app = FastAPI(title="Excel Analysis Service", version="1.0.0")

# Enable CORS for Rails frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/")
async def root():
    return {"message": "Excel Analysis Service is running"}

@app.get("/health")
async def health_check():
    return {"status": "healthy", "service": "excel-analysis"}

@app.post("/api/v1/excel/analyze")
async def analyze_excel_file(file: UploadFile = File(...)):
    """
    Simple Excel file analysis
    """
    try:
        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Invalid file type")
        
        # Save uploaded file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file.filename)[1]) as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_path = tmp_file.name
        
        try:
            # Analyze with openpyxl
            workbook = openpyxl.load_workbook(tmp_path, data_only=False)
            
            analysis_result = {
                "filename": file.filename,
                "sheets": {},
                "summary": {
                    "total_sheets": len(workbook.sheetnames),
                    "has_errors": False,
                    "total_errors": 0
                },
                "errors": [],
                "warnings": []
            }
            
            # Analyze each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Count formulas
                formula_count = 0
                errors = []
                
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.data_type == 'f':  # Formula cell
                            formula_count += 1
                            
                        # Check for error values
                        if isinstance(cell.value, str) and cell.value.startswith('#'):
                            errors.append({
                                "id": f"{sheet_name}_{cell.coordinate}",
                                "type": "formula_error",
                                "location": f"{sheet_name}!{cell.coordinate}",
                                "error_value": cell.value,
                                "description": f"Formula error in cell {cell.coordinate}: {cell.value}"
                            })
                
                analysis_result["sheets"][sheet_name] = {
                    "formula_count": formula_count,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "errors": errors
                }
                
                analysis_result["errors"].extend(errors)
            
            analysis_result["summary"]["total_errors"] = len(analysis_result["errors"])
            analysis_result["summary"]["has_errors"] = len(analysis_result["errors"]) > 0
            
            workbook.close()
            
            return {
                "status": "success",
                "message": "File analysis complete",
                "data": analysis_result
            }
            
        finally:
            # Clean up temporary file
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
                
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Analysis failed: {str(e)}")

@app.post("/api/v1/excel/fix")
async def fix_excel_error():
    """Simple fix endpoint - placeholder"""
    return {
        "success": True,
        "message": "Error fix simulated",
        "updated_data": {}
    }

@app.post("/api/v1/excel/fix-all")
async def fix_all_excel_errors():
    """Simple fix-all endpoint - placeholder"""
    return {
        "success": True,
        "fixed_count": 0,
        "remaining_errors": []
    }

@app.post("/api/v1/image-to-excel")
async def convert_image_to_excel(file: UploadFile = File(...)):
    """Convert image to Excel using OCR"""
    try:
        # Validate file type
        if not file.content_type.startswith('image/'):
            raise HTTPException(status_code=400, detail="File must be an image")
        
        # Save uploaded image temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp_file:
            content = await file.read()
            tmp_file.write(content)
            tmp_path = tmp_file.name
        
        try:
            # Load image using PIL
            image = Image.open(tmp_path)
            
            # Convert to RGB if necessary
            if image.mode != 'RGB':
                image = image.convert('RGB')
            
            # Use OCR to extract text
            try:
                # Try Korean + English OCR
                text = pytesseract.image_to_string(image, lang='kor+eng')
            except:
                # Fallback to English only
                text = pytesseract.image_to_string(image, lang='eng')
            
            # Split text into lines and create a basic table structure
            lines = [line.strip() for line in text.split('\n') if line.strip()]
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Extracted Data"
            
            # Process lines into table format
            for row_idx, line in enumerate(lines, 1):
                # Split by multiple spaces, tabs, or other delimiters
                cells = [cell.strip() for cell in line.replace('\t', ' ').split() if cell.strip()]
                for col_idx, cell_text in enumerate(cells, 1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_text)
            
            # Save to temporary Excel file
            excel_tmp_path = tmp_path.replace('.png', '.xlsx')
            wb.save(excel_tmp_path)
            
            # Read the created file for response
            with open(excel_tmp_path, 'rb') as excel_file:
                excel_content = excel_file.read()
            
            # Encode as base64 for JSON response
            excel_base64 = base64.b64encode(excel_content).decode('utf-8')
            
            return {
                "status": "success",
                "message": "Image converted to Excel successfully",
                "filename": file.filename.replace('.png', '.xlsx').replace('.jpg', '.xlsx').replace('.jpeg', '.xlsx'),
                "excel_file": excel_base64,
                "excel_path": excel_tmp_path,
                "extracted_text": text[:500],  # First 500 chars for preview
                "rows_extracted": len(lines)
            }
            
        finally:
            # Clean up temporary image file
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
                
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Conversion failed: {str(e)}")

if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000, reload=True)