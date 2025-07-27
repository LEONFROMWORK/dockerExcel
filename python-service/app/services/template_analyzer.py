"""
Real Excel template analysis service
Analyzes actual Excel template files to extract structure, formulas, and content
"""
import openpyxl
import pandas as pd
from typing import Dict, Any, List, Optional, Tuple
import re
import logging
from pathlib import Path
import json
from datetime import datetime
import os
import shutil

logger = logging.getLogger(__name__)


class TemplateAnalyzer:
    """Service for analyzing real Excel template files"""
    
    def __init__(self):
        self.template_storage_path = "app/templates/excel/files"
        self.metadata_storage_path = "app/templates/excel/real_templates"
        
        # Ensure storage directories exist
        os.makedirs(self.template_storage_path, exist_ok=True)
        os.makedirs(self.metadata_storage_path, exist_ok=True)
    
    async def analyze_template_file(self, file_path: str, template_name: str = None) -> Dict[str, Any]:
        """
        Analyze a real Excel template file and extract comprehensive metadata
        """
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"Template file not found: {file_path}")
            
            file_path = Path(file_path)
            template_name = template_name or file_path.stem
            
            # Load workbook for analysis
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            
            analysis_result = {
                "template_id": self._generate_template_id(template_name),
                "original_filename": file_path.name,
                "template_name": template_name,
                "file_size": file_path.stat().st_size,
                "analyzed_at": datetime.now().isoformat(),
                "sheets": {},
                "overall_structure": {},
                "i18n_elements": {},
                "formula_patterns": [],
                "chart_definitions": [],
                "conditional_formatting": [],
                "data_validation": [],
                "named_ranges": [],
                "vba_macros": None
            }
            
            # Analyze each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_analysis = await self._analyze_sheet(sheet)
                analysis_result["sheets"][sheet_name] = sheet_analysis
            
            # Extract overall structure
            analysis_result["overall_structure"] = self._extract_overall_structure(workbook)
            
            # Identify internationalization elements
            analysis_result["i18n_elements"] = self._identify_i18n_elements(workbook)
            
            # Extract formula patterns
            analysis_result["formula_patterns"] = self._extract_formula_patterns(workbook)
            
            # Analyze charts
            analysis_result["chart_definitions"] = self._analyze_charts(workbook)
            
            # Extract conditional formatting
            analysis_result["conditional_formatting"] = self._extract_conditional_formatting(workbook)
            
            # Extract data validation
            analysis_result["data_validation"] = self._extract_data_validation(workbook)
            
            # Extract named ranges
            analysis_result["named_ranges"] = self._extract_named_ranges(workbook)
            
            # Check for VBA macros
            analysis_result["vba_macros"] = self._check_vba_macros(workbook)
            
            workbook.close()
            
            # Save analysis result
            await self._save_analysis_result(analysis_result)
            
            return analysis_result
            
        except Exception as e:
            logger.error(f"Error analyzing template file: {str(e)}")
            raise
    
    async def _analyze_sheet(self, sheet) -> Dict[str, Any]:
        """Analyze individual sheet structure and content"""
        from openpyxl.utils import get_column_letter
        max_col_letter = get_column_letter(sheet.max_column) if sheet.max_column else 'A'
        
        analysis = {
            "name": sheet.title,
            "dimensions": f"{sheet.max_row}x{sheet.max_column}",
            "used_range": f"A1:{max_col_letter}{sheet.max_row}",
            "cell_analysis": {
                "total_cells": 0,
                "formula_cells": 0,
                "text_cells": 0,
                "number_cells": 0,
                "empty_cells": 0
            },
            "content_sections": [],
            "table_structures": [],
            "input_areas": [],
            "output_areas": [],
            "formatting_patterns": []
        }
        
        # Analyze cell content and structure
        content_map = {}
        for row in sheet.iter_rows(max_row=min(100, sheet.max_row)):
            for cell in row:
                if cell.value is not None:
                    analysis["cell_analysis"]["total_cells"] += 1
                    content_map[cell.coordinate] = {
                        "value": str(cell.value),
                        "data_type": self._get_cell_data_type(cell),
                        "has_formula": str(cell.value).startswith('=') if isinstance(cell.value, str) else False,
                        "formatting": self._extract_cell_formatting(cell)
                    }
                    
                    if content_map[cell.coordinate]["has_formula"]:
                        analysis["cell_analysis"]["formula_cells"] += 1
                    elif content_map[cell.coordinate]["data_type"] == "text":
                        analysis["cell_analysis"]["text_cells"] += 1
                    elif content_map[cell.coordinate]["data_type"] == "number":
                        analysis["cell_analysis"]["number_cells"] += 1
                else:
                    analysis["cell_analysis"]["empty_cells"] += 1
        
        # Identify content sections
        analysis["content_sections"] = self._identify_content_sections(content_map, sheet)
        
        # Identify table structures
        analysis["table_structures"] = self._identify_table_structures(content_map, sheet)
        
        # Identify input and output areas
        analysis["input_areas"], analysis["output_areas"] = self._identify_input_output_areas(content_map, sheet)
        
        return analysis
    
    def _get_cell_data_type(self, cell) -> str:
        """Determine the data type of a cell"""
        if cell.value is None:
            return "empty"
        elif isinstance(cell.value, str):
            if cell.value.startswith('='):
                return "formula"
            else:
                return "text"
        elif isinstance(cell.value, (int, float)):
            return "number"
        elif hasattr(cell.value, 'date'):
            return "date"
        else:
            return "other"
    
    def _extract_cell_formatting(self, cell) -> Dict[str, Any]:
        """Extract formatting information from a cell"""
        formatting = {}
        
        if cell.font:
            formatting["font"] = {
                "name": cell.font.name,
                "size": cell.font.size,
                "bold": cell.font.bold,
                "italic": cell.font.italic,
                "color": str(cell.font.color.rgb) if cell.font.color and cell.font.color.rgb else None
            }
        
        if cell.fill:
            formatting["fill"] = {
                "type": cell.fill.fill_type,
                "color": str(cell.fill.start_color.rgb) if cell.fill.start_color and cell.fill.start_color.rgb else None
            }
        
        if cell.border:
            formatting["border"] = {
                "top": bool(cell.border.top.style) if cell.border.top else False,
                "bottom": bool(cell.border.bottom.style) if cell.border.bottom else False,
                "left": bool(cell.border.left.style) if cell.border.left else False,
                "right": bool(cell.border.right.style) if cell.border.right else False
            }
        
        if cell.alignment:
            formatting["alignment"] = {
                "horizontal": cell.alignment.horizontal,
                "vertical": cell.alignment.vertical,
                "wrap_text": cell.alignment.wrap_text
            }
        
        if cell.number_format:
            formatting["number_format"] = cell.number_format
        
        return formatting
    
    def _identify_content_sections(self, content_map: Dict, sheet) -> List[Dict]:
        """Identify logical content sections in the sheet"""
        sections = []
        
        # Look for header-like cells (bold, larger font, etc.)
        headers = []
        for coord, cell_info in content_map.items():
            if (cell_info["formatting"].get("font", {}).get("bold") or 
                cell_info["formatting"].get("font", {}).get("size", 0) > 12):
                headers.append({
                    "coordinate": coord,
                    "text": cell_info["value"],
                    "type": "header"
                })
        
        # Group headers into sections
        for header in headers:
            section = {
                "title": header["text"],
                "start_cell": header["coordinate"],
                "type": "section",
                "estimated_range": self._estimate_section_range(header["coordinate"], sheet)
            }
            sections.append(section)
        
        return sections
    
    def _identify_table_structures(self, content_map: Dict, sheet) -> List[Dict]:
        """Identify table-like structures in the sheet"""
        tables = []
        
        # Look for patterns that suggest tables (consecutive cells, headers, data)
        # This is a simplified implementation
        
        return tables
    
    def _identify_input_output_areas(self, content_map: Dict, sheet) -> Tuple[List[Dict], List[Dict]]:
        """Identify areas likely to be user inputs and calculated outputs"""
        input_areas = []
        output_areas = []
        
        for coord, cell_info in content_map.items():
            if cell_info["has_formula"]:
                output_areas.append({
                    "coordinate": coord,
                    "formula": cell_info["value"],
                    "type": "calculated_output"
                })
            elif (cell_info["data_type"] in ["number", "text"] and 
                  not cell_info["formatting"].get("font", {}).get("bold")):
                input_areas.append({
                    "coordinate": coord,
                    "data_type": cell_info["data_type"],
                    "type": "potential_input"
                })
        
        return input_areas, output_areas
    
    def _estimate_section_range(self, start_coord: str, sheet) -> str:
        """Estimate the range covered by a section"""
        # Simple implementation - could be more sophisticated
        cell = sheet[start_coord]
        return f"{start_coord}:{chr(ord(start_coord[0]) + 5)}{cell.row + 10}"
    
    def _extract_overall_structure(self, workbook) -> Dict[str, Any]:
        """Extract overall workbook structure information"""
        return {
            "total_sheets": len(workbook.sheetnames),
            "sheet_names": workbook.sheetnames,
            "has_charts": any(len(sheet._charts) > 0 for sheet in workbook.worksheets),
            "has_images": any(len(sheet._images) > 0 for sheet in workbook.worksheets),
            "has_pivot_tables": self._check_pivot_tables(workbook),
            "complexity_score": self._calculate_complexity_score(workbook)
        }
    
    def _identify_i18n_elements(self, workbook) -> Dict[str, List[str]]:
        """Identify text elements that need internationalization"""
        i18n_elements = {
            "headers": [],
            "labels": [],
            "instructions": [],
            "validation_messages": [],
            "chart_titles": []
        }
        
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(max_row=50):  # Focus on header area
                for cell in row:
                    if isinstance(cell.value, str) and len(cell.value.strip()) > 2:
                        text = cell.value.strip()
                        
                        # Classify text elements
                        if self._is_header_text(cell, text):
                            i18n_elements["headers"].append(text)
                        elif self._is_label_text(cell, text):
                            i18n_elements["labels"].append(text)
                        elif self._is_instruction_text(text):
                            i18n_elements["instructions"].append(text)
        
        # Remove duplicates
        for key in i18n_elements:
            i18n_elements[key] = list(set(i18n_elements[key]))
        
        return i18n_elements
    
    def _is_header_text(self, cell, text: str) -> bool:
        """Check if text appears to be a header"""
        if cell.font and cell.font.bold:
            return True
        if cell.font and cell.font.size and cell.font.size > 12:
            return True
        return False
    
    def _is_label_text(self, cell, text: str) -> bool:
        """Check if text appears to be a label"""
        return ":" in text or text.endswith(":")
    
    def _is_instruction_text(self, text: str) -> bool:
        """Check if text appears to be an instruction"""
        instruction_keywords = ["enter", "input", "select", "choose", "fill", "complete"]
        return any(keyword in text.lower() for keyword in instruction_keywords)
    
    def _extract_formula_patterns(self, workbook) -> List[Dict]:
        """Extract and categorize formula patterns"""
        patterns = []
        
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        pattern = {
                            "sheet": sheet.title,
                            "coordinate": cell.coordinate,
                            "formula": cell.value,
                            "category": self._categorize_formula(cell.value),
                            "references": self._extract_cell_references(cell.value)
                        }
                        patterns.append(pattern)
        
        return patterns
    
    def _categorize_formula(self, formula: str) -> str:
        """Categorize formula by type"""
        formula_upper = formula.upper()
        
        if any(func in formula_upper for func in ['SUM', 'SUMIF', 'SUMIFS']):
            return "aggregation"
        elif any(func in formula_upper for func in ['VLOOKUP', 'HLOOKUP', 'INDEX', 'MATCH']):
            return "lookup"
        elif any(func in formula_upper for func in ['IF', 'AND', 'OR', 'NOT']):
            return "logical"
        elif any(func in formula_upper for func in ['DATE', 'TODAY', 'NOW', 'YEAR', 'MONTH']):
            return "date_time"
        elif any(func in formula_upper for func in ['PMT', 'FV', 'PV', 'NPV', 'IRR']):
            return "financial"
        else:
            return "other"
    
    def _extract_cell_references(self, formula: str) -> List[str]:
        """Extract cell references from formula"""
        # Pattern to match cell references like A1, $A$1, Sheet1!A1
        cell_pattern = r'(?:[A-Za-z0-9_\s]+!)?(?:\$?[A-Z]+\$?[0-9]+)'
        references = re.findall(cell_pattern, formula)
        return list(set(references))
    
    def _analyze_charts(self, workbook) -> List[Dict]:
        """Analyze chart definitions in the workbook"""
        charts = []
        
        for sheet in workbook.worksheets:
            for chart in sheet._charts:
                chart_info = {
                    "sheet": sheet.title,
                    "type": type(chart).__name__,
                    "title": getattr(chart, 'title', None),
                    "data_series": len(getattr(chart, 'series', [])),
                    "position": {
                        "anchor": str(chart.anchor) if hasattr(chart, 'anchor') else None
                    }
                }
                charts.append(chart_info)
        
        return charts
    
    def _extract_conditional_formatting(self, workbook) -> List[Dict]:
        """Extract conditional formatting rules"""
        cf_rules = []
        
        for sheet in workbook.worksheets:
            if hasattr(sheet, 'conditional_formatting'):
                for cf in sheet.conditional_formatting:
                    rule_info = {
                        "sheet": sheet.title,
                        "range": str(cf.sqref) if hasattr(cf, 'sqref') else None,
                        "type": type(cf).__name__,
                        "rules_count": len(cf.rules) if hasattr(cf, 'rules') else 0
                    }
                    cf_rules.append(rule_info)
        
        return cf_rules
    
    def _extract_data_validation(self, workbook) -> List[Dict]:
        """Extract data validation rules"""
        validations = []
        
        for sheet in workbook.worksheets:
            if hasattr(sheet, 'data_validations'):
                for dv in sheet.data_validations.dataValidation:
                    validation_info = {
                        "sheet": sheet.title,
                        "range": str(dv.sqref) if hasattr(dv, 'sqref') else None,
                        "type": dv.type if hasattr(dv, 'type') else None,
                        "formula1": dv.formula1 if hasattr(dv, 'formula1') else None,
                        "error_message": dv.errorTitle if hasattr(dv, 'errorTitle') else None
                    }
                    validations.append(validation_info)
        
        return validations
    
    def _extract_named_ranges(self, workbook) -> List[Dict]:
        """Extract named ranges definition"""
        named_ranges = []
        
        try:
            for name in workbook.defined_names:
                range_info = {
                    "name": name.name if hasattr(name, 'name') else str(name),
                    "refers_to": getattr(name, 'attr_text', str(name)),
                    "scope": "workbook"  # Could be sheet-specific
                }
                named_ranges.append(range_info)
        except Exception as e:
            logger.warning(f"Could not extract named ranges: {e}")
        
        return named_ranges
    
    def _check_vba_macros(self, workbook) -> Optional[Dict]:
        """Check for VBA macros in the workbook"""
        try:
            # Basic check for macro-enabled workbook
            if hasattr(workbook, 'vba_archive') and workbook.vba_archive:
                return {
                    "has_macros": True,
                    "modules_count": "unknown",  # Would need more detailed analysis
                    "security_note": "Macro analysis requires additional tools"
                }
            return {"has_macros": False}
        except:
            return {"has_macros": False, "analysis_failed": True}
    
    def _check_pivot_tables(self, workbook) -> bool:
        """Check if workbook contains pivot tables"""
        for sheet in workbook.worksheets:
            if hasattr(sheet, '_pivots') and len(sheet._pivots) > 0:
                return True
        return False
    
    def _calculate_complexity_score(self, workbook) -> int:
        """Calculate overall complexity score (1-10)"""
        score = 1
        
        # Add points for various features
        sheet_count = len(workbook.sheetnames)
        if sheet_count > 1:
            score += min(sheet_count - 1, 3)
        
        # Check for charts
        has_charts = any(len(sheet._charts) > 0 for sheet in workbook.worksheets)
        if has_charts:
            score += 2
        
        # Check for pivot tables
        if self._check_pivot_tables(workbook):
            score += 2
        
        # Check for complex formulas
        complex_formula_count = 0
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(max_row=100):
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        if len(cell.value) > 50:  # Complex formula heuristic
                            complex_formula_count += 1
        
        if complex_formula_count > 10:
            score += 2
        elif complex_formula_count > 5:
            score += 1
        
        return min(score, 10)
    
    def _generate_template_id(self, template_name: str) -> str:
        """Generate a unique template ID"""
        import re
        # Convert to lowercase and replace spaces/special chars with underscores
        clean_name = re.sub(r'[^a-zA-Z0-9]+', '_', template_name.lower())
        timestamp = datetime.now().strftime('%Y%m%d')
        return f"{clean_name}_{timestamp}"
    
    async def _save_analysis_result(self, analysis_result: Dict[str, Any]):
        """Save analysis result to metadata file"""
        template_id = analysis_result["template_id"]
        metadata_file = os.path.join(self.metadata_storage_path, f"{template_id}.json")
        
        with open(metadata_file, 'w', encoding='utf-8') as f:
            json.dump(analysis_result, f, ensure_ascii=False, indent=2)
        
        logger.info(f"Saved template analysis: {metadata_file}")
    
    async def copy_template_to_storage(self, source_path: str, template_id: str) -> str:
        """Copy template file to storage location"""
        source_path = Path(source_path)
        dest_path = Path(self.template_storage_path) / f"{template_id}{source_path.suffix}"
        
        shutil.copy2(source_path, dest_path)
        logger.info(f"Copied template to storage: {dest_path}")
        
        return str(dest_path)
    
    async def get_template_metadata(self, template_id: str) -> Optional[Dict[str, Any]]:
        """Retrieve template metadata"""
        metadata_file = os.path.join(self.metadata_storage_path, f"{template_id}.json")
        
        if os.path.exists(metadata_file):
            with open(metadata_file, 'r', encoding='utf-8') as f:
                return json.load(f)
        
        return None
    
    async def list_available_templates(self) -> List[Dict[str, Any]]:
        """List all available templates with metadata"""
        templates = []
        
        metadata_files = Path(self.metadata_storage_path).glob("*.json")
        for metadata_file in metadata_files:
            try:
                with open(metadata_file, 'r', encoding='utf-8') as f:
                    metadata = json.load(f)
                    templates.append({
                        "template_id": metadata["template_id"],
                        "template_name": metadata["template_name"],
                        "complexity_score": metadata["overall_structure"]["complexity_score"],
                        "sheet_count": metadata["overall_structure"]["total_sheets"],
                        "analyzed_at": metadata["analyzed_at"],
                        "file_size": metadata["file_size"]
                    })
            except Exception as e:
                logger.error(f"Error reading metadata file {metadata_file}: {e}")
        
        return sorted(templates, key=lambda x: x["analyzed_at"], reverse=True)


# Create singleton instance
template_analyzer = TemplateAnalyzer()