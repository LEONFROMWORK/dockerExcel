"""
Template bridge for integrating AI generation with existing template system
Enables hybrid approach combining templates with AI enhancement
"""

import os
import logging
from typing import Dict, List, Any, Optional
import openpyxl

from ...services.template_selection_service import template_selection_service
from ...services.excel_analyzer import excel_analyzer
from ..structure.excel_schema import (
    ExcelStructure,
    SheetSchema,
    ColumnDefinition,
    DataType,
)

logger = logging.getLogger(__name__)


class TemplateBridge:
    """Bridges AI generation with existing template system"""

    def __init__(self):
        self.template_service = template_selection_service
        self.template_cache = {}

    async def find_relevant_template(
        self, request: str, context: Dict[str, Any]
    ) -> Optional[Dict[str, Any]]:
        """Find the most relevant template for the request"""

        try:
            # Use template selection service
            query = context.get("original_text", request)
            language = context.get("language", "ko")

            recommendations = await self.template_service.get_smart_recommendations(
                query=query, language=language, top_k=5
            )

            if recommendations and recommendations[0]["score"] > 0.7:
                # High confidence match
                best_match = recommendations[0]
                template_info = {
                    "template_id": best_match["id"],
                    "category": best_match["category"],
                    "confidence": best_match["score"],
                    "path": best_match.get("file_path"),
                    "metadata": best_match.get("metadata", {}),
                }

                # Load template structure
                template_structure = await self._load_template_structure(template_info)
                template_info["structure"] = template_structure

                return template_info

            return None

        except Exception as e:
            logger.error(f"Template search failed: {str(e)}")
            return None

    async def enhance_template_with_ai(
        self,
        template: Dict[str, Any],
        context: Dict[str, Any],
        analysis: Dict[str, Any],
    ) -> ExcelStructure:
        """Enhance template structure with AI insights"""

        # Extract template structure
        template_structure = template.get("structure", {})

        # Convert to ExcelStructure
        base_structure = self._convert_template_to_schema(template_structure)

        # AI enhancements based on context
        enhanced_structure = await self._apply_ai_enhancements(
            base_structure, context, analysis
        )

        # Add custom columns based on requirements
        enhanced_structure = self._add_context_specific_columns(
            enhanced_structure, context
        )

        # Optimize structure
        enhanced_structure = self._optimize_structure(enhanced_structure, analysis)

        return enhanced_structure

    async def extract_template_patterns(self, template_path: str) -> Dict[str, Any]:
        """Extract patterns and best practices from template"""

        try:
            # Analyze template file
            analysis = await excel_analyzer.analyze_file(template_path)

            patterns = {
                "structure_patterns": self._extract_structure_patterns(analysis),
                "formula_patterns": self._extract_formula_patterns(analysis),
                "style_patterns": self._extract_style_patterns(analysis),
                "data_patterns": self._extract_data_patterns(analysis),
            }

            return patterns

        except Exception as e:
            logger.error(f"Pattern extraction failed: {str(e)}")
            return {}

    def apply_template_patterns(
        self, generated_structure: ExcelStructure, patterns: Dict[str, Any]
    ) -> ExcelStructure:
        """Apply extracted patterns to generated structure"""

        # Apply structure patterns
        if "structure_patterns" in patterns:
            generated_structure = self._apply_structure_patterns(
                generated_structure, patterns["structure_patterns"]
            )

        # Apply formula patterns
        if "formula_patterns" in patterns:
            generated_structure = self._apply_formula_patterns(
                generated_structure, patterns["formula_patterns"]
            )

        return generated_structure

    async def _load_template_structure(
        self, template_info: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Load template structure from file"""

        template_path = template_info.get("path")
        if not template_path or not os.path.exists(template_path):
            return {}

        try:
            # Cache check
            if template_path in self.template_cache:
                return self.template_cache[template_path]

            # Load and analyze template
            wb = openpyxl.load_workbook(template_path, read_only=True, data_only=False)

            structure = {"sheets": [], "metadata": template_info.get("metadata", {})}

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_structure = {
                    "name": sheet_name,
                    "columns": self._extract_columns_from_sheet(ws),
                    "row_count": ws.max_row,
                    "has_formulas": self._sheet_has_formulas(ws),
                    "has_charts": (
                        len(ws._charts) > 0 if hasattr(ws, "_charts") else False
                    ),
                }
                structure["sheets"].append(sheet_structure)

            wb.close()

            # Cache the structure
            self.template_cache[template_path] = structure

            return structure

        except Exception as e:
            logger.error(f"Failed to load template structure: {str(e)}")
            return {}

    def _convert_template_to_schema(
        self, template_structure: Dict[str, Any]
    ) -> ExcelStructure:
        """Convert template structure to ExcelStructure schema"""

        sheets = []

        for sheet_info in template_structure.get("sheets", []):
            columns = []

            for col_info in sheet_info.get("columns", []):
                column = ColumnDefinition(
                    name=col_info.get("name", "Column"),
                    data_type=self._infer_data_type(col_info),
                    width=col_info.get("width", 15),
                )
                columns.append(column)

            sheet = SheetSchema(
                name=sheet_info.get("name", "Sheet"),
                columns=columns,
                row_count=sheet_info.get("row_count", 100),
                has_header=True,
            )
            sheets.append(sheet)

        structure = ExcelStructure(
            sheets=sheets, metadata=template_structure.get("metadata", {})
        )

        return structure

    async def _apply_ai_enhancements(
        self,
        structure: ExcelStructure,
        context: Dict[str, Any],
        analysis: Dict[str, Any],
    ) -> ExcelStructure:
        """Apply AI-driven enhancements to structure"""

        # Enhance based on intent
        intent = analysis.get("intent", {}).get("primary")

        if intent == "analyze_data":
            # Add analysis-specific sheets
            structure = self._add_analysis_sheets(structure)
        elif intent == "track_metrics":
            # Add dashboard and metrics sheets
            structure = self._add_metrics_sheets(structure)
        elif intent == "plan_forecast":
            # Add forecasting columns and sheets
            structure = self._add_forecast_elements(structure)

        return structure

    def _add_context_specific_columns(
        self, structure: ExcelStructure, context: Dict[str, Any]
    ) -> ExcelStructure:
        """Add columns based on specific context requirements"""

        requirements = context.get("requirements", [])

        for sheet in structure.sheets:
            # Add comparison columns if needed
            if "comparison" in requirements:
                sheet.columns.extend(
                    [
                        ColumnDefinition(
                            name="전기대비", data_type=DataType.PERCENTAGE
                        ),
                        ColumnDefinition(
                            name="목표대비", data_type=DataType.PERCENTAGE
                        ),
                    ]
                )

            # Add trend columns if needed
            if "forecast" in requirements:
                sheet.columns.extend(
                    [
                        ColumnDefinition(name="예측값", data_type=DataType.NUMBER),
                        ColumnDefinition(name="신뢰구간", data_type=DataType.TEXT),
                    ]
                )

        return structure

    def _optimize_structure(
        self, structure: ExcelStructure, analysis: Dict[str, Any]
    ) -> ExcelStructure:
        """Optimize structure based on analysis"""

        complexity = analysis.get("complexity")

        if complexity == "simple":
            # Simplify structure
            for sheet in structure.sheets:
                # Remove non-essential columns
                essential_types = [DataType.TEXT, DataType.NUMBER, DataType.CURRENCY]
                sheet.columns = [
                    col for col in sheet.columns if col.data_type in essential_types
                ][
                    :10
                ]  # Limit columns

        elif complexity == "expert":
            # Add advanced features
            for sheet in structure.sheets:
                sheet.freeze_panes = "B2"
                # Enable advanced features
                if not hasattr(sheet, "conditional_formats"):
                    sheet.conditional_formats = []

        return structure

    def _extract_columns_from_sheet(self, worksheet) -> List[Dict[str, Any]]:
        """Extract column information from worksheet"""

        columns = []

        # Assume first row contains headers
        for col_idx in range(1, worksheet.max_column + 1):
            header_cell = worksheet.cell(row=1, column=col_idx)
            if header_cell.value:
                col_info = {
                    "name": str(header_cell.value),
                    "width": worksheet.column_dimensions[
                        openpyxl.utils.get_column_letter(col_idx)
                    ].width,
                    "sample_values": self._get_column_samples(worksheet, col_idx),
                }
                columns.append(col_info)

        return columns

    def _get_column_samples(
        self, worksheet, col_idx: int, sample_size: int = 5
    ) -> List[Any]:
        """Get sample values from column"""

        samples = []
        for row_idx in range(2, min(worksheet.max_row + 1, sample_size + 2)):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                samples.append(cell.value)

        return samples

    def _sheet_has_formulas(self, worksheet) -> bool:
        """Check if sheet contains formulas"""

        for row in worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row, 100)):
            for cell in row:
                if (
                    cell.value
                    and isinstance(cell.value, str)
                    and cell.value.startswith("=")
                ):
                    return True
        return False

    def _infer_data_type(self, col_info: Dict[str, Any]) -> DataType:
        """Infer data type from column information"""

        samples = col_info.get("sample_values", [])
        col_name = col_info.get("name", "").lower()

        # Check column name patterns
        if any(keyword in col_name for keyword in ["금액", "액", "가격", "비용"]):
            return DataType.CURRENCY
        elif any(keyword in col_name for keyword in ["율", "률", "%"]):
            return DataType.PERCENTAGE
        elif any(keyword in col_name for keyword in ["일자", "날짜", "date"]):
            return DataType.DATE

        # Check sample values
        if samples:
            # Check if all numeric
            try:
                numeric_samples = [float(s) for s in samples if s is not None]
                if len(numeric_samples) == len([s for s in samples if s is not None]):
                    return DataType.NUMBER
            except (ValueError, TypeError):
                pass

        return DataType.TEXT

    def _extract_structure_patterns(self, analysis: Dict[str, Any]) -> Dict[str, Any]:
        """Extract structural patterns from template"""

        return {
            "sheet_organization": analysis.get("sheets", {}).get(
                "organization", "standard"
            ),
            "naming_conventions": analysis.get("naming_patterns", {}),
            "layout_style": analysis.get("layout", "tabular"),
        }

    def _extract_formula_patterns(self, analysis: Dict[str, Any]) -> Dict[str, Any]:
        """Extract formula patterns from template"""

        formulas = analysis.get("formulas", {})
        return {
            "common_functions": formulas.get("functions", []),
            "calculation_patterns": formulas.get("patterns", []),
            "dependencies": formulas.get("dependencies", {}),
        }

    def _extract_style_patterns(self, analysis: Dict[str, Any]) -> Dict[str, Any]:
        """Extract styling patterns from template"""

        return {
            "color_scheme": analysis.get("styles", {}).get("colors", {}),
            "font_preferences": analysis.get("styles", {}).get("fonts", {}),
            "formatting_rules": analysis.get("styles", {}).get("formatting", {}),
        }

    def _extract_data_patterns(self, analysis: Dict[str, Any]) -> Dict[str, Any]:
        """Extract data patterns from template"""

        return {
            "value_ranges": analysis.get("data", {}).get("ranges", {}),
            "data_types": analysis.get("data", {}).get("types", {}),
            "relationships": analysis.get("data", {}).get("relationships", []),
        }

    def _apply_structure_patterns(
        self, structure: ExcelStructure, patterns: Dict[str, Any]
    ) -> ExcelStructure:
        """Apply structure patterns to generated structure"""

        # Apply naming conventions
        naming = patterns.get("naming_conventions", {})
        if naming:
            for sheet in structure.sheets:
                # Apply sheet naming pattern
                sheet.name = self._apply_naming_pattern(sheet.name, naming)

        return structure

    def _apply_formula_patterns(
        self, structure: ExcelStructure, patterns: Dict[str, Any]
    ) -> ExcelStructure:
        """Apply formula patterns to generated structure"""

        # This would add formulas based on patterns
        # Implementation depends on specific requirements

        return structure

    def _apply_naming_pattern(self, name: str, pattern: Dict[str, Any]) -> str:
        """Apply naming pattern to string"""

        # Simple implementation - can be enhanced
        if pattern.get("case") == "title":
            return name.title()
        elif pattern.get("case") == "upper":
            return name.upper()

        return name

    def _add_analysis_sheets(self, structure: ExcelStructure) -> ExcelStructure:
        """Add analysis-specific sheets"""

        # Add summary analysis sheet
        analysis_sheet = SheetSchema(
            name="분석요약",
            columns=[
                ColumnDefinition(name="지표", data_type=DataType.TEXT),
                ColumnDefinition(name="현재값", data_type=DataType.NUMBER),
                ColumnDefinition(name="이전값", data_type=DataType.NUMBER),
                ColumnDefinition(name="변화율", data_type=DataType.PERCENTAGE),
                ColumnDefinition(name="평가", data_type=DataType.TEXT),
            ],
        )
        structure.add_sheet(analysis_sheet)

        return structure

    def _add_metrics_sheets(self, structure: ExcelStructure) -> ExcelStructure:
        """Add metrics tracking sheets"""

        # Add KPI dashboard
        kpi_sheet = SheetSchema(
            name="KPI대시보드",
            columns=[
                ColumnDefinition(name="KPI명", data_type=DataType.TEXT),
                ColumnDefinition(name="목표", data_type=DataType.NUMBER),
                ColumnDefinition(name="실적", data_type=DataType.NUMBER),
                ColumnDefinition(name="달성률", data_type=DataType.PERCENTAGE),
                ColumnDefinition(name="상태", data_type=DataType.TEXT),
            ],
        )
        structure.add_sheet(kpi_sheet)

        return structure

    def _add_forecast_elements(self, structure: ExcelStructure) -> ExcelStructure:
        """Add forecasting elements to structure"""

        # Add forecast columns to relevant sheets
        for sheet in structure.sheets:
            if any(
                col.data_type in [DataType.NUMBER, DataType.CURRENCY]
                for col in sheet.columns
            ):
                sheet.columns.extend(
                    [
                        ColumnDefinition(name="예측_다음달", data_type=DataType.NUMBER),
                        ColumnDefinition(name="예측_3개월", data_type=DataType.NUMBER),
                        ColumnDefinition(
                            name="예측_신뢰도", data_type=DataType.PERCENTAGE
                        ),
                    ]
                )

        return structure
