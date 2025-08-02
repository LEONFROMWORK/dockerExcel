#!/usr/bin/env python3
"""
사용자 Excel 파일 디버깅
손익실적.xlsx 파일 분석
"""
import json
from app.services.excel_to_xspreadsheet import excel_to_xspreadsheet_converter

def debug_user_excel():
    print("🔍 사용자 Excel 파일 디버깅: 손익실적.xlsx")
    
    file_path = "/Users/kevin/Downloads/손익실적.xlsx"
    
    try:
        # 파일 변환 시도
        result = excel_to_xspreadsheet_converter.convert_file(file_path)
        
        if result["success"]:
            print("✅ 변환 성공")
            print(f"시트 수: {result['sheet_count']}")
            print(f"시트 이름: {result['sheet_names']}")
            
            if result["data"]:
                for i, sheet_data in enumerate(result["data"]):
                    print(f"\n=== 시트 {i+1}: '{sheet_data['name']}' ===")
                    
                    # 기본 구조 확인
                    print(f"freeze: '{sheet_data.get('freeze', '')}'")
                    print(f"styles 수: {len(sheet_data.get('styles', []))}")
                    print(f"merges 수: {len(sheet_data.get('merges', []))}")
                    
                    # rows 구조 확인
                    rows = sheet_data.get('rows', {})
                    print(f"rows len: {rows.get('len', 'None')}")
                    
                    # 실제 셀 데이터 개수 확인
                    cell_count = 0
                    data_rows = 0
                    for key, value in rows.items():
                        if key not in ['len', 'height'] and isinstance(value, dict) and 'cells' in value:
                            data_rows += 1
                            cells = value['cells']
                            cell_count += len(cells)
                            
                            # 처음 3개 행의 데이터 출력
                            if data_rows <= 3:
                                print(f"  행 {key}: {len(cells)}개 셀")
                                for cell_key, cell_data in list(cells.items())[:5]:  # 첫 5개 셀만
                                    text = cell_data.get('text', '')
                                    if text:
                                        print(f"    셀 [{key},{cell_key}]: '{text[:30]}{'...' if len(text) > 30 else ''}'")
                    
                    print(f"총 데이터 행: {data_rows}개")
                    print(f"총 셀 데이터: {cell_count}개")
                    
                    # cols 구조 확인
                    cols = sheet_data.get('cols', {})
                    print(f"cols len: {cols.get('len', 'None')}")
                    
                    # 문제 진단
                    if cell_count == 0:
                        print("❌ 문제: 셀 데이터가 없습니다!")
                    elif data_rows == 0:
                        print("❌ 문제: 데이터 행이 없습니다!")
                    else:
                        print("✅ 데이터 구조 정상")
                    
                    # 첫 번째 시트만 JSON으로 저장
                    if i == 0:
                        with open("user_file_debug.json", "w", encoding="utf-8") as f:
                            json.dump(sheet_data, f, ensure_ascii=False, indent=2)
                        print(f"📄 첫 번째 시트 데이터가 user_file_debug.json에 저장되었습니다")
            
            else:
                print("❌ 변환된 데이터가 없습니다")
        
        else:
            print(f"❌ 변환 실패: {result['error']}")
    
    except FileNotFoundError:
        print("❌ 파일을 찾을 수 없습니다")
        print("파일 경로를 확인해주세요: /Users/kevin/Downloads/손익실적.xlsx")
    except Exception as e:
        print(f"❌ 예상치 못한 오류: {str(e)}")
        import traceback
        traceback.print_exc()

def check_openpyxl_direct():
    """OpenPyXL로 직접 파일 읽기 테스트"""
    print("\n🔧 OpenPyXL 직접 읽기 테스트")
    
    file_path = "/Users/kevin/Downloads/손익실적.xlsx"
    
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(file_path, data_only=False)
        print(f"✅ 파일 로드 성공")
        print(f"시트 이름: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            print(f"\n--- 시트 '{sheet_name}' ---")
            print(f"최대 행: {sheet.max_row}")
            print(f"최대 열: {sheet.max_column}")
            
            # 처음 5x5 영역의 데이터 확인
            print("처음 5x5 영역:")
            for row in range(1, min(6, sheet.max_row + 1)):
                row_data = []
                for col in range(1, min(6, sheet.max_column + 1)):
                    cell = sheet.cell(row=row, column=col)
                    value = str(cell.value) if cell.value is not None else ""
                    row_data.append(value[:10] + "..." if len(value) > 10 else value)
                print(f"  행 {row}: {row_data}")
            
            # 실제 데이터가 있는 셀 개수 확인
            data_cells = 0
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        data_cells += 1
            
            print(f"데이터가 있는 셀 수: {data_cells}개")
    
    except Exception as e:
        print(f"❌ OpenPyXL 직접 읽기 실패: {str(e)}")

if __name__ == "__main__":
    debug_user_excel()
    check_openpyxl_direct()