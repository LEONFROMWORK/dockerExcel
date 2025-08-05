"""
Test Rails job execution
"""
import requests
import time
import json
from openpyxl import Workbook
import tempfile
import os

# Test configuration
RAILS_BASE_URL = "http://localhost:3000"

def create_test_excel():
    """Create test Excel file with errors"""
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"

    # Add test data
    ws['A1'] = "Test Data"
    ws['B1'] = 10
    ws['C1'] = 0

    # Add various error formulas
    ws['A2'] = "=B1/C1"          # DIV/0 error
    ws['A3'] = "=VLOOKUP(\"X\",E1:F10,2,FALSE)"  # #N/A error
    ws['A4'] = "=UNKNOWNFUNC()"  # #NAME? error
    ws['A5'] = "=Sheet2!A1"      # #REF! error (no Sheet2)
    ws['A6'] = "=A1+B1"          # #VALUE! error (text + number)
    ws['A7'] = "=A8"             # Circular reference
    ws['A8'] = "=A7"             # Circular reference

    # Save to temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        return tmp.name

def test_rails_job():
    # Create test file
    test_file = create_test_excel()
    print(f"Created test file: {test_file}")

    try:
        # 1. Upload file
        print("\n=== 1. File Upload ===")
        url = f"{RAILS_BASE_URL}/api/v1/excel_analysis/files"

        with open(test_file, 'rb') as f:
            files = {'file': ('test_errors.xlsx', f, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            response = requests.post(url, files=files)

        if response.status_code != 200:
            print(f"Upload failed: {response.status_code}")
            print(response.text)
            return

        result = response.json()
        file_id = result.get('id')
        print(f"✅ File uploaded: ID={file_id}")

        # 2. Start analysis
        print("\n=== 2. Start Analysis ===")
        analyze_url = f"{RAILS_BASE_URL}/api/v1/excel_analysis/analyze"

        response = requests.post(analyze_url, json={
            'file_id': file_id,
            'analysis_type': 'parallel',
            'include_all_tiers': True
        })

        if response.status_code != 200:
            print(f"Analysis start failed: {response.status_code}")
            print(response.text)
            return

        result = response.json()
        session_id = result.get('session_id')
        print(f"✅ Analysis started: session_id={session_id}")
        print(f"Initial errors: {len(result.get('errors', []))}")
        print(f"Status: {result.get('status')}")
        print(f"Message: {result.get('message')}")

        # 3. Check analysis status
        print("\n=== 3. Checking Analysis Status ===")
        status_url = f"{RAILS_BASE_URL}/api/v1/excel_analysis/status/{session_id}"

        # Poll for completion
        for i in range(10):  # Max 10 seconds
            time.sleep(1)
            response = requests.get(status_url)

            if response.status_code == 200:
                status_result = response.json()
                print(f"Attempt {i+1}: Status={status_result.get('status')}, Errors={status_result.get('total_errors', 0)}")

                if status_result.get('analysis_complete'):
                    print("\n✅ Analysis complete!")
                    print(f"Total errors: {status_result.get('total_errors')}")

                    # Get detailed results
                    if 'errors' in status_result:
                        for idx, error in enumerate(status_result['errors'][:5]):  # Show first 5
                            print(f"\nError {idx+1}:")
                            print(f"  Type: {error.get('type')}")
                            print(f"  Cell: {error.get('sheet')}!{error.get('cell')}")
                            print(f"  Message: {error.get('message')}")
                    break
            else:
                print(f"Status check failed: {response.status_code}")

    finally:
        # Clean up
        if os.path.exists(test_file):
            os.unlink(test_file)
            print(f"\nDeleted test file: {test_file}")

if __name__ == "__main__":
    test_rails_job()
