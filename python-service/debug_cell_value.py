#!/usr/bin/env python3
"""
Debug cell value reading issue
"""
import openpyxl


def debug_cell_values():
    print("=== DEBUGGING CELL VALUES ===")

    # Load with different options
    print("\n1. Loading with data_only=True:")
    wb1 = openpyxl.load_workbook("pic.xlsx", data_only=True)
    ws1 = wb1.active
    for row in range(1, 11):
        for col in range(1, 6):
            cell = ws1.cell(row=row, column=col)
            if cell.value is not None:
                print(
                    f"  [{row},{col}] = {cell.value} (type: {type(cell.value).__name__})"
                )
    wb1.close()

    print("\n2. Loading with data_only=False:")
    wb2 = openpyxl.load_workbook("pic.xlsx", data_only=False)
    ws2 = wb2.active
    for row in range(1, 11):
        for col in range(1, 6):
            cell = ws2.cell(row=row, column=col)
            if cell.value is not None:
                print(
                    f"  [{row},{col}] = {cell.value} (type: {type(cell.value).__name__})"
                )
    wb2.close()

    print("\n3. Checking cell attributes:")
    wb3 = openpyxl.load_workbook("pic.xlsx", data_only=False)
    ws3 = wb3.active

    # Check a specific cell with known content
    test_cell = ws3.cell(row=4, column=1)  # "(1) 손익 실적"
    print("\nCell A4 attributes:")
    print(f"  value: {test_cell.value}")
    print(f"  data_type: {test_cell.data_type}")
    print(f"  is_date: {test_cell.is_date}")
    print(
        f"  internal_value: {test_cell._value if hasattr(test_cell, '_value') else 'N/A'}"
    )

    # Check cell styles
    print(f"  has_style: {test_cell.has_style}")
    if test_cell.has_style:
        print(f"  font: {test_cell.font}")
        print(f"  fill: {test_cell.fill}")
        print(f"  border: {test_cell.border}")

    wb3.close()


if __name__ == "__main__":
    debug_cell_values()
