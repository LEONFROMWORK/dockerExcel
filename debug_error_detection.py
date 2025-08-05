"""
Debug error detection issue
"""
import asyncio
import sys
import os
from openpyxl import load_workbook

# Add project to path
sys.path.insert(0, '/Users/kevin/excel-unified/python-service')

from app.services.detection.integrated_error_detector import IntegratedErrorDetector
from app.core.interfaces import DummyProgressReporter

async def debug_detection():
    # Create a test Excel file with known errors
    from openpyxl import Workbook
    import tempfile

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"

    # Add test data
    ws['A1'] = "Test Data"
    ws['B1'] = 10
    ws['C1'] = 0

    # Add various error formulas
    ws['A2'] = "=B1/C1"          # DIV/0 error
    ws['A3'] = "=VLOOKUP(\"X\",E1:F10,2,FALSE)"  # #N/A error
    ws['A4'] = "=UNKNOWNFUNC()"  # #NAME? error
    ws['A5'] = "=Sheet2!A1"      # #REF! error (no Sheet2)
    ws['A6'] = "=A1+B1"          # #VALUE! error (text + number)
    ws['A7'] = "=A8"             # Circular reference
    ws['A8'] = "=A7"             # Circular reference

    # Save to temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        temp_file = tmp.name

    print(f"Created test file: {temp_file}")

    # Test with load_workbook to see actual formulas
    print("\n=== Checking file with openpyxl ===")
    test_wb = load_workbook(temp_file, data_only=False)
    test_ws = test_wb.active

    for row in range(1, 9):
        cell = test_ws[f'A{row}']
        print(f"A{row}: value={cell.value}, type={type(cell.value)}")

    # Now test with error detector
    print("\n=== Running IntegratedErrorDetector ===")
    detector = IntegratedErrorDetector(DummyProgressReporter())

    result = await detector.detect_all_errors(temp_file)

    print(f"\nDetection result status: {result.get('status')}")
    print(f"Total errors found: {len(result.get('errors', []))}")

    if result.get('errors'):
        print("\nErrors detected:")
        for i, error in enumerate(result['errors']):
            print(f"\n{i+1}. {error['type']} at {error['sheet']}!{error['cell']}")
            print(f"   Message: {error['message']}")
            print(f"   Formula: {error.get('formula')}")
            print(f"   Auto-fixable: {error.get('is_auto_fixable')}")
    else:
        print("\nNo errors detected - this is the problem!")

    # Clean up
    os.unlink(temp_file)
    print(f"\nDeleted test file: {temp_file}")

if __name__ == "__main__":
    asyncio.run(debug_detection())
