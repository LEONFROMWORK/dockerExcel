"""
AI 기반 자연어 Excel 생성 엔진
Natural Language to Excel Generation Engine using GPT-4
"""

import json
import logging
from typing import Dict, List, Any, Optional
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import tempfile
import os

from .openai_service import openai_service
from .excel_analyzer import excel_analyzer
from .template_excel_generator import template_excel_generator

logger = logging.getLogger(__name__)


class AIExcelGenerator:
    """AI 기반 Excel 생성기"""
    
    def __init__(self):
        self.style_presets = {
            "modern": {
                "header": {
                    "font": Font(bold=True, color="FFFFFF", size=12, name="Segoe UI"),
                    "fill": PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid"),
                    "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True)
                },
                "subheader": {
                    "font": Font(bold=True, size=11, name="Segoe UI"),
                    "fill": PatternFill(start_color="AED6F1", end_color="AED6F1", fill_type="solid"),
                    "alignment": Alignment(horizontal="center", vertical="center")
                }
            },
            "professional": {
                "header": {
                    "font": Font(bold=True, size=11, name="Arial"),
                    "fill": PatternFill(start_color="333333", end_color="333333", fill_type="solid"),
                    "font_color": "FFFFFF",
                    "alignment": Alignment(horizontal="left", vertical="center")
                }
            }
        }
    
    async def generate_from_natural_language(
        self,
        user_request: str,
        context: Optional[Dict[str, Any]] = None,
        language: str = "ko"
    ) -> Dict[str, Any]:
        """자연어 요청으로부터 Excel 파일 생성"""
        
        try:
            logger.info(f"자연어 Excel 생성 시작: {user_request[:100]}...")
            
            # 1. 사용자 의도 분석
            intent_analysis = await self._analyze_user_intent(user_request, context, language)
            
            # 2. Excel 구조 설계
            excel_structure = await self._design_excel_structure(intent_analysis, language)
            
            # 3. 데이터 스키마 생성
            data_schema = await self._generate_data_schema(excel_structure, intent_analysis)
            
            # 4. Excel 파일 생성
            excel_file = await self._create_excel_file(data_schema, excel_structure, intent_analysis)
            
            # 5. AI 기반 향상
            enhanced_file = await self._enhance_with_ai(excel_file, intent_analysis)
            
            return {
                "status": "success",
                "file_path": enhanced_file,
                "structure": excel_structure,
                "features_applied": self._get_applied_features(excel_structure),
                "ai_insights": intent_analysis.get("insights", [])
            }
            
        except Exception as e:
            logger.error(f"자연어 Excel 생성 실패: {str(e)}")
            return {
                "status": "error",
                "error": str(e),
                "fallback": await self._generate_basic_excel(user_request)
            }
    
    async def _analyze_user_intent(
        self,
        user_request: str,
        context: Optional[Dict[str, Any]],
        language: str
    ) -> Dict[str, Any]:
        """사용자 의도 분석"""
        
        system_prompt = f"""You are an Excel expert assistant fluent in {language}. 
        Analyze the user's request and extract:
        
        1. Primary objective (what they want to achieve)
        2. Data categories and types needed
        3. Required calculations or formulas
        4. Visualization needs (charts, graphs)
        5. Output format preferences
        6. Business domain/context
        7. Complexity level (simple/intermediate/advanced)
        
        Return a structured JSON response with these elements."""
        
        user_prompt = f"""User request: {user_request}
        
        Additional context: {json.dumps(context) if context else 'None'}
        
        Analyze this request and provide a detailed intent analysis."""
        
        response = await openai_service.chat_completion([
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ])
        
        try:
            intent = json.loads(response)
        except:
            # Fallback parsing
            intent = {
                "objective": user_request,
                "data_categories": ["general"],
                "calculations": [],
                "visualizations": [],
                "domain": "general",
                "complexity": "intermediate"
            }
        
        return intent
    
    async def _design_excel_structure(
        self,
        intent_analysis: Dict[str, Any],
        language: str
    ) -> Dict[str, Any]:
        """Excel 구조 설계"""
        
        system_prompt = f"""You are an Excel architect. Based on the user intent analysis,
        design an optimal Excel structure in {language}.
        
        Consider:
        1. Number and purpose of worksheets
        2. Layout and organization of each sheet
        3. Headers and data organization
        4. Formula placement and dependencies
        5. Chart types and locations
        6. Table structures
        7. Data validation rules
        8. Conditional formatting rules
        
        Return a detailed JSON structure specification."""
        
        user_prompt = f"""Intent analysis: {json.dumps(intent_analysis)}
        
        Design a comprehensive Excel structure that best serves this purpose."""
        
        response = await openai_service.chat_completion([
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ])
        
        try:
            structure = json.loads(response)
        except:
            # Fallback structure
            structure = self._get_default_structure(intent_analysis)
        
        return structure
    
    async def _generate_data_schema(
        self,
        excel_structure: Dict[str, Any],
        intent_analysis: Dict[str, Any]
    ) -> Dict[str, Any]:
        """데이터 스키마 생성"""
        
        schema = {
            "sheets": {},
            "data_types": {},
            "relationships": [],
            "validations": [],
            "formulas": []
        }
        
        for sheet_spec in excel_structure.get("sheets", []):
            sheet_name = sheet_spec.get("name", "Sheet1")
            
            # 컬럼 정의
            columns = []
            for col in sheet_spec.get("columns", []):
                column_def = {
                    "name": col.get("name"),
                    "type": col.get("type", "text"),
                    "format": col.get("format"),
                    "validation": col.get("validation"),
                    "formula": col.get("formula"),
                    "width": col.get("width", 15)
                }
                columns.append(column_def)
            
            schema["sheets"][sheet_name] = {
                "columns": columns,
                "sample_data": await self._generate_sample_data(columns, intent_analysis),
                "formatting": sheet_spec.get("formatting", {}),
                "charts": sheet_spec.get("charts", []),
                "tables": sheet_spec.get("tables", [])
            }
        
        return schema
    
    async def _create_excel_file(
        self,
        data_schema: Dict[str, Any],
        excel_structure: Dict[str, Any],
        intent_analysis: Dict[str, Any]
    ) -> str:
        """Excel 파일 생성"""
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # 기본 시트 제거
        
        # 각 시트 생성
        for sheet_name, sheet_data in data_schema["sheets"].items():
            ws = wb.create_sheet(sheet_name)
            
            # 헤더 생성
            headers = [col["name"] for col in sheet_data["columns"]]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                self._apply_style(cell, self.style_presets["modern"]["header"])
                
                # 열 너비 설정
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = sheet_data["columns"][col_idx-1].get("width", 15)
            
            # 샘플 데이터 입력
            sample_data = sheet_data.get("sample_data", [])
            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # 데이터 타입에 따른 포맷 적용
                    col_type = sheet_data["columns"][col_idx-1].get("type", "text")
                    self._apply_data_format(cell, col_type)
            
            # 테이블 생성
            if sample_data:
                table_range = f"A1:{get_column_letter(len(headers))}{len(sample_data) + 1}"
                table = Table(displayName=f"{sheet_name}Table", ref=table_range)
                table.tableStyleInfo = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
                ws.add_table(table)
            
            # 차트 생성
            for chart_spec in sheet_data.get("charts", []):
                await self._create_chart(ws, chart_spec, sample_data)
            
            # 수식 적용
            self._apply_formulas(ws, sheet_data.get("columns", []))
        
        # 파일 저장
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"ai_generated_{timestamp}.xlsx"
        output_path = os.path.join(tempfile.gettempdir(), filename)
        wb.save(output_path)
        
        return output_path
    
    async def _enhance_with_ai(
        self,
        excel_file: str,
        intent_analysis: Dict[str, Any]
    ) -> str:
        """AI 기반 Excel 파일 향상"""
        
        # 파일 분석
        analysis = await excel_analyzer.analyze_file(excel_file)
        
        # AI 기반 개선 제안
        suggestions = await self._get_ai_suggestions(analysis, intent_analysis)
        
        # 개선 사항 적용
        wb = openpyxl.load_workbook(excel_file)
        
        for suggestion in suggestions:
            if suggestion["type"] == "formula":
                self._apply_formula_suggestion(wb, suggestion)
            elif suggestion["type"] == "formatting":
                self._apply_formatting_suggestion(wb, suggestion)
            elif suggestion["type"] == "chart":
                await self._apply_chart_suggestion(wb, suggestion)
        
        # 향상된 파일 저장
        enhanced_path = excel_file.replace('.xlsx', '_enhanced.xlsx')
        wb.save(enhanced_path)
        
        return enhanced_path
    
    async def _get_ai_suggestions(
        self,
        analysis: Dict[str, Any],
        intent_analysis: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        """AI 기반 개선 제안 생성"""
        
        system_prompt = """You are an Excel optimization expert. Based on the file analysis
        and user intent, suggest improvements for:
        
        1. Formula optimizations
        2. Data organization improvements
        3. Visual enhancements
        4. Additional useful features
        5. Best practices implementation
        
        Return specific, actionable suggestions as JSON."""
        
        user_prompt = f"""File analysis: {json.dumps(analysis)}
        User intent: {json.dumps(intent_analysis)}
        
        Provide improvement suggestions."""
        
        response = await openai_service.chat_completion([
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt}
        ])
        
        try:
            suggestions = json.loads(response)
            if isinstance(suggestions, dict):
                suggestions = suggestions.get("suggestions", [])
        except:
            suggestions = []
        
        return suggestions
    
    async def _generate_sample_data(
        self,
        columns: List[Dict[str, Any]],
        intent_analysis: Dict[str, Any]
    ) -> List[List[Any]]:
        """샘플 데이터 생성"""
        
        domain = intent_analysis.get("domain", "general")
        num_rows = 10  # 기본 샘플 행 수
        
        sample_data = []
        
        for _ in range(num_rows):
            row = []
            for col in columns:
                col_type = col.get("type", "text")
                col_name = col.get("name", "")
                
                if col_type == "date":
                    value = datetime.now().strftime("%Y-%m-%d")
                elif col_type == "number":
                    if "price" in col_name.lower() or "amount" in col_name.lower():
                        value = round(50000 + (100000 * hash(col_name) % 10), 0)
                    else:
                        value = hash(col_name) % 100
                elif col_type == "percentage":
                    value = round((hash(col_name) % 100) / 100, 2)
                elif col_type == "currency":
                    value = round(10000 + (90000 * hash(col_name) % 10), 0)
                else:  # text
                    value = f"{col_name} {_ + 1}"
                
                row.append(value)
            
            sample_data.append(row)
        
        return sample_data
    
    async def _create_chart(
        self,
        ws,
        chart_spec: Dict[str, Any],
        data: List[List[Any]]
    ):
        """차트 생성"""
        
        chart_type = chart_spec.get("type", "bar")
        
        if chart_type == "bar":
            chart = BarChart()
        elif chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        else:
            return
        
        chart.title = chart_spec.get("title", "Chart")
        
        # 데이터 범위 설정 (간단한 예시)
        if data:
            data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(data) + 1)
            categories = Reference(ws, min_col=1, min_row=2, max_row=len(data) + 1)
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(categories)
        
        # 차트 위치
        position = chart_spec.get("position", "E5")
        ws.add_chart(chart, position)
    
    def _apply_style(self, cell, style_dict: Dict[str, Any]):
        """셀 스타일 적용"""
        if 'font' in style_dict:
            cell.font = style_dict['font']
        if 'fill' in style_dict:
            cell.fill = style_dict['fill']
        if 'alignment' in style_dict:
            cell.alignment = style_dict['alignment']
        if 'border' in style_dict:
            cell.border = style_dict['border']
    
    def _apply_data_format(self, cell, data_type: str):
        """데이터 타입에 따른 포맷 적용"""
        if data_type == "currency":
            cell.number_format = "#,##0원"
        elif data_type == "percentage":
            cell.number_format = "0.0%"
        elif data_type == "date":
            cell.number_format = "YYYY-MM-DD"
        elif data_type == "number":
            cell.number_format = "#,##0"
    
    def _apply_formulas(self, ws, columns: List[Dict[str, Any]]):
        """수식 적용"""
        for col_idx, col in enumerate(columns, 1):
            if col.get("formula"):
                # 간단한 SUM 예시
                if "sum" in col["formula"].lower():
                    last_row = ws.max_row
                    sum_cell = ws.cell(row=last_row + 1, column=col_idx)
                    sum_cell.value = f"=SUM({get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{last_row})"
                    sum_cell.font = Font(bold=True)
    
    def _get_default_structure(self, intent_analysis: Dict[str, Any]) -> Dict[str, Any]:
        """기본 Excel 구조"""
        return {
            "sheets": [{
                "name": "데이터",
                "columns": [
                    {"name": "항목", "type": "text"},
                    {"name": "값", "type": "number"},
                    {"name": "비고", "type": "text"}
                ],
                "formatting": {
                    "header_style": "modern"
                }
            }]
        }
    
    def _get_applied_features(self, structure: Dict[str, Any]) -> List[str]:
        """적용된 기능 목록"""
        features = ["자동 구조 생성", "스타일링"]
        
        for sheet in structure.get("sheets", []):
            if sheet.get("charts"):
                features.append("차트 생성")
            if sheet.get("tables"):
                features.append("테이블 포맷")
            if any(col.get("formula") for col in sheet.get("columns", [])):
                features.append("수식 자동 생성")
        
        return features
    
    async def _generate_basic_excel(self, user_request: str) -> str:
        """기본 Excel 파일 생성 (폴백)"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        ws['A1'] = "요청사항"
        ws['B1'] = user_request
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"basic_excel_{timestamp}.xlsx"
        output_path = os.path.join(tempfile.gettempdir(), filename)
        wb.save(output_path)
        
        return output_path
    
    def _apply_formula_suggestion(self, wb, suggestion: Dict[str, Any]):
        """수식 제안 적용"""
        sheet_name = suggestion.get("sheet", wb.sheetnames[0])
        ws = wb[sheet_name]
        
        cell_ref = suggestion.get("cell")
        formula = suggestion.get("formula")
        
        if cell_ref and formula:
            ws[cell_ref] = formula
    
    def _apply_formatting_suggestion(self, wb, suggestion: Dict[str, Any]):
        """포맷팅 제안 적용"""
        sheet_name = suggestion.get("sheet", wb.sheetnames[0])
        ws = wb[sheet_name]
        
        range_ref = suggestion.get("range")
        format_type = suggestion.get("format_type")
        
        # 간단한 조건부 서식 예시
        if range_ref and format_type == "highlight_negative":
            from openpyxl.formatting.rule import CellIsRule
            from openpyxl.styles import PatternFill
            
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            rule = CellIsRule(operator='lessThan', formula=['0'], fill=red_fill)
            ws.conditional_formatting.add(range_ref, rule)
    
    async def _apply_chart_suggestion(self, wb, suggestion: Dict[str, Any]):
        """차트 제안 적용"""
        sheet_name = suggestion.get("sheet", wb.sheetnames[0])
        ws = wb[sheet_name]
        
        chart_spec = {
            "type": suggestion.get("chart_type", "bar"),
            "title": suggestion.get("title", "Chart"),
            "position": suggestion.get("position", "E5")
        }
        
        await self._create_chart(ws, chart_spec, [])


# 전역 AI Excel 생성기 인스턴스
ai_excel_generator = AIExcelGenerator()