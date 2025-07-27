"""
Excel 수식 분석기
수식의 복잡도, 의존성, 성능 등을 분석
"""

from typing import Dict, Any, List, Tuple
import re
from dataclasses import dataclass
from enum import Enum
import networkx as nx
from collections import defaultdict
import logging
import openpyxl

logger = logging.getLogger(__name__)


class FormulaComplexity(Enum):
    """수식 복잡도 레벨"""
    SIMPLE = "simple"  # 단순 연산
    MODERATE = "moderate"  # 중간 복잡도
    COMPLEX = "complex"  # 복잡한 수식
    VERY_COMPLEX = "very_complex"  # 매우 복잡한 수식


class FunctionCategory(Enum):
    """함수 카테고리"""
    MATH = "math"  # 수학 함수
    STATISTICAL = "statistical"  # 통계 함수
    LOGICAL = "logical"  # 논리 함수
    LOOKUP = "lookup"  # 조회 함수
    TEXT = "text"  # 텍스트 함수
    DATE_TIME = "date_time"  # 날짜/시간 함수
    FINANCIAL = "financial"  # 재무 함수
    DATABASE = "database"  # 데이터베이스 함수
    ARRAY = "array"  # 배열 함수
    VOLATILE = "volatile"  # 휘발성 함수


@dataclass
class FormulaAnalysis:
    """수식 분석 결과"""
    cell: str
    sheet: str
    formula: str
    complexity: FormulaComplexity
    complexity_score: int
    functions_used: List[str]
    function_categories: List[FunctionCategory]
    referenced_cells: List[str]
    referenced_ranges: List[str]
    dependency_depth: int
    is_array_formula: bool
    is_volatile: bool
    has_external_reference: bool
    has_circular_reference: bool
    performance_impact: str  # low, medium, high
    suggestions: List[str]


@dataclass
class WorkbookAnalysis:
    """워크북 전체 분석 결과"""
    total_formulas: int
    complexity_distribution: Dict[FormulaComplexity, int]
    most_used_functions: List[Tuple[str, int]]
    dependency_graph: nx.DiGraph
    critical_paths: List[List[str]]
    volatile_formulas: List[str]
    array_formulas: List[str]
    external_references: List[str]
    performance_bottlenecks: List[Dict[str, Any]]
    optimization_opportunities: List[Dict[str, Any]]


class FormulaAnalyzer:
    """Excel 수식 분석기"""
    
    def __init__(self):
        # 함수 카테고리 매핑
        self.function_categories = {
            # 수학 함수
            'SUM': FunctionCategory.MATH,
            'AVERAGE': FunctionCategory.MATH,
            'MAX': FunctionCategory.MATH,
            'MIN': FunctionCategory.MATH,
            'ROUND': FunctionCategory.MATH,
            'ABS': FunctionCategory.MATH,
            'POWER': FunctionCategory.MATH,
            'SQRT': FunctionCategory.MATH,
            
            # 통계 함수
            'COUNT': FunctionCategory.STATISTICAL,
            'COUNTA': FunctionCategory.STATISTICAL,
            'COUNTIF': FunctionCategory.STATISTICAL,
            'COUNTIFS': FunctionCategory.STATISTICAL,
            'STDEV': FunctionCategory.STATISTICAL,
            'VAR': FunctionCategory.STATISTICAL,
            
            # 논리 함수
            'IF': FunctionCategory.LOGICAL,
            'IFS': FunctionCategory.LOGICAL,
            'AND': FunctionCategory.LOGICAL,
            'OR': FunctionCategory.LOGICAL,
            'NOT': FunctionCategory.LOGICAL,
            'IFERROR': FunctionCategory.LOGICAL,
            
            # 조회 함수
            'VLOOKUP': FunctionCategory.LOOKUP,
            'HLOOKUP': FunctionCategory.LOOKUP,
            'INDEX': FunctionCategory.LOOKUP,
            'MATCH': FunctionCategory.LOOKUP,
            'XLOOKUP': FunctionCategory.LOOKUP,
            'FILTER': FunctionCategory.LOOKUP,
            
            # 텍스트 함수
            'CONCATENATE': FunctionCategory.TEXT,
            'LEFT': FunctionCategory.TEXT,
            'RIGHT': FunctionCategory.TEXT,
            'MID': FunctionCategory.TEXT,
            'LEN': FunctionCategory.TEXT,
            'TRIM': FunctionCategory.TEXT,
            
            # 날짜/시간 함수
            'TODAY': FunctionCategory.DATE_TIME,
            'NOW': FunctionCategory.DATE_TIME,
            'DATE': FunctionCategory.DATE_TIME,
            'YEAR': FunctionCategory.DATE_TIME,
            'MONTH': FunctionCategory.DATE_TIME,
            'DAY': FunctionCategory.DATE_TIME,
            
            # 재무 함수
            'PMT': FunctionCategory.FINANCIAL,
            'FV': FunctionCategory.FINANCIAL,
            'PV': FunctionCategory.FINANCIAL,
            'RATE': FunctionCategory.FINANCIAL,
            'NPV': FunctionCategory.FINANCIAL,
            'IRR': FunctionCategory.FINANCIAL,
            
            # 데이터베이스 함수
            'DSUM': FunctionCategory.DATABASE,
            'DAVERAGE': FunctionCategory.DATABASE,
            'DCOUNT': FunctionCategory.DATABASE,
            
            # 배열 함수
            'SUMPRODUCT': FunctionCategory.ARRAY,
            'TRANSPOSE': FunctionCategory.ARRAY,
            'UNIQUE': FunctionCategory.ARRAY,
            'SORT': FunctionCategory.ARRAY,
            'SORTBY': FunctionCategory.ARRAY
        }
        
        # 휘발성 함수 목록
        self.volatile_functions = {
            'NOW', 'TODAY', 'RAND', 'RANDBETWEEN', 'INDIRECT', 'OFFSET'
        }
        
        # 성능 영향이 큰 함수
        self.high_impact_functions = {
            'SUMPRODUCT', 'SUMIFS', 'COUNTIFS', 'INDIRECT', 'OFFSET',
            'VLOOKUP', 'HLOOKUP', 'ARRAY formulas'
        }
        
        # 수식 패턴
        self.formula_patterns = {
            'function': re.compile(r'([A-Z]+)\s*\(', re.IGNORECASE),
            'cell_ref': re.compile(r'(?<![A-Z])([A-Z]+\d+)(?![A-Z0-9])', re.IGNORECASE),
            'range_ref': re.compile(r'([A-Z]+\d+:[A-Z]+\d+)', re.IGNORECASE),
            'sheet_ref': re.compile(r"'?([^'!]+)'?!"),
            'external_ref': re.compile(r'\[([^\]]+)\]'),
            'array_formula': re.compile(r'^\s*\{.*\}\s*$')
        }
    
    async def analyze_workbook(self, file_path: str) -> WorkbookAnalysis:
        """워크북 전체 수식 분석"""
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False)
            
            all_analyses = []
            dependency_graph = nx.DiGraph()
            
            # 모든 시트의 수식 분석
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_analyses = await self._analyze_sheet(sheet, sheet_name)
                all_analyses.extend(sheet_analyses)
                
                # 의존성 그래프 구축
                for analysis in sheet_analyses:
                    cell_id = f"{sheet_name}!{analysis.cell}"
                    dependency_graph.add_node(cell_id)
                    
                    for ref_cell in analysis.referenced_cells:
                        ref_id = f"{sheet_name}!{ref_cell}" if '!' not in ref_cell else ref_cell
                        dependency_graph.add_edge(ref_id, cell_id)
            
            # 분석 결과 집계
            return self._aggregate_analyses(all_analyses, dependency_graph)
            
        except Exception as e:
            logger.error(f"워크북 분석 중 오류: {str(e)}")
            raise
    
    async def _analyze_sheet(self, sheet: Any, sheet_name: str) -> List[FormulaAnalysis]:
        """시트 내 모든 수식 분석"""
        
        analyses = []
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    analysis = await self._analyze_formula(
                        cell.value,
                        cell.coordinate,
                        sheet_name,
                        sheet
                    )
                    analyses.append(analysis)
        
        return analyses
    
    async def _analyze_formula(
        self,
        formula: str,
        cell: str,
        sheet: str,
        worksheet: Any
    ) -> FormulaAnalysis:
        """개별 수식 분석"""
        
        # 함수 추출
        functions = self._extract_functions(formula)
        
        # 참조 셀 및 범위 추출
        referenced_cells, referenced_ranges = self._extract_references(formula)
        
        # 복잡도 계산
        complexity, complexity_score = self._calculate_complexity(
            formula, functions, referenced_cells, referenced_ranges
        )
        
        # 함수 카테고리 분류
        function_categories = self._classify_functions(functions)
        
        # 특성 확인
        is_array = bool(self.formula_patterns['array_formula'].match(formula))
        is_volatile = any(func.upper() in self.volatile_functions for func in functions)
        has_external = bool(self.formula_patterns['external_ref'].search(formula))
        
        # 의존성 깊이 계산
        dependency_depth = await self._calculate_dependency_depth(
            cell, sheet, worksheet, referenced_cells
        )
        
        # 성능 영향 평가
        performance_impact = self._evaluate_performance_impact(
            functions, referenced_ranges, is_volatile, is_array
        )
        
        # 최적화 제안
        suggestions = self._generate_suggestions(
            formula, functions, complexity, performance_impact
        )
        
        return FormulaAnalysis(
            cell=cell,
            sheet=sheet,
            formula=formula,
            complexity=complexity,
            complexity_score=complexity_score,
            functions_used=functions,
            function_categories=function_categories,
            referenced_cells=referenced_cells,
            referenced_ranges=referenced_ranges,
            dependency_depth=dependency_depth,
            is_array_formula=is_array,
            is_volatile=is_volatile,
            has_external_reference=has_external,
            has_circular_reference=False,  # 별도 검사 필요
            performance_impact=performance_impact,
            suggestions=suggestions
        )
    
    def _extract_functions(self, formula: str) -> List[str]:
        """수식에서 함수 추출"""
        matches = self.formula_patterns['function'].findall(formula)
        return [match.upper() for match in matches]
    
    def _extract_references(self, formula: str) -> Tuple[List[str], List[str]]:
        """수식에서 셀 참조 추출"""
        # 문자열 리터럴 제거
        formula_clean = re.sub(r'"[^"]*"', '', formula)
        
        # 셀 참조 추출
        cell_refs = self.formula_patterns['cell_ref'].findall(formula_clean)
        
        # 범위 참조 추출
        range_refs = self.formula_patterns['range_ref'].findall(formula_clean)
        
        # 범위에서 개별 셀 제외
        individual_cells = []
        for ref in cell_refs:
            if not any(ref in range_ref for range_ref in range_refs):
                individual_cells.append(ref)
        
        return individual_cells, range_refs
    
    def _calculate_complexity(
        self,
        formula: str,
        functions: List[str],
        cells: List[str],
        ranges: List[str]
    ) -> Tuple[FormulaComplexity, int]:
        """수식 복잡도 계산"""
        
        score = 0
        
        # 길이 기반 점수
        score += len(formula) // 50
        
        # 함수 개수 및 종류
        score += len(functions) * 2
        score += sum(3 for func in functions if func in self.high_impact_functions)
        
        # 중첩 레벨
        nesting_level = self._count_nesting_level(formula)
        score += nesting_level * 5
        
        # 참조 개수
        score += len(cells) + len(ranges) * 2
        
        # 조건문 복잡도
        if_count = formula.upper().count('IF(')
        score += if_count * 3
        
        # 복잡도 레벨 결정
        if score < 10:
            complexity = FormulaComplexity.SIMPLE
        elif score < 25:
            complexity = FormulaComplexity.MODERATE
        elif score < 50:
            complexity = FormulaComplexity.COMPLEX
        else:
            complexity = FormulaComplexity.VERY_COMPLEX
        
        return complexity, score
    
    def _count_nesting_level(self, formula: str) -> int:
        """수식의 중첩 레벨 계산"""
        max_level = 0
        current_level = 0
        
        for char in formula:
            if char == '(':
                current_level += 1
                max_level = max(max_level, current_level)
            elif char == ')':
                current_level -= 1
        
        return max_level
    
    def _classify_functions(self, functions: List[str]) -> List[FunctionCategory]:
        """함수 카테고리 분류"""
        categories = set()
        
        for func in functions:
            if func in self.function_categories:
                categories.add(self.function_categories[func])
            
            # 휘발성 함수 체크
            if func in self.volatile_functions:
                categories.add(FunctionCategory.VOLATILE)
        
        return list(categories)
    
    async def _calculate_dependency_depth(
        self,
        cell: str,
        sheet: str,
        worksheet: Any,
        referenced_cells: List[str]
    ) -> int:
        """의존성 깊이 계산"""
        
        if not referenced_cells:
            return 0
        
        max_depth = 0
        
        for ref_cell in referenced_cells:
            try:
                # 참조 셀의 수식 확인
                ref_cell_obj = worksheet[ref_cell]
                if ref_cell_obj.value and isinstance(ref_cell_obj.value, str) and ref_cell_obj.value.startswith('='):
                    # 재귀적으로 깊이 계산 (간소화된 버전)
                    depth = 1  # 실제로는 재귀 호출 필요
                    max_depth = max(max_depth, depth)
            except:
                pass
        
        return max_depth + 1
    
    def _evaluate_performance_impact(
        self,
        functions: List[str],
        ranges: List[str],
        is_volatile: bool,
        is_array: bool
    ) -> str:
        """성능 영향 평가"""
        
        impact_score = 0
        
        # 휘발성 함수
        if is_volatile:
            impact_score += 5
        
        # 배열 수식
        if is_array:
            impact_score += 4
        
        # 고영향 함수
        for func in functions:
            if func in self.high_impact_functions:
                impact_score += 3
        
        # 큰 범위 참조
        for range_ref in ranges:
            if ':' in range_ref:
                # 범위 크기 추정 (간소화)
                impact_score += 2
        
        # 영향도 레벨 결정
        if impact_score < 5:
            return "low"
        elif impact_score < 10:
            return "medium"
        else:
            return "high"
    
    def _generate_suggestions(
        self,
        formula: str,
        functions: List[str],
        complexity: FormulaComplexity,
        performance_impact: str
    ) -> List[str]:
        """최적화 제안 생성"""
        
        suggestions = []
        
        # 복잡도 관련 제안
        if complexity in [FormulaComplexity.COMPLEX, FormulaComplexity.VERY_COMPLEX]:
            suggestions.append("복잡한 수식을 여러 셀로 분할하여 가독성 향상")
        
        # VLOOKUP 최적화
        if 'VLOOKUP' in functions:
            suggestions.append("VLOOKUP 대신 INDEX/MATCH 또는 XLOOKUP 사용 검토")
        
        # 휘발성 함수 관련
        volatile_found = [f for f in functions if f in self.volatile_functions]
        if volatile_found:
            suggestions.append(f"휘발성 함수 {', '.join(volatile_found)} 사용 최소화")
        
        # 배열 수식 관련
        if 'SUMPRODUCT' in functions:
            suggestions.append("SUMPRODUCT 대신 SUMIFS 사용 가능 여부 검토")
        
        # 중첩 IF 관련
        if_count = formula.upper().count('IF(')
        if if_count > 3:
            suggestions.append("중첩 IF 대신 IFS 함수 또는 VLOOKUP 테이블 사용 검토")
        
        # 성능 관련
        if performance_impact == "high":
            suggestions.append("계산 성능 개선을 위해 수식 구조 재검토 필요")
        
        return suggestions
    
    def _aggregate_analyses(
        self,
        analyses: List[FormulaAnalysis],
        dependency_graph: nx.DiGraph
    ) -> WorkbookAnalysis:
        """분석 결과 집계"""
        
        # 복잡도 분포
        complexity_dist = defaultdict(int)
        for analysis in analyses:
            complexity_dist[analysis.complexity] += 1
        
        # 함수 사용 빈도
        function_count = defaultdict(int)
        for analysis in analyses:
            for func in analysis.functions_used:
                function_count[func] += 1
        
        most_used = sorted(function_count.items(), key=lambda x: x[1], reverse=True)[:10]
        
        # 휘발성 수식
        volatile_formulas = [
            f"{a.sheet}!{a.cell}" for a in analyses if a.is_volatile
        ]
        
        # 배열 수식
        array_formulas = [
            f"{a.sheet}!{a.cell}" for a in analyses if a.is_array_formula
        ]
        
        # 외부 참조
        external_refs = [
            f"{a.sheet}!{a.cell}" for a in analyses if a.has_external_reference
        ]
        
        # 성능 병목
        bottlenecks = [
            {
                'cell': f"{a.sheet}!{a.cell}",
                'impact': a.performance_impact,
                'reason': self._get_bottleneck_reason(a)
            }
            for a in analyses if a.performance_impact == "high"
        ]
        
        # 최적화 기회
        opportunities = self._identify_optimization_opportunities(analyses)
        
        # 크리티컬 패스 (간소화)
        critical_paths = []
        if dependency_graph.number_of_nodes() > 0:
            try:
                # 가장 긴 경로 찾기
                longest_path = nx.dag_longest_path(dependency_graph)
                if longest_path:
                    critical_paths.append(longest_path)
            except:
                pass
        
        return WorkbookAnalysis(
            total_formulas=len(analyses),
            complexity_distribution=dict(complexity_dist),
            most_used_functions=most_used,
            dependency_graph=dependency_graph,
            critical_paths=critical_paths,
            volatile_formulas=volatile_formulas,
            array_formulas=array_formulas,
            external_references=external_refs,
            performance_bottlenecks=bottlenecks,
            optimization_opportunities=opportunities
        )
    
    def _get_bottleneck_reason(self, analysis: FormulaAnalysis) -> str:
        """성능 병목 원인 설명"""
        reasons = []
        
        if analysis.is_volatile:
            reasons.append("휘발성 함수 사용")
        
        if analysis.is_array_formula:
            reasons.append("배열 수식")
        
        high_impact = [f for f in analysis.functions_used if f in self.high_impact_functions]
        if high_impact:
            reasons.append(f"고비용 함수: {', '.join(high_impact)}")
        
        if analysis.complexity == FormulaComplexity.VERY_COMPLEX:
            reasons.append("매우 복잡한 수식")
        
        return "; ".join(reasons) if reasons else "복잡한 계산"
    
    def _identify_optimization_opportunities(
        self,
        analyses: List[FormulaAnalysis]
    ) -> List[Dict[str, Any]]:
        """최적화 기회 식별"""
        
        opportunities = []
        
        # VLOOKUP 사용 패턴
        vlookup_cells = [
            a for a in analyses if 'VLOOKUP' in a.functions_used
        ]
        if len(vlookup_cells) > 5:
            opportunities.append({
                'type': 'vlookup_optimization',
                'description': f"{len(vlookup_cells)}개의 VLOOKUP을 INDEX/MATCH로 대체",
                'impact': 'medium',
                'cells': [f"{a.sheet}!{a.cell}" for a in vlookup_cells[:5]]
            })
        
        # 중복 계산 패턴
        formula_groups = defaultdict(list)
        for a in analyses:
            # 수식 패턴 그룹화 (간소화)
            pattern = re.sub(r'\d+', 'N', a.formula)
            formula_groups[pattern].append(a)
        
        for pattern, group in formula_groups.items():
            if len(group) > 10:
                opportunities.append({
                    'type': 'duplicate_calculation',
                    'description': f"유사한 수식 {len(group)}개를 보조 테이블로 통합",
                    'impact': 'high',
                    'pattern': pattern[:50] + '...' if len(pattern) > 50 else pattern
                })
        
        # 휘발성 함수 최적화
        volatile_count = sum(1 for a in analyses if a.is_volatile)
        if volatile_count > 10:
            opportunities.append({
                'type': 'volatile_reduction',
                'description': f"{volatile_count}개의 휘발성 함수를 정적 값으로 대체",
                'impact': 'high',
                'count': volatile_count
            })
        
        return opportunities[:10]  # 상위 10개만 반환