"""
Excel 차트 분석 및 생성 서비스
Excel Chart Analysis and Generation Service
"""

import logging
import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, AreaChart
from openpyxl.chart.reference import Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Any, Optional
import numpy as np
from datetime import datetime
import json

logger = logging.getLogger(__name__)


class ExcelChartAnalyzer:
    """Excel 차트 분석 및 생성 관리자"""
    
    def __init__(self):
        self.supported_chart_types = {
            'bar': BarChart,
            'column': BarChart, 
            'line': LineChart,
            'pie': PieChart,
            'scatter': ScatterChart,
            'area': AreaChart
        }
    
    def analyze_existing_charts(self, file_path: str) -> Dict[str, Any]:
        """기존 차트 분석"""
        
        try:
            workbook = openpyxl.load_workbook(file_path, keep_vba=True)
            chart_analysis = {
                'total_charts': 0,
                'charts_by_sheet': {},
                'chart_details': []
            }
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_charts = []
                
                # 시트의 모든 차트 분석
                for chart in sheet._charts:
                    chart_info = {
                        'chart_id': str(chart),
                        'chart_type': type(chart).__name__,
                        'title': getattr(chart, 'title', None),
                        'anchor': str(chart.anchor) if hasattr(chart, 'anchor') else None,
                        'data_range': self._extract_chart_data_range(chart),
                        'style': getattr(chart, 'style', None)
                    }
                    
                    sheet_charts.append(chart_info)
                    chart_analysis['chart_details'].append({
                        **chart_info,
                        'sheet': sheet_name
                    })
                
                chart_analysis['charts_by_sheet'][sheet_name] = {
                    'count': len(sheet_charts),
                    'charts': sheet_charts
                }
                chart_analysis['total_charts'] += len(sheet_charts)
            
            # 차트 품질 및 개선 제안
            chart_analysis['recommendations'] = self._generate_chart_recommendations(
                chart_analysis['chart_details']
            )
            
            return chart_analysis
            
        except Exception as e:
            logger.error(f"차트 분석 중 오류: {str(e)}")
            return {
                'error': str(e),
                'total_charts': 0,
                'charts_by_sheet': {},
                'chart_details': []
            }
    
    def suggest_optimal_charts(self, file_path: str, sheet_name: str = None) -> Dict[str, Any]:
        """데이터에 최적화된 차트 제안"""
        
        try:
            # Excel 파일에서 데이터 읽기
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                sheets_to_analyze = [(sheet_name, df)]
            else:
                excel_data = pd.read_excel(file_path, sheet_name=None)
                sheets_to_analyze = list(excel_data.items())
            
            suggestions = {
                'total_suggestions': 0,
                'suggestions_by_sheet': {}
            }
            
            for sheet_name, df in sheets_to_analyze:
                sheet_suggestions = self._analyze_data_for_chart_suggestions(df, sheet_name)
                suggestions['suggestions_by_sheet'][sheet_name] = sheet_suggestions
                suggestions['total_suggestions'] += len(sheet_suggestions.get('suggested_charts', []))
            
            return suggestions
            
        except Exception as e:
            logger.error(f"차트 제안 분석 중 오류: {str(e)}")
            return {
                'error': str(e),
                'total_suggestions': 0,
                'suggestions_by_sheet': {}
            }
    
    def create_chart(self, file_path: str, chart_config: Dict[str, Any]) -> Dict[str, Any]:
        """지정된 설정으로 차트 생성"""
        
        try:
            workbook = openpyxl.load_workbook(file_path, keep_vba=True)
            
            # 차트 설정 파라미터 추출
            sheet_name = chart_config.get('sheet_name')
            chart_type = chart_config.get('chart_type', 'column')
            data_range = chart_config.get('data_range')
            chart_title = chart_config.get('title', 'Chart')
            position = chart_config.get('position', 'E2')
            
            # 시트 선택
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"시트 '{sheet_name}'를 찾을 수 없습니다")
            
            sheet = workbook[sheet_name]
            
            # 차트 생성
            chart_class = self.supported_chart_types.get(chart_type.lower(), BarChart)
            chart = chart_class()
            
            # 데이터 범위 설정
            if isinstance(data_range, dict):
                # 딕셔너리 형태로 범위가 지정된 경우
                data_ref = Reference(
                    sheet,
                    min_col=data_range.get('min_col', 1),
                    min_row=data_range.get('min_row', 1),
                    max_col=data_range.get('max_col', 5),
                    max_row=data_range.get('max_row', 10)
                )
            else:
                # 문자열 범위로 지정된 경우 (예: "A1:E10")
                data_ref = Reference(sheet, range_string=data_range)
            
            chart.add_data(data_ref, titles_from_data=True)
            
            # 차트 스타일 설정
            chart.title = chart_title
            chart.style = chart_config.get('style', 10)
            
            # X축, Y축 레이블 설정
            if 'x_axis_title' in chart_config:
                chart.x_axis.title = chart_config['x_axis_title']
            if 'y_axis_title' in chart_config:
                chart.y_axis.title = chart_config['y_axis_title']
            
            # 차트를 시트에 추가
            sheet.add_chart(chart, position)
            
            # 수정된 파일 저장
            output_path = file_path.replace('.xlsx', '_with_chart.xlsx')
            workbook.save(output_path)
            
            return {
                'status': 'success',
                'message': '차트가 성공적으로 생성되었습니다',
                'chart_type': chart_type,
                'chart_title': chart_title,
                'sheet_name': sheet_name,
                'position': position,
                'output_file': output_path
            }
            
        except Exception as e:
            logger.error(f"차트 생성 중 오류: {str(e)}")
            return {
                'status': 'error',
                'message': f'차트 생성 실패: {str(e)}'
            }
    
    def auto_generate_charts(self, file_path: str, max_charts_per_sheet: int = 3) -> Dict[str, Any]:
        """데이터를 기반으로 자동으로 차트 생성"""
        
        try:
            # 최적 차트 제안 받기
            suggestions = self.suggest_optimal_charts(file_path)
            
            workbook = openpyxl.load_workbook(file_path, keep_vba=True)
            generated_charts = []
            
            for sheet_name, sheet_suggestions in suggestions['suggestions_by_sheet'].items():
                suggested_charts = sheet_suggestions.get('suggested_charts', [])[:max_charts_per_sheet]
                
                for i, chart_suggestion in enumerate(suggested_charts):
                    chart_config = {
                        'sheet_name': sheet_name,
                        'chart_type': chart_suggestion['type'],
                        'data_range': chart_suggestion['data_range'],
                        'title': chart_suggestion['title'],
                        'position': f'{chr(69 + i * 8)}{2 + i * 15}',  # E2, M2, U2...
                        'x_axis_title': chart_suggestion.get('x_axis_title'),
                        'y_axis_title': chart_suggestion.get('y_axis_title')
                    }
                    
                    # 개별 차트 생성 (임시 파일 사용 안함)
                    chart_result = self._create_single_chart(workbook, chart_config)
                    if chart_result['status'] == 'success':
                        generated_charts.append(chart_result)
            
            # 수정된 워크북 저장
            output_path = file_path.replace('.xlsx', '_auto_charts.xlsx')
            workbook.save(output_path)
            
            return {
                'status': 'success',
                'message': f'{len(generated_charts)}개의 차트가 자동 생성되었습니다',
                'generated_charts': generated_charts,
                'output_file': output_path,
                'total_charts_generated': len(generated_charts)
            }
            
        except Exception as e:
            logger.error(f"자동 차트 생성 중 오류: {str(e)}")
            return {
                'status': 'error',
                'message': f'자동 차트 생성 실패: {str(e)}'
            }
    
    def _extract_chart_data_range(self, chart) -> Optional[str]:
        """차트의 데이터 범위 추출"""
        try:
            if hasattr(chart, 'series') and chart.series:
                # 첫 번째 시리즈의 데이터 범위 반환
                series = chart.series[0]
                if hasattr(series, 'xValues') and series.xValues:
                    return str(series.xValues)
                elif hasattr(series, 'values') and series.values:
                    return str(series.values)
            return None
        except Exception:
            return None
    
    def _generate_chart_recommendations(self, chart_details: List[Dict]) -> List[str]:
        """차트 개선 권장사항 생성"""
        recommendations = []
        
        if len(chart_details) == 0:
            recommendations.append("데이터 시각화를 위해 차트 추가를 고려해보세요")
        
        chart_types = [chart['chart_type'] for chart in chart_details]
        if len(set(chart_types)) == 1 and len(chart_types) > 1:
            recommendations.append("다양한 차트 유형을 사용하여 데이터를 더 효과적으로 표현할 수 있습니다")
        
        charts_without_title = [chart for chart in chart_details if not chart.get('title')]
        if charts_without_title:
            recommendations.append(f"{len(charts_without_title)}개의 차트에 제목을 추가하면 가독성이 향상됩니다")
        
        return recommendations
    
    def _analyze_data_for_chart_suggestions(self, df: pd.DataFrame, sheet_name: str) -> Dict[str, Any]:
        """데이터프레임 분석하여 적합한 차트 제안"""
        
        if df.empty or len(df.columns) < 2:
            return {
                'message': '차트 생성에 충분한 데이터가 없습니다',
                'suggested_charts': []
            }
        
        suggestions = []
        
        # 숫자 컬럼과 카테고리 컬럼 식별
        numeric_columns = df.select_dtypes(include=[np.number]).columns.tolist()
        categorical_columns = df.select_dtypes(include=['object', 'string']).columns.tolist()
        datetime_columns = df.select_dtypes(include=['datetime64']).columns.tolist()
        
        # 1. 카테고리별 수치 비교 (막대 차트)
        if len(categorical_columns) >= 1 and len(numeric_columns) >= 1:
            suggestions.append({
                'type': 'column',
                'title': f'{sheet_name} - {categorical_columns[0]} 별 {numeric_columns[0]}',
                'data_range': f'A1:{chr(65 + len([categorical_columns[0]] + numeric_columns[:3]))}{min(len(df) + 1, 20)}',
                'confidence': 0.8,
                'reason': '카테고리별 수치 비교에 적합',
                'x_axis_title': categorical_columns[0],
                'y_axis_title': numeric_columns[0]
            })
        
        # 2. 시계열 데이터 (선 차트)
        if len(datetime_columns) >= 1 and len(numeric_columns) >= 1:
            suggestions.append({
                'type': 'line',
                'title': f'{sheet_name} - 시간별 {numeric_columns[0]} 추이',
                'data_range': f'A1:{chr(65 + len([datetime_columns[0]] + numeric_columns[:2]))}{min(len(df) + 1, 50)}',
                'confidence': 0.9,
                'reason': '시계열 데이터 추이 분석에 적합',
                'x_axis_title': datetime_columns[0],
                'y_axis_title': numeric_columns[0]
            })
        
        # 3. 비율/구성 데이터 (파이 차트)
        if len(categorical_columns) >= 1 and len(numeric_columns) >= 1:
            # 카테고리가 10개 이하일 때만 파이 차트 제안
            unique_categories = df[categorical_columns[0]].nunique()
            if unique_categories <= 10:
                suggestions.append({
                    'type': 'pie',
                    'title': f'{sheet_name} - {categorical_columns[0]} 구성비',
                    'data_range': f'A1:{chr(65 + 1)}{min(unique_categories + 1, 15)}',
                    'confidence': 0.7,
                    'reason': '전체 대비 구성비 표현에 적합',
                    'x_axis_title': categorical_columns[0],
                    'y_axis_title': numeric_columns[0]
                })
        
        # 4. 상관관계 분석 (산점도)
        if len(numeric_columns) >= 2:
            suggestions.append({
                'type': 'scatter',
                'title': f'{sheet_name} - {numeric_columns[0]} vs {numeric_columns[1]} 상관관계',
                'data_range': f'A1:{chr(65 + 1)}{min(len(df) + 1, 100)}',
                'confidence': 0.6,
                'reason': '두 변수 간 상관관계 분석에 적합',
                'x_axis_title': numeric_columns[0],
                'y_axis_title': numeric_columns[1]
            })
        
        # 신뢰도 순으로 정렬
        suggestions.sort(key=lambda x: x['confidence'], reverse=True)
        
        return {
            'data_summary': {
                'total_rows': len(df),
                'total_columns': len(df.columns),
                'numeric_columns': len(numeric_columns),
                'categorical_columns': len(categorical_columns),
                'datetime_columns': len(datetime_columns)
            },
            'suggested_charts': suggestions[:5],  # 최대 5개 제안
            'sheet_name': sheet_name
        }
    
    def _create_single_chart(self, workbook: openpyxl.Workbook, chart_config: Dict[str, Any]) -> Dict[str, Any]:
        """단일 차트 생성 (내부 메서드)"""
        
        try:
            sheet_name = chart_config['sheet_name']
            chart_type = chart_config['chart_type']
            
            if sheet_name not in workbook.sheetnames:
                return {'status': 'error', 'message': f"시트 '{sheet_name}'를 찾을 수 없습니다"}
            
            sheet = workbook[sheet_name]
            
            # 차트 클래스 선택
            chart_class = self.supported_chart_types.get(chart_type.lower(), BarChart)
            chart = chart_class()
            
            # 데이터 범위 설정
            if isinstance(chart_config['data_range'], dict):
                data_ref = Reference(
                    sheet,
                    min_col=chart_config['data_range'].get('min_col', 1),
                    min_row=chart_config['data_range'].get('min_row', 1),
                    max_col=chart_config['data_range'].get('max_col', 3),
                    max_row=chart_config['data_range'].get('max_row', 10)
                )
            else:
                data_ref = Reference(sheet, range_string=chart_config['data_range'])
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.title = chart_config['title']
            
            # 축 제목 설정
            if chart_config.get('x_axis_title'):
                chart.x_axis.title = chart_config['x_axis_title']
            if chart_config.get('y_axis_title'):
                chart.y_axis.title = chart_config['y_axis_title']
            
            # 차트 추가
            sheet.add_chart(chart, chart_config['position'])
            
            return {
                'status': 'success',
                'chart_type': chart_type,
                'title': chart_config['title'],
                'position': chart_config['position']
            }
            
        except Exception as e:
            return {
                'status': 'error',
                'message': f'차트 생성 실패: {str(e)}'
            }


# 전역 차트 분석기 인스턴스
excel_chart_analyzer = ExcelChartAnalyzer()