"""
Excel file analysis service with comprehensive error detection
Excel 파일 분석 및 포괄적인 오류 감지 서비스
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from typing import Dict, Any, List, Optional
import re
from pathlib import Path
import logging
from datetime import datetime
from dataclasses import dataclass

logger = logging.getLogger(__name__)


@dataclass
class ExcelError:
    """Excel 오류 정보를 담는 데이터 클래스"""

    id: str
    error_type: str
    severity: str  # critical, high, medium, low
    location: str  # Sheet!A1
    cell: str  # A1
    sheet: str
    message: str
    formula: Optional[str] = None
    value: Any = None
    suggestion: Optional[str] = None
    auto_fixable: bool = False
    confidence: float = 0.95
    category: str = "critical_error"  # 'critical_error' or 'potential_issue'


class ExcelAnalyzer:
    """Service for analyzing Excel files with comprehensive error detection"""

    def __init__(self):
        self.formula_pattern = re.compile(r"^=")

        # Excel 오류 타입과 설명
        self.error_patterns = {
            "#DIV/0!": "Division by zero error",
            "#N/A": "Value not available error",
            "#NAME?": "Unrecognized formula name",
            "#NULL!": "Null intersection error",
            "#NUM!": "Invalid numeric value",
            "#REF!": "Invalid cell reference",
            "#VALUE!": "Wrong value type",
            "#SPILL!": "Spill range blocked",
            "#CALC!": "Calculation error",
        }

        # 오류 메시지 한글화
        self.error_messages = {
            "#DIV/0!": "0으로 나누기 오류",
            "#N/A": "값을 찾을 수 없음",
            "#NAME?": "이름을 인식할 수 없음",
            "#NULL!": "잘못된 교집합",
            "#NUM!": "숫자가 잘못됨",
            "#REF!": "참조가 잘못됨",
            "#VALUE!": "값 형식이 잘못됨",
            "#SPILL!": "스필 범위 차단",
            "#CALC!": "계산할 수 없음",
        }

        # 수식에서 오류를 발생시킬 수 있는 패턴
        self.risky_patterns = {
            r"/\s*0": "0으로 나누기 가능성",
            r"VLOOKUP": "VLOOKUP 오류 가능성",
            r'1\*"[^"]*"': "텍스트와 숫자 연산",
            r"\^999": "매우 큰 지수 연산",
            r"#REF!": "참조 오류 포함",
            r"SUM\([^)]*#": "오류가 포함된 SUM",
        }

    async def analyze_file(self, file_path: str) -> Dict[str, Any]:
        """Analyze an Excel file and extract metadata with error detection"""
        try:
            file_path = Path(file_path)

            # Basic file info
            file_info = {
                "filename": file_path.name,
                "size": file_path.stat().st_size,
                "modified": datetime.fromtimestamp(
                    file_path.stat().st_mtime
                ).isoformat(),
            }

            # 두 가지 모드로 워크북 열기
            # 1. 수식 포함 (data_only=False)
            workbook_formulas = openpyxl.load_workbook(
                file_path, read_only=False, data_only=False, keep_vba=False
            )
            # 2. 계산된 값 (data_only=True)
            workbook_values = openpyxl.load_workbook(
                file_path, read_only=False, data_only=True, keep_vba=False
            )

            sheets_info = {}
            total_formulas = 0
            total_errors = 0
            all_errors = []

            # 모든 시트 분석
            for sheet_name in workbook_formulas.sheetnames:
                sheet_formulas = workbook_formulas[sheet_name]
                sheet_values = workbook_values[sheet_name]

                sheet_analysis = await self._analyze_sheet(
                    sheet_formulas, sheet_values, sheet_name
                )
                sheets_info[sheet_name] = sheet_analysis
                total_formulas += sheet_analysis.get("formula_count", 0)

                # 오류 수집
                sheet_errors = sheet_analysis.get("errors", [])
                all_errors.extend(sheet_errors)

            # 중복 제거
            unique_errors = []
            seen_errors = set()

            for error in all_errors:
                # 중복 체크를 위한 키 생성 (sheet, cell, error_type 조합)
                error_key = (error.sheet, error.cell, error.error_type)
                if error_key not in seen_errors:
                    seen_errors.add(error_key)
                    unique_errors.append(error)

            all_errors = unique_errors
            total_errors = len(all_errors)

            workbook_formulas.close()
            workbook_values.close()

            # Transform errors to expected format
            formatted_errors = []
            for error in all_errors:
                formatted_errors.append(
                    {
                        "id": error.id,
                        "error": error.error_type,
                        "error_type": error.error_type,
                        "description": error.message,
                        "cell": error.cell,
                        "location": error.location,
                        "sheet": error.sheet,
                        "formula": error.formula,
                        "value": error.value,
                        "severity": error.severity,
                        "category": error.category,
                        "suggestion": error.suggestion,
                        "auto_fixable": error.auto_fixable,
                        "confidence": error.confidence,
                    }
                )

            return {
                "file_info": file_info,
                "sheets": sheets_info,
                "errors": formatted_errors,
                "summary": {
                    "total_sheets": len(sheets_info),
                    "total_formulas": total_formulas,
                    "total_errors": total_errors,
                    "has_errors": total_errors > 0,
                    "error_types": self._summarize_error_types(all_errors),
                },
            }

        except InvalidFileException as e:
            logger.error(f"Invalid Excel file: {str(e)}")
            raise
        except Exception as e:
            logger.error(f"Failed to analyze Excel file: {str(e)}")
            raise

    async def _analyze_sheet(
        self, sheet_formulas, sheet_values, sheet_name: str
    ) -> Dict[str, Any]:
        """Analyze a single sheet with both formula and value views"""
        analysis = {
            "name": sheet_name,
            "rows": sheet_formulas.max_row,
            "columns": sheet_formulas.max_column,
            "used_range": f"A1:{get_column_letter(sheet_formulas.max_column)}{sheet_formulas.max_row}",
            "formulas": [],
            "errors": [],
            "formula_count": 0,
            "merged_cells": [],
        }

        # Get column headers
        if sheet_formulas.max_row > 0:
            headers = []
            for col in range(1, sheet_formulas.max_column + 1):
                cell = sheet_formulas.cell(row=1, column=col)
                headers.append(cell.value if cell.value else f"Column{col}")
            analysis["columns"] = headers

        errors_found = []

        # Analyze all cells
        for row in range(1, sheet_formulas.max_row + 1):
            for col in range(1, sheet_formulas.max_column + 1):
                cell_formulas = sheet_formulas.cell(row=row, column=col)
                cell_values = sheet_values.cell(row=row, column=col)

                # 1. 수식이 있는 셀 검사
                if cell_formulas.data_type == "f" and cell_formulas.value:
                    formula = str(cell_formulas.value)

                    # 수식 정보 수집
                    if len(analysis["formulas"]) < 10:  # Limit stored formulas
                        analysis["formulas"].append(
                            {
                                "cell": cell_formulas.coordinate,
                                "formula": formula,
                                "value": cell_values.value,
                            }
                        )
                    analysis["formula_count"] += 1

                    # 수식 자체의 위험 패턴 검사
                    for pattern, description in self.risky_patterns.items():
                        if re.search(pattern, formula, re.IGNORECASE):
                            # 실제로 오류가 발생했는지 확인
                            calculated_value = cell_values.value
                            error_type = self._check_value_for_error(calculated_value)

                            if error_type:
                                error = self._create_error(
                                    cell_formulas,
                                    sheet_name,
                                    error_type,
                                    formula=formula,
                                    value=calculated_value,
                                )
                                errors_found.append(error)

                # 2. 값 자체가 오류인 경우 검사 (수식이 없어도)
                if cell_values.value is not None:
                    # ArrayFormula 처리
                    actual_value = cell_values.value
                    if (
                        hasattr(actual_value, "__class__")
                        and actual_value.__class__.__name__ == "ArrayFormula"
                    ):
                        # ArrayFormula의 경우 string representation을 체크
                        actual_value = str(actual_value)

                    error_type = self._check_value_for_error(actual_value)
                    if error_type:
                        # 이미 추가된 오류인지 확인
                        already_added = any(
                            e.cell == cell_values.coordinate
                            and e.sheet == sheet_name
                            and e.error_type == error_type
                            for e in errors_found
                        )

                        if not already_added:
                            error = self._create_error(
                                cell_values,
                                sheet_name,
                                error_type,
                                formula=(
                                    str(cell_formulas.value)
                                    if cell_formulas.data_type == "f"
                                    else None
                                ),
                                value=actual_value,
                            )
                            errors_found.append(error)

        # Add errors to analysis
        analysis["errors"] = errors_found

        # Merged cells info
        for merged_range in sheet_formulas.merged_cells.ranges:
            analysis["merged_cells"].append(str(merged_range))

        # Data type analysis
        analysis["data_types"] = self._analyze_data_types(sheet_values)

        return analysis

    def _check_value_for_error(self, value: Any) -> Optional[str]:
        """값이 Excel 오류인지 확인"""
        if value is None:
            return None

        value_str = str(value)

        # ArrayFormula 객체의 경우 str() 변환
        if hasattr(value, "__class__") and value.__class__.__name__ == "ArrayFormula":
            value_str = str(value)

        # 정확한 오류 매칭을 위해 앞뒤 공백 제거
        value_str = value_str.strip()

        # 모든 Excel 오류 타입 확인 (정확한 매칭)
        for error_type in self.error_patterns.keys():
            if value_str == error_type or value_str.startswith(error_type):
                return error_type

        return None

    def _create_error(
        self,
        cell,
        sheet_name: str,
        error_type: str,
        formula: Optional[str] = None,
        value: Any = None,
    ) -> ExcelError:
        """오류 객체 생성"""
        error_id = f"err_{sheet_name}_{cell.coordinate}_{error_type.replace('#', '').replace('!', '').replace('?', '')}"

        # 카테고리 결정: 모든 수식 오류는 명백한 오류
        category = "critical_error"

        return ExcelError(
            id=error_id,
            error_type=error_type,
            severity=self._get_error_severity(error_type),
            location=f"{sheet_name}!{cell.coordinate}",
            cell=cell.coordinate,
            sheet=sheet_name,
            message=self.error_messages.get(
                error_type, self.error_patterns.get(error_type, "Unknown error")
            ),
            formula=formula,
            value=value if value is not None else error_type,
            suggestion=self._get_error_suggestion(error_type),
            auto_fixable=self._is_auto_fixable(error_type),
            confidence=0.95,
            category=category,
        )

    def _get_error_severity(self, error_type: str) -> str:
        """오류 심각도 결정"""
        critical = ["#REF!"]
        high = ["#DIV/0!", "#VALUE!", "#NAME?"]
        medium = ["#N/A", "#NUM!"]

        if error_type in critical:
            return "critical"
        elif error_type in high:
            return "high"
        elif error_type in medium:
            return "medium"
        else:
            return "low"

    def _get_error_suggestion(self, error_type: str) -> str:
        """오류에 대한 수정 제안"""
        suggestions = {
            "#DIV/0!": "분모가 0이 되지 않도록 IF 함수나 IFERROR 함수를 사용하세요",
            "#N/A": "VLOOKUP이나 MATCH 함수에서 값을 찾을 수 없습니다. 데이터 범위나 검색값을 확인하세요",
            "#NAME?": "함수명이나 참조명이 잘못되었습니다. 철자를 확인하세요",
            "#REF!": "참조하는 셀이나 범위가 삭제되었습니다. 올바른 참조로 수정하세요",
            "#VALUE!": "수식에 잘못된 데이터 타입이 사용되었습니다. 인수를 확인하세요",
            "#NUM!": "숫자가 너무 크거나 잘못된 숫자 연산입니다",
            "#NULL!": "범위 연산자(공백)가 잘못 사용되었습니다",
        }

        return suggestions.get(error_type, "수식을 검토하고 수정이 필요합니다")

    def _is_auto_fixable(self, error_type: str) -> bool:
        """자동 수정 가능 여부"""
        auto_fixable_errors = ["#DIV/0!", "#N/A", "#VALUE!"]
        return error_type in auto_fixable_errors

    def _summarize_error_types(self, errors: List[ExcelError]) -> Dict[str, int]:
        """오류 타입별 개수 집계"""
        summary = {}
        for error in errors:
            error_type = error.error_type
            summary[error_type] = summary.get(error_type, 0) + 1
        return summary

    def _analyze_data_types(self, sheet) -> Dict[str, Dict[str, Any]]:
        """Analyze data types in columns"""
        column_types = {}

        for col in range(1, min(sheet.max_column + 1, 20)):  # Limit to first 20 columns
            col_letter = get_column_letter(col)
            types = {"text": 0, "number": 0, "date": 0, "formula": 0, "empty": 0}

            for row in range(2, min(sheet.max_row + 1, 100)):  # Sample first 100 rows
                cell = sheet.cell(row=row, column=col)

                if cell.value is None:
                    types["empty"] += 1
                elif isinstance(cell.value, (int, float)):
                    types["number"] += 1
                elif isinstance(cell.value, datetime):
                    types["date"] += 1
                elif isinstance(cell.value, str):
                    if cell.value.startswith("="):
                        types["formula"] += 1
                    else:
                        types["text"] += 1

            # Determine primary type
            non_empty_types = {k: v for k, v in types.items() if k != "empty" and v > 0}
            if non_empty_types:
                primary_type = max(non_empty_types, key=non_empty_types.get)
                column_types[col_letter] = {
                    "primary": primary_type,
                    "distribution": types,
                    "mixed": len(non_empty_types) > 1,
                }

        return column_types

    async def extract_formulas(self, file_path: str) -> Dict[str, List[Dict[str, Any]]]:
        """Extract all formulas from an Excel file"""
        try:
            workbook = openpyxl.load_workbook(
                file_path, read_only=True, data_only=False
            )
            formulas_by_sheet = {}

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                formulas = []

                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.data_type == "f" and cell.value:
                            formulas.append(
                                {
                                    "cell": cell.coordinate,
                                    "formula": cell.value,
                                    "row": cell.row,
                                    "column": cell.column,
                                }
                            )

                if formulas:
                    formulas_by_sheet[sheet_name] = formulas

            workbook.close()
            return formulas_by_sheet

        except Exception as e:
            logger.error(f"Failed to extract formulas: {str(e)}")
            raise

    def _extract_sheet_data(self, sheet) -> List[List[Dict[str, Any]]]:
        """Extract sheet data with cell formatting"""
        data = []

        for row_idx in range(1, sheet.max_row + 1):
            row_data = []
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell_data = {"value": cell.value, "address": cell.coordinate}

                # Add comprehensive style information
                style_data = {}

                # Font information
                if cell.font:
                    font_data = {
                        "bold": cell.font.bold,
                        "italic": cell.font.italic,
                        "size": cell.font.size,
                        "name": cell.font.name,
                        "underline": cell.font.underline,
                        "strike": cell.font.strike,
                    }

                    # Handle font color
                    if cell.font.color:
                        if cell.font.color.type == "theme":
                            font_data["color"] = f"theme:{cell.font.color.theme}"
                        elif cell.font.color.type == "indexed":
                            font_data["color"] = f"indexed:{cell.font.color.indexed}"
                        elif cell.font.color.rgb:
                            font_data["color"] = f"#{cell.font.color.rgb[2:]}"
                        else:
                            font_data["color"] = "auto"

                    style_data["font"] = font_data

                # Fill (background) information
                if cell.fill and cell.fill.patternType:
                    fill_data = {"type": cell.fill.patternType}

                    if cell.fill.fgColor:
                        if cell.fill.fgColor.type == "theme":
                            fill_data["color"] = f"theme:{cell.fill.fgColor.theme}"
                        elif cell.fill.fgColor.type == "indexed":
                            fill_data["color"] = f"indexed:{cell.fill.fgColor.indexed}"
                        elif cell.fill.fgColor.rgb:
                            fill_data["color"] = f"#{cell.fill.fgColor.rgb[2:]}"

                    style_data["fill"] = fill_data

                # Border information
                if cell.border:
                    border_data = {}
                    for side in ["left", "right", "top", "bottom"]:
                        border_side = getattr(cell.border, side)
                        if border_side and border_side.style:
                            border_data[side] = {"style": border_side.style}
                            if border_side.color and border_side.color.rgb:
                                border_data[side][
                                    "color"
                                ] = f"#{border_side.color.rgb[2:]}"

                    if border_data:
                        style_data["border"] = border_data

                # Alignment information
                if cell.alignment:
                    alignment_data = {}
                    if cell.alignment.horizontal:
                        alignment_data["horizontal"] = cell.alignment.horizontal
                    if cell.alignment.vertical:
                        alignment_data["vertical"] = cell.alignment.vertical
                    if cell.alignment.wrap_text is not None:
                        alignment_data["wrapText"] = cell.alignment.wrap_text
                    if cell.alignment.text_rotation:
                        alignment_data["textRotation"] = cell.alignment.text_rotation

                    if alignment_data:
                        style_data["alignment"] = alignment_data

                # Number format
                if cell.number_format and cell.number_format != "General":
                    style_data["numberFormat"] = cell.number_format

                # Add style data if present
                if style_data:
                    cell_data["style"] = style_data

                row_data.append(cell_data)

            data.append(row_data)

        return data


# Singleton instance
excel_analyzer = ExcelAnalyzer()
