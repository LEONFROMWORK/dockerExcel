"""
Debug error detection issue - simplified version
"""
import asyncio
import sys
import os
from openpyxl import load_workbook

# Add project to path
sys.path.insert(0, '/Users/kevin/excel-unified/python-service')

# Set environment to disable caching
os.environ['DISABLE_CACHE'] = 'true'

async def debug_detection():
    # Create a test Excel file with known errors
    from openpyxl import Workbook
    import tempfile

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"

    # Add test data
    ws['A1'] = "Test Data"
    ws['B1'] = 10
    ws['C1'] = 0

    # Add various error formulas
    ws['A2'] = "=B1/C1"          # DIV/0 error
    ws['A3'] = "=VLOOKUP(\"X\",E1:F10,2,FALSE)"  # #N/A error
    ws['A4'] = "=UNKNOWNFUNC()"  # #NAME? error
    ws['A5'] = "=Sheet2!A1"      # #REF! error (no Sheet2)
    ws['A6'] = "=A1+B1"          # #VALUE! error (text + number)
    ws['A7'] = "=A8"             # Circular reference
    ws['A8'] = "=A7"             # Circular reference

    # Save to temp file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        temp_file = tmp.name

    print(f"Created test file: {temp_file}")

    # Test directly with formula error detector
    print("\n=== Testing FormulaErrorDetector directly ===")
    from app.services.detection.strategies.formula_error_detector import FormulaErrorDetector
    from app.services.workbook_loader import OpenpyxlWorkbookLoader

    # Load workbook
    loader = OpenpyxlWorkbookLoader()
    workbook = await loader.load_workbook(temp_file)

    # Create detector and detect errors
    formula_detector = FormulaErrorDetector()
    errors = await formula_detector.detect(workbook)

    print(f"\nFormulaErrorDetector found {len(errors)} errors")
    for i, error in enumerate(errors):
        print(f"{i+1}. {error.type} at {error.sheet}!{error.cell}: {error.message}")

    # Now test with IntegratedErrorDetector but skip caching
    print("\n=== Testing IntegratedErrorDetector (no cache) ===")
    from app.services.detection.integrated_error_detector import IntegratedErrorDetector
    from app.core.interfaces import DummyProgressReporter

    detector = IntegratedErrorDetector(DummyProgressReporter())

    # Directly call the internal method to skip caching
    all_errors = await detector._run_detectors_parallel(workbook)

    print(f"\nIntegratedErrorDetector found {len(all_errors)} errors (before deduplication)")

    # Deduplicate
    unique_errors = detector._deduplicate_errors(all_errors)
    print(f"After deduplication: {len(unique_errors)} errors")

    for i, error in enumerate(unique_errors):
        print(f"{i+1}. {error.type} at {error.sheet}!{error.cell}: {error.message}")

    # Clean up
    os.unlink(temp_file)
    print(f"\nDeleted test file: {temp_file}")

if __name__ == "__main__":
    asyncio.run(debug_detection())
