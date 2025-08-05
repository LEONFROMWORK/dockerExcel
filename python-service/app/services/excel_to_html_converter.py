"""
Excel to HTML converter with proper merged cell support
"""

import openpyxl
from openpyxl.utils import get_column_letter
from typing import Dict, Any


class ExcelToHtmlConverter:
    def __init__(self):
        self.default_col_width = 64  # pixels
        self.default_row_height = 25  # pixels

    def convert_file_to_html(
        self, file_path: str, sheet_name: str = None
    ) -> Dict[str, Any]:
        """Convert Excel file to HTML with proper merged cell support"""
        wb = openpyxl.load_workbook(file_path, data_only=True)

        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        return self.convert_sheet_to_html(ws)

    def convert_sheet_to_html(self, worksheet) -> Dict[str, Any]:
        """Convert worksheet to HTML structure"""
        # Get dimensions
        max_row = worksheet.max_row or 100
        max_col = worksheet.max_column or 26

        # Extract column widths
        col_widths = {}
        for col in range(1, max_col + 1):
            col_letter = get_column_letter(col)
            width = worksheet.column_dimensions[col_letter].width
            if width:
                # Excel width to pixels: width * 7 + 5
                col_widths[col] = int(width * 7 + 5)
            else:
                col_widths[col] = self.default_col_width

        # Extract row heights
        row_heights = {}
        for row in range(1, max_row + 1):
            height = worksheet.row_dimensions[row].height
            if height:
                # Excel points to pixels: points * 1.33
                row_heights[row] = int(height * 1.33)
            else:
                row_heights[row] = self.default_row_height

        # Process merged cells
        merged_cells = []
        merged_map = {}  # Map of cells that are part of merges

        for merge_range in worksheet.merged_cells.ranges:
            start_row = merge_range.min_row
            start_col = merge_range.min_col
            end_row = merge_range.max_row
            end_col = merge_range.max_col

            merged_cells.append(
                {
                    "start": {"row": start_row - 1, "col": start_col - 1},
                    "end": {"row": end_row - 1, "col": end_col - 1},
                    "colspan": end_col - start_col + 1,
                    "rowspan": end_row - start_row + 1,
                }
            )

            # Mark all cells in merge (except master) as hidden
            for r in range(start_row, end_row + 1):
                for c in range(start_col, end_col + 1):
                    if r != start_row or c != start_col:
                        merged_map[f"{r},{c}"] = True

        # Build HTML table
        html_rows = []

        for row in range(1, max_row + 1):
            html_cols = []

            for col in range(1, max_col + 1):
                # Skip cells that are part of a merge (but not the master)
                if f"{row},{col}" in merged_map:
                    continue

                cell = worksheet.cell(row=row, column=col)

                # Cell attributes
                cell_attrs = {
                    "class": "excel-cell",
                    "style": self._get_cell_style(
                        cell, col_widths[col], row_heights[row]
                    ),
                }

                # Check if this is a merged cell master
                for merge in merged_cells:
                    if (
                        merge["start"]["row"] == row - 1
                        and merge["start"]["col"] == col - 1
                    ):
                        cell_attrs["colspan"] = merge["colspan"]
                        cell_attrs["rowspan"] = merge["rowspan"]
                        cell_attrs["class"] += " merged-cell"

                        # Calculate total width for merged cell
                        total_width = sum(
                            col_widths[c] for c in range(col, col + merge["colspan"])
                        )
                        total_height = sum(
                            row_heights[r] for r in range(row, row + merge["rowspan"])
                        )

                        cell_attrs[
                            "style"
                        ] += f" width: {total_width}px; height: {total_height}px;"
                        break

                # Cell value
                value = cell.value if cell.value is not None else ""

                html_cols.append({"value": str(value), "attrs": cell_attrs})

            html_rows.append(html_cols)

        return {
            "html_rows": html_rows,
            "col_widths": col_widths,
            "row_heights": row_heights,
            "merged_cells": merged_cells,
            "max_row": max_row,
            "max_col": max_col,
        }

    def _get_cell_style(self, cell, width: int, height: int) -> str:
        """Get CSS style for cell"""
        styles = []

        # Size
        styles.append(f"width: {width}px")
        styles.append(f"min-width: {width}px")
        styles.append(f"height: {height}px")

        # Alignment
        if cell.alignment:
            if cell.alignment.horizontal:
                styles.append(f"text-align: {cell.alignment.horizontal}")
            if cell.alignment.vertical:
                if cell.alignment.vertical == "top":
                    styles.append("vertical-align: top")
                elif cell.alignment.vertical == "center":
                    styles.append("vertical-align: middle")
                elif cell.alignment.vertical == "bottom":
                    styles.append("vertical-align: bottom")

        # Font
        if cell.font:
            if cell.font.bold:
                styles.append("font-weight: bold")
            if cell.font.size:
                styles.append(f"font-size: {cell.font.size}pt")
            if cell.font.color and cell.font.color.rgb:
                styles.append(f"color: #{cell.font.color.rgb[2:]}")

        # Background
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
            styles.append(f"background-color: #{cell.fill.fgColor.rgb[2:]}")

        # Border
        if cell.border:
            border_styles = []
            if cell.border.top and cell.border.top.style:
                border_styles.append("border-top: 1px solid #d0d0d0")
            if cell.border.bottom and cell.border.bottom.style:
                border_styles.append("border-bottom: 1px solid #d0d0d0")
            if cell.border.left and cell.border.left.style:
                border_styles.append("border-left: 1px solid #d0d0d0")
            if cell.border.right and cell.border.right.style:
                border_styles.append("border-right: 1px solid #d0d0d0")
            styles.extend(border_styles)
        else:
            styles.append("border: 1px solid #e0e0e0")

        return "; ".join(styles)
