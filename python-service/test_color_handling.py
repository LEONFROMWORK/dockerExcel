#!/usr/bin/env python3
"""
Test color handling in Excel file
"""

from openpyxl import load_workbook

file_path = "/Users/kevin/excel-unified/rails-app/public/손익실적.xlsx"
wb = load_workbook(file_path)
ws = wb.active

print("=== COLOR ANALYSIS ===\n")

# Check first few cells with styling
for row in range(1, 10):
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col)
        if cell.value:
            print(f"\nCell [{row},{col}]: {cell.value}")
            
            # Font color
            if cell.font and cell.font.color:
                print(f"  Font color type: {type(cell.font.color)}")
                print(f"  Font color RGB: {cell.font.color.rgb}")
                print(f"  Font color theme: {cell.font.color.theme}")
                print(f"  Font color tint: {cell.font.color.tint}")
                print(f"  Font color auto: {cell.font.color.auto}")
                print(f"  Font color indexed: {cell.font.color.indexed}")
            
            # Background color
            if cell.fill and cell.fill.fgColor:
                print(f"  Fill color type: {type(cell.fill.fgColor)}")
                print(f"  Fill color RGB: {cell.fill.fgColor.rgb}")
                print(f"  Fill color theme: {cell.fill.fgColor.theme}")
                print(f"  Fill color tint: {cell.fill.fgColor.tint}")
                print(f"  Fill color auto: {cell.fill.fgColor.auto}")
                print(f"  Fill color indexed: {cell.fill.fgColor.indexed}")
                
            if row >= 5:
                break