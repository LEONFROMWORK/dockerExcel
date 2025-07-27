"""
Style builder - specialized for Excel styling operations
"""

from typing import Dict, List, Any, Optional
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, IconSetRule,
    CellIsRule, FormulaRule
)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import Color

from .configs import CellStyleConfig, RangeStyleConfig, AlternateRowConfig, ConditionalFormatConfig


class StyleBuilder:
    """Specialized builder for Excel styling operations"""
    
    def apply_cell_style(self, cell, config: CellStyleConfig) -> None:
        """Apply multiple styles to a cell"""
        if config.font:
            cell.font = config.font
        if config.fill:
            cell.fill = config.fill
        if config.alignment:
            cell.alignment = config.alignment
        if config.border:
            cell.border = config.border
        if config.number_format:
            cell.number_format = config.number_format
    
    def apply_cell_style_legacy(self, cell, **kwargs) -> None:
        """Legacy method for backward compatibility"""
        config = CellStyleConfig(**kwargs)
        self.apply_cell_style(cell, config)
    
    def apply_range_style(self, worksheet: Worksheet, config: RangeStyleConfig) -> None:
        """Apply style to a range of cells"""
        for row in range(config.start_row, config.end_row + 1):
            for col in range(config.start_col, config.end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                self.apply_cell_style(cell, config.style)
    
    def apply_range_style_legacy(
        self, worksheet: Worksheet,
        start_row: int, start_col: int,
        end_row: int, end_col: int,
        **style_kwargs
    ) -> None:
        """Legacy method for backward compatibility"""
        style = CellStyleConfig(**style_kwargs)
        config = RangeStyleConfig(
            start_row=start_row,
            start_col=start_col,
            end_row=end_row,
            end_col=end_col,
            style=style
        )
        self.apply_range_style(worksheet, config)
    
    def merge_and_style(
        self,
        worksheet: Worksheet,
        start_cell: str,
        end_cell: str,
        value: str,
        **style_kwargs
    ) -> None:
        """Merge cells and apply style"""
        worksheet.merge_cells(f"{start_cell}:{end_cell}")
        merged_cell = worksheet[start_cell]
        merged_cell.value = value
        self.apply_cell_style(merged_cell, **style_kwargs)
    
    def add_conditional_formatting(self, worksheet: Worksheet, config: ConditionalFormatConfig) -> None:
        """Add conditional formatting rules"""
        if config.rule_type == 'color_scale':
            rule = self._create_color_scale_rule(**config.params)
        elif config.rule_type == 'data_bar':
            rule = self._create_data_bar_rule(**config.params)
        elif config.rule_type == 'icon_set':
            rule = self._create_icon_set_rule(**config.params)
        elif config.rule_type == 'cell_is':
            rule = self._create_cell_is_rule(**config.params)
        elif config.rule_type == 'formula':
            rule = self._create_formula_rule(**config.params)
        else:
            return
        
        worksheet.conditional_formatting.add(config.cell_range, rule)
    
    def _create_color_scale_rule(self, **kwargs) -> ColorScaleRule:
        """Create a color scale rule"""
        return ColorScaleRule(
            start_type=kwargs.get('start_type', 'min'),
            start_color=kwargs.get('start_color', 'FF0000'),
            mid_type=kwargs.get('mid_type'),
            mid_color=kwargs.get('mid_color'),
            end_type=kwargs.get('end_type', 'max'),
            end_color=kwargs.get('end_color', '00FF00')
        )
    
    def _create_data_bar_rule(self, **kwargs) -> DataBarRule:
        """Create a data bar rule"""
        return DataBarRule(
            start_type=kwargs.get('start_type', 'min'),
            end_type=kwargs.get('end_type', 'max'),
            color=kwargs.get('color', '638EC6'),
            showValue=kwargs.get('show_value', True),
            minLength=kwargs.get('min_length'),
            maxLength=kwargs.get('max_length')
        )
    
    def _create_icon_set_rule(self, **kwargs) -> IconSetRule:
        """Create an icon set rule"""
        return IconSetRule(
            icon_style=kwargs.get('icon_style', '3Arrows'),
            type=kwargs.get('type', 'percent'),
            values=kwargs.get('values', [0, 33, 67]),
            showValue=kwargs.get('show_value', True)
        )
    
    def _create_cell_is_rule(self, **kwargs) -> CellIsRule:
        """Create a cell value rule"""
        dxf = DifferentialStyle(
            font=kwargs.get('font'),
            fill=kwargs.get('fill'),
            border=kwargs.get('border')
        )
        
        return CellIsRule(
            operator=kwargs.get('operator', 'equal'),
            formula=kwargs.get('formula', ['0']),
            dxf=dxf
        )
    
    def _create_formula_rule(self, **kwargs) -> FormulaRule:
        """Create a formula-based rule"""
        dxf = DifferentialStyle(
            font=kwargs.get('font'),
            fill=kwargs.get('fill'),
            border=kwargs.get('border')
        )
        
        return FormulaRule(
            formula=kwargs.get('formula', ['True']),
            dxf=dxf
        )
    
    def apply_alternate_row_coloring(self, worksheet: Worksheet, config: AlternateRowConfig) -> None:
        """Apply alternate row coloring"""
        for row in range(config.start_row, config.end_row + 1):
            fill_color = config.color1 if (row - config.start_row) % 2 == 0 else config.color2
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            for col in range(config.start_col, config.end_col + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.fill = fill